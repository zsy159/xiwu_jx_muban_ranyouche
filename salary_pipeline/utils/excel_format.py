"""Shared openpyxl number formats for pipeline Excel exports."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable, Protocol

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from salary_pipeline.calculators.non_frontline.classification import (
    expand_highlight_columns,
    highlight_column_for_row,
)
from salary_pipeline.data_ingestion.data_loader import normalize_header, normalize_name

TWO_DECIMAL_FORMAT = "#,##0.00"

# Light amber — visible against white, keeps black text readable.
PARITY_MISMATCH_FILL = PatternFill(
    start_color="FFFFEB9C",
    end_color="FFFFEB9C",
    fill_type="solid",
)
PARITY_MISMATCH_FILL_RGB = "FFFFEB9C"
PARITY_MISMATCH_FILL_COMMENT = "金标准与系统数值不一致"

# Light blue — formula cells with manual inputs / parity-deferred (wa_parity_deferred).
MANUAL_DEFERRED_FILL = PatternFill(
    start_color="FFBDD7EE",
    end_color="FFBDD7EE",
    fill_type="solid",
)
MANUAL_DEFERRED_FILL_RGB = "FFBDD7EE"
MANUAL_DEFERRED_FILL_COMMENT = "公式含手工录入，对账暂缓"

# Light gray — golden 提成汇总 cells with no formula (direct static value).
GOLDEN_STATIC_FILL = PatternFill(
    start_color="FFD9D9D9",
    end_color="FFD9D9D9",
    fill_type="solid",
)
GOLDEN_STATIC_FILL_RGB = "FFD9D9D9"

STATIC_FILL_COMMENT = "金标准直接填数，无公式"

_HIGHLIGHT_FILL_RGBS = frozenset(
    {MANUAL_DEFERRED_FILL_RGB, GOLDEN_STATIC_FILL_RGB}
)

# Light orange — formula anomaly / 公式形态异常（区别于琥珀 mismatch、蓝色 deferred）
FORMULA_ANOMALY_FILL = PatternFill(
    start_color="FFFCE4D6",
    end_color="FFFCE4D6",
    fill_type="solid",
)
FORMULA_ANOMALY_FILL_RGB = "FFFCE4D6"

_COMMISSION_SUMMARY_LEGEND_ITEMS: tuple[tuple[PatternFill, str], ...] = (
    (GOLDEN_STATIC_FILL, "浅灰 #D9D9D9：金标准直接填数"),
    (MANUAL_DEFERRED_FILL, "浅蓝 #BDD7EE：公式含手工"),
    (PARITY_MISMATCH_FILL, "琥珀 #FFEB9C：数值不一致"),
    (FORMULA_ANOMALY_FILL, "浅橙 #FCE4D6：公式形态异常"),
)


class ParityMismatchCell(Protocol):
    join_values: tuple[tuple[str, Any], ...]
    column: str
    golden_value: float | None
    computed_value: float | None
    root_cause: str | None


class CommissionSummaryAnnotation(Protocol):
    name: str
    column: str

    def comment_text(self) -> str: ...

# Count / ID columns — leave default display (not forced to two decimals).
INTEGER_DISPLAY_HEADERS: frozenset[str] = frozenset(
    {
        "序号",
        "人数",
        "考核量",
        "实际销量",
        "台数",
        "订单号",
    }
)


def apply_two_decimal_format(
    worksheet: Worksheet,
    columns: Iterable[str],
    *,
    header_row: int,
    integer_columns: frozenset[str] | None = None,
) -> None:
    """Apply ``#,##0.00`` to numeric data cells in listed columns."""
    skip = integer_columns if integer_columns is not None else INTEGER_DISPLAY_HEADERS
    col_names = list(columns)
    first_data_row = header_row + 1

    for col_idx, name in enumerate(col_names, start=1):
        if name in skip:
            continue
        for row_idx in range(first_data_row, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            value = cell.value
            if isinstance(value, bool):
                continue
            if isinstance(value, (int, float)):
                cell.number_format = TWO_DECIMAL_FORMAT


def format_writer_sheet(
    writer,
    sheet_name: str,
    columns: Iterable[str],
    *,
    header_row: int,
    integer_columns: frozenset[str] | None = None,
) -> None:
    """Format numeric cells on a sheet written via ``pd.ExcelWriter``."""
    apply_two_decimal_format(
        writer.sheets[sheet_name],
        columns,
        header_row=header_row,
        integer_columns=integer_columns,
    )


def _header_column_map(
    worksheet: Worksheet, header_row: int
) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col_idx in range(1, worksheet.max_column + 1):
        header = normalize_header(worksheet.cell(row=header_row, column=col_idx).value)
        if header:
            mapping[header] = col_idx
    return mapping


def _join_key_tuple(
    worksheet: Worksheet,
    row_idx: int,
    join_keys: list[str],
    col_map: dict[str, int],
) -> tuple[tuple[str, Any], ...]:
    values: list[tuple[str, Any]] = []
    for key in join_keys:
        col_idx = col_map.get(key)
        if col_idx is None:
            values.append((key, None))
            continue
        raw = worksheet.cell(row=row_idx, column=col_idx).value
        values.append((key, normalize_name(raw)))
    return tuple(values)


def _build_excel_row_index(
    worksheet: Worksheet,
    join_keys: list[str],
    *,
    header_row: int,
    data_start_row: int,
) -> dict[tuple[tuple[str, Any], ...], int]:
    col_map = _header_column_map(worksheet, header_row)
    index: dict[tuple[tuple[str, Any], ...], int] = {}
    for row_idx in range(data_start_row, worksheet.max_row + 1):
        key = _join_key_tuple(worksheet, row_idx, join_keys, col_map)
        if any(value is None for _, value in key):
            continue
        index[key] = row_idx
    return index


def format_mismatch_comment_text(
    *,
    golden_value: float | None = None,
    computed_value: float | None = None,
    root_cause: str | None = None,
    headline: str = PARITY_MISMATCH_FILL_COMMENT,
) -> str:
    """Build amber mismatch cell comment: golden / computed / root cause."""
    lines = [headline]
    golden = golden_value
    computed = computed_value
    if golden is not None and computed is not None:
        diff = computed - golden
        lines.append(f"金标准: {golden:g} | 系统: {computed:g} | 差: {diff:+g}")
    if root_cause:
        lines.append(f"原因: {root_cause.strip()}")
    return "\n".join(lines)


def _mismatch_comment(mismatch: ParityMismatchCell) -> str:
    return format_mismatch_comment_text(
        golden_value=mismatch.golden_value,
        computed_value=mismatch.computed_value,
        root_cause=getattr(mismatch, "root_cause", None),
    )


def _legend_row_present(worksheet: Worksheet, row: int) -> bool:
    for col in range(1, worksheet.max_column + 1):
        value = worksheet.cell(row=row, column=col).value
        if isinstance(value, str) and "数值不一致" in value:
            return True
    return False


def add_commission_summary_color_legend(
    workbook_path: Path,
    sheet_name: str,
    *,
    insert_at_row: int = 2,
) -> bool:
    """Insert a row of color swatches and Chinese labels for reconcile highlighting.

    Returns True when a new legend row was inserted, False when one already exists.
    """
    wb = load_workbook(workbook_path)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"sheet {sheet_name!r} not found in {workbook_path}")
    ws = wb[sheet_name]
    if _legend_row_present(ws, insert_at_row):
        wb.close()
        return False
    ws.insert_rows(insert_at_row)
    col = 1
    for fill, label in _COMMISSION_SUMMARY_LEGEND_ITEMS:
        swatch = ws.cell(row=insert_at_row, column=col, value="")
        swatch.fill = fill
        ws.cell(row=insert_at_row, column=col + 1, value=label)
        col += 2
    wb.save(workbook_path)
    return True


def highlight_commission_summary_mismatches(
    workbook_path: Path,
    sheet_name: str,
    mismatches: Iterable[ParityMismatchCell],
    join_keys: list[str],
    compare_columns: Iterable[str],
    *,
    header_row: int = 2,
    data_start_row: int = 3,
) -> int:
    """Highlight parity mismatch cells in an exported 提成汇总 workbook."""
    mismatch_list = list(mismatches)
    compare_col_set = expand_highlight_columns(compare_columns)

    wb = load_workbook(workbook_path)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"sheet {sheet_name!r} not found in {workbook_path}")
    ws = wb[sheet_name]
    col_map = _header_column_map(ws, header_row)
    row_index = _build_excel_row_index(
        ws, join_keys, header_row=header_row, data_start_row=data_start_row
    )

    for row_idx in range(data_start_row, ws.max_row + 1):
        for col_name in compare_col_set:
            col_idx = col_map.get(col_name)
            if col_idx is None:
                continue
            cell = ws.cell(row=row_idx, column=col_idx)
            if (
                cell.fill
                and cell.fill.fill_type == "solid"
                and getattr(cell.fill.start_color, "rgb", None) == PARITY_MISMATCH_FILL_RGB
            ):
                cell.fill = PatternFill()

    highlighted = 0
    for mismatch in mismatch_list:
        row_idx = row_index.get(mismatch.join_values)
        if row_idx is None:
            continue
        col_idx = col_map.get(mismatch.column)
        if col_idx is None:
            continue
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = PARITY_MISMATCH_FILL
        cell.comment = Comment(_mismatch_comment(mismatch), "对账")
        highlighted += 1

    wb.save(workbook_path)
    return highlighted


def highlight_commission_summary_deferred_cells(
    workbook_path: Path,
    sheet_name: str,
    deferred_cells: dict[str, frozenset[str]],
    *,
    static_cells: dict[tuple[str, str], frozenset[str]] | None = None,
    deferred_reasons: dict[str, dict[str, str]] | None = None,
    role_title: str = "销售顾问",
    header_row: int = 2,
    data_start_row: int = 3,
    static_comment: str = STATIC_FILL_COMMENT,
    deferred_comment: str = MANUAL_DEFERRED_FILL_COMMENT,
) -> int:
    """Highlight parity-deferred and golden static (直接填数) cells in 提成汇总.

    Rows are matched by 姓名 + 职务 (not 店别). ``deferred_cells`` is keyed by 姓名
    and filtered to ``role_title``; ``static_cells`` is keyed by (姓名, 职务).

    Fill colors: gray = 金标准直填（无公式）；blue = 公式含手工（wa_parity_deferred、二网 AH、topology 尾项常数）。
    Deferred wins when a cell appears in both maps.
    """
    static_cells = static_cells or {}
    if not deferred_cells and not static_cells:
        return 0

    wb = load_workbook(workbook_path)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"sheet {sheet_name!r} not found in {workbook_path}")
    ws = wb[sheet_name]
    col_map = _header_column_map(ws, header_row)

    name_col = col_map.get("姓名")
    role_col = col_map.get("职务")
    shop_col = col_map.get("店别")
    if name_col is None or role_col is None:
        wb.save(workbook_path)
        return 0

    deferred_columns = {col for cols in deferred_cells.values() for col in cols}
    static_columns = {col for cols in static_cells.values() for col in cols}
    highlight_columns = expand_highlight_columns(deferred_columns | static_columns)

    for row_idx in range(data_start_row, ws.max_row + 1):
        for col_name in highlight_columns:
            col_idx = col_map.get(col_name)
            if col_idx is None:
                continue
            cell = ws.cell(row=row_idx, column=col_idx)
            if (
                cell.fill
                and cell.fill.fill_type == "solid"
                and getattr(cell.fill.start_color, "rgb", None) in _HIGHLIGHT_FILL_RGBS
            ):
                cell.fill = PatternFill()

    highlighted = 0
    for row_idx in range(data_start_row, ws.max_row + 1):
        name = normalize_name(ws.cell(row=row_idx, column=name_col).value)
        role = normalize_name(ws.cell(row=row_idx, column=role_col).value) or ""
        shop = (
            normalize_name(ws.cell(row=row_idx, column=shop_col).value)
            if shop_col is not None
            else None
        )
        if name is None:
            continue

        deferred_cols: set[str] = set()
        if role == role_title:
            deferred_cols.update(deferred_cells.get(name, frozenset()))
        static_cols = set(static_cells.get((name, role), frozenset()))
        static_only = static_cols - deferred_cols

        for col_name in static_only:
            display_col = highlight_column_for_row(shop, role, col_name)
            col_idx = col_map.get(display_col)
            if col_idx is None:
                continue
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = GOLDEN_STATIC_FILL
            cell.comment = Comment(static_comment, "对账")
            highlighted += 1

        for col_name in deferred_cols:
            display_col = highlight_column_for_row(shop, role, col_name)
            col_idx = col_map.get(display_col)
            if col_idx is None:
                continue
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = MANUAL_DEFERRED_FILL
            reason = (deferred_reasons or {}).get(name, {}).get(col_name)
            comment_text = (
                f"{deferred_comment}\n原因: {reason.strip()}"
                if reason
                else deferred_comment
            )
            cell.comment = Comment(comment_text, "对账")
            highlighted += 1

    wb.save(workbook_path)
    return highlighted


def add_commission_summary_annotations(
    workbook_path: Path,
    sheet_name: str,
    annotations: Iterable[CommissionSummaryAnnotation],
    *,
    role_title: str = "销售顾问",
    header_row: int = 2,
    data_start_row: int = 3,
) -> int:
    """Highlight formula-anomaly cells and attach Excel comments (批注) in 提成汇总."""
    ann_list = list(annotations)

    wb = load_workbook(workbook_path)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"sheet {sheet_name!r} not found in {workbook_path}")
    ws = wb[sheet_name]
    col_map = _header_column_map(ws, header_row)

    name_col = col_map.get("姓名")
    role_col = col_map.get("职务")
    if name_col is None or role_col is None:
        wb.save(workbook_path)
        return 0

    # Clear stale orange fills on all advisor rows (not only columns in this batch).
    for row_idx in range(data_start_row, ws.max_row + 1):
        role = normalize_name(ws.cell(row=row_idx, column=role_col).value)
        if role != role_title:
            continue
        for col_name in col_map:
            col_idx = col_map[col_name]
            cell = ws.cell(row=row_idx, column=col_idx)
            if (
                cell.fill
                and cell.fill.fill_type == "solid"
                and getattr(cell.fill.start_color, "rgb", None) == FORMULA_ANOMALY_FILL_RGB
            ):
                cell.fill = PatternFill()
                cell.comment = None

    if not ann_list:
        wb.save(workbook_path)
        return 0

    by_name: dict[str, list[CommissionSummaryAnnotation]] = {}
    for ann in ann_list:
        by_name.setdefault(ann.name, []).append(ann)

    applied = 0
    for row_idx in range(data_start_row, ws.max_row + 1):
        name = normalize_name(ws.cell(row=row_idx, column=name_col).value)
        role = normalize_name(ws.cell(row=row_idx, column=role_col).value)
        if name is None or role != role_title:
            continue
        for ann in by_name.get(name, []):
            col_idx = col_map.get(ann.column)
            if col_idx is None:
                continue
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = FORMULA_ANOMALY_FILL
            cell.comment = Comment(ann.comment_text(), "对账")
            applied += 1

    wb.save(workbook_path)
    return applied
