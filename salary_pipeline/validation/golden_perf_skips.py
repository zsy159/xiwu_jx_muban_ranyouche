"""Parity skips for golden 绩效整理表 quirks (e.g. D 列含二网 → AH 公式留空)."""

from __future__ import annotations

from functools import lru_cache
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

# Hub columns that SUMIFS 绩效整理表 AH by 销售顾问
HUB_COLUMNS_FROM_AH = frozenset({"权限结余绩效"})

ERWANG_CHANNEL_MARKERS = ("二网",)


def is_golden_erwang_channel(channel: object) -> bool:
    """True when golden 绩效整理表 D 列触发 AH 公式留空（含二网）。"""
    text = str(channel).strip()
    if not text:
        return False
    return any(marker in text for marker in ERWANG_CHANNEL_MARKERS)


def _is_blank(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and pd.isna(value):
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _to_float(value: object) -> float | None:
    if _is_blank(value):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def load_computed_ah_by_vin(
    computed_perf_path: Path,
    *,
    sheet_name: str = "绩效整理表",
    header_row: int = 2,
) -> dict[str, float]:
    """Map VIN → computed 整车超额 from exported performance sheet."""
    if not computed_perf_path.exists():
        return {}
    df = pd.read_excel(
        computed_perf_path,
        sheet_name=sheet_name,
        header=header_row - 1,
        engine="openpyxl",
    )
    vin_col = "VIN码" if "VIN码" in df.columns else "O"
    ah_col = "整车超额" if "整车超额" in df.columns else "AH"
    if vin_col not in df.columns or ah_col not in df.columns:
        return {}
    out: dict[str, float] = {}
    for vin, ah in zip(df[vin_col], df[ah_col], strict=False):
        if _is_blank(vin):
            continue
        key = str(vin).strip()
        num = _to_float(ah)
        if num is not None:
            out[key] = num
    return out


def load_golden_erwang_blank_ah_adjustments(
    golden_workbook: Path,
    computed_ah_by_vin: dict[str, float],
    *,
    perf_sheet: str = "绩效整理表",
    data_start_row: int = 3,
) -> dict[str, float]:
    """Per-advisor sum of computed AH where golden D is 二网 and golden AH is blank.

    When golden AH formula sees D 列二网 it returns blank, but order_context may
    derive a different D and the system correctly computes AH. Hub 权限结余绩效
    parity should not flag the resulting aggregate difference.
    """
    if not golden_workbook.exists() or not computed_ah_by_vin:
        return {}

    wb = load_workbook(golden_workbook, data_only=True, read_only=True)
    try:
        if perf_sheet not in wb.sheetnames:
            return {}
        ws = wb[perf_sheet]
        adjustments: dict[str, float] = {}
        for row in range(data_start_row, ws.max_row + 1):
            vin = ws.cell(row=row, column=15).value  # O
            advisor = ws.cell(row=row, column=16).value  # P
            channel = ws.cell(row=row, column=4).value  # D
            golden_ah = ws.cell(row=row, column=34).value  # AH
            if _is_blank(vin) or _is_blank(advisor):
                continue
            if not is_golden_erwang_channel(channel):
                continue
            if not _is_blank(golden_ah):
                continue
            vin_key = str(vin).strip()
            computed_ah = computed_ah_by_vin.get(vin_key)
            if computed_ah is None:
                continue
            name = str(advisor).strip()
            adjustments[name] = adjustments.get(name, 0.0) + computed_ah
        return adjustments
    finally:
        wb.close()


@lru_cache(maxsize=8)
def _cached_adjustments(
    golden_workbook: str,
    computed_perf_path: str,
) -> dict[str, float]:
    computed = load_computed_ah_by_vin(Path(computed_perf_path))
    return load_golden_erwang_blank_ah_adjustments(
        Path(golden_workbook),
        computed,
    )


def erwang_blank_ah_adjustments_for_paths(
    golden_workbook: Path | None,
    computed_perf_path: Path | None,
) -> dict[str, float]:
    if golden_workbook is None or computed_perf_path is None:
        return {}
    if not golden_workbook.exists() or not computed_perf_path.exists():
        return {}
    return _cached_adjustments(str(golden_workbook.resolve()), str(computed_perf_path.resolve()))


def hub_parity_skip_erwang_blank_ah(
    name: str,
    column: str,
    golden_val: object,
    computed_val: object,
    adjustments: dict[str, float],
    *,
    tolerance: float = 1e-6,
) -> bool:
    """Return True when Hub diff is fully explained by golden 二网-blank AH rows."""
    if column not in HUB_COLUMNS_FROM_AH:
        return False
    adj = adjustments.get(str(name).strip())
    if adj is None or adj == 0:
        return False
    g_num = _to_float(golden_val)
    c_num = _to_float(computed_val)
    if g_num is None or c_num is None:
        return False
    return abs((c_num - g_num) - adj) <= tolerance


def erwang_blank_ah_deferred_cells(
    adjustments: dict[str, float],
) -> dict[str, frozenset[str]]:
    """Map advisor → Hub columns to highlight blue (golden D 二网致 AH 留空)."""
    return {
        name: frozenset(HUB_COLUMNS_FROM_AH)
        for name, adj in adjustments.items()
        if adj != 0
    }
