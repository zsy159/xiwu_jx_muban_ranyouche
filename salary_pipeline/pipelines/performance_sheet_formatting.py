"""Post-export 绩效整理表 formatting: golden header reconcile highlights."""

from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.utils import column_index_from_string

from salary_pipeline.data_ingestion.performance_sheet_golden import DATA_START_ROW, PERF_SHEET
from salary_pipeline.data_ingestion.performance_sheet_golden_scan import (
    PerfManualCell,
    load_golden_column_headers,
    scan_golden_manual_cells,
)
from salary_pipeline.pipelines.performance_sheet_builder import IMPLEMENTED_COLUMNS
from salary_pipeline.pipelines.performance_sheet_column_sources import (
    canonical_perf_letter,
    is_implemented_perf_column,
    unimplemented_header_labels,
)
from salary_pipeline.pipelines.performance_sheet_export import (
    HEADER_ROW as EXPORT_HEADER_ROW,
    PERF_COLUMN_LABELS,
)
from salary_pipeline.pipelines.performance_sheet_paths import load_performance_sheet_frame
from salary_pipeline.utils.excel_format import (
    GOLDEN_STATIC_FILL,
    MANUAL_DEFERRED_FILL,
    PARITY_MISMATCH_FILL,
    PARITY_MISMATCH_FILL_RGB,
    STATIC_FILL_COMMENT,
    _COMMISSION_SUMMARY_LEGEND_ITEMS,
    _legend_row_present,
    format_mismatch_comment_text,
)

logger = logging.getLogger(__name__)

PERF_JOIN_HEADER = "VIN码"
PERF_MANUAL_DEFERRED_COMMENT = "金标准公式含手工录入"
DATETIME_COMPARE_COLUMNS = frozenset({"M"})


@dataclass
class PerfHighlightStats:
    mismatches: int = 0
    manual_marked: int = 0
    unimplemented_marked: int = 0


@dataclass
class PerfMismatch:
    vin: str
    column: str
    golden_value: float | None
    computed_value: float | None
    root_cause: str | None = None


def implemented_value_columns() -> frozenset[str]:
    skip = frozenset({"O", "P", "K", "G"})
    return frozenset(c for c in IMPLEMENTED_COLUMNS if c not in skip)


def resolve_perf_golden_path(month_config: dict[str, Any]) -> Path | None:
    from salary_pipeline.data_ingestion.data_loader import (
        resolve_parity_golden_workbook,
        workbook_has_sheet,
    )
    from salary_pipeline.paths import resolve_project_path

    path = resolve_parity_golden_workbook(month_config, sheet_name=PERF_SHEET)
    if path is not None:
        return path
    sales_rel = (month_config.get("workbooks") or {}).get("sales")
    if sales_rel:
        sales_path = resolve_project_path(sales_rel)
        if sales_path.is_file() and workbook_has_sheet(sales_path, PERF_SHEET):
            return sales_path
    return None


def collect_performance_sheet_mismatches(
    computed_frame: pd.DataFrame,
    golden_path: Path,
    *,
    columns: frozenset[str] | None = None,
    tolerance: float = 1e-4,
) -> list[PerfMismatch]:
    """Compare system frame vs golden 绩效整理表 by VIN (read-only golden)."""
    if computed_frame.empty or "O" not in computed_frame.columns:
        return []

    compare_cols = columns or implemented_value_columns()
    header_spec = load_golden_column_headers(golden_path)
    label_by_letter = {letter: label for letter, label in header_spec}
    label_by_letter.update(PERF_COLUMN_LABELS)

    from salary_pipeline.data_ingestion.data_loader import WorkbookLoader

    loader = WorkbookLoader(golden_path)
    mismatches: list[PerfMismatch] = []

    for letter in sorted(compare_cols, key=column_index_from_string):
        if letter not in computed_frame.columns:
            continue
        golden_raw = loader.read_sheet_columns(
            PERF_SHEET, {"O": "O", letter: letter}, label=f"parity {letter}"
        )
        golden_rows = golden_raw.iloc[DATA_START_ROW - 1 :].reset_index(drop=True)
        golden_rows["O"] = golden_rows["O"].astype(str).str.strip()
        golden_rows = golden_rows[~golden_rows["O"].isin(("", "nan", "None", "<NA>", "VIN码"))]

        built = computed_frame[["O", letter]].copy()
        built["O"] = built["O"].astype(str).str.strip()
        built = built[~built["O"].isin(("", "nan", "None", "<NA>", "VIN码"))]
        if built.empty or golden_rows.empty:
            continue

        if letter in DATETIME_COMPARE_COLUMNS:
            golden_rows[letter] = pd.to_datetime(golden_rows[letter], errors="coerce")
            built[letter] = pd.to_datetime(built[letter], errors="coerce")
            merged = built.merge(
                golden_rows[["O", letter]],
                on="O",
                suffixes=("_sys", "_gold"),
                how="inner",
            )
            for idx, row in merged.iterrows():
                sys_v = row[f"{letter}_sys"]
                gold_v = row[f"{letter}_gold"]
                if pd.isna(sys_v) and pd.isna(gold_v):
                    continue
                if pd.isna(sys_v) or pd.isna(gold_v) or sys_v != gold_v:
                    mismatches.append(
                        PerfMismatch(
                            vin=str(row["O"]),
                            column=label_by_letter.get(letter, letter),
                            golden_value=gold_v.timestamp() if pd.notna(gold_v) else None,
                            computed_value=sys_v.timestamp() if pd.notna(sys_v) else None,
                        )
                    )
            continue

        golden_rows[letter] = pd.to_numeric(golden_rows[letter], errors="coerce")
        built[letter] = pd.to_numeric(built[letter], errors="coerce")
        merged = built.merge(
            golden_rows[["O", letter]],
            on="O",
            suffixes=("_sys", "_gold"),
            how="inner",
        )
        diff = (
            merged[f"{letter}_sys"].fillna(0) - merged[f"{letter}_gold"].fillna(0)
        ).abs()
        for idx in merged.index[diff > tolerance]:
            row = merged.loc[idx]
            vin = str(row["O"])
            mismatches.append(
                PerfMismatch(
                    vin=vin,
                    column=label_by_letter.get(letter, letter),
                    golden_value=float(row[f"{letter}_gold"])
                    if pd.notna(row[f"{letter}_gold"])
                    else None,
                    computed_value=float(row[f"{letter}_sys"])
                    if pd.notna(row[f"{letter}_sys"])
                    else None,
                )
            )

    return mismatches


def _header_column_map(ws, header_row: int) -> dict[str, int]:
    from salary_pipeline.data_ingestion.data_loader import normalize_header

    mapping: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        header = normalize_header(ws.cell(row=header_row, column=col_idx).value)
        if header:
            mapping[header] = col_idx
    return mapping


def _build_vin_row_index(ws, *, header_row: int, data_start_row: int) -> dict[str, int]:
    col_map = _header_column_map(ws, header_row)
    vin_col = col_map.get(PERF_JOIN_HEADER)
    if vin_col is None:
        return {}
    index: dict[str, int] = {}
    for row_idx in range(data_start_row, ws.max_row + 1):
        raw = ws.cell(row=row_idx, column=vin_col).value
        if raw is None:
            continue
        vin = str(raw).strip()
        if vin and vin not in ("VIN码", "nan"):
            index[vin] = row_idx
    return index


def add_performance_sheet_color_legend(
    workbook_path: Path,
    sheet_name: str = PERF_SHEET,
    *,
    insert_at_row: int = 2,
) -> bool:
    wb = load_workbook(workbook_path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return False
    ws = wb[sheet_name]
    if _legend_row_present(ws, insert_at_row):
        wb.close()
        return False
    ws.insert_rows(insert_at_row)
    col = 1
    for fill, label in _COMMISSION_SUMMARY_LEGEND_ITEMS:
        swatch = ws.cell(row=insert_at_row, column=col, value="")
        swatch.fill = fill
        ws.cell(row=insert_at_row, column=col + 1, value=label)
        col += 2
    wb.save(workbook_path)
    return True


def highlight_performance_sheet_mismatches(
    workbook_path: Path,
    mismatches: list[PerfMismatch],
    *,
    sheet_name: str = PERF_SHEET,
    header_row: int = 3,
    data_start_row: int = 4,
) -> int:
    if not mismatches:
        return 0
    wb = load_workbook(workbook_path)
    ws = wb[sheet_name]
    col_map = _header_column_map(ws, header_row)
    row_index = _build_vin_row_index(ws, header_row=header_row, data_start_row=data_start_row)
    count = 0
    for mm in mismatches:
        row_idx = row_index.get(mm.vin)
        col_idx = col_map.get(mm.column)
        if row_idx is None or col_idx is None:
            continue
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = PARITY_MISMATCH_FILL
        cell.comment = Comment(
            format_mismatch_comment_text(
                golden_value=mm.golden_value,
                computed_value=mm.computed_value,
                root_cause=mm.root_cause,
            ),
            "对账",
        )
        count += 1
    wb.save(workbook_path)
    return count


def highlight_performance_sheet_manual_cells(
    workbook_path: Path,
    manual_cells: list[PerfManualCell],
    *,
    sheet_name: str = PERF_SHEET,
    header_row: int = 3,
    data_start_row: int = 4,
) -> int:
    if not manual_cells:
        return 0
    wb = load_workbook(workbook_path)
    ws = wb[sheet_name]
    col_map = _header_column_map(ws, header_row)
    row_index = _build_vin_row_index(ws, header_row=header_row, data_start_row=data_start_row)
    count = 0
    for mc in manual_cells:
        # Static golden fills: skip implemented cols (system computes them);
        # unimplemented cols are grayed via highlight_unimplemented_columns.
        if mc.pattern in ("直接填数", "常数公式", "纯算术公式"):
            continue
        fill = MANUAL_DEFERRED_FILL
        headline = PERF_MANUAL_DEFERRED_COMMENT
        row_idx = row_index.get(mc.vin) if mc.vin else None
        if row_idx is None and mc.vin is None and mc.advisor:
            # advisor-only tail row (e.g. 熊俊杰)
            p_col = col_map.get("销售顾问")
            if p_col:
                for r in range(data_start_row, ws.max_row + 1):
                    if str(ws.cell(row=r, column=p_col).value or "").strip() == mc.advisor:
                        row_idx = r
                        break
        col_idx = col_map.get(mc.header)
        if row_idx is None or col_idx is None:
            continue
        cell = ws.cell(row=row_idx, column=col_idx)
        if (
            cell.fill
            and cell.fill.fill_type == "solid"
            and getattr(cell.fill.start_color, "rgb", None) == PARITY_MISMATCH_FILL_RGB
        ):
            continue
        cell.fill = fill
        detail = f"金标准手填: {mc.pattern}"
        if mc.detail:
            detail = f"{detail} ({mc.detail})"
        cell.comment = Comment(detail, "对账")
        count += 1
    wb.save(workbook_path)
    return count


def highlight_unimplemented_columns(
    workbook_path: Path,
    unimplemented_headers: frozenset[str],
    *,
    sheet_name: str = PERF_SHEET,
    header_row: int = 3,
    data_start_row: int = 4,
) -> int:
    if not unimplemented_headers:
        return 0
    wb = load_workbook(workbook_path)
    ws = wb[sheet_name]
    col_map = _header_column_map(ws, header_row)
    count = 0
    for header in unimplemented_headers:
        col_idx = col_map.get(header)
        if col_idx is None:
            continue
        for row_idx in range(data_start_row, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None and cell.value != "":
                continue
            cell.fill = GOLDEN_STATIC_FILL
            if cell.comment is None:
                cell.comment = Comment(STATIC_FILL_COMMENT, "对账")
            count += 1
    wb.save(workbook_path)
    return count


def apply_performance_sheet_highlighting(
    month_config: dict[str, Any],
    computed_path: Path,
    *,
    golden_path: Path | None = None,
    computed_frame: pd.DataFrame | None = None,
) -> PerfHighlightStats:
    """Apply legend, mismatch amber, manual blue/gray, and unimplemented column gray."""
    golden = golden_path or resolve_perf_golden_path(month_config)
    if golden is None or not golden.exists():
        logger.info("Skipping performance sheet highlighting: no golden workbook")
        return PerfHighlightStats()

    parity_cfg = month_config.get("parity", {})
    tolerance = float(parity_cfg.get("numeric_tolerance", 1e-4))
    header_spec = load_golden_column_headers(golden)
    if not header_spec:
        return PerfHighlightStats()

    label_by_letter = {letter: label for letter, label in header_spec}
    unimplemented_headers = unimplemented_header_labels(header_spec)

    frame = computed_frame
    if frame is None:
        frame = load_performance_sheet_frame(computed_path)
    mismatches = collect_performance_sheet_mismatches(
        frame, golden, tolerance=tolerance
    )
    manual_cells = [
        mc
        for mc in scan_golden_manual_cells(golden)
        if is_implemented_perf_column(mc.letter, mc.header)
        and canonical_perf_letter(mc.letter, mc.header) not in ("O", "P", "K", "G")
        and mc.pattern not in ("直接填数", "常数公式", "纯算术公式")
    ]
    manual_by_key = {
        (mc.vin, mc.header): mc for mc in manual_cells if mc.vin
    }
    for mm in mismatches:
        mc = manual_by_key.get((mm.vin, mm.column))
        if mc is None:
            continue
        if mc.pattern == "公式+尾项":
            mm.root_cause = (
                f"金标准{mc.detail}（系统未应用手工尾项，仅标差异）"
            )
        elif mc.pattern in ("直接填数", "常数公式", "纯算术公式"):
            mm.root_cause = f"金标准{mc.pattern}，系统按公式重算"

    legend_inserted = add_performance_sheet_color_legend(computed_path)
    header_row = EXPORT_HEADER_ROW + 1 if legend_inserted else EXPORT_HEADER_ROW
    data_start_row = header_row + 1

    mismatch_count = highlight_performance_sheet_mismatches(
        computed_path,
        mismatches,
        header_row=header_row,
        data_start_row=data_start_row,
    )
    manual_count = highlight_performance_sheet_manual_cells(
        computed_path,
        manual_cells,
        header_row=header_row,
        data_start_row=data_start_row,
    )
    unimpl_count = highlight_unimplemented_columns(
        computed_path,
        unimplemented_headers,
        header_row=header_row,
        data_start_row=data_start_row,
    )

    stats = PerfHighlightStats(
        mismatches=mismatch_count,
        manual_marked=manual_count,
        unimplemented_marked=unimpl_count,
    )
    logger.info(
        "Performance sheet highlighting -> %s (mismatches=%s manual=%s unimplemented_cells=%s)",
        computed_path,
        stats.mismatches,
        stats.manual_marked,
        stats.unimplemented_marked,
    )
    return stats
