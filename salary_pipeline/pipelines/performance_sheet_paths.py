"""Resolve 绩效整理表 paths — prefer 财务确认版 over 系统生成."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from salary_pipeline.data_ingestion.data_loader import normalize_header
from salary_pipeline.paths import resolve_project_path
from salary_pipeline.pipelines.performance_sheet_export import (
    HEADER_ROW as EXPORT_HEADER_ROW,
    PERF_COLUMN_LABELS,
    SOURCE_ANNOTATION_ROW,
)
from salary_pipeline.utils.excel_format import _legend_row_present

SYSTEM_PERF_FILENAME = "绩效整理表-系统生成.xlsx"
CONFIRMED_PERF_FILENAME = "绩效整理表-财务确认版.xlsx"
DEFAULT_PERF_SHEET = "绩效整理表"

_LABEL_TO_COLUMN: dict[str, str] = {label: col for col, label in PERF_COLUMN_LABELS.items()}
_PERF_HEADER_MARKERS = frozenset({"VIN码", "订单号", "销售顾问"})


def _perf_sheet_header_present(worksheet: Any, row: int) -> bool:
    for col in range(1, worksheet.max_column + 1):
        header = normalize_header(worksheet.cell(row=row, column=col).value)
        if header in _PERF_HEADER_MARKERS:
            return True
    return False


def resolve_performance_sheet_read_rows(
    workbook_path: Path,
    sheet_name: str = DEFAULT_PERF_SHEET,
    *,
    header_row: int = EXPORT_HEADER_ROW,
    legend_insert_row: int = SOURCE_ANNOTATION_ROW,
) -> int:
    """Return Excel header row for exported 绩效整理表 (accounts for reconcile legend)."""
    path = Path(workbook_path)
    if not path.exists():
        return header_row
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return header_row
        ws = wb[sheet_name]
        if _perf_sheet_header_present(ws, header_row):
            return header_row
        if _legend_row_present(ws, legend_insert_row) and _perf_sheet_header_present(
            ws, header_row + 1
        ):
            return header_row + 1
        return header_row
    finally:
        wb.close()


def performance_sheet_output_dir(month_config: dict[str, Any]) -> Path:
    """Directory holding system / confirmed performance sheet exports."""
    outputs = month_config.get("outputs", {})
    perf_raw = outputs.get("performance_sheet_file")
    if perf_raw:
        return resolve_project_path(perf_raw).parent
    commission_raw = outputs.get("commission_summary_file")
    if commission_raw:
        return resolve_project_path(commission_raw).parent
    month = month_config.get("month", "")
    if month:
        from salary_pipeline.paths import output_month_dir

        return output_month_dir(month)
    return resolve_project_path("output")


def resolve_system_performance_sheet_path(month_config: dict[str, Any]) -> Path:
    """Path to system-computed 绩效整理表 (may not exist yet)."""
    outputs = month_config.get("outputs", {})
    perf_raw = outputs.get("performance_sheet_file")
    if perf_raw:
        return resolve_project_path(perf_raw)
    return performance_sheet_output_dir(month_config) / SYSTEM_PERF_FILENAME


def resolve_confirmed_performance_sheet_path(month_config: dict[str, Any]) -> Path:
    """Path to finance-confirmed 绩效整理表 (may not exist yet)."""
    return performance_sheet_output_dir(month_config) / CONFIRMED_PERF_FILENAME


def resolve_performance_sheet_path(month_config: dict[str, Any]) -> Path:
    """Prefer 财务确认版 over 系统生成 for downstream computation reads."""
    confirmed = resolve_confirmed_performance_sheet_path(month_config)
    if confirmed.exists():
        return confirmed
    return resolve_system_performance_sheet_path(month_config)


def _normalize_import_columns(columns: pd.Index) -> dict[Any, str]:
    rename: dict[Any, str] = {}
    for col in columns:
        text = str(col).strip()
        if text in _LABEL_TO_COLUMN:
            rename[col] = _LABEL_TO_COLUMN[text]
        elif len(text) <= 3 and text.isalpha():
            rename[col] = text.upper()
        else:
            rename[col] = text
    return rename


def load_performance_sheet_frame(
    path: Path,
    *,
    sheet_name: str = DEFAULT_PERF_SHEET,
) -> pd.DataFrame:
    """Load exported 绩效整理表 (row1 title, row2 sources, row3 headers) into letter-keyed frame."""
    path = Path(path)
    if not path.exists():
        return pd.DataFrame()
    header_row = resolve_performance_sheet_read_rows(path, sheet_name)
    raw = pd.read_excel(
        path,
        sheet_name=sheet_name,
        header=header_row - 1,
        engine="openpyxl",
    )
    if raw.empty:
        return raw
    return raw.rename(columns=_normalize_import_columns(raw.columns))


def load_resolved_performance_frame(
    month_config: dict[str, Any],
    fallback: pd.DataFrame | None = None,
) -> pd.DataFrame | None:
    """Load perf frame from resolved on-disk path when present."""
    path = resolve_performance_sheet_path(month_config)
    if not path.exists():
        return fallback
    frame = load_performance_sheet_frame(path)
    if frame.empty:
        return fallback
    return frame


def dataframe_to_letter_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Map Chinese export headers back to Excel letter column keys."""
    if df.empty:
        return df.copy()
    return df.rename(columns=_normalize_import_columns(df.columns))
