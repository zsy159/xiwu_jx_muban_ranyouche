from __future__ import annotations

import json
import logging
import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from salary_pipeline.data_ingestion.data_loader import WorkbookLoader, normalize_name
from salary_pipeline.calculators.performance_sheet.order_context import (
    enrich_order_context,
)
from salary_pipeline.data_ingestion.performance_sheet_golden import _normalize_vin
from salary_pipeline.ops.basic import ratio_with_cap, sumif_by_key
from salary_pipeline.ops.lookup import lookup_match_index, sumifs_by_keys

logger = logging.getLogger(__name__)

# W–AI 绩效列按岗位族分治，见 config/hub_performance.yaml（对齐 提成依据）。
# 本引擎仅回放 Excel 拓扑公式；岗位专属算法应逐步迁入 modules/ + commission_rules。

HUB_SHEET = "提成汇总"
TASK_SHEET = "销售任务及完成率"
PERF_SHEET = "绩效整理表"

# Excel column letter -> 提成汇总 header (row 2)
HUB_COLUMN_MAP: dict[str, str] = {
    "F": "考核量",
    "G": "实际销量",
    "H": "销量完成率",
    "I": "集客达成率",
    "J": "加装额",
    "K": "加装销量完成率",
    "L": "保险渗透率",
    "M": "整车毛利",
    "N": "加装毛利",
    "O": "保险毛利",
    "P": "按揭毛利",
    "W": "整车绩效",
    "X": "权限结余绩效",
    "Y": "加装绩效",
    "Z": "保险绩效",
    "AA": "金融绩效",
    "AB": "爱车宝绩效",
    "AC": "上户绩效",
    "AD": "盈利产品绩效",
    "AE": "延保提成",
    "AF": "特殊车型+指定车型",
    "AG": "座位险提成",
    "AH": "二手车提成",
    "AI": "玻碎险提成",
    "AK": "整车完成考核",
    "AM": "综合项",
    "AN": "04月活动",
    "AO": "超期",
}

INVITE_SHEET = "邀约专员提成"

CELL_REF = re.compile(r"^([A-Z]{1,3}\d+)$", re.IGNORECASE)
SHEET_CELL = re.compile(r"^(?:'([^']+)'|([^'!]+))!([A-Z]{1,3}\d+)$", re.IGNORECASE)
SUMIF_RE = re.compile(
    r"^=SUMIF\("
    r"(?:'(?P<qsheet>[^']+)'|(?P<usheet>[^'!]+))!"
    r"(?P<keycol>[A-Z]{1,3}):(?P=keycol),"
    r"(?P<crit>(?:提成汇总!)?D\d+|#[A-Z/!]+),"
    r"(?:'(?P<qsheet2>[^']+)'|(?P<usheet2>[^P'!][^'!]*|绩效整理表|销售任务及完成率|指标汇总))!"
    r"(?P<valcol>[A-Z]{1,3}):(?P=valcol)\)$",
    re.IGNORECASE,
)
# Simpler SUMIF parser via split
SUMIF_SIMPLE = re.compile(
    r"^=SUMIF\((.+)\)$",
    re.IGNORECASE,
)
SUM_RE = re.compile(r"^=SUM\((.+)\)$", re.IGNORECASE)
DIV_RE = re.compile(r"^=([A-Z]{1,3})(\d+)/([A-Z]{1,3})(\d+)$", re.IGNORECASE)
SUB_ARITH_RE = re.compile(
    r"^=([A-Z]{1,3})(\d+)-(?:'(?P<qsheet>[^']+)'|(?P<usheet>[^'!]+))!"
    r"(?P<addr>[A-Z]{1,3}\d+)$",
    re.IGNORECASE,
)
ADD_ARITH_RE = re.compile(
    r"^=指标汇总!(?P<a>[A-Z]+\d+)\+(?:'(?P<qsheet>[^']+)'|(?P<usheet>[^'!]+))!(?P<b>[A-Z]+\d+)$",
    re.IGNORECASE,
)
H_IF_CAP = re.compile(
    r"^=IF\(F(\d+)<>0,IF\(G\1/F\1>(?P<cap>\d+)%,(?P=cap)%,G\1/F\1\),0\)$",
    re.IGNORECASE,
)
INDEX_MATCH_RE = re.compile(
    r"^=IFERROR\(INDEX\("
    r"(?:'(?P<qsheet>[^']+)'|(?P<usheet>[^'!]+))!"
    r"(?P<vcol>[A-Z]{1,3}):(?P=vcol),"
    r"MATCH\(D(?P<row>\d+),(?:'(?P<qsheet2>[^']+)'|(?P<usheet2>[^'!]+))!"
    r"(?P<kcol>[A-Z]{1,3}):(?P=kcol),0\)\),0\)$",
    re.IGNORECASE,
)
K_RATIO_RE = re.compile(
    r"^=IF\(\(G(\d+)\*1500\)<>0,J\1/\(G\1\*1500\),0\)$",
    re.IGNORECASE,
)
L_PERF_RE = re.compile(
    r"^=IFERROR\(SUMIFS\("
    r"绩效整理表!K:K,绩效整理表!AB:AB,\">0\",绩效整理表!P:P,提成汇总!D(?P<row>\d+)\)"
    r"/SUMIF\(绩效整理表!P:P,提成汇总!D(?P=row),绩效整理表!K:K\),0\)$",
    re.IGNORECASE,
)
SUMIFS_PERF_RE = re.compile(
    r"^=SUMIFS\("
    r"绩效整理表!(?P<vcol>[A-Z]{1,3}):(?P=vcol),"
    r"绩效整理表!P:P,(?:提成汇总!)?D(?P<row>\d+)\)$",
    re.IGNORECASE,
)
SUMIFS_ADD_CONST_RE = re.compile(
    r"^=SUMIFS\((.+)\)\+(?P<tail>-?\d+(?:\.\d+)?)$",
    re.IGNORECASE,
)
MUL_CONST_RE = re.compile(
    r"^=([A-Z]{1,3})(\d+)\*(-?\d+(?:\.\d+)?)$",
    re.IGNORECASE,
)
MUL_CELLS_RE = re.compile(
    r"^=([A-Z]{1,3})(\d+)\*([A-Z]{1,3})(\d+)$",
    re.IGNORECASE,
)
MUL_EXPR_RE = re.compile(
    r"^=(.+)\*([A-Z]{1,3})(\d+)$",
    re.IGNORECASE,
)
SUB_SHEET_CELL_RE = re.compile(
    r"^=(?:'(?P<qsheet>[^']+)'|(?P<usheet>[^'!]+))!(?P<col>[A-Z]{1,3})(\d+)"
    r"-(?:'(?P<qsheet2>[^']+)'|(?P<usheet2>[^'!]+))!(?P<col2>[A-Z]{1,3})(\d+)$",
    re.IGNORECASE,
)
BIG_CUSTOMER_SHEET = "大客户"
NEW_MEDIA_SHEET = "新媒体"
RECRUIT_SHEET = "招聘"
USED_CAR_SHEET = "二手置换 "  # trailing space matches Excel sheet name
METRICS_SHEET = "指标汇总"

SUB_SUM_MINUS_RE = re.compile(
    r"^=SUM\((?P<body>.+)\)-(?:'(?P<qsheet>[^']+)'|(?P<usheet>[^'!]+))!"
    r"(?P<addr>[A-Z]{1,3}\d+)$",
    re.IGNORECASE,
)
SUB_SUM_CHAIN_RE = re.compile(
    r"^=SUM\((?P<body>[^)]+)\)(?P<tail>(?:-(?:'[^']+'|[^'!]+)![A-Z]{1,3}\d+)+)$",
    re.IGNORECASE,
)
SUB_SHEET_TAIL_RE = re.compile(
    r"-(?:'(?P<qsheet>[^']+)'|(?P<usheet>[^'!]+))!(?P<addr>[A-Z]{1,3}\d+)",
    re.IGNORECASE,
)
HUB_CELL_CRIT_RE = re.compile(r"^([A-Z]{1,3})(\d+)$", re.IGNORECASE)


def _col_to_index(letter: str) -> int:
    n = 0
    for ch in letter.upper():
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n - 1


def _index_to_col(idx: int) -> str:
    idx += 1
    parts: list[str] = []
    while idx:
        idx, rem = divmod(idx - 1, 26)
        parts.append(chr(rem + ord("A")))
    return "".join(reversed(parts))


def _is_unrecoverable_ref(formula: str) -> bool:
    """Skip only broken formulas; SUMIF(...,#REF!,...) still evaluates to 0."""
    text = formula.strip().upper()
    if text in ("#REF!", "=REF!", "=#REF!"):
        return True
    if "SUMIF(" in text or "SUMIFS(" in text:
        return False
    return "#REF!" in text


class HubFormulaEngine:
    """Evaluate 提成汇总 formula cells in topology execution order."""

    def __init__(
        self,
        topology_path: Path,
        loader: WorkbookLoader,
        *,
        computed_perf_frame: pd.DataFrame | None = None,
        use_golden_perf_sheet: bool = True,
        bootstrap_from_golden: bool = False,
    ) -> None:
        topo = json.loads(topology_path.read_text(encoding="utf-8"))
        self.cells: dict[str, dict[str, Any]] = topo["cells"]
        allowed_cols = set(HUB_COLUMN_MAP)
        self.execution_order = [
            key
            for key in topo["execution_order"]
            if key.startswith(f"{HUB_SHEET}!")
            and _column_letters(key.split("!", 1)[1]) in allowed_cols
        ]
        self.loader = loader
        self._computed_perf_frame = computed_perf_frame
        self._use_golden_perf_sheet = use_golden_perf_sheet
        self._bootstrap_from_golden = bootstrap_from_golden
        self.values: dict[str, float] = {}
        self.warnings: list[str] = []
        self._sheet_cache: dict[str, pd.DataFrame] = {}
        self._row_roles: dict[int, str] = {}
        self._row_names: dict[int, str] = {}

    def apply(self, summary: pd.DataFrame) -> pd.DataFrame:
        if summary.empty or "_excel_row" not in summary.columns:
            raise ValueError("summary requires _excel_row column")

        row_names = {
            int(row["_excel_row"]): normalize_name(row["姓名"])
            for _, row in summary.iterrows()
        }
        self._row_roles = {
            int(row["_excel_row"]): str(row.get("职务", ""))
            for _, row in summary.iterrows()
        }
        self._row_names = dict(row_names)
        self.values.clear()
        self.warnings.clear()

        for _pass in range(12):
            before = len(self.values)
            for full_key in self.execution_order:
                formula = self.cells[full_key].get("formula", "")
                if _is_unrecoverable_ref(formula):
                    if _pass == 0:
                        self.warnings.append(f"skip #REF!: {full_key} {formula}")
                    continue
                coord = full_key.split("!", 1)[1].upper()
                row_num = int(re.search(r"\d+", coord).group())
                name = row_names.get(row_num)
                try:
                    value = self._eval(formula, row_num, name)
                except Exception as exc:
                    if _pass == 0:
                        self.warnings.append(f"eval fail {full_key}: {exc}")
                    continue
                if value is not None:
                    self.values[coord] = 0.0 if pd.isna(value) else float(value)
            if len(self.values) == before:
                break

        out = summary.copy()
        for col_name in HUB_COLUMN_MAP.values():
            if col_name not in out.columns:
                out[col_name] = pd.NA
        for idx, row in out.iterrows():
            excel_row = int(row["_excel_row"])
            for letter, col_name in HUB_COLUMN_MAP.items():
                key = f"{letter}{excel_row}"
                if key in self.values:
                    out.at[idx, col_name] = self.values[key]
                else:
                    static = self._bootstrap_cell(letter, excel_row)
                    if static is not None:
                        out.at[idx, col_name] = static

        logger.info(
            "Hub formula engine: computed=%s warnings=%s",
            len(self.values),
            len(self.warnings),
        )
        if self.warnings:
            logger.warning("Formula warnings (first 5): %s", self.warnings[:5])
        return out.drop(columns=["_excel_row"], errors="ignore")

    def _eval(self, formula: str, row: int, name: Any) -> float | None:
        formula = formula.strip()
        if not formula.startswith("="):
            return None

        m = SUMIF_SIMPLE.match(formula)
        if m and formula.upper().startswith("=SUMIF("):
            body = m.group(1)
            if "+" in body.upper() and "SUMIF(" in body.upper():
                return self._eval_sumif_chain(body, row, name)
            return self._eval_sumif(body, row, name)

        m = MUL_EXPR_RE.match(formula)
        if m:
            left_part = m.group(1).strip()
            mul_ref = f"{m.group(2).upper()}{m.group(3)}"
            if left_part.upper().startswith("SUMIFS("):
                left_val = self._eval_sumifs_generic(left_part, row, name)
            else:
                left_val = self._eval(f"={left_part}", row, name)
            if left_val is None:
                return None
            return left_val * (self._cell(mul_ref) or 0.0)

        if formula.upper().startswith("=SUMIFS("):
            m_add = SUMIFS_ADD_CONST_RE.match(formula)
            if m_add:
                base = self._eval_sumifs_generic(f"SUMIFS({m_add.group(1)})", row, name)
                if base is None:
                    return None
                return float(base) + float(m_add.group("tail"))
            result = self._eval_sumifs_generic(formula[1:], row, name)
            if result is not None:
                return result

        m = SUM_RE.match(formula)
        if m:
            return self._eval_sum(m.group(1))

        m = CELL_REF.match(formula[1:])
        if m:
            return self._cell(m.group(1).upper())

        m = SUB_ARITH_RE.match(formula)
        if m:
            left = self._cell(f"{m.group(1).upper()}{m.group(2)}")
            sheet = m.group("qsheet") or m.group("usheet")
            right = self._sheet_cell(sheet, m.group("addr").upper())
            return left - right

        m = SUB_SUM_MINUS_RE.match(formula)
        if m:
            left = self._eval_sum(m.group("body"))
            sheet = m.group("qsheet") or m.group("usheet")
            right = self._sheet_cell(sheet, m.group("addr").upper())
            return left - right

        m = SUB_SUM_CHAIN_RE.match(formula)
        if m:
            return self._eval_sum_minus_chain(m.group("body"), m.group("tail"))

        m = SHEET_CELL.match(formula[1:])
        if m:
            sheet = m.group(1) or m.group(2)
            addr = m.group(3).upper()
            return self._sheet_cell(sheet, addr)

        m = DIV_RE.match(formula)
        if m:
            num = self._cell(f"{m.group(1).upper()}{m.group(2)}")
            den = self._cell(f"{m.group(3).upper()}{m.group(4)}")
            if den == 0:
                return 0.0
            return num / den

        m = ADD_ARITH_RE.match(formula)
        if m:
            left = self._sheet_cell("指标汇总", m.group("a").upper())
            sheet = m.group("qsheet") or m.group("usheet")
            right = self._sheet_cell(sheet, m.group("b").upper())
            return left + right

        m = H_IF_CAP.match(formula)
        if m:
            cap = float(m.group("cap")) / 100.0
            f_val = self._cell(f"F{row}") or 0.0
            g_val = self._cell(f"G{row}") or 0.0
            return float(ratio_with_cap(g_val, f_val, cap=cap))

        m = INDEX_MATCH_RE.match(formula)
        if m:
            return self._eval_index_match(m, name)

        m = K_RATIO_RE.match(formula)
        if m:
            r = int(m.group(1))
            g_val = self._cell(f"G{r}") or 0.0
            j_val = self._cell(f"J{r}") or 0.0
            denom = g_val * 1500
            return j_val / denom if denom else 0.0

        m = L_PERF_RE.match(formula)
        if m:
            return self._eval_insurance_penetration(name)

        if "#REF!" in formula.upper() and (
            formula.upper().startswith("=SUMIF(")
            or formula.upper().startswith("=SUMIFS(")
            or formula.upper().startswith("=IFERROR(SUMIFS(")
        ):
            return 0.0

        m = SUMIFS_PERF_RE.match(formula)
        if m:
            return self._eval_sumifs_perf(m, name)

        m = MUL_CONST_RE.match(formula)
        if m:
            left = self._cell(f"{m.group(1).upper()}{m.group(2)}")
            return left * float(m.group(3))

        m = MUL_CELLS_RE.match(formula)
        if m:
            left = self._cell(f"{m.group(1).upper()}{m.group(2)}")
            right = self._cell(f"{m.group(3).upper()}{m.group(4)}")
            return left * right

        m = SUB_SHEET_CELL_RE.match(formula)
        if m:
            sheet = m.group("qsheet") or m.group("usheet")
            left = self._sheet_cell(sheet, f"{m.group('col').upper()}{m.group(4)}")
            sheet2 = m.group("qsheet2") or m.group("usheet2")
            right = self._cell(f"{m.group('col2').upper()}{m.group(5)}")
            return left - right

        return None

    def _bootstrap_cell(self, letter: str, excel_row: int) -> float | None:
        if not self._bootstrap_from_golden:
            return None
        full_key = f"{HUB_SHEET}!{letter}{excel_row}"
        cell = self.cells.get(full_key, {})
        if cell.get("formula"):
            return None
        value = self.loader.read_cell_value(HUB_SHEET, f"{letter}{excel_row}")
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return None
        if isinstance(value, str) and not value.strip():
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    def _eval_sum_minus_chain(self, sum_body: str, tail: str) -> float:
        total = self._eval_sum(sum_body)
        for part in SUB_SHEET_TAIL_RE.finditer(tail):
            sheet = part.group("qsheet") or part.group("usheet")
            total -= self._sheet_cell(sheet, part.group("addr").upper())
        return total

    def _eval_sumif_chain(self, body: str, row: int, name: Any) -> float:
        """Evaluate SUMIF(a)+SUMIF(b)+… when the outer formula starts with =SUMIF(…).

        After stripping the leading ``=SUMIF(``, the first segment has no ``SUMIF`` prefix
        (e.g. ``绩效整理表!P:P,D92,绩效整理表!AN:AN)``); only later segments do.
        """
        total = 0.0
        for idx, part in enumerate(
            re.split(r"\+(?=SUMIF\()", body, flags=re.IGNORECASE)
        ):
            part = part.strip()
            if part.upper().startswith("SUMIF("):
                inner = part[6:].rstrip(")")
            elif idx == 0:
                inner = part.rstrip(")")
            else:
                continue
            total += self._eval_sumif(inner, row, name)
        return total

    def _eval_sumifs_perf(self, match: re.Match[str], name: Any) -> float:
        vcol = match.group("vcol").upper()
        frame = self._sheet_frame(PERF_SHEET)
        result = sumifs_by_keys(frame, vcol, [("P", name)])
        return float(result)

    def _eval_sumifs_generic(self, expr: str, row: int, name: Any) -> float | None:
        expr = expr.strip()
        if not expr.upper().startswith("SUMIFS(") or not expr.endswith(")"):
            return None
        body = expr[7:-1]
        parts = self._split_args(body)
        if len(parts) < 3:
            return None
        sheet, v_range = self._parse_range_ref(parts[0])
        vcol = v_range.split(":")[0].replace("$", "")
        frame = self._sheet_frame(sheet.strip() or PERF_SHEET)
        keys: list[tuple[str, Any]] = []
        idx = 1
        while idx + 1 < len(parts):
            _, key_range = self._parse_range_ref(parts[idx])
            key_col = key_range.split(":")[0].replace("$", "")
            crit_raw = parts[idx + 1].strip().strip('"')
            if key_col == "P" and (
                crit_raw in {f"D{row}", f"提成汇总!D{row}"}
                or crit_raw.startswith("D") and crit_raw[1:].isdigit()
            ):
                keys.append(("P", name))
            elif key_col == "H" and crit_raw.startswith("<>"):
                exclude = crit_raw[2:]
                keys.append((key_col, lambda s, ex=exclude: s.astype(str) != ex))
            else:
                keys.append((key_col, crit_raw))
            idx += 2
        if not keys:
            return None
        return float(sumifs_by_keys(frame, vcol, keys))

    def _eval_sumif(self, body: str, row: int, name: Any) -> float:
        parts = self._split_args(body)
        if len(parts) != 3:
            raise ValueError(f"SUMIF arity {len(parts)}")
        sheet1, key_range = self._parse_range_ref(parts[0])
        crit = parts[1].strip()
        sheet2, val_range = self._parse_range_ref(parts[2])
        sheet = sheet1 or sheet2
        key_col = key_range.split(":")[0].replace("$", "")
        val_col = val_range.split(":")[0].replace("$", "")

        if crit.upper().startswith("#REF"):
            self.warnings.append(f"SUMIF #REF criteria row {row}")
            return 0.0

        if crit.startswith("提成汇总!D") or crit.startswith("D"):
            criteria = name
        elif m := HUB_CELL_CRIT_RE.match(crit):
            criteria = self._cell(f"{m.group(1).upper()}{m.group(2)}")
        else:
            criteria = self._resolve_criteria(crit, row, name)

        frame = self._sheet_frame(sheet)
        return float(sumif_by_key(frame, key_col, val_col, str(criteria)))

    def _eval_index_match(self, match: re.Match[str], name: Any) -> float:
        sheet = match.group("qsheet") or match.group("usheet")
        vcol = match.group("vcol").upper()
        kcol = match.group("kcol").upper()
        frame = self._sheet_frame(sheet)
        result = lookup_match_index(
            pd.Series([name]),
            frame[kcol],
            frame[vcol],
        )
        return float(result.iloc[0])

    def _eval_insurance_penetration(self, name: Any) -> float:
        frame = self._sheet_frame(PERF_SHEET)
        numer = sumifs_by_keys(
            frame,
            "K",
            [("AB", lambda s: pd.to_numeric(s, errors="coerce") > 0), ("P", name)],
        )
        denom = float(sumif_by_key(frame, "P", "K", str(name)))
        return numer / denom if denom else 0.0

    def _eval_sum(self, body: str) -> float:
        total = 0.0
        for part in self._split_args(body):
            part = part.strip()
            if ":" in part:
                total += self._sum_range(part)
            else:
                ref = part.replace("$", "").replace(f"{HUB_SHEET}!", "").upper()
                total += self._cell(ref) or 0.0
        return total

    def _sum_range(self, range_ref: str) -> float:
        range_ref = range_ref.replace("$", "").replace(f"{HUB_SHEET}!", "")
        if ":" not in range_ref:
            return self._cell(range_ref.upper()) or 0.0
        start, end = range_ref.split(":")
        m1 = re.match(r"([A-Z]+)(\d+)", start, re.I)
        m2 = re.match(r"([A-Z]+)(\d+)", end, re.I)
        if not m1 or not m2:
            return 0.0
        c1, r1 = m1.group(1).upper(), int(m1.group(2))
        c2, r2 = m2.group(1).upper(), int(m2.group(2))
        c_lo, c_hi = sorted((_col_to_index(c1), _col_to_index(c2)))
        r_lo, r_hi = sorted((r1, r2))
        total = 0.0
        for excel_row in range(r_lo, r_hi + 1):
            for col_idx in range(c_lo, c_hi + 1):
                total += self._cell(f"{_index_to_col(col_idx)}{excel_row}") or 0.0
        return total

    def _cell(self, ref: str) -> float:
        ref = ref.upper()
        if ref in self.values:
            return self.values[ref]
        lazy = self._lazy_cell(ref)
        if lazy is not None:
            self.values[ref] = lazy
            return lazy
        if not self._bootstrap_from_golden:
            return 0.0
        value = self.loader.read_cell_value(HUB_SHEET, ref)
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return 0.0
        try:
            return float(value)
        except (TypeError, ValueError):
            return 0.0

    def _lazy_cell(self, ref: str) -> float | None:
        """Evaluate hub dependency cells (BA, H, …) omitted from execution_order."""
        match = re.match(r"([A-Z]{1,3})(\d+)", ref.upper())
        if not match:
            return None
        letter, row_num = match.group(1), int(match.group(2))
        full_key = f"{HUB_SHEET}!{letter}{row_num}"
        cell = self.cells.get(full_key, {})
        formula = cell.get("formula", "")
        name = self._row_names.get(row_num)
        if formula:
            try:
                value = self._eval(formula, row_num, name)
            except Exception:
                return None
            if value is None:
                return None
            return 0.0 if pd.isna(value) else float(value)
        if not self._bootstrap_from_golden:
            return None
        value = self.loader.read_cell_value(HUB_SHEET, ref)
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return None
        if isinstance(value, str) and not value.strip():
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    def _sheet_cell(self, sheet: str, address: str) -> float:
        sheet = self._resolve_sheet_name(sheet)
        value = self.loader.read_cell_value(sheet, address)
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return 0.0
        return float(value)

    def _resolve_sheet_name(self, sheet: str) -> str:
        sheet = sheet or ""
        if (
            sheet.strip() == PERF_SHEET
            and self._computed_perf_frame is not None
            and not self._computed_perf_frame.empty
            and not self._use_golden_perf_sheet
        ):
            return PERF_SHEET
        names = self.loader._workbook().sheetnames
        if sheet in names:
            return sheet
        stripped = sheet.strip()
        if stripped in names:
            return stripped
        if stripped == "二手置换" and USED_CAR_SHEET in names:
            return USED_CAR_SHEET
        if stripped in self.loader.sheet_paths:
            return stripped
        if sheet in self.loader.sheet_paths:
            return sheet
        raise KeyError(f"Worksheet {sheet} does not exist.")

    def _sheet_frame(self, sheet: str) -> pd.DataFrame:
        if (
            sheet.strip() == PERF_SHEET
            and self._computed_perf_frame is not None
            and not self._computed_perf_frame.empty
            and not self._use_golden_perf_sheet
        ):
            resolved = PERF_SHEET
        else:
            resolved = self._resolve_sheet_name(sheet.strip())
        sheet = resolved
        if sheet not in self._sheet_cache:
            if sheet == TASK_SHEET:
                task_letters = {"C": "C", "F": "F", "Y": "Y", "Z": "Z"}
                raw = self.loader._read_raw_sheet(sheet)
                if raw.shape[1] >= column_index_from_string("AG"):
                    task_letters["AG"] = "AG"
                frame = self.loader.read_sheet_columns(
                    sheet,
                    task_letters,
                    label=TASK_SHEET,
                )
                frame["C"] = frame["C"].map(normalize_name)
                for col in task_letters:
                    if col == "C":
                        continue
                    frame[col] = pd.to_numeric(frame[col], errors="coerce")
            elif sheet == PERF_SHEET:
                if (
                    self._computed_perf_frame is not None
                    and not self._computed_perf_frame.empty
                    and not self._use_golden_perf_sheet
                ):
                    frame = self._perf_frame_from_computed()
                else:
                    letters = [
                        "P", "S", "K", "AB", "BG", "BI", "AC", "H", "O", "Q",
                        "AH", "AK", "AM", "AN", "AS", "AL", "AT",
                        "AO", "AQ", "AR", "AP", "AU", "AG", "AI", "AJ",
                    ]
                    frame = self.loader.read_sheet_columns(
                        sheet,
                        {letter: letter for letter in letters},
                        label=PERF_SHEET,
                    )
                    frame["P"] = frame["P"].map(normalize_name)
                    for col in letters:
                        if col != "P":
                            frame[col] = pd.to_numeric(frame[col], errors="coerce")
                    frame = self._overlay_computed_perf(frame)
            elif sheet == INVITE_SHEET:
                frame = self.loader.read_sheet_columns(
                    sheet,
                    {"C": "C", "AF": "AF", "AE": "AE"},
                    label=INVITE_SHEET,
                )
                frame["C"] = frame["C"].map(normalize_name)
                frame["AE"] = frame["AE"].map(
                    lambda v: normalize_name(v) if isinstance(v, str) else v
                )
                frame["AF"] = pd.to_numeric(frame["AF"], errors="coerce")
            elif sheet == NEW_MEDIA_SHEET:
                frame = self.loader.read_sheet_columns(
                    sheet,
                    {"Y": "Y", "AB": "AB"},
                    label=NEW_MEDIA_SHEET,
                )
                frame["Y"] = frame["Y"].map(normalize_name)
                frame["AB"] = pd.to_numeric(frame["AB"], errors="coerce")
            elif sheet == RECRUIT_SHEET:
                frame = self.loader.read_sheet_columns(
                    sheet,
                    {"Q": "Q", "W": "W"},
                    label=RECRUIT_SHEET,
                )
                frame["Q"] = frame["Q"].map(normalize_name)
                frame["W"] = pd.to_numeric(frame["W"], errors="coerce")
            elif sheet == METRICS_SHEET:
                frame = self.loader.read_sheet_columns(
                    sheet,
                    {"E": "E"},
                    label=METRICS_SHEET,
                )
                frame["E"] = pd.to_numeric(frame["E"], errors="coerce")
            elif sheet == BIG_CUSTOMER_SHEET:
                wb = load_workbook(
                    self.loader.workbook_path, read_only=True, data_only=True
                )
                try:
                    ws = wb[BIG_CUSTOMER_SHEET]
                    max_col = int(ws.max_column or 0)
                    max_row = int(ws.max_row or 0)
                    r_idx = _col_to_index("R") + 1
                    x_idx = _col_to_index("X") + 1
                    rows: list[dict[str, Any]] = []
                    for row_num in range(1, max_row + 1):
                        r_val = (
                            ws.cell(row=row_num, column=r_idx).value
                            if r_idx <= max_col
                            else None
                        )
                        x_val = (
                            ws.cell(row=row_num, column=x_idx).value
                            if x_idx <= max_col
                            else 0
                        )
                        rows.append({"R": r_val, "X": x_val})
                finally:
                    wb.close()
                frame = pd.DataFrame(rows)
                frame["R"] = frame["R"].map(normalize_name)
                frame["X"] = pd.to_numeric(frame["X"], errors="coerce").fillna(0)
            else:
                frame = self._load_sumif_sheet(sheet)
            self._sheet_cache[sheet] = frame
        return self._sheet_cache[sheet]

    def _overlay_computed_perf(self, frame: pd.DataFrame) -> pd.DataFrame:
        """Replace golden 绩效整理表 columns with Phase B builder output (by VIN)."""
        overlay = self._computed_perf_frame
        if overlay is None or overlay.empty or "O" not in overlay.columns:
            return frame
        skip = {"O", "P", "K", "G", "_excel_row"}
        computed_cols = [
            col
            for col in overlay.columns
            if col in frame.columns and col not in skip
        ]
        if not computed_cols:
            return frame

        keyed = overlay.dropna(subset=["O"]).drop_duplicates("O", keep="last").set_index("O")
        vin_keys = frame["O"].map(_normalize_vin)
        out = frame.copy()
        for col in computed_cols:
            mapped = vin_keys.map(keyed[col])
            mask = mapped.notna()
            if mask.any():
                out.loc[mask, col] = pd.to_numeric(mapped[mask], errors="coerce")
        return out

    def _perf_frame_from_computed(self) -> pd.DataFrame:
        """Build Hub SUMIF source entirely from ``computed_perf_frame`` (no golden 绩效整理表)."""
        frame = self._computed_perf_frame.copy()
        if "P" in frame.columns:
            frame["P"] = frame["P"].map(normalize_name)
        if "O" in frame.columns:
            frame["O"] = frame["O"].map(_normalize_vin)
        skeleton_cols = [c for c in ("O", "P", "K", "G") if c in frame.columns]
        if skeleton_cols:
            ctx = enrich_order_context(frame[skeleton_cols], self.loader)
            for col in ("H", "S", "I", "R", "L", "A", "D"):
                if col in ctx.columns:
                    frame[col] = ctx[col].values
        for col in ("Q", "BG", "BI", "AC", "AU"):
            if col not in frame.columns:
                frame[col] = pd.NA
        for col in frame.columns:
            if str(col).startswith("_") or col in {"P", "O", "G"}:
                continue
            frame[col] = pd.to_numeric(frame[col], errors="coerce")
        return frame

    def _load_sumif_sheet(self, sheet: str) -> pd.DataFrame:
        """Load a workbook sheet with string-like columns normalized for SUMIF keys."""
        wb = load_workbook(self.loader.workbook_path, read_only=True, data_only=True)
        try:
            ws = wb[sheet]
            max_col = int(ws.max_column or 0)
            max_row = int(ws.max_row or 0)
            rows: list[dict[str, Any]] = []
            for row_num in range(1, max_row + 1):
                row_data: dict[str, Any] = {}
                for col_idx in range(1, max_col + 1):
                    row_data[_index_to_col(col_idx - 1)] = ws.cell(
                        row=row_num, column=col_idx
                    ).value
                rows.append(row_data)
        finally:
            wb.close()
        frame = pd.DataFrame(rows)
        for col in frame.columns:
            if not pd.api.types.is_numeric_dtype(frame[col]):
                frame[col] = frame[col].map(
                    lambda v: normalize_name(v) if isinstance(v, str) else v
                )
            else:
                frame[col] = pd.to_numeric(frame[col], errors="coerce")
        return frame

    @staticmethod
    def _split_args(body: str) -> list[str]:
        parts: list[str] = []
        depth = 0
        current: list[str] = []
        for char in body:
            if char == "(":
                depth += 1
            elif char == ")":
                depth -= 1
            if char == "," and depth == 0:
                parts.append("".join(current).strip())
                current = []
            else:
                current.append(char)
        if current:
            parts.append("".join(current).strip())
        return parts

    @staticmethod
    def _parse_range_ref(token: str) -> tuple[str, str]:
        token = token.strip()
        if token.startswith("'"):
            bang = token.index("!")
            sheet = token[1 : bang - 1] if token[bang - 1] == "'" else token[1:bang].strip("'")
            return sheet, token[bang + 1 :]
        if "!" in token:
            sheet, range_ref = token.split("!", 1)
            return sheet.strip(), range_ref
        return HUB_SHEET, token

    @staticmethod
    def _resolve_criteria(crit: str, row: int, name: Any) -> Any:
        crit = crit.strip()
        if crit == f"D{row}" or crit == f"提成汇总!D{row}":
            return name
        return crit.strip('"')


def _column_letters(coord: str) -> str:
    match = re.match(r"([A-Z]{1,3})", coord.upper())
    return match.group(1) if match else ""
