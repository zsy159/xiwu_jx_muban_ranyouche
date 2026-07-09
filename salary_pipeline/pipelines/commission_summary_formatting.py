"""Post-export 提成汇总 formatting: parity highlights and formula annotations."""

from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.comments import Comment

from salary_pipeline.calculators.sales_advisor.parity_annotations import (
    annotations_for_workbook,
    enrich_cell_mismatches,
    parity_values_for_annotations,
)
from salary_pipeline.calculators.sales_advisor.registry import (
    build_reconcile_deferred_cells,
    wa_parity_deferred_cells,
    wa_parity_deferred_reasons,
)
from salary_pipeline.calculators.sales_advisor.topology_specs import (
    collect_topology_static_fill_cells,
)
from salary_pipeline.pipelines.performance_sheet_paths import (
    resolve_system_performance_sheet_path,
)
from salary_pipeline.pipelines.commission_summary import (
    EXPORT_HEADER_ROW,
    LEGEND_INSERT_ROW,
    computed_highlight_rows,
)
from salary_pipeline.pipelines.commission_summary_column_sources import (
    manual_column_headers,
)
from salary_pipeline.utils.excel_format import (
    GOLDEN_STATIC_FILL,
    STATIC_FILL_COMMENT,
    add_commission_summary_annotations,
    add_commission_summary_color_legend,
    commission_summary_legend_present,
    highlight_commission_summary_deferred_cells,
    highlight_commission_summary_mismatches,
)
from salary_pipeline.validation.parity import (
    CommissionSummaryParity,
    resolve_hub_compare_columns,
)

logger = logging.getLogger(__name__)


@dataclass
class HighlightStats:
    mismatches: int = 0
    deferred: int = 0
    annotated: int = 0
    manual_marked: int = 0


def parity_highlight_mode(parity_cfg: dict[str, Any]) -> str:
    """Return ``mismatch_only`` (fast amber fills) or ``full`` (deferred + annotations)."""
    if parity_cfg.get("lightweight_highlight") or parity_cfg.get("skip_root_cause"):
        return "mismatch_only"
    return str(parity_cfg.get("highlight_mode", "mismatch_only"))


def resolve_highlight_golden_path(month_config: dict[str, Any]) -> Path | None:
    """Golden workbook used for reconcile highlighting (not upload merge)."""
    from salary_pipeline.data_ingestion.data_loader import (
        resolve_parity_golden_workbook,
    )

    sheet = month_config.get("outputs", {}).get("commission_summary_sheet", "提成汇总")
    return resolve_parity_golden_workbook(month_config, sheet_name=sheet)


def _header_column_map(ws, header_row: int) -> dict[str, int]:
    from salary_pipeline.data_ingestion.data_loader import normalize_header

    mapping: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        header = normalize_header(ws.cell(row=header_row, column=col_idx).value)
        if header:
            mapping[header] = col_idx
    return mapping


def highlight_manual_hub_columns(
    workbook_path: Path,
    manual_headers: frozenset[str],
    *,
    sheet_name: str = "提成汇总",
    header_row: int = EXPORT_HEADER_ROW,
    data_start_row: int = EXPORT_HEADER_ROW + 1,
) -> int:
    """Light gray for empty cells in manual / non-computed columns."""
    if not manual_headers:
        return 0
    wb = load_workbook(workbook_path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return 0
    ws = wb[sheet_name]
    col_map = _header_column_map(ws, header_row)
    count = 0
    for header in manual_headers:
        col_idx = col_map.get(header)
        if col_idx is None:
            continue
        for row_idx in range(data_start_row, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None and cell.value != "":
                continue
            cell.fill = GOLDEN_STATIC_FILL
            if cell.comment is None:
                cell.comment = Comment(STATIC_FILL_COMMENT, "对账")
            count += 1
    wb.save(workbook_path)
    return count


def apply_commission_summary_highlighting(
    month_config: dict[str, Any],
    computed_path: Path,
    *,
    golden_path: Path | None = None,
) -> HighlightStats:
    """
    Apply color legend, parity mismatch fills, and manual-column gray.

    Intended to run after ``CommissionSummaryBuilder.export_excel`` on both
    ``compute`` and upload trial paths so generated workbooks match reconcile output.
    """
    parity_cfg = month_config.get("parity", {})
    if parity_cfg.get("auto_highlight", True) is False:
        return HighlightStats()

    golden = golden_path or resolve_highlight_golden_path(month_config)
    if golden is None:
        logger.info("Skipping commission summary highlighting: no golden workbook")
        return HighlightStats()

    sheet = month_config["outputs"].get("commission_summary_sheet", "提成汇总")
    golden_header_row = int(parity_cfg.get("header_row", 2))
    golden_data_start_row = int(parity_cfg.get("data_start_row", 3))
    compare_columns = resolve_hub_compare_columns(parity_cfg)
    if not compare_columns:
        return HighlightStats()

    perf_columns = frozenset(parity_cfg.get("performance_columns") or [])

    highlight_mode = parity_highlight_mode(parity_cfg)
    lightweight = highlight_mode != "full"

    legend_present = commission_summary_legend_present(
        computed_path, sheet, insert_at_row=LEGEND_INSERT_ROW
    )
    legend_inserted = add_commission_summary_color_legend(
        computed_path, sheet, insert_at_row=LEGEND_INSERT_ROW
    )
    highlight_header_row, highlight_data_start = computed_highlight_rows(
        legend_inserted=legend_inserted,
        legend_present=legend_present,
    )

    perf_path = resolve_system_performance_sheet_path(month_config)
    highlight_checker = CommissionSummaryParity(
        join_keys=parity_cfg.get("join_keys", ["店别", "职务", "姓名"]),
        numeric_tolerance=float(parity_cfg.get("numeric_tolerance", 1e-6)),
        columns=compare_columns,
        deferred_cells={},
        golden_workbook=golden,
        computed_perf_path=perf_path,
        performance_columns=perf_columns if perf_columns else None,
        treat_empty_as_zero=True,
    )
    mismatches = highlight_checker.collect_mismatches_from_files(
        computed_path,
        golden,
        sheet,
        header_row=highlight_header_row,
        data_start_row=highlight_data_start,
        golden_header_row=golden_header_row,
        golden_data_start_row=golden_data_start_row,
    )
    if not lightweight:
        deferred_for_highlight = build_reconcile_deferred_cells(
            golden,
            perf_path=perf_path,
            header_row=golden_header_row,
            data_start_row=golden_data_start_row,
        )
        highlight_checker.deferred_cells = deferred_for_highlight
        mismatches = enrich_cell_mismatches(
            mismatches,
            golden_workbook=golden,
            golden_data_start_row=golden_data_start_row,
        )

    write_comments = bool(parity_cfg.get("write_comments", not lightweight))
    mismatch_count = highlight_commission_summary_mismatches(
        computed_path,
        sheet,
        mismatches,
        parity_cfg.get("join_keys", ["店别", "职务", "姓名"]),
        compare_columns,
        header_row=highlight_header_row,
        data_start_row=highlight_data_start,
        write_comments=write_comments,
    )

    deferred_count = 0
    manual_count = 0
    if not lightweight:
        static_cells = collect_topology_static_fill_cells(
            golden_workbook_path=golden,
            header_row=golden_header_row,
            data_start_row=golden_data_start_row,
        )
        deferred_count = highlight_commission_summary_deferred_cells(
            computed_path,
            sheet,
            deferred_for_highlight,
            static_cells=static_cells,
            deferred_reasons=wa_parity_deferred_reasons(),
            header_row=highlight_header_row,
            data_start_row=highlight_data_start,
        )

        wb = load_workbook(computed_path, read_only=True, data_only=True)
        try:
            ws = wb[sheet]
            headers = [
                str(ws.cell(row=highlight_header_row, column=c).value or "").strip()
                for c in range(1, ws.max_column + 1)
            ]
        finally:
            wb.close()
        manual_count = highlight_manual_hub_columns(
            computed_path,
            manual_column_headers([h for h in headers if h]),
            sheet_name=sheet,
            header_row=highlight_header_row,
            data_start_row=highlight_data_start,
        )

    annotated_count = 0
    if not lightweight:
        base_annotations = annotations_for_workbook(
            deferred_cells=deferred_for_highlight,
            golden_workbook=golden,
            golden_data_start_row=golden_data_start_row,
        )
        parity_values = parity_values_for_annotations(
            computed_path,
            golden,
            sheet,
            [ann.key() for ann in base_annotations],
            join_keys=parity_cfg.get("join_keys", ["店别", "职务", "姓名"]),
            header_row=highlight_header_row,
            data_start_row=highlight_data_start,
            golden_header_row=golden_header_row,
            golden_data_start_row=golden_data_start_row,
        )
        annotations = annotations_for_workbook(
            parity_values=parity_values,
            deferred_cells=deferred_for_highlight,
            golden_workbook=golden,
            golden_data_start_row=golden_data_start_row,
        )
        annotated_count = add_commission_summary_annotations(
            computed_path,
            sheet,
            annotations,
            header_row=highlight_header_row,
            data_start_row=highlight_data_start,
        )

    stats = HighlightStats(
        mismatches=mismatch_count,
        deferred=deferred_count,
        annotated=annotated_count,
        manual_marked=manual_count,
    )
    logger.info(
        "Commission summary highlighting -> %s "
        "(mismatches=%s, deferred=%s, manual=%s, annotated=%s)",
        computed_path,
        stats.mismatches,
        stats.deferred,
        stats.manual_marked,
        stats.annotated,
    )
    return stats
