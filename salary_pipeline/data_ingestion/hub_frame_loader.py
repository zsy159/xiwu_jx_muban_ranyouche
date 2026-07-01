"""Build 提成汇总 column frames for payout SUMIF (computed hub only)."""

from __future__ import annotations

import logging
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from salary_pipeline.data_ingestion.data_loader import (
    _column_sort_key,
    _log_frame_shape,
    normalize_header,
    normalize_name,
)
from salary_pipeline.pipelines.hub_formula_engine import HUB_COLUMN_MAP

logger = logging.getLogger(__name__)

# Excel letter → 提成汇总 header (golden layout); used to read computed hub by name.
PAYOUT_LETTER_TO_HEADER: dict[str, str] = {
    "D": "姓名",
    **HUB_COLUMN_MAP,
    "AK": "整车完成考核",
    "AL": "加装完成考核",
    "AM": "综合项",
    "AN": "04月活动",
    "AO": "超期",
    "AP": "（已发放奖励）",
    "AQ": "交车支出",
    "AR": "保客考核",
    "AT": "提成合计",
    "AX": "计提单台",
    "AY": "计提金额",
}

# Columns XW提成-发 pulls from 提成汇总 via SUMIF
PAYOUT_HUB_LETTERS = [
    "D",
    "F",
    "G",
    "W",
    "X",
    "Y",
    "Z",
    "AA",
    "AB",
    "AC",
    "AD",
    "AE",
    "AF",
    "AG",
    "AH",
    "AI",
    "AK",
    "AL",
    "AM",
    "AN",
    "AO",
    "AP",
    "AQ",
    "AR",
    "AT",
    "AX",
    "AY",
]


def _col_to_index(letter: str) -> int:
    n = 0
    for ch in letter.upper():
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n - 1


def _sheet_max_column(workbook_path: Path, sheet_name: str) -> int:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        return int(wb[sheet_name].max_column or 0)
    finally:
        wb.close()


def _letters_within_bounds(
    workbook_path: Path, sheet_name: str, letters: list[str]
) -> list[str]:
    max_col = _sheet_max_column(workbook_path, sheet_name)
    return [letter for letter in letters if _col_to_index(letter) + 1 <= max_col]


def detect_hub_data_start_row(
    workbook_path: Path,
    sheet_name: str = "提成汇总",
    *,
    default: int = 3,
) -> int:
    """Return first data row by locating the 姓名 header (handles legend row offset)."""
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return default
        ws = wb[sheet_name]
        for row_idx in range(1, min(ws.max_row, 12) + 1):
            for col_idx in range(1, min(ws.max_column, 20) + 1):
                if normalize_header(ws.cell(row=row_idx, column=col_idx).value) == "姓名":
                    return row_idx + 1
    finally:
        wb.close()
    return default


def read_hub_columns_by_letter(
    workbook_path: Path,
    sheet_name: str = "提成汇总",
    letters: list[str] | None = None,
    *,
    data_start_row: int | None = None,
) -> pd.DataFrame:
    """Load hub metrics by Excel column letter (values, not formulas)."""
    letters = letters or PAYOUT_HUB_LETTERS
    available = _letters_within_bounds(workbook_path, sheet_name, letters)
    if not available:
        return pd.DataFrame()

    start_row = (
        data_start_row
        if data_start_row is not None
        else detect_hub_data_start_row(workbook_path, sheet_name)
    )
    usecols = ",".join(sorted(set(available), key=_column_sort_key))
    raw = pd.read_excel(
        workbook_path,
        sheet_name=sheet_name,
        usecols=usecols,
        header=None,
        skiprows=start_row - 1,
        engine="openpyxl",
    )
    letter_order = sorted(set(available), key=_column_sort_key)
    letter_to_pos = {letter.upper(): idx for idx, letter in enumerate(letter_order)}
    out = pd.DataFrame()
    for letter in letter_order:
        out[letter] = raw.iloc[:, letter_to_pos[letter.upper()]]
    if "D" in out.columns:
        out["D"] = out["D"].map(normalize_name)
    for letter in letter_order:
        if letter != "D":
            out[letter] = pd.to_numeric(out[letter], errors="coerce")
    return _log_frame_shape(out, f"hub columns {workbook_path.name}!{sheet_name}")


def read_hub_columns_mapped(
    workbook_path: Path,
    sheet_name: str = "提成汇总",
    letters: list[str] | None = None,
    *,
    data_start_row: int | None = None,
) -> pd.DataFrame:
    """Load hub metrics by header name, output keyed by Excel letter for SUMIF."""
    letters = letters or PAYOUT_HUB_LETTERS
    start_row = (
        data_start_row
        if data_start_row is not None
        else detect_hub_data_start_row(workbook_path, sheet_name)
    )
    header_row = start_row - 1
    raw = pd.read_excel(
        workbook_path,
        sheet_name=sheet_name,
        header=header_row - 1,
        engine="openpyxl",
    )
    col_by_header = {
        normalize_header(col): col for col in raw.columns if normalize_header(col)
    }
    out = pd.DataFrame()
    for letter in letters:
        header = normalize_header(PAYOUT_LETTER_TO_HEADER.get(letter, ""))
        src = col_by_header.get(header) if header else None
        if src is None:
            out[letter] = pd.NA
        else:
            out[letter] = raw[src].values
    if "D" in out.columns:
        out["D"] = out["D"].map(normalize_name)
    for letter in letters:
        if letter != "D":
            out[letter] = pd.to_numeric(out[letter], errors="coerce")
    return _log_frame_shape(
        out, f"hub columns (mapped) {workbook_path.name}!{sheet_name}"
    )


def _empty_hub_frame(
    golden_workbook: Path,
    *,
    sheet_name: str = "提成汇总",
    letters: list[str] | None = None,
    data_start_row: int = 3,
) -> pd.DataFrame:
    """Name column only — numeric SUMIF source columns left empty."""
    letters = letters or PAYOUT_HUB_LETTERS
    frame = read_hub_columns_by_letter(
        golden_workbook,
        sheet_name,
        ["D"],
        data_start_row=data_start_row,
    )
    for letter in letters:
        if letter != "D":
            frame[letter] = pd.NA
    return frame


def build_hub_sumif_frame(
    golden_workbook: Path,
    *,
    computed_workbook: Path | None = None,
    sheet_name: str = "提成汇总",
    letters: list[str] | None = None,
    data_start_row: int = 3,
) -> pd.DataFrame:
    """
    Load hub columns for payout SUMIF from computed 提成汇总 only.

  ``golden_workbook`` is retained for API compatibility and is used only to
  supply the D-column name skeleton when the computed hub file is missing.
  Golden W–AR values are never merged into the SUMIF source frame.
    """
    letters = letters or PAYOUT_HUB_LETTERS
    if computed_workbook is not None and computed_workbook.exists():
        return read_hub_columns_mapped(
            computed_workbook,
            sheet_name,
            letters,
        )

    logger.warning(
        "Computed hub %s missing; payout SUMIF numeric columns will be empty "
        "(D-column skeleton from %s)",
        computed_workbook,
        golden_workbook.name,
    )
    return _empty_hub_frame(
        golden_workbook,
        sheet_name=sheet_name,
        letters=letters,
        data_start_row=data_start_row,
    )
