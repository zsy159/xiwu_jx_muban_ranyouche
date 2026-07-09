from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any

import pandas as pd
import yaml
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from salary_pipeline.modules.base import PERSONNEL_FILENAME, PERSONNEL_SHEET
from salary_pipeline.modules.base import SUMMARY_KEY_COLUMNS
from salary_pipeline.paths import CONFIG_DIR, resolve_project_path

logger = logging.getLogger(__name__)

EXCEL_COL_RE = re.compile(r"^([A-Z]{1,3})(\d+)$", re.IGNORECASE)


def normalize_name(value: Any) -> str | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    text = str(value).strip()
    return text or None


def normalize_header(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\n", "").strip()
    return text or None


def read_golden_summary_sheet(
    workbook_path: Path,
    sheet_name: str,
    *,
    header_row: int = 2,
    data_start_row: int | None = None,
) -> pd.DataFrame:
    """Read the reference 提成汇总 sheet (computed values, not formulas)."""
    df = pd.read_excel(
        workbook_path,
        sheet_name=sheet_name,
        header=header_row - 1,
        engine="openpyxl",
    )
    df.columns = [normalize_header(c) for c in df.columns]
    if data_start_row and data_start_row > header_row + 1:
        skip = data_start_row - header_row - 2
        if skip > 0:
            df = df.iloc[skip:].reset_index(drop=True)
    df = _clean_summary_frame(df)
    logger.info(
        "Loaded golden summary %s!%s shape=%s",
        workbook_path.name,
        sheet_name,
        df.shape,
    )
    return df


def read_aftersales_metric_frame(
    workbook_path: Path,
    sheet_name: str,
    column_map: dict[str, str],
    *,
    data_start_row: int = 5,
) -> pd.DataFrame:
    """Read aftersales anchor metrics by Excel column letters (avoids merged-header ambiguity)."""
    key_letters = {"B": "店别", "C": "姓名"}
    all_letters = {**key_letters, **{letter: name for letter, name in column_map.items()}}
    letters = sorted(all_letters.keys(), key=_column_sort_key)
    usecols = ",".join(letters)
    raw = pd.read_excel(
        workbook_path,
        sheet_name=sheet_name,
        usecols=usecols,
        header=None,
        skiprows=data_start_row - 1,
        engine="openpyxl",
    )
    letter_to_pos = {letter.upper(): idx for idx, letter in enumerate(letters)}
    out = pd.DataFrame()
    for letter, name in all_letters.items():
        out[name] = raw.iloc[:, letter_to_pos[letter.upper()]]
    out["店别"] = out["店别"].ffill().map(normalize_name)
    out["姓名"] = out["姓名"].map(normalize_name)
    out = out[out["姓名"].notna()]
    out = out[~out["姓名"].isin(["小计", "空白", "0", 0])]
    return _log_frame_shape(out.reset_index(drop=True), f"aftersales metrics {sheet_name}")


def read_computed_aftersales_excel(
    workbook_path: Path,
    sheet_name: str,
    column_map: dict[str, str],
) -> pd.DataFrame:
    """Read pipeline-exported aftersales xlsx (header at row 4 / startrow=3)."""
    df = pd.read_excel(
        workbook_path,
        sheet_name=sheet_name,
        header=3,
        engine="openpyxl",
    )
    df.columns = [normalize_header(c) for c in df.columns]
    if "姓名" not in df.columns and "姓 名" in df.columns:
        df = df.rename(columns={"姓 名": "姓名"})
    if "店别" not in df.columns and "部门" in df.columns:
        df["店别"] = df["部门"].ffill().map(normalize_name)
    df["姓名"] = df["姓名"].map(normalize_name)
    return filter_comparable_rows(_clean_summary_frame(df))


def read_payout_metric_frame(
    workbook_path: Path,
    sheet_name: str,
    column_map: dict[str, str],
    *,
    data_start_row: int = 3,
) -> pd.DataFrame:
    """Read payout output metrics by Excel column letters."""
    key_letters = {"B": "店别", "C": "职务", "D": "姓名"}
    all_letters = {**key_letters, **{letter: name for letter, name in column_map.items()}}
    letters = sorted(all_letters.keys(), key=_column_sort_key)
    usecols = ",".join(letters)
    raw = pd.read_excel(
        workbook_path,
        sheet_name=sheet_name,
        usecols=usecols,
        header=None,
        skiprows=data_start_row - 1,
        engine="openpyxl",
    )
    letter_to_pos = {letter.upper(): idx for idx, letter in enumerate(letters)}
    out = pd.DataFrame()
    for letter, name in all_letters.items():
        out[name] = raw.iloc[:, letter_to_pos[letter.upper()]]
    out["店别"] = out["店别"].ffill().map(normalize_name)
    out["职务"] = out["职务"].map(normalize_name)
    out["姓名"] = out["姓名"].map(normalize_name)
    out = out[out["姓名"].notna()]
    out = out[~out["姓名"].isin(["小计", "空白", "0", 0])]
    return _log_frame_shape(out.reset_index(drop=True), f"payout metrics {sheet_name}")


def resolve_computed_payout_read_rows(
    workbook_path: Path,
    sheet_name: str,
    *,
    header_row: int = 3,
    data_start_row: int = 4,
    legend_insert_row: int = 2,
) -> tuple[int, int]:
    """Return header/data rows for computed payout, accounting for reconcile legend row."""
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return header_row, data_start_row
        ws = wb[sheet_name]
        if _summary_sheet_header_present(ws, header_row):
            return header_row, data_start_row
        if (
            _legend_row_present_in_sheet(ws, legend_insert_row)
            and _summary_sheet_header_present(ws, header_row + 1)
        ):
            return header_row + 1, data_start_row + 1
        return header_row, data_start_row
    finally:
        wb.close()


def read_computed_payout_excel(
    workbook_path: Path,
    sheet_name: str,
    *,
    header_row: int = 3,
    adjust_for_legend: bool = True,
) -> pd.DataFrame:
    """Read pipeline-exported payout xlsx (default header at Excel row 3)."""
    data_start_row = header_row + 1
    if adjust_for_legend:
        header_row, data_start_row = resolve_computed_payout_read_rows(
            workbook_path,
            sheet_name,
            header_row=header_row,
            data_start_row=data_start_row,
        )
    df = pd.read_excel(
        workbook_path,
        sheet_name=sheet_name,
        header=header_row - 1,
        engine="openpyxl",
    )
    df.columns = [normalize_header(c) for c in df.columns]
    return filter_comparable_rows(_clean_summary_frame(df))


def _summary_sheet_header_present(worksheet: Any, row: int) -> bool:
    for col in range(1, worksheet.max_column + 1):
        if normalize_header(worksheet.cell(row=row, column=col).value) == "店别":
            return True
    return False


def _legend_row_present_in_sheet(worksheet: Any, row: int) -> bool:
    for col in range(1, worksheet.max_column + 1):
        value = worksheet.cell(row=row, column=col).value
        if isinstance(value, str) and "数值不一致" in value:
            return True
    return False


def resolve_computed_summary_read_rows(
    workbook_path: Path,
    sheet_name: str,
    *,
    header_row: int = 3,
    data_start_row: int = 4,
    legend_insert_row: int = 2,
) -> tuple[int, int]:
    """Return header/data rows for computed 提成汇总, accounting for reconcile legend row."""
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return header_row, data_start_row
        ws = wb[sheet_name]
        if _summary_sheet_header_present(ws, header_row):
            return header_row, data_start_row
        if (
            _legend_row_present_in_sheet(ws, legend_insert_row)
            and _summary_sheet_header_present(ws, header_row + 1)
        ):
            return header_row + 1, data_start_row + 1
        return header_row, data_start_row
    finally:
        wb.close()


def read_computed_summary_excel(
    workbook_path: Path,
    sheet_name: str = "提成汇总",
    *,
    header_row: int = 3,
    data_start_row: int = 4,
    adjust_for_legend: bool = True,
) -> pd.DataFrame:
    """Read pipeline-exported 提成汇总.xlsx (default header at row 3)."""
    if adjust_for_legend:
        header_row, data_start_row = resolve_computed_summary_read_rows(
            workbook_path,
            sheet_name,
            header_row=header_row,
            data_start_row=data_start_row,
        )
    return read_golden_summary_sheet(
        workbook_path,
        sheet_name,
        header_row=header_row,
        data_start_row=data_start_row,
    )


def filter_comparable_rows(df: pd.DataFrame) -> pd.DataFrame:
    """Exclude subtotal / placeholder rows from parity comparison.

    Keeps hub-linked alias rows with blank 序号 (e.g. 余才万3/渠道) that still
    appear in the system skeleton and carry golden 整车绩效 values.
    """
    return filter_skeleton_rows(df)


def filter_skeleton_rows(df: pd.DataFrame) -> pd.DataFrame:
    """Rows needed for hub formula evaluation (SUM ranges, SUMIF keys).

    Unlike ``filter_comparable_rows``, keeps rows with blank 序号 such as
    duplicate 区域顾问 entries (余才万2 / 余才万5) that still participate in
    store-block ``SUM(G*:G*)`` subtotals.
    """
    if df.empty:
        return df
    out = df.copy()
    if "姓名" in out.columns:
        out = out[out["姓名"].notna()]
        out = out[~out["姓名"].isin(["空白", "0", "小计", 0])]
    if "职务" in out.columns:
        out = out[out["职务"].notna()]
        out = out[~out["职务"].astype(str).isin(["小计", "0"])]
    return out.reset_index(drop=True)


def _clean_summary_frame(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if "店别" in out.columns:
        out["店别"] = out["店别"].ffill()
    if "姓名" in out.columns:
        out["姓名"] = out["姓名"].map(normalize_name)
    for col in ("店别", "职务"):
        if col in out.columns:
            out[col] = out[col].map(normalize_name)
    return out


def summary_frame_from_builder(df: pd.DataFrame) -> pd.DataFrame:
    return _clean_summary_frame(df)


def load_month_config(config_dir: Path | None = None) -> dict[str, Any]:
    path = (config_dir or CONFIG_DIR) / "month.yaml"
    with path.open(encoding="utf-8") as handle:
        return yaml.safe_load(handle)


class WorkbookLoader:
    """Read registered sheets from the monthly sales workbook."""

    def __init__(
        self,
        workbook_path: Path,
        *,
        sheet_paths: dict[str, Path] | None = None,
    ) -> None:
        self.workbook_path = Path(workbook_path)
        self.sheet_paths = {
            name: Path(path) for name, path in (sheet_paths or {}).items()
        }
        self._wb = None
        self._cell_cache: dict[tuple[str, str], Any] = {}
        self._raw_sheet_cache: dict[tuple[Path, str], pd.DataFrame] = {}
        self._delegates: dict[Path, WorkbookLoader] = {}

    def _path_for_sheet(self, sheet_name: str) -> Path:
        return self.sheet_paths.get(sheet_name, self.workbook_path)

    def _resolved_sheet_name(self, sheet_name: str) -> str:
        path = self._path_for_sheet(sheet_name)
        if path.resolve() == self.workbook_path.resolve():
            return sheet_name
        from salary_pipeline.ingestion_upload.file_intake import scan_workbook_sheets
        from salary_pipeline.ingestion_upload.manifest import resolve_sheet_from_upload

        inner_names = scan_workbook_sheets(path)
        resolved = resolve_sheet_from_upload(
            sheet_name,
            filename=path.name,
            sheet_names=inner_names,
        )
        return resolved or sheet_name

    def _delegate_for_sheet(self, sheet_name: str) -> WorkbookLoader:
        path = self._path_for_sheet(sheet_name)
        if path.resolve() == self.workbook_path.resolve():
            return self
        delegate = self._delegates.get(path)
        if delegate is None:
            delegate = WorkbookLoader(path)
            self._delegates[path] = delegate
        return delegate

    def worksheet(self, sheet_name: str):
        """Return openpyxl worksheet, including supplemental sheet_sources uploads."""
        delegate = self._delegate_for_sheet(sheet_name)
        if delegate is not self:
            resolved = delegate._resolved_sheet_name(sheet_name)
            return delegate._workbook()[resolved]
        resolved = self._resolved_sheet_name(sheet_name)
        return self._workbook()[resolved]

    def has_sheet(self, sheet_name: str) -> bool:
        """True when sheet exists in merged workbook or supplemental uploads."""
        if sheet_name in self.sheet_paths:
            return True
        delegate = self._delegate_for_sheet(sheet_name)
        if delegate is not self:
            resolved = delegate._resolved_sheet_name(sheet_name)
            return resolved in delegate._workbook().sheetnames
        if sheet_name in self._workbook().sheetnames:
            return True
        resolved = self._resolved_sheet_name(sheet_name)
        return resolved in self._workbook().sheetnames

    def _workbook(self):
        if self._wb is None:
            self._wb = load_workbook(
                self.workbook_path, read_only=True, data_only=True
            )
        return self._wb

    def _read_cell_value_local(self, sheet_name: str, address: str) -> Any:
        address = address.strip().upper()
        resolved_sheet = self._resolved_sheet_name(sheet_name)
        cache_key = (resolved_sheet, address)
        if cache_key in self._cell_cache:
            return self._cell_cache[cache_key]
        match = EXCEL_COL_RE.match(address)
        if not match:
            raise ValueError(f"Invalid cell address: {address}")
        col_letter, row = match.group(1).upper(), int(match.group(2))
        ws = self._workbook()[resolved_sheet]
        value = ws.cell(
            row=row, column=column_index_from_string(col_letter)
        ).value
        self._cell_cache[cache_key] = value
        return value

    def read_cell_value(self, sheet_name: str, address: str) -> Any:
        delegate = self._delegate_for_sheet(sheet_name)
        if delegate is self:
            return self._read_cell_value_local(sheet_name, address)
        return delegate.read_cell_value(sheet_name, address)

    def _read_raw_sheet(self, sheet_name: str) -> pd.DataFrame:
        """Load a full sheet once per loader instance (shared across column slices)."""
        path = self._path_for_sheet(sheet_name)
        resolved_name = self._resolved_sheet_name(sheet_name)
        cache_key = (path.resolve(), resolved_name)
        if cache_key not in self._raw_sheet_cache:
            logger.debug(
                "Caching raw sheet %s from %s",
                resolved_name,
                path.name,
            )
            self._raw_sheet_cache[cache_key] = pd.read_excel(
                path,
                sheet_name=resolved_name,
                header=None,
                engine="openpyxl",
            )
        return self._raw_sheet_cache[cache_key]

    def read_sheet_columns(
        self,
        sheet_name: str,
        columns: dict[str, str],
        *,
        label: str | None = None,
    ) -> pd.DataFrame:
        """
        Load selected columns by Excel letter.

        columns: logical_name -> column letter (e.g. {"姓名": "C", "考核量": "Y"})
        """
        raw = self._read_raw_sheet(sheet_name)
        out = pd.DataFrame()
        for logical_name, letter in columns.items():
            col_idx = column_index_from_string(letter.upper()) - 1
            out[logical_name] = raw.iloc[:, col_idx]
        out = _log_frame_shape(out, label or f"{self.workbook_path.name}!{sheet_name}")
        return out

    def read_sales_task_sheet(self) -> pd.DataFrame:
        columns: dict[str, str] = {
            "姓名": "C",
            "考核量": "Y",
            "实际销量": "Z",
        }
        raw = self._read_raw_sheet("销售任务及完成率")
        if raw.shape[1] >= column_index_from_string("F"):
            columns["集客达成率"] = "F"
        if raw.shape[1] >= column_index_from_string("AG"):
            columns["合并完成率"] = "AG"
        frame = self.read_sheet_columns(
            "销售任务及完成率",
            columns,
            label="销售任务及完成率",
        )
        frame["姓名"] = frame["姓名"].map(normalize_name)
        frame["考核量"] = pd.to_numeric(frame["考核量"], errors="coerce")
        frame["实际销量"] = pd.to_numeric(frame["实际销量"], errors="coerce")
        if "集客达成率" in frame.columns:
            frame["集客达成率"] = pd.to_numeric(frame["集客达成率"], errors="coerce")
        if "合并完成率" in frame.columns:
            frame["合并完成率"] = pd.to_numeric(frame["合并完成率"], errors="coerce")
        return frame


def lookup_combined_completion_rate(loader: WorkbookLoader, name: Any) -> float | None:
    """Hub BA 合并完成率：销售任务及完成率 AG 列，按姓名（C 列）匹配。"""
    frame = loader.read_sales_task_sheet()
    if "合并完成率" not in frame.columns:
        return None
    key = normalize_name(name)
    matches = frame.loc[frame["姓名"] == key, "合并完成率"]
    if matches.empty:
        return None
    value = matches.iloc[0]
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    return float(value)


def build_workbook_loader(context: dict[str, Any]) -> WorkbookLoader:
    config = context["month_config"]
    sales_path = resolve_project_path(config["workbooks"]["sales"])
    sheet_paths = context.get("sheet_sources")
    if not sheet_paths:
        rel = config.get("workbooks", {}).get("sheet_sources_file")
        if rel:
            from salary_pipeline.ingestion_upload.sheet_merge import load_sheet_sources

            sheet_paths = load_sheet_sources(resolve_project_path(rel))
    if not sheet_paths:
        from salary_pipeline.ingestion_upload.sheet_merge import (
            load_sheet_sources_for_workbook,
        )

        sheet_paths = load_sheet_sources_for_workbook(sales_path)
    from salary_pipeline.ingestion_upload.sheet_merge import supplement_sheet_sources

    sheet_paths = supplement_sheet_sources(sales_path, sheet_paths)
    return WorkbookLoader(sales_path, sheet_paths=sheet_paths)


def workbook_has_sheet(workbook_path: Path, sheet_name: str) -> bool:
    """Return True when workbook_path contains sheet_name."""
    path = Path(workbook_path)
    if not path.is_file():
        return False
    wb = load_workbook(path, read_only=True)
    try:
        return sheet_name in wb.sheetnames
    finally:
        wb.close()


def resolve_parity_golden_workbook(
    month_config: dict[str, Any],
    *,
    sheet_name: str = "提成汇总",
) -> Path | None:
    """
    Workbook containing sheet_name for reconcile / highlight reads.

    Upload-only merged sales workbooks lack 提成汇总; never treat them as golden.
    """
    parity = month_config.get("parity", {})
    for key in (
        "reference_golden_workbook",
        "highlight_golden_workbook",
        "golden_workbook",
    ):
        raw = parity.get(key)
        if not raw:
            continue
        path = resolve_project_path(str(raw))
        if path.is_file() and workbook_has_sheet(path, sheet_name):
            return path
    return resolve_canonical_skeleton_workbook()


def resolve_canonical_skeleton_workbook() -> Path | None:
    """Canonical month workbook with 提成汇总 layout (structure-only reads)."""
    from salary_pipeline.ingestion_upload.default_rules import (
        canonical_month_label,
        load_default_rules,
    )

    cfg = load_default_rules()
    rel = cfg.get("skeleton_reference_workbook")
    if rel:
        path = resolve_project_path(rel)
        if path.is_file() and workbook_has_sheet(path, "提成汇总"):
            return path

    month, _ = canonical_month_label()
    raw_dir = resolve_project_path(f"data/raw/{month}")
    if not raw_dir.is_dir():
        return None
    for candidate in sorted(
        raw_dir.glob("*.xlsx"),
        key=lambda p: p.stat().st_size,
        reverse=True,
    ):
        if workbook_has_sheet(candidate, "提成汇总"):
            return candidate
    return None


def resolve_month_sheet_sources(month_config: dict[str, Any]) -> dict[str, Path]:
    """Load supplemental sheet paths from context or month config."""
    rel = (month_config.get("workbooks") or {}).get("sheet_sources_file")
    if not rel:
        return {}
    from salary_pipeline.ingestion_upload.sheet_merge import load_sheet_sources

    return load_sheet_sources(resolve_project_path(rel))


def is_personnel_workbook(workbook_path: Path) -> bool:
    """True when workbook supplies 人员信息 skeleton keys."""
    path = Path(workbook_path)
    if not path.is_file():
        return False
    if path.name == PERSONNEL_FILENAME:
        return True
    return workbook_has_sheet(path, PERSONNEL_SHEET)


def resolve_personnel_workbook(
    month_config: dict[str, Any],
) -> Path | None:
    """Locate optional 人员信息 upload for skeleton row keys."""
    sheet_sources = resolve_month_sheet_sources(month_config)
    personnel = sheet_sources.get(PERSONNEL_SHEET)
    if personnel is not None and is_personnel_workbook(personnel):
        return personnel

    sales_rel = (month_config.get("workbooks") or {}).get("sales")
    if sales_rel:
        sales_path = resolve_project_path(sales_rel)
        if workbook_has_sheet(sales_path, PERSONNEL_SHEET):
            return sales_path
        if is_personnel_workbook(sales_path):
            return sales_path
        from salary_pipeline.ingestion_upload.sheet_merge import supplement_sheet_sources

        supplemented = supplement_sheet_sources(sales_path, sheet_sources)
        personnel = supplemented.get(PERSONNEL_SHEET)
        if personnel is not None and is_personnel_workbook(personnel):
            return personnel
    return None


def resolve_personnel_sheet_name(
    workbook_path: Path,
    *,
    explicit: str | None = None,
) -> str:
    """Pick the worksheet inside a personnel workbook."""
    if explicit:
        return explicit
    wb = load_workbook(workbook_path, read_only=True)
    try:
        names = list(wb.sheetnames)
    finally:
        wb.close()
    if not names:
        raise ValueError(f"{workbook_path.name} 无工作表")
    resolved = PERSONNEL_SHEET if PERSONNEL_SHEET in names else None
    if resolved is None and workbook_path.name == PERSONNEL_FILENAME:
        return names[0]
    if resolved is None:
        raise ValueError(
            f"{workbook_path.name} 缺少工作表「{PERSONNEL_SHEET}」；"
            f"或将文件命名为 {PERSONNEL_FILENAME}"
        )
    return resolved


def _detect_personnel_header_row(preview: pd.DataFrame) -> tuple[int, bool]:
    """Return (1-based header row, has_named_headers)."""
    for row_idx in range(min(3, len(preview))):
        row_vals = [normalize_header(v) for v in preview.iloc[row_idx]]
        named = {v for v in row_vals if v}
        if set(SUMMARY_KEY_COLUMNS).issubset(named):
            return row_idx + 1, True
    return 1, False


def read_personnel_skeleton_keys(
    workbook_path: Path,
    sheet_name: str | None = None,
    *,
    data_start_row: int = 3,
) -> pd.DataFrame:
    """Read 店别/职务/姓名 row keys from 人员信息 (structure only)."""
    resolved_sheet = resolve_personnel_sheet_name(workbook_path, explicit=sheet_name)
    preview = pd.read_excel(
        workbook_path,
        sheet_name=resolved_sheet,
        header=None,
        nrows=5,
        engine="openpyxl",
    )
    header_row, has_headers = _detect_personnel_header_row(preview)
    if has_headers:
        df = pd.read_excel(
            workbook_path,
            sheet_name=resolved_sheet,
            header=header_row - 1,
            engine="openpyxl",
        )
        df.columns = [normalize_header(c) for c in df.columns]
    else:
        df = pd.read_excel(
            workbook_path,
            sheet_name=resolved_sheet,
            usecols="A:C",
            header=None,
            engine="openpyxl",
        )
        df.columns = list(SUMMARY_KEY_COLUMNS)
    df = summary_frame_from_builder(df)
    df = filter_skeleton_rows(df)
    df["_excel_row"] = range(data_start_row, data_start_row + len(df))
    skeleton = df[[*SUMMARY_KEY_COLUMNS, "_excel_row"]].copy()
    return _log_frame_shape(skeleton, f"personnel skeleton {resolved_sheet}")


def resolve_summary_skeleton_workbook(
    month_config: dict[str, Any],
    *,
    sheet_name: str = "提成汇总",
) -> tuple[Path | None, str]:
    """
    Pick a workbook that contains the 提成汇总 sheet for row-key bootstrap.

    Upload-only merged workbooks often lack this sheet; fall back to the
    canonical reference workbook (structure keys only, no metric values).
    """
    path, source, _ = resolve_summary_skeleton_source(month_config, sheet_name=sheet_name)
    return path, source


def resolve_summary_skeleton_source(
    month_config: dict[str, Any],
    *,
    sheet_name: str = "提成汇总",
) -> tuple[Path | None, str, str]:
    """
    Resolve skeleton row-key source.

    Returns (workbook_path, source_label, read_sheet_name).
    read_sheet_name is 提成汇总 or 人员信息.
    """
    parity = month_config.get("parity", {})
    sales_rel = (month_config.get("workbooks") or {}).get("sales")
    if sales_rel:
        sales_path = resolve_project_path(sales_rel)
        if sales_path.is_file() and workbook_has_sheet(sales_path, sheet_name):
            return sales_path, "sales", sheet_name

    personnel_path = resolve_personnel_workbook(month_config)
    if personnel_path is not None:
        return personnel_path, "personnel", PERSONNEL_SHEET

    for key in (
        "golden_workbook",
        "reference_golden_workbook",
        "highlight_golden_workbook",
    ):
        raw = parity.get(key)
        if not raw:
            continue
        path = resolve_project_path(str(raw))
        if path.is_file() and workbook_has_sheet(path, sheet_name):
            return path, key, sheet_name

    path = resolve_canonical_skeleton_workbook()
    if path is not None:
        return path, "canonical", sheet_name

    return None, "", ""


def read_summary_skeleton_keys(
    workbook_path: Path,
    sheet_name: str = "提成汇总",
    *,
    header_row: int = 2,
    data_start_row: int = 3,
) -> pd.DataFrame:
    """Iteration-1 bootstrap: row keys and Excel row numbers only (no metric columns)."""
    golden = read_golden_summary_sheet(
        workbook_path,
        sheet_name,
        header_row=header_row,
        data_start_row=data_start_row,
    )
    golden = summary_frame_from_builder(golden)
    golden["_excel_row"] = range(data_start_row, data_start_row + len(golden))
    golden = filter_skeleton_rows(golden)
    skeleton = golden[[*SUMMARY_KEY_COLUMNS, "_excel_row"]].copy()
    return _log_frame_shape(skeleton, f"skeleton keys {sheet_name}")


def _log_frame_shape(frame: pd.DataFrame, label: str) -> pd.DataFrame:
    logger.info("Loaded %s shape=%s", label, frame.shape)
    return frame


def _column_sort_key(letter: str) -> int:
    letter = letter.upper()
    value = 0
    for char in letter:
        value = value * 26 + (ord(char) - ord("A") + 1)
    return value
