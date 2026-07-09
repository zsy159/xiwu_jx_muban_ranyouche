"""Read-only scan of golden 绩效整理表 — headers and manual-fill patterns."""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from salary_pipeline.calculators.sales_advisor.topology_specs import (
    is_manual_formula_adjustment,
    is_pure_direct_fill_formula,
)
from salary_pipeline.data_ingestion.data_loader import normalize_name
from salary_pipeline.data_ingestion.performance_sheet_golden import (
    DATA_START_ROW,
    PERF_SHEET,
)
from salary_pipeline.pipelines.performance_sheet_column_sources import (
    is_implemented_perf_column,
)

_TAIL_CONST_RE = re.compile(r"([+\-*]\d+(?:\.\d+)?)\s*\)?\s*$")


@dataclass(frozen=True)
class PerfManualCell:
    """Golden 绩效整理表 cell classified as manual entry (read-only metadata)."""

    row: int
    letter: str
    header: str
    vin: str | None
    advisor: str | None
    pattern: str
    formula: str | None
    detail: str


def _cell_formula(raw: Any) -> str:
    if hasattr(raw, "text"):
        text = str(raw.text).strip()
        return text if text.startswith("=") else ""
    if isinstance(raw, str) and raw.strip().startswith("="):
        return raw.strip()
    return ""


def describe_manual_pattern(formula: str | None, *, plain_value: Any = None) -> tuple[str, str]:
    """Return (pattern_type, detail) for a golden manual cell."""
    if formula:
        if is_manual_formula_adjustment(formula):
            m = _TAIL_CONST_RE.search(formula)
            tail = m.group(1) if m else "?"
            return "公式+尾项", f"尾项{tail}"
        if is_pure_direct_fill_formula(formula):
            if re.match(r"^=\s*[-+]?\d", formula.strip()):
                return "常数公式", formula.strip()
            return "纯算术公式", formula.strip()[:80]
    if plain_value is not None and isinstance(plain_value, (int, float)):
        return "直接填数", f"常数 {plain_value:g}"
    return "直接填数", ""


def load_golden_column_headers(
    golden_path: Path,
    *,
    sheet_name: str = PERF_SHEET,
    header_row: int = 2,
) -> tuple[tuple[str, str], ...]:
    """Ordered (Excel letter, row-2 header text) from golden workbook."""
    wb = load_workbook(golden_path, read_only=True, data_only=False)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return ()
    ws = wb[sheet_name]
    header_cells = next(ws.iter_rows(min_row=header_row, max_row=header_row))
    pairs: list[tuple[str, str]] = []
    for cell in header_cells:
        if cell.value:
            pairs.append((cell.column_letter, str(cell.value).strip()))
    wb.close()
    return tuple(pairs)


def scan_golden_manual_cells(
    golden_path: Path,
    *,
    sheet_name: str = PERF_SHEET,
    header_row: int = 2,
    data_start_row: int = DATA_START_ROW,
    value_columns: frozenset[str] | None = None,
) -> list[PerfManualCell]:
    """Detect manual-fill patterns in golden 绩效整理表 (values not copied)."""
    headers = dict(load_golden_column_headers(golden_path, sheet_name=sheet_name, header_row=header_row))
    if not headers:
        return []

    wb = load_workbook(golden_path, read_only=True, data_only=False)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]
    max_col = max(column_index_from_string(letter) for letter in headers)

    def row_map(row_cells: tuple[Any, ...]) -> dict[str, Any]:
        m: dict[str, Any] = {}
        for i, c in enumerate(row_cells):
            m[get_column_letter(i + 1)] = c
        return m

    out: list[PerfManualCell] = []
    for row_cells in ws.iter_rows(min_row=data_start_row, min_col=1, max_col=max_col):
        cell_map = row_map(row_cells)
        row_idx = row_cells[0].row
        o_cell = cell_map.get("O")
        p_cell = cell_map.get("P")
        vin = str(o_cell.value).strip() if o_cell and o_cell.value else None
        if vin in (None, "", "VIN码", "nan"):
            vin = None
        advisor = normalize_name(p_cell.value) if p_cell else None

        for letter, header in headers.items():
            if value_columns is not None and letter not in value_columns:
                continue
            if letter in (
                "A", "B", "C", "D", "F", "G", "H", "I", "J", "K",
                "M", "N", "O", "P", "Q", "R", "T",
            ):
                continue
            cell = cell_map.get(letter)
            if cell is None:
                continue
            val = cell.value
            formula = _cell_formula(val)
            if val is None or (isinstance(val, str) and not str(val).strip()):
                continue

            pattern: str | None = None
            detail = ""
            if formula:
                if is_manual_formula_adjustment(formula):
                    pattern, detail = describe_manual_pattern(formula)
                elif is_pure_direct_fill_formula(formula):
                    if is_implemented_perf_column(letter, header):
                        continue
                    pattern, detail = describe_manual_pattern(formula)
            elif isinstance(val, (int, float)):
                if is_implemented_perf_column(letter, header):
                    continue
                pattern, detail = describe_manual_pattern(None, plain_value=val)

            if pattern is None:
                continue
            out.append(
                PerfManualCell(
                    row=row_idx,
                    letter=letter,
                    header=header,
                    vin=vin,
                    advisor=advisor,
                    pattern=pattern,
                    formula=formula or None,
                    detail=detail,
                )
            )

    wb.close()
    return out


def golden_headers_by_letter(golden_path: Path) -> dict[str, str]:
    return dict(load_golden_column_headers(golden_path))
