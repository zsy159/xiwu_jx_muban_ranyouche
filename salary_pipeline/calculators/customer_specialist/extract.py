"""从主账套「客户部提成」抽取计算器输入。"""

from __future__ import annotations

from typing import Any

from salary_pipeline.calculators.customer_specialist.types import (
    ActivityRowInput,
    BaokeMetricRow,
    BaokeStoreInput,
    CustomerSpecialistInput,
    LeftLineItemsInput,
    LineItem,
)
from salary_pipeline.data_ingestion.data_loader import WorkbookLoader

_BAOKE_TYPE_BY_LABEL = {
    "电话回访": "phone_callback",
    "基盘客户转介绍": "referral",
    "保客挖掘置换/增购": "mining",
    "全员营销": "all_staff",
}


def _num(value: Any) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _optional_num(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, str) and not value.replace(".", "", 1).isdigit():
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _read_line_items(ws: Any, start: int, end: int) -> list[LineItem]:
    items: list[LineItem] = []
    for row in range(start, end + 1):
        category = str(ws.cell(row, 1).value or "")
        item_name = str(ws.cell(row, 2).value or "")
        if not item_name and not category:
            continue
        items.append(
            LineItem(
                category=category,
                item_name=item_name,
                achievement_rate=_optional_num(ws.cell(row, 3).value),
                coefficient=_num(ws.cell(row, 4).value),
                qty_dengfang=_num(ws.cell(row, 5).value),
                qty_zhangbaozhen=_num(ws.cell(row, 7).value),
            )
        )
    return items


def _read_baoke_metrics(ws: Any, start: int, end: int) -> list[BaokeMetricRow]:
    metrics: list[BaokeMetricRow] = []
    for row in range(start, end + 1):
        label = str(ws.cell(row, 40).value or "")
        metric_type = _BAOKE_TYPE_BY_LABEL.get(label, "all_staff")
        baseline = ws.cell(row, 41).value
        actual = ws.cell(row, 42).value
        improvement = ws.cell(row, 43).value
        flat = _num(ws.cell(row, 46).value) if metric_type == "phone_callback" else 0.0
        metrics.append(
            BaokeMetricRow(
                metric_type=metric_type,
                label=label,
                baseline_rate=_optional_num(baseline),
                actual_rate=_optional_num(actual),
                improvement_pct=_optional_num(improvement),
                delivery_count=_num(ws.cell(row, 44).value),
                flat_amount=flat,
            )
        )
    return metrics


def _read_activity_row(ws: Any, row: int) -> ActivityRowInput:
    return ActivityRowInput(
        prospect_callbacks=_num(ws.cell(row, 12).value),
        five_day_callbacks=_num(ws.cell(row, 13).value),
        thirty_day_callbacks=_num(ws.cell(row, 14).value),
        defeat_callbacks=_num(ws.cell(row, 15).value),
        visit_count=_num(ws.cell(row, 17).value),
        group_chat_count=_num(ws.cell(row, 19).value),
        birthday_count=_num(ws.cell(row, 21).value),
        reputation_posts=_num(ws.cell(row, 23).value),
        complaint_handling=_num(ws.cell(row, 25).value),
        satisfaction_score=_num(ws.cell(row, 26).value),
        satisfaction_bonus=_num(ws.cell(row, 27).value),
        baoke_marketing_flat=_num(ws.cell(row, 28).value),
    )


def extract_role_inputs(loader: WorkbookLoader, role_name: str) -> CustomerSpecialistInput:
    from salary_pipeline.calculators.customer_specialist.registry import get_role

    role = get_role(role_name)
    if role is None:
        raise KeyError(role_name)
    ws = loader._workbook()["客户部提成"]
    template = role["template"]

    if template == "left_line_items":
        person = role.get("person_column", "zhangbaozhen")
        rows = role.get("left_rows", [4, 41])
        return CustomerSpecialistInput(
            template=template,
            left=LeftLineItemsInput(
                person=person,
                line_items=_read_line_items(ws, int(rows[0]), int(rows[1])),
                fixed_vehicle_performance=float(
                    role.get("defaults", {}).get("fixed_vehicle_performance", 0)
                ),
            ),
        )

    if template == "left_and_baoke":
        block = role["baoke_block"]
        mrows = block["metric_rows"]
        return CustomerSpecialistInput(
            template=template,
            left=LeftLineItemsInput(
                person=role.get("person_column", "dengfang"),
                line_items=_read_line_items(ws, 4, 41),
            ),
            baoke=BaokeStoreInput(
                store_label=str(block.get("store_label", "")),
                metrics=_read_baoke_metrics(ws, int(mrows[0]), int(mrows[1])),
            ),
        )

    if template == "activity_summary":
        block = role.get("baoke_block")
        baoke = None
        if block:
            mrows = block["metric_rows"]
            baoke = BaokeStoreInput(
                store_label=str(block.get("store_label", "")),
                metrics=_read_baoke_metrics(ws, int(mrows[0]), int(mrows[1])),
            )
        return CustomerSpecialistInput(
            template=template,
            activity=_read_activity_row(ws, int(role.get("activity_row", 3))),
            baoke=baoke,
        )

    if template == "baoke_store":
        block = role["baoke_block"]
        mrows = block["metric_rows"]
        return CustomerSpecialistInput(
            template=template,
            baoke=BaokeStoreInput(
                store_label=str(block.get("store_label", "")),
                metrics=_read_baoke_metrics(ws, int(mrows[0]), int(mrows[1])),
            ),
        )

    raise ValueError(template)


def lookup_golden_cells(loader: WorkbookLoader, role_name: str) -> dict[str, float]:
    from salary_pipeline.calculators.customer_specialist.registry import get_role

    role = get_role(role_name)
    if not role:
        return {}
    ws = loader._workbook()["客户部提成"]
    out: dict[str, float] = {}
    for label, ref in role.get("golden_cells", {}).items():
        if isinstance(ref, str):
            from openpyxl.utils import column_index_from_string

            col_letters = "".join(ch for ch in ref if ch.isalpha())
            row_digits = "".join(ch for ch in ref if ch.isdigit())
            val = ws.cell(int(row_digits), column_index_from_string(col_letters)).value
        else:
            val = ref
        try:
            out[label] = float(val)
        except (TypeError, ValueError):
            continue
    hub_map = role.get("hub_mapping", {})
    for hub_col, spec in hub_map.items():
        if spec.get("type") == "fixed":
            out[hub_col] = float(spec.get("value", 0))
        elif spec.get("type") == "cell":
            val = ws.cell(int(spec["row"]), int(spec["col"])).value
            try:
                out[hub_col] = float(val)
            except (TypeError, ValueError):
                pass
    return out
