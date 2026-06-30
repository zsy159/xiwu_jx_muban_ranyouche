"""Collect formula-anomaly annotations for 提成汇总 reconcile Excel output."""

from __future__ import annotations

import re
from dataclasses import dataclass, replace
from pathlib import Path
from typing import Any, Iterable

import yaml

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from salary_pipeline.calculators.sales_advisor.registry import (
    is_hub_linked,
    load_role_registry,
    wa_parity_deferred_cells,
    wa_parity_deferred_reasons,
)
from salary_pipeline.calculators.non_frontline.classification import (
    column_mapping_for_tier,
    non_frontline_tier,
)
from salary_pipeline.data_ingestion.data_loader import normalize_name
from salary_pipeline.calculators.sales_advisor.topology_specs import (
    HUB_LETTERS_W_AI,
    _topology_cells,
    hub_column_name,
)
from salary_pipeline.pipelines.hub_formula_engine import HUB_COLUMN_MAP
from salary_pipeline.paths import CONFIG_DIR
from salary_pipeline.validation.parity import CellMismatch

_ANNOTATION_PATH = CONFIG_DIR / "parity_annotation.yaml"
_PARITY_VALUE_TOLERANCE = 1e-6

_F_P_LETTERS = frozenset("FGHIJKLMNOP")
_F_P_COLUMNS = frozenset(hub_column_name(letter) for letter in _F_P_LETTERS)
_W_AI_COLUMNS = frozenset(hub_column_name(letter) for letter in HUB_LETTERS_W_AI)
_COLUMN_TO_LETTER = {name: letter for letter, name in HUB_COLUMN_MAP.items()}

_PERF_COLUMN_LABELS: dict[str, str] = {
    "S": "装饰底价(S)",
    "AG": "单台绩效(AG)",
    "AH": "整车超额(AH)",
    "AI": "加装绩效(AI)",
    "AJ": "保险提成(AJ)",
    "AK": "按揭提成(AK)",
    "AN": "上户(AN)",
    "AS": "上户(AS)",
    "AO": "座位险(AO)",
    "AQ": "交车奖励(AQ)",
    "AR": "保客(AR)",
    "AL": "盈利产品(AL)",
}

_SUMIF_VAL_COL = re.compile(
    r"绩效整理表!(?P<vcol>[A-Z]{1,3}):(?P=vcol)",
    re.IGNORECASE,
)


@dataclass(frozen=True)
class HubCellAnnotation:
    """One annotated Hub cell (姓名 + 列名)."""

    name: str
    column: str
    reason: str
    source: str = "registry"
    golden_value: float | None = None
    computed_value: float | None = None

    def key(self) -> tuple[str, str]:
        return (self.name.strip(), self.column.strip())

    def comment_text(self) -> str:
        lines = [self.reason.strip()]
        if self.golden_value is not None and self.computed_value is not None:
            diff = self.computed_value - self.golden_value
            lines.append(
                f"金标准={self.golden_value:g}  系统={self.computed_value:g}  差={diff:+g}"
            )
        return "\n".join(lines)


def load_annotation_registry(path: Path | None = None) -> list[HubCellAnnotation]:
    cfg_path = path or _ANNOTATION_PATH
    if not cfg_path.exists():
        return []
    with cfg_path.open(encoding="utf-8") as fh:
        data = yaml.safe_load(fh) or {}
    out: list[HubCellAnnotation] = []
    for entry in data.get("hub_annotations") or []:
        name = str(entry.get("name", "")).strip()
        column = str(entry.get("column", "")).strip()
        reason = str(entry.get("reason", "")).strip()
        if name and column and reason:
            out.append(HubCellAnnotation(name=name, column=column, reason=reason))
    return out


def _deferred_as_annotations(registry: dict[str, Any] | None = None) -> list[HubCellAnnotation]:
    reg = registry or load_role_registry()
    out: list[HubCellAnnotation] = []
    for entry in reg.get("wa_parity_deferred") or []:
        name = str(entry.get("name", "")).strip()
        cols = entry.get("columns") or {}
        for column, reason in cols.items():
            column = str(column).strip()
            reason = str(reason).strip()
            if name and column and reason:
                out.append(
                    HubCellAnnotation(
                        name=name,
                        column=column,
                        reason=reason,
                        source="deferred",
                    )
                )
    return out


def golden_hub_names_by_row(
    golden_workbook: Path,
    *,
    sheet: str = "提成汇总",
    data_start_row: int = 3,
) -> dict[int, str]:
    """Map Hub 提成汇总 Excel row → 姓名 (column D) from the golden workbook."""
    wb = load_workbook(golden_workbook, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb[sheet]
    name_col = column_index_from_string("D")
    out: dict[int, str] = {}
    for row in range(data_start_row, ws.max_row + 1):
        name = normalize_name(ws.cell(row=row, column=name_col).value)
        if name:
            out[row] = name
    wb.close()
    return out


def _parity_values_match(
    golden: float | None,
    computed: float | None,
    *,
    tolerance: float = _PARITY_VALUE_TOLERANCE,
) -> bool:
    if golden is None and computed is None:
        return True
    if golden is None or computed is None:
        return False
    return abs(golden - computed) <= tolerance


def detect_topology_formula_anomalies(
    registry: dict[str, Any] | None = None,
    *,
    golden_workbook: Path | None = None,
    golden_data_start_row: int = 3,
) -> list[HubCellAnnotation]:
    """Scan golden Hub topology for broken refs (tail constants → blue deferred)."""
    reg = registry or load_role_registry()
    cells = _topology_cells()
    advisors = [
        (str(r["name"]).strip(), int(r["hub_excel_row"]))
        for r in reg.get("roles") or []
        if is_hub_linked(r) and r.get("hub_excel_row")
    ]
    if not advisors:
        return []

    row_names: dict[int, str] | None = None
    if golden_workbook is not None:
        row_names = golden_hub_names_by_row(
            golden_workbook,
            data_start_row=golden_data_start_row,
        )

    out: list[HubCellAnnotation] = []
    for letter in HUB_LETTERS_W_AI:
        column = hub_column_name(letter)
        for name, excel_row in advisors:
            if row_names is not None and row_names.get(excel_row) != name:
                continue
            key = f"提成汇总!{letter}{excel_row}"
            cell = cells.get(key, {})
            formula = str(cell.get("formula") or "").strip()

            reason: str | None = None
            if formula and "#REF!" in formula.upper():
                reason = (
                    f"金标准 {column} 公式含 #REF! 断裂引用：{formula[:100]}"
                )

            if reason:
                out.append(
                    HubCellAnnotation(
                        name=name,
                        column=column,
                        reason=reason,
                        source="topology",
                    )
                )
    return out


def collect_hub_cell_annotations(
    *,
    registry: dict[str, Any] | None = None,
    parity_values: dict[tuple[str, str], tuple[float | None, float | None]] | None = None,
    include_topology: bool = True,
    include_deferred: bool = False,
    golden_workbook: Path | None = None,
    golden_data_start_row: int = 3,
) -> list[HubCellAnnotation]:
    """Merge YAML registry + optional topology auto-detect; enrich with parity diff."""
    merged: dict[tuple[str, str], HubCellAnnotation] = {}

    for ann in load_annotation_registry():
        merged[ann.key()] = ann

    if include_topology:
        for ann in detect_topology_formula_anomalies(
            registry,
            golden_workbook=golden_workbook,
            golden_data_start_row=golden_data_start_row,
        ):
            merged.setdefault(ann.key(), ann)

    if include_deferred:
        for ann in _deferred_as_annotations(registry):
            merged.setdefault(ann.key(), ann)

    if parity_values:
        enriched: dict[tuple[str, str], HubCellAnnotation] = {}
        for key, ann in merged.items():
            golden, computed = parity_values.get(key, (None, None))
            if golden is not None or computed is not None:
                enriched[key] = replace(
                    ann,
                    golden_value=golden,
                    computed_value=computed,
                )
            else:
                enriched[key] = ann
        merged = enriched

    return list(merged.values())


def parity_values_for_annotations(
    computed_path: Path,
    golden_workbook: Path,
    golden_sheet: str,
    annotation_keys: Iterable[tuple[str, str]],
    *,
    join_keys: list[str] | None = None,
    header_row: int = 2,
    data_start_row: int = 3,
    golden_header_row: int | None = None,
    golden_data_start_row: int | None = None,
) -> dict[tuple[str, str], tuple[float | None, float | None]]:
    """Load golden vs computed values for annotated (姓名, 列) pairs."""
    from salary_pipeline.data_ingestion.data_loader import (
        read_computed_summary_excel,
        read_golden_summary_sheet,
        summary_frame_from_builder,
    )
    from salary_pipeline.validation.parity import filter_comparable_rows

    keys = list(annotation_keys)
    if not keys:
        return {}

    join_keys = join_keys or ["店别", "职务", "姓名"]
    columns = sorted({col for _, col in keys})
    computed = filter_comparable_rows(
        summary_frame_from_builder(
            read_computed_summary_excel(
                computed_path,
                header_row=header_row,
                data_start_row=data_start_row,
            )
        )
    )
    golden = filter_comparable_rows(
        summary_frame_from_builder(
            read_golden_summary_sheet(
                golden_workbook,
                golden_sheet,
                header_row=golden_header_row if golden_header_row is not None else header_row,
                data_start_row=(
                    golden_data_start_row
                    if golden_data_start_row is not None
                    else data_start_row
                ),
            )
        )
    )
    merged = golden.merge(
        computed,
        on=join_keys,
        how="inner",
        suffixes=("_golden", "_computed"),
    )
    if merged.empty:
        return {}

    out: dict[tuple[str, str], tuple[float | None, float | None]] = {}
    name_col = "姓名_golden" if "姓名_golden" in merged.columns else "姓名"
    for name, col in keys:
        g_col = f"{col}_golden" if f"{col}_golden" in merged.columns else col
        c_col = f"{col}_computed" if f"{col}_computed" in merged.columns else col
        if g_col not in merged.columns or c_col not in merged.columns:
            continue
        rows = merged[merged[name_col].astype(str).str.strip() == name.strip()]
        if rows.empty:
            continue
        row = rows.iloc[0]
        g_val = row[g_col]
        c_val = row[c_col]
        try:
            g_num = float(g_val) if g_val == g_val else None  # NaN check
        except (TypeError, ValueError):
            g_num = None
        try:
            c_num = float(c_val) if c_val == c_val else None
        except (TypeError, ValueError):
            c_num = None
        out[(name, col)] = (g_num, c_num)
    return out


def annotations_for_workbook(
    *,
    registry: dict[str, Any] | None = None,
    parity_values: dict[tuple[str, str], tuple[float | None, float | None]] | None = None,
    deferred_cells: dict[str, frozenset[str]] | None = None,
    golden_workbook: Path | None = None,
    golden_data_start_row: int = 3,
) -> list[HubCellAnnotation]:
    """Annotations for reconcile: registry + topology; skip deferred / manual-formula cells."""
    reg = registry or load_role_registry()
    deferred = deferred_cells if deferred_cells is not None else wa_parity_deferred_cells(reg)
    annotations = collect_hub_cell_annotations(
        registry=reg,
        parity_values=parity_values,
        include_topology=True,
        include_deferred=False,
        golden_workbook=golden_workbook,
        golden_data_start_row=golden_data_start_row,
    )
    filtered: list[HubCellAnnotation] = []
    for ann in annotations:
        if ann.column in deferred.get(ann.name, frozenset()):
            continue
        if ann.source == "topology" and parity_values:
            golden, computed = parity_values.get(ann.key(), (None, None))
            if _parity_values_match(golden, computed):
                continue
        filtered.append(ann)
    return filtered


def _join_lookup(
    mismatch: CellMismatch,
) -> tuple[str | None, Any, Any]:
    joined = mismatch.join_dict()
    name = normalize_name(joined.get("姓名"))
    shop = joined.get("店别")
    role = joined.get("职务")
    return name, shop, role


def _hub_excel_row(name: str, registry: dict[str, Any]) -> int | None:
    for role in registry.get("roles") or []:
        if str(role.get("name", "")).strip() == name.strip():
            row = role.get("hub_excel_row")
            if row:
                return int(row)
    return None


def _perf_column_label(col: str) -> str:
    return _PERF_COLUMN_LABELS.get(col.upper(), f"绩效整理表 {col.upper()} 列")


def _extract_sumif_value_column(formula: str) -> str | None:
    matches = list(_SUMIF_VAL_COL.finditer(formula.replace(" ", "")))
    if len(matches) >= 2:
        return matches[-1].group("vcol").upper()
    if matches:
        return matches[0].group("vcol").upper()
    return None


def _describe_golden_topology_formula(
    name: str,
    column: str,
    *,
    registry: dict[str, Any],
    golden_workbook: Path | None = None,
    golden_data_start_row: int = 3,
) -> str | None:
    row = _hub_excel_row(name, registry)
    letter = _COLUMN_TO_LETTER.get(column)
    if row is None or letter is None:
        return None

    if golden_workbook is not None:
        row_names = golden_hub_names_by_row(
            golden_workbook,
            data_start_row=golden_data_start_row,
        )
        actual = row_names.get(row)
        if actual and actual != name.strip():
            return (
                f"hub_excel_row 登记行 {row} 对应姓名「{actual}」，"
                f"与当前「{name}」不一致，topology 公式可能错位"
            )

    key = f"提成汇总!{letter}{row}"
    cell = _topology_cells().get(key, {})
    formula = str(cell.get("formula") or "").strip()
    if not formula:
        return "金标准直接填数，系统按公式/数据源重算"

    formula_upper = formula.upper()
    if "#REF!" in formula_upper:
        val_col = _extract_sumif_value_column(formula)
        if val_col and "绩效整理表" in formula:
            label = _perf_column_label(val_col)
            snippet = formula[:90] + ("…" if len(formula) > 90 else "")
            return (
                f"金标准公式含 #REF! 断裂引用（{snippet}）；"
                f"系统按姓名汇总{label}，与金标准 Excel 求值不一致"
            )
        snippet = formula[:100] + ("…" if len(formula) > 100 else "")
        return f"金标准公式含 #REF! 断裂引用：{snippet}"

    if "绩效整理表" in formula and "SUMIF" in formula_upper:
        val_col = _extract_sumif_value_column(formula)
        if val_col:
            label = _perf_column_label(val_col)
            return (
                f"金标准 SUMIF 汇总{label} 与系统数据源/语义不一致"
            )

    if column in _W_AI_COLUMNS and "SUMIFS" in formula_upper:
        m = re.search(
            r"绩效整理表!(?P<vcol>[A-Z]{1,3}):",
            formula,
            re.IGNORECASE,
        )
        if m:
            label = _perf_column_label(m.group("vcol").upper())
            return f"W–AI 绩效层：金标准 SUMIFS({label}) 与系统汇总源不一致"

    if column in _F_P_COLUMNS:
        snippet = formula[:80] + ("…" if len(formula) > 80 else "")
        return f"F–P 验收层：金标准 Hub 公式（{snippet}）与系统回放不一致"

    return None


def _non_frontline_mismatch_note(
    shop: Any,
    role: Any,
    display_column: str,
    *,
    config: dict[str, Any] | None = None,
) -> str | None:
    tier = non_frontline_tier(shop, role, config=config)
    if tier is None:
        return None
    mapping = column_mapping_for_tier(tier, config=config)
    for physical, semantic in mapping.items():
        if semantic == display_column and physical != semantic:
            return f"非一线列映射：金标准 {physical} ↔ 系统 {semantic}"
    return None


def lookup_mismatch_root_cause(
    mismatch: CellMismatch,
    *,
    registry: dict[str, Any] | None = None,
    golden_workbook: Path | None = None,
    golden_data_start_row: int = 3,
) -> str:
    """Resolve categorized root-cause text for one amber mismatch cell."""
    reg = registry or load_role_registry()
    name, shop, role = _join_lookup(mismatch)
    column = mismatch.column

    if name:
        for ann in load_annotation_registry():
            if ann.name == name and ann.column == column:
                return ann.reason.strip()
        deferred_reason = wa_parity_deferred_reasons(reg).get(name, {}).get(column)
        if deferred_reason:
            return deferred_reason.strip()

    if name:
        topo = _describe_golden_topology_formula(
            name,
            column,
            registry=reg,
            golden_workbook=golden_workbook,
            golden_data_start_row=golden_data_start_row,
        )
        if topo:
            return topo

    nf_note = _non_frontline_mismatch_note(shop, role, column)
    if nf_note:
        return nf_note

    if column in _F_P_COLUMNS:
        return f"F–P 验收层：Hub 公式回放与金标准不一致（列 {column}）"
    if column in _W_AI_COLUMNS:
        return f"W–AI 绩效层：SUMIFS 来源与金标准公式不一致（列 {column}）"
    return "数值不一致，根因待登记（见 parity_report 或 parity_annotation.yaml）"


def enrich_cell_mismatches(
    mismatches: Iterable[CellMismatch],
    *,
    registry: dict[str, Any] | None = None,
    golden_workbook: Path | None = None,
    golden_data_start_row: int = 3,
) -> list[CellMismatch]:
    """Attach ``root_cause`` to parity mismatch cells for Excel comments."""
    reg = registry or load_role_registry()
    out: list[CellMismatch] = []
    for mismatch in mismatches:
        if mismatch.root_cause:
            out.append(mismatch)
            continue
        cause = lookup_mismatch_root_cause(
            mismatch,
            registry=reg,
            golden_workbook=golden_workbook,
            golden_data_start_row=golden_data_start_row,
        )
        out.append(replace(mismatch, root_cause=cause))
    return out
