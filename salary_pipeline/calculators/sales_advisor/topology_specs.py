"""从 Hub topology 解析销售顾问 W–AI 公式规格。"""

from __future__ import annotations

import json
import re
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from salary_pipeline.calculators.non_frontline.classification import (
    column_mapping_for_tier,
    load_non_frontline_config,
)
from salary_pipeline.calculators.sales_advisor.types import HubColumnFormula
from salary_pipeline.data_ingestion.data_loader import normalize_header, normalize_name
from salary_pipeline.pipelines.hub_formula_engine import HUB_COLUMN_MAP, HUB_SHEET
from salary_pipeline.paths import CONFIG_DIR, resolve_project_path
from salary_pipeline.pipelines.commission_summary import load_month_config

HUB_LETTERS_W_AI = (
    "W",
    "X",
    "Y",
    "Z",
    "AA",
    "AB",
    "AC",
    "AD",
    "AE",
    "AF",
    "AG",
    "AH",
    "AI",
)

SUMIFS_PERF = re.compile(
    r"^=SUMIFS\("
    r"绩效整理表!(?P<vcol>[A-Z]{1,3}):(?P=vcol),"
    r"绩效整理表!P:P,(?:提成汇总!)?D(?P<row>\d+)"
    r"(?:,绩效整理表!H:H,\"<>(?P<exclude>[^\"]+)\")?"
    r"\)"
    r"(?:\*(?P<mul>[A-Z]+\d+))?"
    r"(?:\+(?P<add>-?\d+(?:\.\d+)?))?$",
    re.IGNORECASE,
)
SUMIF_SINGLE = re.compile(
    r"^=SUMIF\("
    r"绩效整理表!(?P<kcol>[A-Z]{1,3}):(?P=kcol),"
    r"(?:提成汇总!)?(?P<crit>D\d+|[A-Z]+\d+),"
    r"(?:'绩效整理表'|绩效整理表)!"
    r"(?P<vcol>[A-Z]{1,3}):(?P=vcol)\)$",
    re.IGNORECASE,
)
SUMIF_CHAIN = re.compile(
    r"^=SUMIF\(.+\+SUMIF\(.+\)$",
    re.IGNORECASE,
)

_HUB_NAME_BY_LETTER = {letter: name for letter, name in HUB_COLUMN_MAP.items()}


@lru_cache(maxsize=1)
def _topology_cells() -> dict[str, dict[str, Any]]:
    config = load_month_config(CONFIG_DIR)
    topo_path = resolve_project_path(config["topology"]["sales"])
    topo = json.loads(topo_path.read_text(encoding="utf-8"))
    return topo.get("cells", {})


def hub_column_name(letter: str) -> str:
    return _HUB_NAME_BY_LETTER[letter.upper()]


def parse_hub_formula(formula: str, *, hub_letter: str) -> HubColumnFormula | None:
    """将 Excel 公式解析为可复算结构。"""
    formula = formula.strip()
    col_name = hub_column_name(hub_letter)

    m = SUMIFS_PERF.match(formula)
    if m:
        return HubColumnFormula(
            hub_column=col_name,
            kind="sumifs",
            perf_columns=(m.group("vcol").upper(),),
            multiply_ref=m.group("mul").upper() if m.group("mul") else None,
            add_const=float(m.group("add") or 0),
            exclude_vehicle=m.group("exclude"),
        )

    if SUMIF_CHAIN.match(formula):
        parts = re.split(r"\+(?=SUMIF\()", formula[1:], flags=re.IGNORECASE)
        cols: list[str] = []
        for part in parts:
            part = part.strip()
            if part.upper().startswith("SUMIF("):
                part = part[6:].rstrip(")")
            m_single = re.match(
                r"绩效整理表![A-Z]{1,3}:[A-Z]{1,3},"
                r"(?:提成汇总!)?D\d+,"
                r"绩效整理表!(?P<vcol>[A-Z]{1,3}):",
                part,
                re.IGNORECASE,
            )
            if m_single:
                cols.append(m_single.group("vcol").upper())
        if cols:
            return HubColumnFormula(
                hub_column=col_name,
                kind="sumif_chain",
                perf_columns=tuple(cols),
            )

    m = SUMIF_SINGLE.match(formula)
    if m:
        crit = m.group("crit").upper()
        criteria_ref = crit if re.match(r"[A-Z]+\d+", crit) and not crit.startswith("D") else None
        return HubColumnFormula(
            hub_column=col_name,
            kind="sumif",
            perf_columns=(m.group("vcol").upper(),),
            sumif_key_col=m.group("kcol").upper(),
            sumif_criteria_ref=criteria_ref,
        )

    return None



_SECTION_LABEL_STRINGS = frozenset(
    {
        "岗位分值",
        "业绩分值",
        "岗位绩效",
        "业绩绩效",
        "业绩绩效1",
        "业绩绩效2",
        "新能源专项",
        "配件外销",
        "售后总产值",
    }
)


def _load_topology_cells(topology_path: Path | None = None) -> dict[str, dict[str, Any]]:
    if topology_path is not None:
        topo = json.loads(topology_path.read_text(encoding="utf-8"))
        return topo.get("cells", {})
    return _topology_cells()


_PURE_CONST_FORMULA_RE = re.compile(r"^=\s*[-+]?\d+(?:\.\d+)?\s*$", re.IGNORECASE)
_MANUAL_TAIL_ARITH_RE = re.compile(r"[+\-]\d+(?:\.\d+)?\s*\)?\s*$")
_PURE_ARITH_FORMULA_RE = re.compile(
    r"^=\s*\(?\s*[-+]?\d+(?:\.\d+)?(?:\s*[-+*/]\s*\(?\s*[-+]?\d+(?:\.\d+)?\s*\)?)*\s*\)?\s*$",
    re.IGNORECASE,
)
_SHEET_REF_RE = re.compile(r"(?:'[^']+'|[^'!\s,()]+)!", re.IGNORECASE)
_CELL_REF_RE = re.compile(r"(?<![A-Z])\b[A-Z]{1,3}\d+\b", re.IGNORECASE)
_COL_RANGE_REF_RE = re.compile(r"\b[A-Z]{1,3}:[A-Z]{1,3}\b", re.IGNORECASE)


def _strip_formula_string_literals(formula: str) -> str:
    return re.sub(r'"[^"]*"', "", formula)


def _formula_has_cell_or_range_reference(formula: str) -> bool:
    text = _strip_formula_string_literals(formula)
    if _SHEET_REF_RE.search(text):
        return True
    if _CELL_REF_RE.search(text):
        return True
    if _COL_RANGE_REF_RE.search(text):
        return True
    return False


def is_pure_direct_fill_formula(formula: str) -> bool:
    """True for golden formulas that are only constants or pure arithmetic.

    Examples: ``=100``, ``=-140``, ``=189*20``, ``=462+101+77+102``.
    These are treated as 金标准直接填数 (gray), not formula-with-manual (blue).
    """
    text = formula.strip()
    if not text.startswith("="):
        return False
    if _PURE_CONST_FORMULA_RE.match(text):
        return True
    if not _formula_has_cell_or_range_reference(text) and _PURE_ARITH_FORMULA_RE.match(
        text
    ):
        return True
    return False


def is_manual_formula_adjustment(formula: str) -> bool:
    """True when golden formula embeds a manual numeric constant on a real formula.

    Covers trailing arithmetic such as ``=SUMIFS(...)-100`` or ``=AH80-100``.
    Pure constants / pure arithmetic without references are static fills, not deferred.
    """
    text = formula.strip()
    if not text.startswith("="):
        return False
    if is_pure_direct_fill_formula(text):
        return False
    if "#REF!" in text.upper():
        return False
    if _MANUAL_TAIL_ARITH_RE.search(text):
        return True
    return False


def _cell_formula_text(raw_value: Any, topo_cell: dict[str, Any]) -> str:
    formula = str(topo_cell.get("formula") or "").strip()
    if formula:
        return formula
    if isinstance(raw_value, str) and raw_value.strip().startswith("="):
        return raw_value.strip()
    if hasattr(raw_value, "text"):
        return str(raw_value.text).strip()
    return ""


def _cell_has_formula(raw_value: Any, topo_cell: dict[str, Any]) -> bool:
    if str(topo_cell.get("formula") or "").strip():
        return True
    if raw_value is None:
        return False
    if hasattr(raw_value, "text"):
        return True
    if isinstance(raw_value, str) and raw_value.strip().startswith("="):
        return True
    return False


def _is_direct_fill_value(value: Any) -> bool:
    if value is None or isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return not (isinstance(value, float) and value != value)
    if isinstance(value, str):
        text = value.strip()
        if not text or text in _SECTION_LABEL_STRINGS:
            return False
        if text.startswith("="):
            return is_pure_direct_fill_formula(text)
        return True
    return False


def _static_fill_allowed_columns() -> frozenset[str]:
    """Hub W–AI / F–P columns plus non-frontline physical columns (M–U, etc.)."""
    cols = set(HUB_COLUMN_MAP.values())
    cfg = load_non_frontline_config()
    for tier in ("management", "support"):
        cols.update(column_mapping_for_tier(tier, config=cfg).keys())
    return frozenset(cols)


def _resolve_scan_column_index(
    col_name: str,
    col_map: dict[str, int],
    letter_by_column: dict[str, str],
) -> int | None:
    if col_name in col_map:
        return col_map[col_name]
    letter = letter_by_column.get(col_name)
    if letter is None:
        return None
    return column_index_from_string(letter)


def _topo_key_for_column(
    sheet_name: str,
    row_idx: int,
    col_name: str,
    col_idx: int,
    letter_by_column: dict[str, str],
) -> str:
    letter = letter_by_column.get(col_name) or get_column_letter(col_idx)
    return f"{sheet_name}!{letter}{row_idx}"


def _cell_needs_manual_fill(topo_cell: dict[str, Any]) -> bool:
    """True when topology has no evaluable formula (manual entry required)."""
    formula = str(topo_cell.get("formula") or "").strip()
    if not formula:
        return True
    return is_pure_direct_fill_formula(formula)


def _cell_is_golden_static_fill(raw_value: Any, topo_cell: dict[str, Any]) -> bool:
    """Topology-only: manual fill when no evaluable formula (golden value ignored)."""
    return _cell_needs_manual_fill(topo_cell)


def collect_topology_static_fill_cells(
    *,
    topology_path: Path | None = None,
    golden_workbook_path: Path | None = None,
    sheet_name: str = HUB_SHEET,
    header_row: int = 2,
    data_start_row: int = 3,
    data_columns: frozenset[str] | None = None,
    letter_by_column: dict[str, str] | None = None,
) -> dict[tuple[str, str], frozenset[str]]:
    """Detect 提成汇总 cells that require manual entry (topology-only).

    Marks cells with no evaluable formula or pure constant formulas. Does not
    copy golden values into computed output — only drives gray highlight.
    """
    config = load_month_config(CONFIG_DIR)
    topo_path = topology_path or resolve_project_path(config["topology"]["sales"])
    golden_path = golden_workbook_path or resolve_project_path(
        config["workbooks"]["sales"]
    )
    topo_cells = _load_topology_cells(topo_path)
    allowed_columns = data_columns or _static_fill_allowed_columns()

    wb = load_workbook(golden_path, data_only=False, read_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb[sheet_name]

    col_map: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        header = normalize_header(ws.cell(row=header_row, column=col_idx).value)
        if header:
            col_map[header] = col_idx

    name_col = col_map.get("姓名")
    role_col = col_map.get("职务")
    if name_col is None or role_col is None:
        wb.close()
        return {}

    letter_map = letter_by_column or {
        name: letter for letter, name in HUB_COLUMN_MAP.items()
    }
    out: dict[tuple[str, str], set[str]] = {}

    for row_idx in range(data_start_row, ws.max_row + 1):
        name = normalize_name(ws.cell(row=row_idx, column=name_col).value)
        if name is None or name == "空白":
            continue
        role = normalize_name(ws.cell(row=row_idx, column=role_col).value) or ""

        for col_name in allowed_columns:
            col_idx = _resolve_scan_column_index(col_name, col_map, letter_map)
            if col_idx is None:
                continue
            raw = ws.cell(row=row_idx, column=col_idx).value
            topo_key = _topo_key_for_column(
                sheet_name, row_idx, col_name, col_idx, letter_map
            )
            if not _cell_is_golden_static_fill(raw, topo_cells.get(topo_key, {})):
                continue
            out.setdefault((name, role), set()).add(col_name)

    wb.close()
    return {key: frozenset(cols) for key, cols in out.items()}


def collect_topology_manual_formula_cells(
    *,
    topology_path: Path | None = None,
    golden_workbook_path: Path | None = None,
    sheet_name: str = HUB_SHEET,
    header_row: int = 2,
    data_start_row: int = 3,
    data_columns: frozenset[str] | None = None,
) -> dict[tuple[str, str], frozenset[str]]:
    """Detect golden 提成汇总 cells whose formula embeds manual numeric entry."""
    config = load_month_config(CONFIG_DIR)
    topo_path = topology_path or resolve_project_path(config["topology"]["sales"])
    golden_path = golden_workbook_path or resolve_project_path(
        config["workbooks"]["sales"]
    )
    topo_cells = _load_topology_cells(topo_path)
    allowed_columns = data_columns or _static_fill_allowed_columns()

    wb = load_workbook(golden_path, data_only=False, read_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb[sheet_name]

    col_map: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        header = normalize_header(ws.cell(row=header_row, column=col_idx).value)
        if header:
            col_map[header] = col_idx

    name_col = col_map.get("姓名")
    role_col = col_map.get("职务")
    if name_col is None or role_col is None:
        wb.close()
        return {}

    letter_by_column = {name: letter for letter, name in HUB_COLUMN_MAP.items()}
    out: dict[tuple[str, str], set[str]] = {}

    for row_idx in range(data_start_row, ws.max_row + 1):
        name = normalize_name(ws.cell(row=row_idx, column=name_col).value)
        if name is None or name == "空白":
            continue
        role = normalize_name(ws.cell(row=row_idx, column=role_col).value) or ""

        for col_name in allowed_columns:
            col_idx = _resolve_scan_column_index(col_name, col_map, letter_by_column)
            if col_idx is None:
                continue
            raw = ws.cell(row=row_idx, column=col_idx).value
            topo_key = _topo_key_for_column(
                sheet_name, row_idx, col_name, col_idx, letter_by_column
            )
            topo_cell = topo_cells.get(topo_key, {})
            if not _cell_has_formula(raw, topo_cell):
                continue
            formula = _cell_formula_text(raw, topo_cell)
            if is_pure_direct_fill_formula(formula):
                continue
            if not is_manual_formula_adjustment(formula):
                continue
            out.setdefault((name, role), set()).add(col_name)

    wb.close()
    return {key: frozenset(cols) for key, cols in out.items()}


def load_row_specs(excel_row: int) -> dict[str, HubColumnFormula]:
    """按 Hub Excel 行号加载 W–AI 公式规格。"""
    cells = _topology_cells()
    specs: dict[str, HubColumnFormula] = {}
    for letter in HUB_LETTERS_W_AI:
        key = f"提成汇总!{letter}{excel_row}"
        formula = cells.get(key, {}).get("formula", "")
        if not formula:
            continue
        spec = parse_hub_formula(formula, hub_letter=letter)
        if spec is not None:
            specs[spec.hub_column] = spec
    return specs
