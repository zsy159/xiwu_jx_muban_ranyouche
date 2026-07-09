"""Upload intake: validate, scan sheets, match manifest, stage files."""

from __future__ import annotations

import io
import shutil
import zipfile
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import BinaryIO

from openpyxl import load_workbook

from salary_pipeline.ingestion_upload.manifest import (
    RequiredSheet,
    build_required_sheet_manifest,
    filename_implies_sheet,
    is_mandatory_input,
    normalize_sheet_name,
    required_input_sheets,
    resolve_sheet_alias,
    resolve_sheet_from_upload,
)
from salary_pipeline.modules.base import PERSONNEL_FILENAME, PERSONNEL_SHEET
from salary_pipeline.paths import PROJECT_ROOT, raw_month_dir

MAX_FILE_BYTES = 80 * 1024 * 1024  # 80 MB per file
MAX_TOTAL_BYTES = 200 * 1024 * 1024
ALLOWED_SUFFIXES = {".xlsx", ".zip"}

# Streamlit file_uploader accept list: extensions + MIME types for macOS Finder.
STREAMLIT_UPLOAD_ACCEPT = [
    "xlsx",
    "zip",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/zip",
    "application/x-zip-compressed",
]


class SheetMatchStatus(str, Enum):
    READY = "ready"
    MISSING = "missing"
    CONFLICT = "conflict"
    NOTE = "note"  # optional formula sheet present


@dataclass
class UploadedFile:
    filename: str
    path: Path
    size_bytes: int
    sheet_names: list[str] = field(default_factory=list)


@dataclass
class SheetMatch:
    required: RequiredSheet
    status: SheetMatchStatus
    sources: list[str] = field(default_factory=list)
    resolved_name: str | None = None
    detail: str | None = None


@dataclass
class IntakeResult:
    month_id: str
    staging_dir: Path
    uploads: list[UploadedFile]
    matches: list[SheetMatch]
    sales_workbook: Path | None = None
    rules_workbook: Path | None = None
    consolidated_workbook: Path | None = None
    sheet_sources_path: Path | None = None
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def all_required_ready(self) -> bool:
        required = [m for m in self.matches if is_mandatory_input(m.required)]
        return all(m.status == SheetMatchStatus.READY for m in required)

    @property
    def missing_role_family_sheets(self) -> list[str]:
        return [
            m.required.name
            for m in self.matches
            if m.required.optional_role_family
            and m.status == SheetMatchStatus.MISSING
        ]

    @property
    def missing_sheets(self) -> list[str]:
        return [
            m.required.name
            for m in self.matches
            if is_mandatory_input(m.required) and m.status == SheetMatchStatus.MISSING
        ]

    @property
    def conflict_sheets(self) -> list[str]:
        return [
            m.required.name
            for m in self.matches
            if m.status == SheetMatchStatus.CONFLICT
        ]

    def can_proceed(
        self, conflict_resolutions: dict[str, str] | None = None
    ) -> bool:
        """True when no required sheet is missing and every conflict has a source."""
        return not self.proceed_blockers(conflict_resolutions)

    def proceed_blockers(
        self, conflict_resolutions: dict[str, str] | None = None
    ) -> list[str]:
        """Human-readable reasons merge/trial buttons stay disabled."""
        conflict_resolutions = conflict_resolutions or {}
        blockers: list[str] = []
        missing = self.missing_sheets
        if missing:
            preview = "、".join(missing[:5])
            if len(missing) > 5:
                preview += "…"
            blockers.append(f"尚有 {len(missing)} 张必需表缺失（{preview}）")
        unresolved = [
            name
            for name in self.conflict_sheets
            if not conflict_resolutions.get(name)
        ]
        if unresolved:
            preview = "、".join(unresolved[:5])
            if len(unresolved) > 5:
                preview += "…"
            blockers.append(
                f"请先为 {len(unresolved)} 处冲突工作表选择来源（{preview}）"
            )
        return blockers


def display_match_icon(
    match: SheetMatch,
    conflict_resolutions: dict[str, str] | None = None,
) -> str:
    """UI status glyph; optional role-family sheets show warning when missing."""
    status, _ = display_match_status(match, conflict_resolutions)
    if match.required.optional_role_family and status == SheetMatchStatus.MISSING:
        return "⚠️"
    icons = {
        SheetMatchStatus.READY: "✅",
        SheetMatchStatus.MISSING: "⬜",
        SheetMatchStatus.CONFLICT: "⚠️",
        SheetMatchStatus.NOTE: "ℹ️",
    }
    return icons.get(status, "•")


def display_match_status(
    match: SheetMatch,
    conflict_resolutions: dict[str, str] | None = None,
) -> tuple[SheetMatchStatus, list[str]]:
    """Status/sources for UI after the user picks a conflict source."""
    conflict_resolutions = conflict_resolutions or {}
    if match.status == SheetMatchStatus.CONFLICT:
        chosen = conflict_resolutions.get(match.required.name)
        if chosen and chosen in match.sources:
            return SheetMatchStatus.READY, [chosen]
    return match.status, list(match.sources)


def preferred_conflict_source_index(
    sources: list[str],
    *,
    sales_workbook: Path | None = None,
) -> int:
    """Prefer the sales workbook when multiple uploads contain the same sheet."""
    if sales_workbook is not None:
        sales_name = sales_workbook.name
        if sales_name in sources:
            return sources.index(sales_name)
    for idx, name in enumerate(sources):
        lower = name.lower()
        if "售后" in name:
            continue
        if "销售" in name or "西物" in lower:
            return idx
    return 0


def scan_workbook_sheets(path: Path) -> list[str]:
    """Read sheet names with openpyxl read_only."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def _validate_upload(name: str, data: bytes) -> str | None:
    suffix = Path(name).suffix.lower()
    if suffix not in ALLOWED_SUFFIXES:
        return f"{name}: 仅支持 .xlsx / .zip"
    if len(data) == 0:
        return f"{name}: 文件为空（0 字节），请重新选择并上传"
    if len(data) > MAX_FILE_BYTES:
        return f"{name}: 单文件超过 {MAX_FILE_BYTES // (1024 * 1024)} MB"
    if suffix == ".xlsx" and not data[:2] == b"PK":
        return f"{name}: 不是有效的 xlsx (ZIP) 文件"
    return None


def _extract_zip(data: bytes, dest: Path) -> list[tuple[str, Path]]:
    extracted: list[tuple[str, Path]] = []
    try:
        zf_ctx = zipfile.ZipFile(io.BytesIO(data))
    except zipfile.BadZipFile as exc:
        raise ValueError("ZIP 文件无效或已损坏") from exc
    with zf_ctx as zf:
        for info in zf.infolist():
            if info.is_dir():
                continue
            inner_name = Path(info.filename).name
            if not inner_name.lower().endswith(".xlsx"):
                continue
            if info.file_size > MAX_FILE_BYTES:
                raise ValueError(f"ZIP 内 {inner_name} 超过大小限制")
            target = dest / inner_name
            with zf.open(info) as src, target.open("wb") as out:
                shutil.copyfileobj(src, out)
            extracted.append((inner_name, target))
    return extracted


def _save_uploads(
    uploads: list[tuple[str, bytes]],
    staging_dir: Path,
) -> list[Path]:
    upload_dir = staging_dir / "uploads"
    upload_dir.mkdir(parents=True, exist_ok=True)
    saved: list[Path] = []
    total = 0
    for name, data in uploads:
        err = _validate_upload(name, data)
        if err:
            raise ValueError(err)
        total += len(data)
        if total > MAX_TOTAL_BYTES:
            raise ValueError(f"上传总量超过 {MAX_TOTAL_BYTES // (1024 * 1024)} MB")
        suffix = Path(name).suffix.lower()
        if suffix == ".zip":
            saved.extend(path for _, path in _extract_zip(data, upload_dir))
        else:
            target = upload_dir / Path(name).name
            target.write_bytes(data)
            saved.append(target)
    return saved


def _build_file_index(paths: list[Path]) -> list[UploadedFile]:
    files: list[UploadedFile] = []
    for path in paths:
        try:
            sheets = scan_workbook_sheets(path)
        except Exception as exc:
            raise ValueError(f"无法读取 {path.name}: {exc}") from exc
        files.append(
            UploadedFile(
                filename=path.name,
                path=path,
                size_bytes=path.stat().st_size,
                sheet_names=sheets,
            )
        )
    return files


def _resolve_upload_sheet_name(req_name: str, uf: UploadedFile) -> str | None:
    """Resolve manifest sheet name to an actual worksheet inside an upload file."""
    return resolve_sheet_from_upload(
        req_name,
        filename=uf.filename,
        sheet_names=uf.sheet_names,
    )


def _missing_match_detail(req_name: str, files: list[UploadedFile]) -> str:
    for uf in files:
        if not filename_implies_sheet(uf.filename, req_name):
            continue
        if uf.size_bytes == 0:
            return f"已上传 {uf.filename} 但文件为空（0 字节），请重新上传"
        if not uf.sheet_names:
            return f"已上传 {uf.filename} 但无法读取工作表"
        inner = "、".join(uf.sheet_names)
        return (
            f"已上传 {uf.filename}，内部工作表为「{inner}」，"
            f"与必需名「{req_name}」不一致"
        )
    return "未在任何上传文件中找到同名工作表"


def _ready_match_detail(
    req_name: str,
    source_file: str,
    resolved: str | None,
) -> str | None:
    if resolved and resolved != req_name:
        return f"按文件名「{source_file}」匹配，使用工作表「{resolved}」"
    return None


def _append_personnel_file_sources(
    files: list[UploadedFile],
    sources: list[str],
    resolved: str | None,
) -> tuple[list[str], str | None]:
    """Match 人员信息.xlsx by filename when the inner sheet is not named 人员信息."""
    seen = set(sources)
    out = list(sources)
    resolved_out = resolved
    for uf in files:
        if uf.filename != PERSONNEL_FILENAME:
            continue
        if uf.filename not in seen:
            seen.add(uf.filename)
            out.append(uf.filename)
        if resolved_out is None and uf.sheet_names:
            resolved_out = resolve_sheet_alias(PERSONNEL_SHEET, set(uf.sheet_names))
            if resolved_out is None:
                resolved_out = uf.sheet_names[0]
    return out, resolved_out


def _match_sheets(
    manifest: list[RequiredSheet],
    files: list[UploadedFile],
) -> list[SheetMatch]:
    # sheet_name -> list of source filenames
    index: dict[str, list[str]] = {}
    alias_index: dict[str, list[str]] = {}
    for uf in files:
        for sheet in uf.sheet_names:
            index.setdefault(sheet, []).append(uf.filename)
            alias_index.setdefault(normalize_sheet_name(sheet), []).append(uf.filename)

    matches: list[SheetMatch] = []
    for req in manifest:
        resolved: str | None = None
        sources: list[str] = []
        for uf in files:
            found = _resolve_upload_sheet_name(req.name, uf)
            if found:
                resolved = found
                sources.append(uf.filename)
        # dedupe sources preserving order
        seen: set[str] = set()
        unique_sources = []
        for src in sources:
            if src not in seen:
                seen.add(src)
                unique_sources.append(src)
        sources = unique_sources

        if req.optional_skeleton and req.name == PERSONNEL_SHEET:
            sources, resolved = _append_personnel_file_sources(files, sources, resolved)

        if req.optional_note or req.optional_skeleton:
            if sources:
                status = SheetMatchStatus.NOTE
            else:
                continue  # don't show absent optional sheets
            detail = _ready_match_detail(req.name, sources[0], resolved)
            matches.append(
                SheetMatch(
                    required=req,
                    status=status,
                    sources=sources,
                    resolved_name=resolved,
                    detail=detail,
                )
            )
            continue

        if req.optional_input and not sources:
            continue  # optional topology inputs — absent is OK, hidden from UI

        if req.optional_role_family and not sources:
            status = SheetMatchStatus.MISSING
            detail = _missing_match_detail(req.name, files)
            matches.append(
                SheetMatch(
                    required=req,
                    status=status,
                    sources=[],
                    resolved_name=None,
                    detail=detail,
                )
            )
            continue

        if not sources:
            status = SheetMatchStatus.MISSING
            detail = _missing_match_detail(req.name, files)
        elif len(sources) > 1:
            status = SheetMatchStatus.CONFLICT
            detail = f"在 {len(sources)} 个文件中找到：{'、'.join(sources)}"
        else:
            status = SheetMatchStatus.READY
            detail = _ready_match_detail(req.name, sources[0], resolved)

        matches.append(
            SheetMatch(
                required=req,
                status=status,
                sources=sources,
                resolved_name=resolved,
                detail=detail,
            )
        )
    return matches


def _score_sales_candidate(uf: UploadedFile, required_names: set[str]) -> int:
    score = 0
    available = set(uf.sheet_names)
    for name in required_names:
        if _resolve_upload_sheet_name(name, uf):
            score += 1
    # Prefer filenames hinting at sales workbook
    lower = uf.filename.lower()
    if "销售" in lower or "西物" in lower:
        score += 3
    if "提成依据" in uf.filename or "依据" in uf.filename:
        score -= 5
    if "售后" in lower:
        score -= 10
    return score


def _infer_workbooks(
    files: list[UploadedFile],
    matches: list[SheetMatch],
) -> tuple[Path | None, Path | None]:
    path_by_name = {uf.filename: uf.path for uf in files}
    required_names = {
        m.required.name for m in matches if is_mandatory_input(m.required)
    }

    sales_file: str | None = None
    best_score = -1
    for uf in files:
        score = _score_sales_candidate(uf, required_names)
        if score > best_score:
            best_score = score
            sales_file = uf.filename

    rules_file: str | None = None
    for uf in files:
        if "提成依据" in uf.filename or "依据" in uf.filename:
            rules_file = uf.filename
            break

    sales_path = path_by_name.get(sales_file) if sales_file else None
    rules_path = path_by_name.get(rules_file) if rules_file else None
    return sales_path, rules_path


def intake_uploads(
    month_id: str,
    uploads: list[tuple[str, bytes]],
    *,
    staging_root: Path | None = None,
    manifest: list[RequiredSheet] | None = None,
) -> IntakeResult:
    """
    Validate uploads, scan sheet names, match manifest, write to staging.

    uploads: list of (filename, raw_bytes) from Streamlit file_uploader.
    """
    manifest = manifest or build_required_sheet_manifest()
    staging_dir = staging_root or (raw_month_dir(month_id) / ".staging")
    staging_dir.mkdir(parents=True, exist_ok=True)

    result = IntakeResult(
        month_id=month_id,
        staging_dir=staging_dir,
        uploads=[],
        matches=[],
    )

    try:
        saved_paths = _save_uploads(uploads, staging_dir)
    except ValueError as exc:
        result.errors.append(str(exc))
        return result

    if not saved_paths:
        result.errors.append("未找到有效的 .xlsx 文件")
        return result

    try:
        result.uploads = _build_file_index(saved_paths)
    except ValueError as exc:
        result.errors.append(str(exc))
        return result

    result.matches = _match_sheets(manifest, result.uploads)
    sales_path, rules_path = _infer_workbooks(result.uploads, result.matches)
    result.sales_workbook = sales_path
    result.rules_workbook = rules_path

    if result.conflict_sheets:
        result.warnings.append(
            "以下工作表在多个文件中出现，需人工选择来源后再合并: "
            + ", ".join(result.conflict_sheets)
        )

    missing_role_families = result.missing_role_family_sheets
    if missing_role_families:
        preview = "、".join(missing_role_families[:6])
        if len(missing_role_families) > 6:
            preview += "…"
        result.warnings.append(
            f"以下岗位族专用表未上传（不阻断试算，对应岗位绩效将留空）: {preview}"
        )

    return result


def read_upload_bytes(upload: BinaryIO | bytes, filename: str) -> tuple[str, bytes]:
    data = upload.read() if hasattr(upload, "read") else upload
    return filename, data


def normalize_streamlit_upload_files(files) -> list:
    """Normalize file_uploader return value to a list (single or multi mode)."""
    if files is None:
        return []
    if isinstance(files, list):
        return files
    return [files]


def format_upload_size(size_bytes: int) -> str:
    """Human-readable upload size for UI."""
    if size_bytes <= 0:
        return "0 字节（空文件）"
    if size_bytes < 1024:
        return f"{size_bytes} B"
    if size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} KB"
    return f"{size_bytes / (1024 * 1024):.2f} MB"


def pairs_from_streamlit_files(files) -> list[tuple[str, bytes]]:
    """Build (filename, bytes) pairs from Streamlit UploadedFile objects."""
    pairs: list[tuple[str, bytes]] = []
    for f in normalize_streamlit_upload_files(files):
        data = f.getvalue()
        if len(data) == 0:
            raise ValueError(
                f"{f.name}: 文件为空（0 字节）。"
                "若上传列表显示 0.0MB，说明浏览器未传回内容，请重新选择文件。"
            )
        pairs.append((f.name, data))
    return pairs


def _is_rules_workbook(path: Path) -> bool:
    return "提成依据" in path.name or path.name.startswith("依据")


def _is_sales_workbook(path: Path) -> bool:
    lower = path.name.lower()
    if not lower.endswith(".xlsx"):
        return False
    if "售后" in path.name:
        return False
    return "销售提成" in path.name or "西物超市" in path.name or (
        "销售" in path.name and "西物" in path.name
    )


def discover_local_raw_workbooks(month_id: str) -> tuple[Path | None, Path | None]:
    """
    Locate sales and rules workbooks under data/raw/<month_id>/.

    Sales: filename hints 西物/销售提成 (excludes 售后提成).
    Rules: 提成依据.xlsx or similar.
    """
    month_dir = raw_month_dir(month_id)
    if not month_dir.is_dir():
        return None, None

    sales: Path | None = None
    rules: Path | None = None
    for path in sorted(month_dir.glob("*.xlsx")):
        if _is_rules_workbook(path):
            rules = path
            continue
        if _is_sales_workbook(path):
            if sales is None or "西物" in path.name:
                sales = path
    return sales, rules


def intake_local_raw(
    month_id: str,
    *,
    include_rules_workbook: bool = False,
    sales_path: Path | None = None,
    rules_path: Path | None = None,
    staging_root: Path | None = None,
    manifest: list[RequiredSheet] | None = None,
) -> IntakeResult:
    """
    Simulate upload intake using workbooks already on disk in data/raw/<month>/.

    Copies files into staging/uploads/ and runs the same matching as intake_uploads.
    """
    month_dir = raw_month_dir(month_id)
    discovered_sales, discovered_rules = discover_local_raw_workbooks(month_id)
    sales = sales_path or discovered_sales
    rules = rules_path or discovered_rules

    staging_dir = staging_root or (raw_month_dir(month_id) / ".staging")
    result = IntakeResult(
        month_id=month_id,
        staging_dir=staging_dir,
        uploads=[],
        matches=[],
    )

    if sales is None or not sales.exists():
        result.errors.append(
            f"未在 {month_dir.relative_to(PROJECT_ROOT)} 找到销售提成工作簿"
            "（文件名需含「销售提成」或「西物超市」，且非售后）"
        )
        return result

    uploads: list[tuple[str, bytes]] = [(sales.name, sales.read_bytes())]
    if include_rules_workbook:
        if rules is None or not rules.exists():
            result.warnings.append(
                f"已勾选包含提成依据，但 {month_dir.relative_to(PROJECT_ROOT)} 下未找到"
            )
        else:
            uploads.append((rules.name, rules.read_bytes()))

    intake = intake_uploads(
        month_id,
        uploads,
        staging_root=staging_dir,
        manifest=manifest,
    )
    intake.warnings = result.warnings + intake.warnings
    return intake
