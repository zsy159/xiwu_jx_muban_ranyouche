"""Merge required sheets from multiple uploads into one sales workbook."""

from __future__ import annotations

import json
import logging
import shutil
from copy import copy
from pathlib import Path

from openpyxl import load_workbook

from salary_pipeline.ingestion_upload.file_intake import IntakeResult, SheetMatchStatus
from salary_pipeline.ingestion_upload.manifest import resolve_sheet_alias
from salary_pipeline.paths import PROJECT_ROOT, resolve_project_path

logger = logging.getLogger(__name__)

SHEET_SOURCES_FILENAME = "sheet_sources.json"

# Formula sheets kept out of the merged workbook (openpyxl corrupts cached values).
# Hub F/G/H still read them via supplemental upload paths.
HUB_SUPPLEMENTAL_SHEETS: tuple[str, ...] = ("销售任务及完成率",)


def _copy_sheet(source_ws, target_wb, sheet_name: str) -> None:
    if sheet_name in target_wb.sheetnames:
        del target_wb[sheet_name]
    target_ws = target_wb.create_sheet(sheet_name)
    for row in source_ws.iter_rows():
        for cell in row:
            new_cell = target_ws.cell(
                row=cell.row,
                column=cell.column,
                value=cell.value,
            )
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
    target_ws.sheet_properties = copy(source_ws.sheet_properties)
    target_ws.sheet_format = copy(source_ws.sheet_format)
    target_ws.merged_cells = copy(source_ws.merged_cells)


def _find_source_path(intake: IntakeResult, filename: str) -> Path:
    for uf in intake.uploads:
        if uf.filename == filename:
            return uf.path
    raise FileNotFoundError(f"Upload not found: {filename}")


def _relative_project_path(path: Path) -> str:
    try:
        return str(path.relative_to(PROJECT_ROOT))
    except ValueError:
        return str(path)


def _resolve_match_source(
    intake: IntakeResult,
    match,
    conflict_resolutions: dict[str, str],
) -> tuple[str, Path, str | None]:
    manifest_name = match.required.name
    if match.status == SheetMatchStatus.CONFLICT:
        source_file = conflict_resolutions.get(manifest_name)
        if not source_file:
            raise ValueError(f"工作表 {manifest_name} 存在冲突，需指定来源文件")
    else:
        source_file = match.sources[0]

    source_path = _find_source_path(intake, source_file)
    resolved = match.resolved_name
    if resolved is None:
        resolved = resolve_sheet_alias(manifest_name, set(scan_sheets(source_path)))
    return manifest_name, source_path, resolved


def plan_sheet_sources(
    intake: IntakeResult,
    *,
    conflict_resolutions: dict[str, str] | None = None,
) -> dict[str, Path]:
    """
    Map manifest sheet names to the upload file that should supply them.

  When every required sheet already lives in the sales workbook, this is empty.
    """
    if intake.sales_workbook is None:
        return {}

    conflict_resolutions = conflict_resolutions or {}
    base_path = intake.sales_workbook.resolve()
    sheet_sources: dict[str, Path] = {}

    for match in intake.matches:
        if match.required.optional_note or match.status == SheetMatchStatus.MISSING:
            continue
        manifest_name, source_path, resolved = _resolve_match_source(
            intake, match, conflict_resolutions
        )
        if resolved is None:
            continue
        if source_path.resolve() != base_path:
            sheet_sources[manifest_name] = source_path

    return sheet_sources


def needs_openpyxl_merge(
    intake: IntakeResult,
    *,
    conflict_resolutions: dict[str, str] | None = None,
) -> bool:
    """True when sheets must be physically renamed inside the base workbook."""
    if intake.sales_workbook is None:
        return False

    conflict_resolutions = conflict_resolutions or {}
    base_sheets = set(scan_sheets(intake.sales_workbook))

    for match in intake.matches:
        if match.required.optional_note or match.status == SheetMatchStatus.MISSING:
            continue
        manifest_name, source_path, resolved = _resolve_match_source(
            intake, match, conflict_resolutions
        )
        if resolved is None:
            continue
        if source_path.resolve() != intake.sales_workbook.resolve():
            continue
        if resolved != manifest_name and manifest_name not in base_sheets:
            return True
    return False


def write_sheet_sources(
    sheet_sources: dict[str, Path],
    staging_dir: Path,
) -> Path | None:
    if not sheet_sources:
        return None
    rel_map = {
        sheet: _relative_project_path(path)
        for sheet, path in sorted(sheet_sources.items())
    }
    out_path = staging_dir / SHEET_SOURCES_FILENAME
    out_path.write_text(
        json.dumps(rel_map, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return out_path


def load_sheet_sources(path: Path | None) -> dict[str, Path]:
    if path is None or not path.exists():
        return {}
    raw = json.loads(path.read_text(encoding="utf-8"))
    return {
        sheet: resolve_project_path(rel)
        for sheet, rel in raw.items()
    }


def supplement_sheet_sources(
    sales_workbook_path: Path,
    sheet_sources: dict[str, Path] | None = None,
) -> dict[str, Path]:
    """Attach upload workbooks for hub formula sheets missing from the merged base."""
    out = dict(sheet_sources or {})
    base_path = Path(sales_workbook_path)
    try:
        base_sheets = set(scan_sheets(base_path))
    except OSError:
        return out

    raw_dir = base_path.parent
    for sheet_name in HUB_SUPPLEMENTAL_SHEETS:
        if sheet_name in out or sheet_name in base_sheets:
            continue
        for candidate in (
            raw_dir / "uploads" / f"{sheet_name}.xlsx",
            raw_dir / ".staging" / "uploads" / f"{sheet_name}.xlsx",
        ):
            if candidate.is_file():
                out[sheet_name] = candidate.resolve()
                logger.info(
                    "Supplemental sheet %s -> %s",
                    sheet_name,
                    candidate.name,
                )
                break
    return out


def build_consolidated_workbook(
    intake: IntakeResult,
    output_path: Path,
    *,
    conflict_resolutions: dict[str, str] | None = None,
) -> Path:
    """
    Assemble the sales workbook for trial compute.

    Prefer a byte-level copy of the sales workbook so cached cell values in
    formula source sheets (e.g. 销售任务及完成率) stay intact. openpyxl save
    round-trips corrupt those values and zeros out 一线/销售顾问 Hub metrics.

    Sheets supplied by supplemental uploads are read via sheet_sources.json
  instead of being copied into the base workbook with openpyxl.
    """
    conflict_resolutions = conflict_resolutions or {}
    if intake.sales_workbook is None:
        raise ValueError("无法推断销售账套主文件")

    base_path = intake.sales_workbook
    sheet_sources = plan_sheet_sources(
        intake, conflict_resolutions=conflict_resolutions
    )

    if needs_openpyxl_merge(intake, conflict_resolutions=conflict_resolutions):
        logger.warning(
            "Sheet rename inside base workbook requires openpyxl merge; "
            "some cached source values may be lost"
        )
        return _build_with_openpyxl(
            intake,
            output_path,
            conflict_resolutions=conflict_resolutions,
            sheet_sources=sheet_sources,
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(base_path, output_path)
    sources_path = write_sheet_sources(sheet_sources, output_path.parent)
    intake.sheet_sources_path = sources_path
    intake.consolidated_workbook = output_path
    logger.info(
        "Consolidated workbook (copy) -> %s supplemental_sheets=%s",
        output_path,
        len(sheet_sources),
    )
    return output_path


def _build_with_openpyxl(
    intake: IntakeResult,
    output_path: Path,
    *,
    conflict_resolutions: dict[str, str],
    sheet_sources: dict[str, Path],
) -> Path:
    base_path = intake.sales_workbook
    assert base_path is not None
    wb = load_workbook(base_path)
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
        wb.remove(wb["Sheet"])

    for match in intake.matches:
        if match.required.optional_note or match.status == SheetMatchStatus.MISSING:
            continue
        manifest_name, source_path, resolved = _resolve_match_source(
            intake, match, conflict_resolutions
        )
        if resolved is None:
            continue
        if manifest_name in sheet_sources:
            continue
        if source_path.resolve() == base_path.resolve() and resolved in wb.sheetnames:
            if resolved != manifest_name and manifest_name not in wb.sheetnames:
                wb[resolved].title = manifest_name
            continue

        src_wb = load_workbook(source_path, read_only=False, data_only=False)
        try:
            src_ws = src_wb[resolved]
            target_name = manifest_name if manifest_name in src_wb.sheetnames else resolved
            if target_name != resolved and manifest_name not in src_wb.sheetnames:
                target_name = resolved
            _copy_sheet(src_ws, wb, target_name)
        finally:
            src_wb.close()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()
    sources_path = write_sheet_sources(sheet_sources, output_path.parent)
    intake.sheet_sources_path = sources_path
    intake.consolidated_workbook = output_path
    logger.info("Consolidated workbook (openpyxl) -> %s", output_path)
    return output_path


def scan_sheets(path: Path) -> list[str]:
    from salary_pipeline.ingestion_upload.file_intake import scan_workbook_sheets

    return scan_workbook_sheets(path)


def prepend_generated_sheets(
    archived_workbook: Path,
    generated: dict[str, Path],
    output_path: Path | None = None,
) -> Path:
    """
    Prepend generated sheets (绩效整理表, 提成汇总) to front of archived workbook.

    generated: sheet_name -> path to xlsx containing that sheet (single-sheet files ok).
    """
    target_path = output_path or archived_workbook
    wb = load_workbook(archived_workbook)

    existing = set(wb.sheetnames)
    for sheet_name in reversed(list(generated.keys())):
        gen_path = generated[sheet_name]
        gen_wb = load_workbook(gen_path, read_only=False, data_only=False)
        try:
            if sheet_name in gen_wb.sheetnames:
                src_ws = gen_wb[sheet_name]
            else:
                src_ws = gen_wb.active
            if sheet_name in existing:
                del wb[sheet_name]
            _copy_sheet(src_ws, wb, sheet_name)
        finally:
            gen_wb.close()

    # Reorder: generated sheets first
    desired_order = list(generated.keys()) + [
        s for s in wb.sheetnames if s not in generated
    ]
    wb._sheets = [wb[s] for s in desired_order if s in wb.sheetnames]

    target_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target_path)
    wb.close()
    logger.info("Prepended generated sheets -> %s", target_path)
    return target_path
