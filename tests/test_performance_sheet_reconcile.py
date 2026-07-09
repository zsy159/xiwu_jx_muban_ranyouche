"""Tests for 绩效整理表 golden header alignment and reconcile highlighting."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from salary_pipeline.calculators.sales_advisor.topology_specs import (
    is_manual_formula_adjustment,
)
from salary_pipeline.data_ingestion.data_loader import load_month_config
from salary_pipeline.data_ingestion.performance_sheet_golden_scan import (
    describe_manual_pattern,
    load_golden_column_headers,
    scan_golden_manual_cells,
)
from salary_pipeline.paths import CONFIG_DIR, PROJECT_ROOT, resolve_project_path
from salary_pipeline.pipelines.performance_sheet_builder import (
    IMPLEMENTED_COLUMNS,
    PerformanceSheetBuilder,
)
from salary_pipeline.pipelines.performance_sheet_export import (
    export_computed_performance_sheet,
    prepare_export_frame,
    resolve_export_column_spec,
    unimplemented_column_headers,
)
from salary_pipeline.pipelines.performance_sheet_formatting import (
    collect_performance_sheet_mismatches,
    implemented_value_columns,
)

GOLDEN = PROJECT_ROOT / "data/raw/2026-05/燃油车-2026年05月西物超市销售提成(终)(1).xlsx"
XUE_VIN = "L6T798NE4TT750321"


class PerformanceSheetGoldenScanTest(unittest.TestCase):
    def test_describe_manual_tail_pattern(self) -> None:
        formula = (
            "=IF(AND(A461=\"武侯自有店\",H461=\"星越L\"),200,"
            "SUMIFS(提成标准!F:F,提成标准!E:E,H461)*K461)-1200"
        )
        self.assertTrue(is_manual_formula_adjustment(formula))
        pattern, detail = describe_manual_pattern(formula)
        self.assertEqual(pattern, "公式+尾项")
        self.assertIn("-1200", detail)

    @unittest.skipUnless(GOLDEN.exists(), "golden workbook missing")
    def test_golden_headers_count(self) -> None:
        spec = load_golden_column_headers(GOLDEN)
        self.assertGreaterEqual(len(spec), 60)
        letters = [letter for letter, _ in spec]
        self.assertEqual(letters[0], "A")
        self.assertIn("AG", letters)
        self.assertIn("VIN码", [label for _, label in spec])

    @unittest.skipUnless(GOLDEN.exists(), "golden workbook missing")
    def test_scan_finds_xuexiangjian_ag_tail(self) -> None:
        cells = scan_golden_manual_cells(GOLDEN)
        xue_ag = [
            c
            for c in cells
            if c.vin == XUE_VIN and c.letter == "AG" and c.pattern == "公式+尾项"
        ]
        self.assertEqual(len(xue_ag), 1)
        self.assertIn("-1200", xue_ag[0].detail)


class PerformanceSheetExportWithoutGoldenTest(unittest.TestCase):
    def test_export_without_golden_aligns_source_row_with_extras(self) -> None:
        """Trial compute exports without golden_path; Slice 8 extras must not break row 2."""
        from salary_pipeline.pipelines.performance_sheet_export import (
            HEADER_ROW,
            SOURCE_ANNOTATION_ROW,
        )

        data = {col: [1.0] for col in IMPLEMENTED_COLUMNS}
        frame = pd.DataFrame(data)
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "perf.xlsx"
            export_computed_performance_sheet(frame, path, title="trial-no-golden")
            wb = load_workbook(path, read_only=True)
            ws = wb["绩效整理表"]
            sources = [
                c.value
                for c in next(
                    ws.iter_rows(
                        min_row=SOURCE_ANNOTATION_ROW, max_row=SOURCE_ANNOTATION_ROW
                    )
                )
            ]
            headers = [
                c.value
                for c in next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW))
            ]
            wb.close()
        self.assertEqual(len(sources), len(headers))
        self.assertGreater(len(sources), len(resolve_export_column_spec(None)))


@unittest.skipUnless(GOLDEN.exists(), "golden workbook missing")
class PerformanceSheetExportAlignmentTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        config = load_month_config(CONFIG_DIR)
        cls.config = config
        cls.builder = PerformanceSheetBuilder(
            __import__(
                "salary_pipeline.data_ingestion.data_loader", fromlist=["WorkbookLoader"]
            ).WorkbookLoader(resolve_project_path(config["workbooks"]["sales"]))
        )
        cls.frame = cls.builder.build()
        cls.golden_spec = resolve_export_column_spec(GOLDEN)

    def test_export_matches_golden_column_count(self) -> None:
        export_frame = prepare_export_frame(self.frame, column_spec=self.golden_spec)
        self.assertEqual(len(export_frame.columns), len(self.golden_spec))

    def test_unimplemented_columns_are_marked_empty(self) -> None:
        export_frame = prepare_export_frame(self.frame, column_spec=self.golden_spec)
        unimpl = unimplemented_column_headers(self.golden_spec)
        self.assertIn("促销奖", unimpl)  # BD — no source formula in golden
        self.assertTrue(export_frame["促销奖"].isna().all())
        self.assertNotIn("整车节约", unimpl)
        self.assertFalse(export_frame["整车节约"].isna().all())

    def test_export_xlsx_has_golden_headers(self) -> None:
        from salary_pipeline.pipelines.performance_sheet_export import HEADER_ROW

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "perf.xlsx"
            export_computed_performance_sheet(
                self.frame, path, golden_path=GOLDEN, title="test"
            )
            wb = load_workbook(path, read_only=True)
            ws = wb["绩效整理表"]
            headers = [c.value for c in next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW))]
            wb.close()
            golden_labels = [label for _, label in self.golden_spec]
            self.assertEqual(headers, golden_labels)

    def test_export_has_source_annotation_row(self) -> None:
        from salary_pipeline.pipelines.performance_sheet_column_sources import (
            build_source_annotation_row,
            source_annotation_for_column,
        )
        from salary_pipeline.pipelines.performance_sheet_export import (
            SOURCE_ANNOTATION_ROW,
        )

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "perf.xlsx"
            export_computed_performance_sheet(
                self.frame, path, golden_path=GOLDEN, title="test"
            )
            wb = load_workbook(path, read_only=True)
            ws = wb["绩效整理表"]
            sources = [
                c.value for c in next(ws.iter_rows(min_row=SOURCE_ANNOTATION_ROW, max_row=SOURCE_ANNOTATION_ROW))
            ]
            wb.close()
            expected = build_source_annotation_row(self.golden_spec)
            self.assertEqual(sources, expected)
            self.assertNotIn("金标准", "".join(str(s) for s in sources if s))
            header_by_letter = {letter: label for letter, label in self.golden_spec}
            self.assertEqual(
                source_annotation_for_column("L", header_by_letter["L"]),
                "系统销售毛利!AO",
            )
            self.assertEqual(
                source_annotation_for_column("S", header_by_letter["S"]),
                "系统销售毛利!BQ",
            )
            l_idx = next(i for i, (_, h) in enumerate(self.golden_spec) if h == "订单合计(含税)")
            s_idx = next(i for i, (_, h) in enumerate(self.golden_spec) if h == "精品最低价金额")
            self.assertEqual(sources[l_idx], "系统销售毛利!AO")
            self.assertEqual(sources[s_idx], "系统销售毛利!BQ")

    def test_source_annotation_uses_header_semantics(self) -> None:
        from salary_pipeline.pipelines.performance_sheet_column_sources import (
            build_source_annotation_row,
        )

        # Letter positions wrong but Chinese headers identify canonical columns.
        spec = (("I", "订单合计(含税)"), ("R", "精品最低价金额"))
        sources = build_source_annotation_row(spec)
        self.assertEqual(sources, ["系统销售毛利!AO", "系统销售毛利!BQ"])

    def test_system_sourced_columns_not_scanned_as_direct_fill(self) -> None:
        cells = scan_golden_manual_cells(GOLDEN)
        l_direct = [c for c in cells if c.header == "订单合计(含税)" and c.pattern == "直接填数"]
        s_direct = [c for c in cells if c.header == "精品最低价金额" and c.pattern == "直接填数"]
        self.assertEqual(l_direct, [])
        self.assertEqual(s_direct, [])

    def test_reconcile_does_not_gray_computed_l_s_columns(self) -> None:
        from salary_pipeline.data_ingestion.data_loader import load_month_config
        from salary_pipeline.pipelines.performance_sheet_formatting import (
            apply_performance_sheet_highlighting,
        )
        from salary_pipeline.utils.excel_format import GOLDEN_STATIC_FILL_RGB

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "perf.xlsx"
            export_computed_performance_sheet(
                self.frame, path, golden_path=GOLDEN, title="test"
            )
            config = load_month_config(CONFIG_DIR)
            apply_performance_sheet_highlighting(
                config, path, golden_path=GOLDEN, computed_frame=self.frame
            )
            wb = load_workbook(path)
            ws = wb["绩效整理表"]
            header_row = 4
            col_map = {
                str(ws.cell(header_row, c).value).strip(): c
                for c in range(1, ws.max_column + 1)
                if ws.cell(header_row, c).value
            }
            for name in ("订单合计(含税)", "精品最低价金额"):
                col_idx = col_map[name]
                gray = 0
                for r in range(header_row + 1, ws.max_row + 1):
                    cell = ws.cell(r, col_idx)
                    rgb = (
                        getattr(cell.fill.start_color, "rgb", None)
                        if cell.fill and cell.fill.fill_type == "solid"
                        else None
                    )
                    if rgb == GOLDEN_STATIC_FILL_RGB:
                        gray += 1
                self.assertEqual(gray, 0, f"{name} should not be gray-filled")
            promo_idx = col_map["促销奖"]
            promo_gray = sum(
                1
                for r in range(header_row + 1, ws.max_row + 1)
                if getattr(ws.cell(r, promo_idx).fill.start_color, "rgb", None)
                == GOLDEN_STATIC_FILL_RGB
            )
            self.assertEqual(promo_gray, len(self.frame))
            wb.close()

    def test_reconcile_no_gray_on_any_implemented_column(self) -> None:
        from salary_pipeline.data_ingestion.data_loader import load_month_config
        from salary_pipeline.pipelines.performance_sheet_formatting import (
            apply_performance_sheet_highlighting,
        )
        from salary_pipeline.pipelines.performance_sheet_column_sources import (
            is_implemented_perf_column,
        )
        from salary_pipeline.utils.excel_format import GOLDEN_STATIC_FILL_RGB

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "perf.xlsx"
            export_computed_performance_sheet(
                self.frame, path, golden_path=GOLDEN, title="test"
            )
            config = load_month_config(CONFIG_DIR)
            apply_performance_sheet_highlighting(
                config, path, golden_path=GOLDEN, computed_frame=self.frame
            )
            wb = load_workbook(path)
            ws = wb["绩效整理表"]
            header_row = 4
            for letter, label in self.golden_spec:
                if not is_implemented_perf_column(letter, label):
                    continue
                col_idx = None
                for c in range(1, ws.max_column + 1):
                    if str(ws.cell(header_row, c).value).strip() == label:
                        col_idx = c
                        break
                self.assertIsNotNone(col_idx, label)
                gray = sum(
                    1
                    for r in range(header_row + 1, ws.max_row + 1)
                    if getattr(ws.cell(r, col_idx).fill.start_color, "rgb", None)
                    == GOLDEN_STATIC_FILL_RGB
                )
                self.assertEqual(gray, 0, f"{label} ({letter}) must not be gray")
            wb.close()

@unittest.skipUnless(GOLDEN.exists(), "golden workbook missing")
class PerformanceSheetMismatchTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        config = load_month_config(CONFIG_DIR)
        loader = __import__(
            "salary_pipeline.data_ingestion.data_loader", fromlist=["WorkbookLoader"]
        ).WorkbookLoader(resolve_project_path(config["workbooks"]["sales"]))
        cls.frame = PerformanceSheetBuilder(loader).build()

    def test_xuexiangjian_ag_mismatch_detected(self) -> None:
        mismatches = collect_performance_sheet_mismatches(
            self.frame, GOLDEN, columns=frozenset({"AG"})
        )
        xue = [m for m in mismatches if m.vin == XUE_VIN]
        self.assertEqual(len(xue), 1)
        self.assertAlmostEqual(xue[0].golden_value or 0, 800.0, places=2)
        self.assertAlmostEqual(xue[0].computed_value or 0, 2000.0, places=2)

    def test_settlement_date_matches_system_sales_gross(self) -> None:
        vin_frame = self.frame[self.frame["O"].notna()].copy()
        mismatches = collect_performance_sheet_mismatches(
            vin_frame, GOLDEN, columns=frozenset({"M"}), tolerance=0
        )
        total = len(vin_frame)
        self.assertGreater(total, 400)
        self.assertEqual(len(mismatches), 0, mismatches[:3])

    def test_implemented_value_columns_subset(self) -> None:
        impl = implemented_value_columns()
        self.assertIn("AG", impl)
        self.assertIn("O", IMPLEMENTED_COLUMNS)
        self.assertNotIn("O", impl)


if __name__ == "__main__":
    unittest.main()
