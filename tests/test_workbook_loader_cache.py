"""WorkbookLoader sheet-level read cache."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import pandas as pd
from openpyxl import Workbook

from salary_pipeline.data_ingestion.data_loader import WorkbookLoader


def _make_fixture_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "测试表"
    ws["A1"] = "VIN"
    ws["B1"] = "姓名"
    ws["A2"] = "VIN001"
    ws["B2"] = "张三"
    ws["A3"] = "VIN002"
    ws["B3"] = "李四"
    wb.save(path)


class TestWorkbookLoaderCache(unittest.TestCase):
    def setUp(self) -> None:
        self.tmp = tempfile.TemporaryDirectory()
        self.workbook = Path(self.tmp.name) / "fixture.xlsx"
        _make_fixture_workbook(self.workbook)
        self.loader = WorkbookLoader(self.workbook)

    def tearDown(self) -> None:
        self.tmp.cleanup()

    @patch("salary_pipeline.data_ingestion.data_loader.pd.read_excel")
    def test_read_sheet_columns_hits_sheet_cache(self, mock_read_excel: unittest.mock.MagicMock) -> None:
        mock_read_excel.return_value = pd.DataFrame(
            [["VIN", "姓名"], ["VIN001", "张三"], ["VIN002", "李四"]]
        )

        first = self.loader.read_sheet_columns("测试表", {"A": "A"})
        second = self.loader.read_sheet_columns("测试表", {"B": "B"})

        mock_read_excel.assert_called_once()
        self.assertEqual(list(first["A"]), ["VIN", "VIN001", "VIN002"])
        self.assertEqual(list(second["B"]), ["姓名", "张三", "李四"])

    def test_same_column_reads_share_cache(self) -> None:
        a = self.loader.read_sheet_columns("测试表", {"A": "A"})
        b = self.loader.read_sheet_columns("测试表", {"B": "B"})
        self.assertIn("测试表", self.loader._raw_sheet_cache)
        self.assertEqual(len(self.loader._raw_sheet_cache), 1)
        self.assertEqual(list(a["A"]), ["VIN", "VIN001", "VIN002"])
        self.assertEqual(list(b["B"]), ["姓名", "张三", "李四"])


if __name__ == "__main__":
    unittest.main()
