"""Tests for upload intake and worksheet-level matching."""

from __future__ import annotations

import io
import json
import shutil
import tempfile
import unittest
import zipfile
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from salary_pipeline.ingestion_upload.file_intake import (
    SheetMatchStatus,
    display_match_status,
    intake_uploads,
    preferred_conflict_source_index,
    scan_workbook_sheets,
)
from salary_pipeline.ingestion_upload.manifest import (
    FAMILY_SALES,
    ROLE_FAMILY_INPUTS,
    build_required_sheet_manifest,
    group_manifest_by_family,
    is_mandatory_input,
    required_input_sheets,
    resolve_sheet_alias,
)
from salary_pipeline.data_ingestion.data_loader import WorkbookLoader, build_workbook_loader
from salary_pipeline.pipelines.hub_formula_engine import HubFormulaEngine
from salary_pipeline.modules.base import PERSONNEL_FILENAME, PERSONNEL_SHEET
from salary_pipeline.ingestion_upload.month_config import write_month_config
from salary_pipeline.ingestion_upload.overrides import (
    apply_overrides,
    load_overrides,
    store_sheet_override,
)
from salary_pipeline.ingestion_upload.sheet_merge import (
    build_consolidated_workbook,
    load_sheet_sources_for_workbook,
    needs_openpyxl_merge,
    plan_sheet_sources,
    resolve_sheet_sources_path,
    supplement_sheet_sources,
)


def _make_workbook(path: Path, sheets: dict[str, list[list]]) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    for name, rows in sheets.items():
        ws = wb.create_sheet(name)
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
    wb.save(path)


class UploadIntakeTest(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.tmp = Path(self._tmp.name)

    def tearDown(self) -> None:
        self._tmp.cleanup()

    def test_manifest_includes_registry_and_closure_sheets(self) -> None:
        manifest = build_required_sheet_manifest()
        names = {s.name for s in manifest if is_mandatory_input(s)}
        self.assertIn("终端明细表", names)
        self.assertIn("系统销售毛利", names)
        self.assertNotIn("翼真考核", names)
        self.assertNotIn("按揭绩效", names)
        self.assertIn("按揭原表", names)
        self.assertIn("指标汇总", names)
        self.assertIn("保客考核明细", names)
        self.assertIn("直营店交车", names)
        self.assertIn("综合表", names)
        self.assertIn("比对表", names)
        self.assertIn("延保提成", names)
        self.assertIn("工厂购进", names)
        self.assertIn("系统二手车降价", names)
        self.assertNotIn("产品经理提成核算", names)
        self.assertNotIn("销售管理岗提成依据 新标", names)
        self.assertNotIn("绩效整理表", names)

    def test_manifest_topology_replay_sheets_are_optional(self) -> None:
        manifest = build_required_sheet_manifest()
        optional_topology = {s.name for s in manifest if s.optional_input}
        self.assertEqual(
            optional_topology,
            {
                "产品经理提成核算",
                "销售管理岗提成依据 新标",
                "翼真考核",
                "按揭绩效",
            },
        )
        mandatory = {s.name for s in manifest if is_mandatory_input(s)}
        self.assertNotIn("产品经理提成核算", mandatory)
        self.assertNotIn("销售管理岗提成依据 新标", mandatory)

    def test_manifest_formula_notes_exclude_generated_perf_sheet(self) -> None:
        manifest = build_required_sheet_manifest()
        all_names = {s.name for s in manifest}
        self.assertNotIn("绩效整理表", all_names)
        optional = {s.name for s in manifest if s.optional_note}
        self.assertIn("销售任务及完成率", optional)
        skeleton = {s.name for s in manifest if s.optional_skeleton}
        self.assertEqual(skeleton, {PERSONNEL_SHEET})

    def test_intake_matches_factory_purchase_by_filename(self) -> None:
        sales = self.tmp / "销售.xlsx"
        required = required_input_sheets()
        sheet_data = {
            s.name: [["x"]]
            for s in required[:5]
            if s.name != "工厂购进"
        }
        _make_workbook(sales, sheet_data)
        factory = self.tmp / "工厂购进.xlsx"
        _make_workbook(factory, {"数据": [["VIN", "公司"], ["VIN1", "A"]]})

        intake = intake_uploads(
            "2099-01b",
            [(sales.name, sales.read_bytes()), (factory.name, factory.read_bytes())],
            staging_root=self.tmp / "staging-factory",
        )
        factory_match = next(
            m for m in intake.matches if m.required.name == "工厂购进"
        )
        self.assertEqual(factory_match.status, SheetMatchStatus.READY)
        self.assertEqual(factory_match.sources, ["工厂购进.xlsx"])
        self.assertEqual(factory_match.resolved_name, "数据")
        self.assertIn("按文件名", factory_match.detail or "")

    def test_intake_matches_used_car_discount_by_filename(self) -> None:
        sales = self.tmp / "销售.xlsx"
        required = required_input_sheets()
        sheet_data = {
            s.name: [["x"]]
            for s in required[:5]
            if s.name != "系统二手车降价"
        }
        _make_workbook(sales, sheet_data)
        discount = self.tmp / "系统二手车降价.xlsx"
        _make_workbook(discount, {"Sheet1": [["VIN", "佣金"], ["VIN1", 100]]})

        intake = intake_uploads(
            "2099-01c",
            [(sales.name, sales.read_bytes()), (discount.name, discount.read_bytes())],
            staging_root=self.tmp / "staging-discount",
        )
        discount_match = next(
            m for m in intake.matches if m.required.name == "系统二手车降价"
        )
        self.assertEqual(discount_match.status, SheetMatchStatus.READY)
        self.assertEqual(discount_match.sources, ["系统二手车降价.xlsx"])
        self.assertEqual(discount_match.resolved_name, "Sheet1")
        self.assertIn("按文件名", discount_match.detail or "")

    def test_intake_matches_personnel_filename_without_sheet_name(self) -> None:
        sales = self.tmp / "销售.xlsx"
        required = required_input_sheets()
        sheet_data = {s.name: [["x"]] for s in required[:3]}
        _make_workbook(sales, sheet_data)
        personnel = self.tmp / PERSONNEL_FILENAME
        _make_workbook(personnel, {"数据": [["店别", "职务", "姓名"], ["A", "B", "C"]]})

        intake = intake_uploads(
            "2099-01",
            [(sales.name, sales.read_bytes()), (personnel.name, personnel.read_bytes())],
            staging_root=self.tmp / "staging",
        )
        personnel_match = next(
            m for m in intake.matches if m.required.name == PERSONNEL_SHEET
        )
        self.assertEqual(personnel_match.status, SheetMatchStatus.NOTE)
        self.assertEqual(personnel_match.sources, [PERSONNEL_FILENAME])
        self.assertIsNotNone(personnel_match.resolved_name)

    def test_empty_upload_rejected(self) -> None:
        intake = intake_uploads(
            "2099-08",
            [("空.xlsx", b"")],
            staging_root=self.tmp / "staging-empty",
        )
        self.assertTrue(intake.errors)
        self.assertIn("0 字节", intake.errors[0])

    def test_missing_match_detail_when_absent(self) -> None:
        sales = self.tmp / "销售.xlsx"
        _make_workbook(sales, {"终端明细表": [["x"]]})

        intake = intake_uploads(
            "2099-08b",
            [("销售.xlsx", sales.read_bytes())],
            staging_root=self.tmp / "staging-detail",
        )
        factory_match = next(
            m for m in intake.matches if m.required.name == "工厂购进"
        )
        self.assertEqual(factory_match.status, SheetMatchStatus.MISSING)
        self.assertIn("未在任何上传文件", factory_match.detail or "")

    def test_build_consolidated_workbook_includes_factory_purchase_source(self) -> None:
        import json

        base = self.tmp / "base.xlsx"
        factory = self.tmp / "工厂购进.xlsx"
        _make_workbook(base, {"终端明细表": [["a"]]})
        _make_workbook(factory, {"工厂购进": [["b"]]})

        intake = intake_uploads(
            "2099-08c",
            [
                ("base.xlsx", base.read_bytes()),
                ("工厂购进.xlsx", factory.read_bytes()),
            ],
            staging_root=self.tmp / "staging-factory-merge",
        )
        out = self.tmp / "merged-factory.xlsx"
        build_consolidated_workbook(intake, out)
        self.assertIsNotNone(intake.sheet_sources_path)
        sources = json.loads(intake.sheet_sources_path.read_text(encoding="utf-8"))
        self.assertIn("工厂购进", sources)

    def test_build_consolidated_workbook_includes_used_car_discount_source(self) -> None:
        import json

        base = self.tmp / "base.xlsx"
        discount = self.tmp / "系统二手车降价.xlsx"
        _make_workbook(base, {"终端明细表": [["a"]]})
        _make_workbook(discount, {"系统二手车降价": [["b"]]})

        intake = intake_uploads(
            "2099-08d",
            [
                ("base.xlsx", base.read_bytes()),
                ("系统二手车降价.xlsx", discount.read_bytes()),
            ],
            staging_root=self.tmp / "staging-discount-merge",
        )
        out = self.tmp / "merged-discount.xlsx"
        build_consolidated_workbook(intake, out)
        self.assertIsNotNone(intake.sheet_sources_path)
        sources = json.loads(intake.sheet_sources_path.read_text(encoding="utf-8"))
        self.assertIn("系统二手车降价", sources)

    def test_build_consolidated_workbook_registers_role_family_sources(self) -> None:
        import json

        role_sheets = [name for _, name, _ in ROLE_FAMILY_INPUTS]
        base = self.tmp / "base.xlsx"
        _make_workbook(base, {"终端明细表": [["a"]]})
        uploads = [("base.xlsx", base.read_bytes())]
        for sheet_name in role_sheets:
            upload = self.tmp / f"{sheet_name}.xlsx"
            _make_workbook(upload, {sheet_name: [["x"]]})
            uploads.append((upload.name, upload.read_bytes()))

        intake = intake_uploads(
            "2099-08e",
            uploads,
            staging_root=self.tmp / "staging-role-family",
        )
        out = self.tmp / "merged-role-family.xlsx"
        build_consolidated_workbook(intake, out)
        merged_sheets = set(scan_workbook_sheets(out))
        for sheet_name in role_sheets:
            self.assertNotIn(
                sheet_name,
                merged_sheets,
                msg=f"{sheet_name} should stay supplemental, not physical merge",
            )
        self.assertIsNotNone(intake.sheet_sources_path)
        sources = json.loads(intake.sheet_sources_path.read_text(encoding="utf-8"))
        for sheet_name in role_sheets:
            self.assertIn(sheet_name, sources)

    def test_resolve_sheet_sources_path_adjacent_to_workbook(self) -> None:
        base = self.tmp / "merged.xlsx"
        _make_workbook(base, {"终端明细表": [["a"]]})
        sources_path = self.tmp / "sheet_sources.json"
        sources_path.write_text(
            '{"邀约专员提成": "uploads/邀约专员提成.xlsx"}',
            encoding="utf-8",
        )
        resolved = resolve_sheet_sources_path(base)
        self.assertEqual(resolved, sources_path.resolve())

    def test_workbook_loader_reads_role_family_from_sheet_sources(self) -> None:
        base = self.tmp / "base.xlsx"
        invite = self.tmp / "邀约专员提成.xlsx"
        _make_workbook(base, {"终端明细表": [["a"]]})
        _make_workbook(
            invite,
            {
                "邀约专员提成": [
                    ["", "", "姓名", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "发放"],
                    ["", "", "张三", *([""] * 28), 1200],
                ]
            },
        )
        sources_path = self.tmp / "sheet_sources.json"
        sources_path.write_text(
            json.dumps({"邀约专员提成": str(invite.resolve())}, ensure_ascii=False),
            encoding="utf-8",
        )
        loader = WorkbookLoader(
            base,
            sheet_paths=load_sheet_sources_for_workbook(base, explicit=sources_path),
        )
        self.assertTrue(loader.has_sheet("邀约专员提成"))
        frame = loader.read_sheet_columns(
            "邀约专员提成",
            {"C": "C", "AF": "AF"},
            label="邀约专员提成",
        )
        zhang = frame[frame["C"] == "张三"]
        self.assertFalse(zhang.empty)
        self.assertEqual(float(zhang.iloc[0]["AF"]), 1200.0)

    def test_hub_resolve_role_family_sheet_from_sheet_sources(self) -> None:
        base = self.tmp / "base.xlsx"
        invite = self.tmp / "邀约专员提成.xlsx"
        _make_workbook(base, {"终端明细表": [["a"]]})
        _make_workbook(
            invite,
            {
                "邀约专员提成": [
                    ["", "", "姓名", *([""] * 29), "发放"],
                    ["", "", "李四", *([""] * 29), 500],
                ]
            },
        )
        sources_path = self.tmp / "sheet_sources.json"
        sources_path.write_text(
            json.dumps({"邀约专员提成": str(invite.resolve())}, ensure_ascii=False),
            encoding="utf-8",
        )
        loader = WorkbookLoader(
            base,
            sheet_paths=load_sheet_sources_for_workbook(base, explicit=sources_path),
        )
        topo = self.tmp / "topology.json"
        topo.write_text(
            json.dumps({"cells": {}, "execution_order": []}),
            encoding="utf-8",
        )
        engine = HubFormulaEngine(topo, loader)
        self.assertEqual(engine._resolve_sheet_name("邀约专员提成"), "邀约专员提成")

    def test_manifest_includes_role_family_sheets(self) -> None:
        manifest = build_required_sheet_manifest()
        by_name = {s.name: s for s in manifest}
        role_sheets = {
            "新媒体": ("新媒体",),
            "翼真考核": (FAMILY_SALES, "新媒体"),
            "邀约专员提成": ("邀约专员",),
            "客户部提成": ("客户专员",),
            "直营店经理提成 (财务)": ("直营店经理",),
            "招聘": ("招聘",),
            "按揭绩效": (FAMILY_SALES, "按揭内勤"),
        }
        for sheet_name, families in role_sheets.items():
            self.assertIn(sheet_name, by_name, msg=f"missing {sheet_name}")
            sheet = by_name[sheet_name]
            self.assertTrue(sheet.optional_role_family, msg=sheet_name)
            self.assertFalse(is_mandatory_input(sheet), msg=sheet_name)
            self.assertEqual(sheet.families, families)
            if sheet_name not in {"按揭绩效", "翼真考核"}:
                self.assertEqual(sheet.source, "role")
        self.assertEqual(by_name["按揭绩效"].source, "registry")

    def test_manifest_groups_by_family(self) -> None:
        groups = dict(group_manifest_by_family())
        self.assertIn(FAMILY_SALES, groups)
        self.assertIn("新媒体", groups)
        self.assertIn("招聘", groups)
        self.assertIn("按揭内勤", groups)
        sales_names = {s.name for s in groups[FAMILY_SALES] if is_mandatory_input(s)}
        self.assertIn("终端明细表", sales_names)
        self.assertNotIn("新媒体", sales_names)
        self.assertNotIn("按揭绩效", sales_names)
        new_media_names = {s.name for s in groups["新媒体"]}
        self.assertEqual(new_media_names, {"新媒体", "翼真考核"})
        mortgage_names = {s.name for s in groups["按揭内勤"]}
        self.assertEqual(mortgage_names, {"按揭绩效"})

    def test_scan_workbook_sheets(self) -> None:
        path = self.tmp / "a.xlsx"
        _make_workbook(path, {"终端明细表": [["h"]], "整车成本": [["h"]]})
        sheets = scan_workbook_sheets(path)
        self.assertEqual(set(sheets), {"终端明细表", "整车成本"})

    def test_sheet_level_matching_ready(self) -> None:
        sales = self.tmp / "销售.xlsx"
        required = required_input_sheets()
        sheet_data = {s.name: [["x"]] for s in required[:5]}
        _make_workbook(sales, sheet_data)

        data = sales.read_bytes()
        intake = intake_uploads(
            "2099-01",
            [("销售.xlsx", data)],
            staging_root=self.tmp / "staging",
        )
        self.assertFalse(intake.errors)
        matched = {
            m.required.name: m.status
            for m in intake.matches
            if is_mandatory_input(m.required)
        }
        for name in sheet_data:
            self.assertEqual(matched[name], SheetMatchStatus.READY)
        self.assertIsNotNone(intake.sales_workbook)

    def test_missing_and_conflict_detection(self) -> None:
        f1 = self.tmp / "a.xlsx"
        f2 = self.tmp / "b.xlsx"
        _make_workbook(f1, {"终端明细表": [["1"]]})
        _make_workbook(f2, {"终端明细表": [["2"]]})

        intake = intake_uploads(
            "2099-02",
            [
                ("a.xlsx", f1.read_bytes()),
                ("b.xlsx", f2.read_bytes()),
            ],
            staging_root=self.tmp / "staging2",
        )
        terminal = next(m for m in intake.matches if m.required.name == "终端明细表")
        self.assertEqual(terminal.status, SheetMatchStatus.CONFLICT)
        self.assertIn("终端明细表", intake.conflict_sheets)
        self.assertFalse(intake.all_required_ready)
        self.assertFalse(intake.can_proceed())
        self.assertFalse(intake.can_proceed({"终端明细表": "a.xlsx"}))
        blockers = intake.proceed_blockers()
        self.assertTrue(any("冲突" in b for b in blockers))
        resolved_blockers = intake.proceed_blockers({"终端明细表": "a.xlsx"})
        self.assertFalse(any("冲突" in b for b in resolved_blockers))
        self.assertTrue(any("缺失" in b for b in resolved_blockers))

    def test_can_proceed_missing_sheets(self) -> None:
        sales = self.tmp / "partial.xlsx"
        _make_workbook(sales, {"终端明细表": [["x"]]})
        intake = intake_uploads(
            "2099-06",
            [("partial.xlsx", sales.read_bytes())],
            staging_root=self.tmp / "staging6",
        )
        self.assertFalse(intake.can_proceed())
        blockers = intake.proceed_blockers()
        self.assertTrue(any("缺失" in b for b in blockers))

    def test_role_family_missing_does_not_block_proceed(self) -> None:
        from salary_pipeline.ingestion_upload.file_intake import display_match_icon

        sales = self.tmp / "sales-only.xlsx"
        required = required_input_sheets()
        sheet_data = {s.name: [["x"]] for s in required}
        _make_workbook(sales, sheet_data)
        intake = intake_uploads(
            "2099-06b",
            [("sales-only.xlsx", sales.read_bytes())],
            staging_root=self.tmp / "staging6b",
        )
        self.assertTrue(intake.can_proceed())
        self.assertEqual(intake.proceed_blockers(), [])
        missing_rf = intake.missing_role_family_sheets
        self.assertIn("新媒体", missing_rf)
        self.assertTrue(any("岗位族专用表未上传" in w for w in intake.warnings))
        new_media = next(m for m in intake.matches if m.required.name == "新媒体")
        self.assertEqual(new_media.status, SheetMatchStatus.MISSING)
        self.assertEqual(display_match_icon(new_media), "⚠️")

    def test_new_media_missing_sheet_does_not_crash(self) -> None:
        from salary_pipeline.modules.new_media_performance import NewMediaPerformanceModule

        base = self.tmp / "base-new-media.xlsx"
        _make_workbook(base, {"终端明细表": [["a"]]})
        ctx = {
            "month_config": {"workbooks": {"sales": str(base)}},
            "summary_skeleton": pd.DataFrame(
                [
                    {
                        "店别": "新媒体销售部",
                        "职务": "主播",
                        "姓名": "测试",
                    }
                ]
            ),
        }
        result = NewMediaPerformanceModule().run(ctx)
        self.assertFalse(result.metadata.get("sheet_available"))
        row = result.metrics[result.metrics["姓名"] == "测试"].iloc[0]
        self.assertEqual(float(row["整车绩效"]), 0.0)

    def test_zip_upload_extraction(self) -> None:
        inner = self.tmp / "inner.xlsx"
        _make_workbook(inner, {"整车成本": [["v"]]})
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as zf:
            zf.writestr("inner.xlsx", inner.read_bytes())
        intake = intake_uploads(
            "2099-03",
            [("bundle.zip", buf.getvalue())],
            staging_root=self.tmp / "staging3",
        )
        self.assertFalse(intake.errors)
        cost = next(m for m in intake.matches if m.required.name == "整车成本")
        self.assertEqual(cost.status, SheetMatchStatus.READY)

    def test_resolve_sheet_alias_trailing_space(self) -> None:
        available = {"二手置换 ", "比对表"}
        self.assertEqual(resolve_sheet_alias("二手置换", available), "二手置换 ")
        self.assertEqual(resolve_sheet_alias("二手置换 ", available), "二手置换 ")

    def test_display_match_status_reflects_conflict_resolution(self) -> None:
        from salary_pipeline.ingestion_upload.file_intake import SheetMatch
        from salary_pipeline.ingestion_upload.manifest import RequiredSheet

        match = SheetMatch(
            required=RequiredSheet(name="新媒体", families=("新媒体",)),
            status=SheetMatchStatus.CONFLICT,
            sources=["销售.xlsx", "提成依据.xlsx"],
        )
        status, sources = display_match_status(match, {})
        self.assertEqual(status, SheetMatchStatus.CONFLICT)
        self.assertEqual(sources, ["销售.xlsx", "提成依据.xlsx"])

        status, sources = display_match_status(
            match, {"新媒体": "销售.xlsx"}
        )
        self.assertEqual(status, SheetMatchStatus.READY)
        self.assertEqual(sources, ["销售.xlsx"])

    def test_preferred_conflict_source_index_prefers_sales_workbook(self) -> None:
        sources = ["提成依据.xlsx", "西物销售提成.xlsx"]
        idx = preferred_conflict_source_index(
            sources,
            sales_workbook=Path("西物销售提成.xlsx"),
        )
        self.assertEqual(idx, 1)
        idx2 = preferred_conflict_source_index(sources)
        self.assertEqual(idx2, 1)

    def test_build_consolidated_workbook(self) -> None:
        import json

        base = self.tmp / "base.xlsx"
        extra = self.tmp / "extra.xlsx"
        _make_workbook(base, {"终端明细表": [["a"]]})
        _make_workbook(extra, {"比对表": [["b"]]})

        intake = intake_uploads(
            "2099-04",
            [
                ("base.xlsx", base.read_bytes()),
                ("extra.xlsx", extra.read_bytes()),
            ],
            staging_root=self.tmp / "staging4",
        )
        out = self.tmp / "merged.xlsx"
        build_consolidated_workbook(intake, out)
        sheets = set(scan_workbook_sheets(out))
        self.assertIn("终端明细表", sheets)
        self.assertNotIn("比对表", sheets)
        self.assertIsNotNone(intake.sheet_sources_path)
        sources = json.loads(intake.sheet_sources_path.read_text(encoding="utf-8"))
        self.assertIn("比对表", sources)

    def test_write_month_config_staging_paths(self) -> None:
        cfg_dir = self.tmp / "cfg"
        path = write_month_config(
            "2099-05",
            sales_workbook="data/raw/2099-05/sales.xlsx",
            sales_topology="data/topology/2099-05/sales.topology.json",
            sheet_sources_file="data/raw/2099-05/.staging/sheet_sources.json",
            staging=True,
            config_dir=cfg_dir,
        )
        self.assertTrue(path.exists())
        text = path.read_text(encoding="utf-8")
        self.assertIn(".staging", text)
        self.assertIn("sheet_sources.json", text)
        self.assertIn("2099-05", text)

    def test_single_file_merge_uses_copy_not_openpyxl(self) -> None:
        sales = self.tmp / "销售.xlsx"
        required = required_input_sheets()
        sheet_data = {s.name: [["x"]] for s in required[:5]}
        _make_workbook(sales, sheet_data)
        intake = intake_uploads(
            "2099-04b",
            [("销售.xlsx", sales.read_bytes())],
            staging_root=self.tmp / "staging4b",
        )
        self.assertFalse(needs_openpyxl_merge(intake))
        self.assertEqual(plan_sheet_sources(intake), {})
        out = self.tmp / "merged-copy.xlsx"
        build_consolidated_workbook(intake, out)
        self.assertEqual(out.read_bytes(), sales.read_bytes())

    def test_trial_preview_uses_staging_commission_path(self) -> None:
        import yaml

        cfg_dir = self.tmp / "cfg-trial"
        config_path = write_month_config(
            "2099-07",
            sales_workbook="data/raw/2099-07/.staging/销售账套-合并.xlsx",
            sales_topology="data/topology/2099-07/销售.topology.json",
            staging=True,
            config_dir=cfg_dir,
        )
        cfg = yaml.safe_load(config_path.read_text(encoding="utf-8"))
        self.assertEqual(
            cfg["outputs"]["commission_summary_file"],
            "output/2099-07/.staging/提成汇总.xlsx",
        )
        self.assertNotEqual(
            cfg["outputs"]["commission_summary_file"],
            "output/2099-07/提成汇总.xlsx",
        )
        self.assertTrue(config_path.exists())

    def test_supplement_sheet_sources_finds_task_upload(self) -> None:
        import yaml

        from salary_pipeline.data_ingestion.data_loader import build_workbook_loader
        from salary_pipeline.paths import resolve_project_path

        config = yaml.safe_load(
            (Path(__file__).resolve().parents[1] / "salary_pipeline/config/month.yaml").read_text(
                encoding="utf-8"
            )
        )
        sales_path = resolve_project_path(config["workbooks"]["sales"])
        task_upload = sales_path.parent / "uploads" / "销售任务及完成率.xlsx"
        if not task_upload.is_file():
            self.skipTest("2026-02 销售任务及完成率 upload not present")

        supplemented = supplement_sheet_sources(sales_path, {})
        self.assertIn("销售任务及完成率", supplemented)
        self.assertEqual(supplemented["销售任务及完成率"].resolve(), task_upload.resolve())

        loader = build_workbook_loader({"month_config": config})
        frame = loader.read_sales_task_sheet()
        tang = frame[frame["姓名"] == "唐鹏"]
        self.assertFalse(tang.empty)
        self.assertEqual(float(tang.iloc[0]["考核量"]), 5.0)
        self.assertEqual(float(tang.iloc[0]["实际销量"]), 3.0)

    def test_supplement_sheet_sources_finds_mortgage_perf_upload(self) -> None:
        from salary_pipeline.data_ingestion.data_loader import WorkbookLoader
        from salary_pipeline.modules.mortgage_clerk_performance import (
            MORTGAGE_PERF_SHEET,
            MortgageClerkPerformanceModule,
        )

        base = self.tmp / "base.xlsx"
        mortgage = self.tmp / "按揭绩效.xlsx"
        _make_workbook(base, {"终端明细表": [["a"]]})
        af_row = [""] * 31 + [3985.8325]
        _make_workbook(
            mortgage,
            {
                MORTGAGE_PERF_SHEET: [[] for _ in range(12)] + [af_row, af_row],
            },
        )
        uploads_dir = self.tmp / "uploads"
        uploads_dir.mkdir()
        shutil.copy2(mortgage, uploads_dir / "按揭绩效.xlsx")

        supplemented = supplement_sheet_sources(base, {})
        self.assertIn(MORTGAGE_PERF_SHEET, supplemented)

        loader = WorkbookLoader(base, sheet_paths=supplemented)
        self.assertTrue(loader.has_sheet(MORTGAGE_PERF_SHEET))
        self.assertEqual(float(loader.read_cell_value(MORTGAGE_PERF_SHEET, "AF13")), 3985.8325)

        ctx = {
            "month_config": {"workbooks": {"sales": str(base)}},
            "sheet_sources": supplemented,
            "summary_skeleton": pd.DataFrame(
                [
                    {"店别": "西物", "职务": "按揭内勤", "姓名": "熊宇"},
                ]
            ),
        }
        result = MortgageClerkPerformanceModule().run(ctx)
        self.assertEqual(result.metadata.get("source_sheet"), MORTGAGE_PERF_SHEET)
        self.assertTrue(result.metadata.get("sheet_available"))
        row = result.metrics[result.metrics["姓名"] == "熊宇"].iloc[0]
        self.assertAlmostEqual(float(row["加装绩效"]), 3985.8325, places=2)

    def test_mortgage_clerk_missing_sheet_does_not_crash(self) -> None:
        from salary_pipeline.modules.mortgage_clerk_performance import (
            MortgageClerkPerformanceModule,
        )

        base = self.tmp / "base.xlsx"
        _make_workbook(base, {"终端明细表": [["a"]]})
        ctx = {
            "month_config": {"workbooks": {"sales": str(base)}},
            "summary_skeleton": pd.DataFrame(
                [
                    {"店别": "西物", "职务": "按揭内勤", "姓名": "熊宇"},
                ]
            ),
        }
        result = MortgageClerkPerformanceModule().run(ctx)
        self.assertFalse(result.metadata.get("sheet_available"))
        row = result.metrics[result.metrics["姓名"] == "熊宇"].iloc[0]
        self.assertEqual(float(row["加装绩效"]), 0.0)

    def test_overrides_roundtrip(self) -> None:
        import pandas as pd

        df = pd.DataFrame(
            [{"店别": "A", "职务": "顾问", "姓名": "张三", "整车绩效": 100.0}]
        )
        ov_path = self.tmp / "overrides.json"
        store_sheet_override(ov_path, "提成汇总", df)
        loaded = apply_overrides(df, "提成汇总", load_overrides(ov_path))
        self.assertEqual(float(loaded.iloc[0]["整车绩效"]), 100.0)


if __name__ == "__main__":
    unittest.main()
