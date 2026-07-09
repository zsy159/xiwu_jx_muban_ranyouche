"""Skeleton row-key resolution when upload workbook lacks 提成汇总."""

from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from salary_pipeline.data_ingestion.data_loader import (
    read_personnel_skeleton_keys,
    resolve_personnel_workbook,
    resolve_summary_skeleton_source,
    resolve_summary_skeleton_workbook,
)
from salary_pipeline.modules.base import PERSONNEL_FILENAME, PERSONNEL_SHEET
from salary_pipeline.ingestion_upload.trial_run import resolve_trial_sales_topology
from salary_pipeline.modules.summary_skeleton import SummarySkeletonModule
from salary_pipeline.paths import PROJECT_ROOT


def _make_workbook(path: Path, sheets: dict[str, list[list]]) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    for name, rows in sheets.items():
        ws = wb.create_sheet(name)
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
    wb.save(path)


class SummarySkeletonFallbackTests(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.tmp = Path(self._tmp.name)

    def tearDown(self) -> None:
        self._tmp.cleanup()

    def test_resolve_prefers_upload_when_sheet_present(self) -> None:
        golden = (
            PROJECT_ROOT
            / "data/raw/2026-05/燃油车-2026年05月西物超市销售提成(终)(1).xlsx"
        )
        if not golden.exists():
            self.skipTest("canonical golden workbook not available")

        cfg = {
            "workbooks": {"sales": str(golden)},
            "parity": {"golden_workbook": str(golden)},
            "outputs": {"commission_summary_sheet": "提成汇总"},
        }
        path, source = resolve_summary_skeleton_workbook(cfg)
        self.assertIsNotNone(path)
        self.assertEqual(source, "sales")

    def test_resolve_falls_back_to_canonical_for_merge_workbook(self) -> None:
        merged = PROJECT_ROOT / "data/raw/2026-05/销售账套-合并-2026-05.xlsx"
        golden = (
            PROJECT_ROOT
            / "data/raw/2026-05/燃油车-2026年05月西物超市销售提成(终)(1).xlsx"
        )
        if not merged.exists() or not golden.exists():
            self.skipTest("2026-05 workbooks not available")

        cfg = {
            "workbooks": {"sales": str(merged)},
            "parity": {"golden_workbook": str(merged)},
            "outputs": {"commission_summary_sheet": "提成汇总"},
        }
        path, source = resolve_summary_skeleton_workbook(cfg)
        self.assertEqual(path.resolve(), golden.resolve())
        self.assertEqual(source, "canonical")

    def test_resolve_prefers_personnel_over_canonical(self) -> None:
        merged = PROJECT_ROOT / "data/raw/2026-05/销售账套-合并-2026-05.xlsx"
        if not merged.exists():
            self.skipTest("merged workbook not available")

        personnel = self.tmp / PERSONNEL_FILENAME
        _make_workbook(
            personnel,
            {
                "Sheet1": [
                    ["店别", "职务", "姓名"],
                    ["金牛", "销售顾问", "张三"],
                    ["金牛", "销售顾问", "李四"],
                ]
            },
        )
        sources_path = self.tmp / "sheet_sources.json"
        sources_path.write_text(
            json.dumps({PERSONNEL_SHEET: str(personnel)}, ensure_ascii=False),
            encoding="utf-8",
        )

        cfg = {
            "workbooks": {
                "sales": str(merged),
                "sheet_sources_file": str(sources_path),
            },
            "parity": {"golden_workbook": None},
            "outputs": {"commission_summary_sheet": "提成汇总"},
        }
        path, source, read_sheet = resolve_summary_skeleton_source(cfg)
        self.assertEqual(path.resolve(), personnel.resolve())
        self.assertEqual(source, "personnel")
        self.assertEqual(read_sheet, PERSONNEL_SHEET)

    def test_read_personnel_skeleton_with_abc_columns(self) -> None:
        personnel = self.tmp / PERSONNEL_FILENAME
        _make_workbook(
            personnel,
            {"数据": [["金牛", "销售顾问", "王五"], ["高新", "邀约专员", "赵六"]]},
        )
        skeleton = read_personnel_skeleton_keys(personnel, data_start_row=3)
        self.assertEqual(len(skeleton), 2)
        self.assertEqual(list(skeleton.columns), ["店别", "职务", "姓名", "_excel_row"])
        self.assertEqual(int(skeleton.iloc[0]["_excel_row"]), 3)
        self.assertEqual(skeleton.iloc[0]["姓名"], "王五")

    def test_module_uses_personnel_upload(self) -> None:
        merged = PROJECT_ROOT / "data/raw/2026-05/销售账套-合并-2026-05.xlsx"
        if not merged.exists():
            self.skipTest("merged workbook not available")

        personnel = self.tmp / PERSONNEL_FILENAME
        _make_workbook(
            personnel,
            {
                PERSONNEL_SHEET: [
                    ["店别", "职务", "姓名"],
                    ["金牛", "销售顾问", "测试甲"],
                ]
            },
        )
        sources_path = self.tmp / "sheet_sources.json"
        sources_path.write_text(
            json.dumps({PERSONNEL_SHEET: str(personnel)}, ensure_ascii=False),
            encoding="utf-8",
        )

        cfg = {
            "workbooks": {
                "sales": str(merged),
                "sheet_sources_file": str(sources_path),
            },
            "parity": {
                "golden_workbook": None,
                "header_row": 2,
                "data_start_row": 3,
            },
            "outputs": {"commission_summary_sheet": "提成汇总"},
        }
        result = SummarySkeletonModule().run({"month_config": cfg})
        self.assertEqual(result.metadata["bootstrap"], "personnel")
        self.assertEqual(result.metadata["rows"], 1)
        self.assertEqual(result.metrics.iloc[0]["姓名"], "测试甲")

    def test_module_runs_with_upload_only_workbook(self) -> None:
        merged = PROJECT_ROOT / "data/raw/2026-05/销售账套-合并-2026-05.xlsx"
        if not merged.exists():
            self.skipTest("merged workbook not available")

        cfg = {
            "workbooks": {"sales": str(merged)},
            "parity": {
                "golden_workbook": str(merged),
                "header_row": 2,
                "data_start_row": 3,
            },
            "outputs": {"commission_summary_sheet": "提成汇总"},
        }
        result = SummarySkeletonModule().run({"month_config": cfg})
        self.assertGreater(result.metadata["rows"], 0)
        self.assertIn(result.metadata["bootstrap"], ("canonical", "reference_golden_workbook"))

    def test_null_golden_workbook_uses_canonical(self) -> None:
        merged = PROJECT_ROOT / "data/raw/2026-05/销售账套-合并-2026-05.xlsx"
        if not merged.exists():
            self.skipTest("merged workbook not available")

        cfg = {
            "workbooks": {"sales": str(merged)},
            "parity": {
                "golden_workbook": None,
                "header_row": 2,
                "data_start_row": 3,
            },
            "outputs": {"commission_summary_sheet": "提成汇总"},
        }
        result = SummarySkeletonModule().run({"month_config": cfg})
        self.assertGreater(result.metadata["rows"], 0)

    def test_resolve_personnel_from_uploads_dir(self) -> None:
        sales = self.tmp / "销售账套-合并.xlsx"
        _make_workbook(sales, {"终端明细表": [["h"]]})
        uploads = self.tmp / "uploads"
        uploads.mkdir()
        personnel = uploads / PERSONNEL_FILENAME
        _make_workbook(
            personnel,
            {PERSONNEL_SHEET: [["店别", "职务", "姓名"], ["店", "职务", "人"]]},
        )
        cfg = {"workbooks": {"sales": str(sales)}}
        path = resolve_personnel_workbook(cfg)
        self.assertIsNotNone(path)
        self.assertEqual(path.resolve(), personnel.resolve())


class TrialTopologyFallbackTests(unittest.TestCase):
    def test_upload_topology_without_hub_uses_canonical(self) -> None:
        upload_topo = (
            "data/topology/2026-05/销售账套-合并-2026-05.topology.json"
        )
        canonical = (
            "data/topology/2026-05/燃油车-2026年05月西物超市销售提成(终)(1).topology.json"
        )
        if not (PROJECT_ROOT / upload_topo).exists():
            self.skipTest("upload topology not available")
        if not (PROJECT_ROOT / canonical).exists():
            self.skipTest("canonical topology not available")

        rel, source = resolve_trial_sales_topology(upload_topo, "2026-05")
        self.assertEqual(source, "canonical")
        self.assertEqual(rel, canonical)

    def test_canonical_topology_kept_when_upload_has_hub(self) -> None:
        canonical = (
            "data/topology/2026-05/燃油车-2026年05月西物超市销售提成(终)(1).topology.json"
        )
        if not (PROJECT_ROOT / canonical).exists():
            self.skipTest("canonical topology not available")

        rel, source = resolve_trial_sales_topology(canonical, "2026-05")
        self.assertEqual(source, "upload")
        self.assertEqual(rel, canonical)


if __name__ == "__main__":
    unittest.main()
