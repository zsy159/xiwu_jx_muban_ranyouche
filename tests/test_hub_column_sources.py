"""Tests for 提成汇总 source annotation row."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from salary_pipeline.pipelines.commission_summary import (
    EXPORT_HEADER_ROW,
    EXPORT_SOURCE_ROW,
    CommissionSummaryBuilder,
)
from salary_pipeline.pipelines.commission_summary_column_sources import (
    HUB_PERF_SUMIF_MAP,
    build_source_annotation_row,
    manual_column_headers,
    source_annotation_for_header,
)


class HubColumnSourceTests(unittest.TestCase):
    def test_w_ai_perf_sumif_labels(self) -> None:
        self.assertEqual(
            source_annotation_for_header("整车绩效"),
            "绩效整理表!AG SUMIF(P,D)×完成率(门店块×BA合并完成率/个人块×H销量完成率，见拓扑)",
        )
        self.assertEqual(
            source_annotation_for_header("权限结余绩效"),
            "绩效整理表!AH SUMIF(P,D)",
        )
        self.assertEqual(
            source_annotation_for_header("加装绩效"),
            "绩效整理表!AI SUMIF(P,D)×H",
        )
        self.assertEqual(
            source_annotation_for_header("上户绩效"),
            "绩效整理表!AN+AS SUMIF(P,D)",
        )
        self.assertIn("整车绩效", HUB_PERF_SUMIF_MAP)

    def test_f_p_margin_and_task_labels(self) -> None:
        self.assertEqual(
            source_annotation_for_header("集客达成率"),
            "销售任务及完成率!F INDEX-MATCH(C,D)",
        )
        self.assertEqual(
            source_annotation_for_header("加装额"),
            "绩效整理表!S SUMIF(P,D)",
        )
        self.assertEqual(
            source_annotation_for_header("整车毛利"),
            "绩效整理表!BG SUMIF(P,D)",
        )
        self.assertEqual(
            source_annotation_for_header("保险渗透率"),
            "绩效整理表!K SUMIFS(AB>0,P)/SUMIF(K,P)",
        )

    def test_adjustment_and_ar_labels(self) -> None:
        self.assertEqual(
            source_annotation_for_header("综合项"),
            "综合表!L SUMIF(B,D)",
        )
        self.assertEqual(
            source_annotation_for_header("04月活动"),
            "'重功超期+活动'!X SUMIF(Q,D)",
        )
        self.assertEqual(
            source_annotation_for_header("超期"),
            "绩效整理表!AU SUMIF(P,D)",
        )
        self.assertEqual(
            source_annotation_for_header("（已发放奖励）"),
            "综合表!J SUMIF(B,D)",
        )
        self.assertEqual(
            source_annotation_for_header("保客考核"),
            "保客考核明细!J SUMIF(E,D)",
        )
        self.assertNotIn(
            "HubFormulaEngine",
            " ".join(build_source_annotation_row(["整车毛利", "综合项", "保客考核"])),
        )

    def test_build_source_row_no_golden_reference(self) -> None:
        labels = build_source_annotation_row(["姓名", "整车绩效", "预算单台"])
        self.assertNotIn("金标准", " ".join(labels))
        self.assertEqual(labels[2], "需手工填入")

    def test_non_frontline_manual_columns_marked_for_gray_highlight(self) -> None:
        """岗位绩效/业绩绩效(1/2)/新能源专项：金标准验证为人工填入，
        应纳入 manual_column_headers 以便导出时对空单元格加灰色标注
        （见 hub_column_rules.yaml 非一线管理 delegate: manual_semantic）。"""
        headers = [
            "岗位绩效",
            "业绩绩效",
            "业绩绩效1",
            "业绩绩效2",
            "新能源专项",
            "台次",
            "整车绩效",
        ]
        manual = manual_column_headers(headers)
        for h in ("岗位绩效", "业绩绩效", "业绩绩效1", "业绩绩效2", "新能源专项"):
            self.assertIn(h, manual)
        # 台次/整车绩效 由系统计算（毛利派生 / HubRuleEngine），不应标记为手工
        self.assertNotIn("台次", manual)
        self.assertNotIn("整车绩效", manual)

    def test_export_includes_source_row_above_headers(self) -> None:
        df = pd.DataFrame(
            {
                "店别": ["西物"],
                "职务": ["销售顾问"],
                "姓名": ["张三"],
                "整车绩效": [100.0],
            }
        )
        builder = CommissionSummaryBuilder(
            template_columns=["店别", "职务", "姓名", "整车绩效"]
        )
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "提成汇总.xlsx"
            builder.export_excel(df, path)
            wb = load_workbook(path)
            ws = wb["提成汇总"]
            self.assertIn("绩效整理表!AG", str(ws.cell(row=EXPORT_SOURCE_ROW, column=4).value))
            self.assertEqual(
                ws.cell(row=EXPORT_HEADER_ROW, column=4).value,
                "整车绩效",
            )


if __name__ == "__main__":
    unittest.main()
