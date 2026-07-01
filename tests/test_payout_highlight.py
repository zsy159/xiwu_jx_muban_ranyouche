"""Payout parity Excel highlighting tests."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from salary_pipeline.data_ingestion.data_loader import read_computed_payout_excel
from salary_pipeline.pipelines.payout_formatting import (
    PAYOUT_COMPUTED_DATA_START_ROW,
    PAYOUT_COMPUTED_HEADER_ROW,
    PAYOUT_LEGEND_INSERT_ROW,
    apply_payout_highlighting,
    resolve_payout_compare_columns,
)
from salary_pipeline.pipelines.xw_payout import payout_export_columns
from salary_pipeline.pipelines.xw_payout_formula_engine import XW_COLUMN_MAP
from salary_pipeline.utils.excel_format import (
    PARITY_MISMATCH_FILL_COMMENT,
    PARITY_MISMATCH_FILL_RGB,
    highlight_commission_summary_mismatches,
)
from salary_pipeline.validation.parity import CellMismatch, CommissionSummaryParity


def _export_computed_payout(frame: pd.DataFrame, path: Path, sheet: str = "XW提成-发") -> None:
    columns = payout_export_columns(XW_COLUMN_MAP)
    export = pd.DataFrame(columns=columns)
    for col in columns:
        export[col] = frame[col] if col in frame.columns else pd.NA
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        title = pd.DataFrame([[sheet]], columns=[sheet])
        title.to_excel(writer, sheet_name=sheet, index=False, header=False, startrow=0)
        export.to_excel(writer, sheet_name=sheet, index=False, startrow=2)


def _export_golden_payout(frame: pd.DataFrame, path: Path, sheet: str = "XW提成-发") -> None:
    """Golden layout: row 2 headers, row 3+ data keyed by Excel letters."""
    from openpyxl import Workbook

    wb = Workbook()
    if wb.active is not None:
        wb.remove(wb.active)
    ws = wb.create_sheet(sheet)
    ws["A1"] = "title"
    headers = ["序号", "店别", "职务", "姓名", "人数", *XW_COLUMN_MAP.values()]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=2, column=col_idx, value=header)
    letter_by_name = {name: letter for letter, name in XW_COLUMN_MAP.items()}
    for row_offset, (_, row) in enumerate(frame.iterrows(), start=3):
        ws.cell(row=row_offset, column=1, value=row_offset - 2)
        ws.cell(row=row_offset, column=2, value=row.get("店别"))
        ws.cell(row=row_offset, column=3, value=row.get("职务"))
        ws.cell(row=row_offset, column=4, value=row.get("姓名"))
        ws.cell(row=row_offset, column=5, value=1)
        for col_name, value in row.items():
            if col_name in {"店别", "职务", "姓名"}:
                continue
            letter = letter_by_name.get(col_name)
            if letter is None:
                continue
            ws.cell(
                row=row_offset,
                column=column_index_from_string(letter),
                value=value,
            )
    wb.save(path)
    wb.close()


class PayoutHighlightTests(unittest.TestCase):
    def test_highlight_mismatch_cells_with_comments(self) -> None:
        join_keys = ["店别", "职务", "姓名"]
        df = pd.DataFrame(
            [
                {
                    "店别": "武侯展厅",
                    "职务": "销售顾问",
                    "姓名": "张三",
                    "考核量": 8,
                    "整车绩效": 101.0,
                },
            ]
        )
        mismatch = CellMismatch(
            join_values=(
                ("店别", "武侯展厅"),
                ("职务", "销售顾问"),
                ("姓名", "张三"),
            ),
            column="整车绩效",
            golden_value=100.0,
            computed_value=101.0,
        )
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "XW提成-发.xlsx"
            _export_computed_payout(df, path)
            highlighted = highlight_commission_summary_mismatches(
                path,
                "XW提成-发",
                [mismatch],
                join_keys,
                ["整车绩效"],
                header_row=PAYOUT_COMPUTED_HEADER_ROW,
                data_start_row=PAYOUT_COMPUTED_DATA_START_ROW,
            )
            self.assertEqual(highlighted, 1)

            wb = load_workbook(path)
            ws = wb["XW提成-发"]
            perf_col = payout_export_columns(XW_COLUMN_MAP).index("整车绩效") + 1
            cell = ws.cell(row=PAYOUT_COMPUTED_DATA_START_ROW, column=perf_col)
            self.assertEqual(cell.fill.start_color.rgb, PARITY_MISMATCH_FILL_RGB)
            assert cell.comment is not None
            self.assertIn(PARITY_MISMATCH_FILL_COMMENT, cell.comment.text)
            self.assertIn("金标准: 100", cell.comment.text)
            self.assertIn("系统: 101", cell.comment.text)

    def test_apply_payout_highlighting_inserts_legend_and_marks_mismatch(self) -> None:
        month_config = {
            "parity": {},
            "payout": {
                "xw": {
                    "anchor_sheet": "XW提成-发",
                },
            },
            "payout_parity": {
                "join_keys": ["店别", "职务", "姓名"],
                "columns": ["整车绩效"],
                "numeric_tolerance": 1e-4,
                "auto_highlight": True,
            },
        }
        golden_df = pd.DataFrame(
            [
                {
                    "店别": "武侯展厅",
                    "职务": "销售顾问",
                    "姓名": "张三",
                    "整车绩效": 100.0,
                },
            ]
        )
        computed_df = golden_df.copy()
        computed_df.loc[0, "整车绩效"] = 200.0

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            golden_path = tmp_path / "golden.xlsx"
            computed_path = tmp_path / "computed.xlsx"
            _export_golden_payout(golden_df, golden_path)
            _export_computed_payout(computed_df, computed_path)

            stats = apply_payout_highlighting(
                month_config, computed_path, "xw", golden_path=golden_path
            )
            self.assertGreaterEqual(stats.mismatches, 1)

            wb = load_workbook(computed_path)
            ws = wb["XW提成-发"]
            self.assertIn("数值不一致", ws.cell(row=PAYOUT_LEGEND_INSERT_ROW, column=6).value)
            data_row = PAYOUT_COMPUTED_DATA_START_ROW + 1
            perf_col = payout_export_columns(XW_COLUMN_MAP).index("整车绩效") + 1
            cell = ws.cell(row=data_row, column=perf_col)
            self.assertEqual(cell.fill.start_color.rgb, PARITY_MISMATCH_FILL_RGB)

    def test_collect_payout_mismatches_from_files(self) -> None:
        join_keys = ["店别", "职务", "姓名"]
        golden = pd.DataFrame(
            [
                {
                    "店别": "武侯展厅",
                    "职务": "销售顾问",
                    "姓名": "张三",
                    "整车绩效": 100.0,
                },
            ]
        )
        computed = golden.copy()
        computed.loc[0, "整车绩效"] = 150.0

        with tempfile.TemporaryDirectory() as tmp:
            golden_path = Path(tmp) / "golden.xlsx"
            computed_path = Path(tmp) / "computed.xlsx"
            _export_golden_payout(golden, golden_path)
            _export_computed_payout(computed, computed_path)

            checker = CommissionSummaryParity(
                join_keys=join_keys,
                columns=["整车绩效"],
                literal_columns=True,
            )
            mismatches = checker.collect_payout_mismatches_from_files(
                computed_path,
                golden_path,
                "XW提成-发",
                XW_COLUMN_MAP,
                data_start_row=3,
            )
            self.assertEqual(len(mismatches), 1)
            self.assertEqual(mismatches[0].column, "整车绩效")
            self.assertEqual(mismatches[0].golden_value, 100.0)
            self.assertEqual(mismatches[0].computed_value, 150.0)

    def test_read_computed_payout_adjusts_for_legend_row(self) -> None:
        df = pd.DataFrame(
            [
                {
                    "店别": "武侯展厅",
                    "职务": "销售顾问",
                    "姓名": "张三",
                    "整车绩效": 100.0,
                },
            ]
        )
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "computed.xlsx"
            _export_computed_payout(df, path)
            month_config = {
                "payout": {"xw": {"anchor_sheet": "XW提成-发"}},
                "payout_parity": {
                    "columns": ["整车绩效"],
                    "auto_highlight": True,
                },
            }
            golden_path = Path(tmp) / "golden.xlsx"
            _export_golden_payout(df, golden_path)
            apply_payout_highlighting(
                month_config, path, "xw", golden_path=golden_path
            )
            loaded = read_computed_payout_excel(path, "XW提成-发")
            self.assertEqual(len(loaded), 1)
            self.assertIn("店别", loaded.columns)

    def test_resolve_payout_compare_columns_includes_extended_metrics(self) -> None:
        parity_cfg = {
            "columns": ["整车绩效", "提成合计"],
            "compare_all_metrics": True,
        }
        column_map = {"F": "整车绩效", "I": "权限结余绩效", "U": "提成合计"}
        cols = resolve_payout_compare_columns(parity_cfg, column_map)
        self.assertEqual(cols[0], "整车绩效")
        self.assertIn("权限结余绩效", cols)
        self.assertIn("提成合计", cols)

    def test_apply_payout_highlighting_marks_permission_balance_mismatch(self) -> None:
        month_config = {
            "payout": {"xw": {"anchor_sheet": "XW提成-发"}},
            "payout_parity": {
                "join_keys": ["店别", "职务", "姓名"],
                "columns": ["整车绩效"],
                "compare_all_metrics": True,
                "numeric_tolerance": 1e-4,
                "auto_highlight": True,
            },
        }
        golden_df = pd.DataFrame(
            [
                {
                    "店别": "武侯DCC",
                    "职务": "销售顾问",
                    "姓名": "蒲喜",
                    "整车绩效": 553.0,
                    "权限结余绩效": -747.3054,
                },
            ]
        )
        computed_df = golden_df.copy()
        computed_df.loc[0, "权限结余绩效"] = -964.0254

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            golden_path = tmp_path / "golden.xlsx"
            computed_path = tmp_path / "computed.xlsx"
            _export_golden_payout(golden_df, golden_path)
            _export_computed_payout(computed_df, computed_path)

            stats = apply_payout_highlighting(
                month_config, computed_path, "xw", golden_path=golden_path
            )
            self.assertGreaterEqual(stats.mismatches, 1)

            wb = load_workbook(computed_path)
            ws = wb["XW提成-发"]
            data_row = PAYOUT_COMPUTED_DATA_START_ROW + 1
            perf_col = payout_export_columns(XW_COLUMN_MAP).index("权限结余绩效") + 1
            cell = ws.cell(row=data_row, column=perf_col)
            self.assertEqual(cell.fill.start_color.rgb, PARITY_MISMATCH_FILL_RGB)
            assert cell.comment is not None
            self.assertIn("金标准", cell.comment.text)
            self.assertIn("系统", cell.comment.text)


if __name__ == "__main__":
    unittest.main()
