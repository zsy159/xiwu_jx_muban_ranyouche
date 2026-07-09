"""Tests for Phase B 绩效整理表 builder (Slice 1–4)."""

from __future__ import annotations

import re
import unittest

import pandas as pd

from salary_pipeline.data_ingestion.data_loader import WorkbookLoader, load_month_config
from salary_pipeline.data_ingestion.performance_sheet_golden import DATA_START_ROW
from salary_pipeline.modules.summary_skeleton import SummarySkeletonModule
from salary_pipeline.ops.basic import sumif_by_key
from salary_pipeline.paths import CONFIG_DIR, PROJECT_ROOT, resolve_project_path
from salary_pipeline.pipelines.hub_formula_engine import HubFormulaEngine
from salary_pipeline.pipelines.performance_sheet_builder import (
    PerformanceSheetBuilder,
    SLICE_2_COLUMNS,
    SLICE_3_COLUMNS,
    SLICE_4_COLUMNS,
    sumif_advisor_performance,
)

GOLDEN = PROJECT_ROOT / "data/raw/2026-05/燃油车-2026年05月西物超市销售提成(终)(1).xlsx"
TOLERANCE = 1e-4
ORDER_G_PATTERN = re.compile(r"^DC\d+")

# 何宇 — Hub AC 上户绩效相关；AJ 列顾问级 SUMIF 与金标准一致
HE_YU_AJ_SUMIF = 787.386

ADVISOR_SUMIF_COLUMNS_SLICE_2 = ("AJ", "AK", "AO", "AP")
ADVISOR_SUMIF_COLUMNS_SLICE_3 = ADVISOR_SUMIF_COLUMNS_SLICE_2 + ("AL",)
VEHICLE_COST_COLUMNS = ("AW", "AX", "AY", "AZ", "BA", "BB")


def _golden_perf_frame(loader: WorkbookLoader, columns: tuple[str, ...]) -> pd.DataFrame:
    cols = {"P": "P", **{c: c for c in columns}}
    raw = loader.read_sheet_columns("绩效整理表", cols, label="golden perf advisor gate")
    frame = raw.iloc[DATA_START_ROW - 1 :].reset_index(drop=True)
    for col in columns:
        frame[col] = pd.to_numeric(frame[col], errors="coerce")
    return frame


ORDER_KEY_COLUMNS = ("O", "P", "K", "G")


def _golden_order_skeleton(loader: WorkbookLoader) -> pd.DataFrame:
    from salary_pipeline.data_ingestion.performance_sheet_golden import (
        load_performance_order_skeleton,
    )

    return load_performance_order_skeleton(loader, key_cols=ORDER_KEY_COLUMNS)


@unittest.skipUnless(GOLDEN.exists(), "golden workbook missing")
class PerformanceSheetBuilderTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        config = load_month_config(CONFIG_DIR)
        cls.config = config
        cls.loader = WorkbookLoader(resolve_project_path(config["workbooks"]["sales"]))
        cls.builder = PerformanceSheetBuilder(cls.loader)
        cls.topology = resolve_project_path(config["topology"]["sales"])
        skeleton = SummarySkeletonModule().run({"month_config": config}).metrics
        cls.advisor_skeleton = skeleton[skeleton["职务"] == "销售顾问"].copy()
        cls.advisors = cls.advisor_skeleton["姓名"].tolist()

    def test_slice_1_ab_ak_full_match(self) -> None:
        built = self.builder.build_slice_1()
        for col in ("AB", "AK"):
            with self.subTest(column=col):
                golden = self.builder.lookup_golden_column(col)
                merged = built[["O", col]].merge(golden, on="O", suffixes=("_built", "_golden"))
                diff = (
                    merged[f"{col}_built"].fillna(0)
                    - merged[f"{col}_golden"].fillna(0)
                ).abs()
                mismatches = (diff > TOLERANCE).sum()
                self.assertEqual(
                    mismatches,
                    0,
                    f"{col}: {mismatches} mismatches / {len(merged)}",
                )

    def test_slice_1_aj_mostly_matches(self) -> None:
        built = self.builder.build_slice_1()
        golden = self.builder.lookup_golden_column("AJ")
        merged = built[["O", "AJ"]].merge(golden, on="O", suffixes=("_built", "_golden"))
        diff = (
            merged["AJ_built"].fillna(0) - merged["AJ_golden"].fillna(0)
        ).abs()
        mismatches = int((diff > TOLERANCE).sum())
        self.assertLessEqual(mismatches, 4, f"AJ mismatches={mismatches}, expected <=4")
        match_rate = 1.0 - mismatches / len(merged)
        self.assertGreaterEqual(match_rate, 0.99)

    def test_slice_2_ao_ap_full_match(self) -> None:
        built = self.builder.build_slice_2()
        for col in ("AO", "AP"):
            with self.subTest(column=col):
                golden = self.builder.lookup_golden_column(col)
                merged = built[["O", col]].merge(golden, on="O", suffixes=("_built", "_golden"))
                diff = (
                    merged[f"{col}_built"].fillna(0)
                    - merged[f"{col}_golden"].fillna(0)
                ).abs()
                mismatches = int((diff > TOLERANCE).sum())
                self.assertEqual(
                    mismatches,
                    0,
                    f"{col}: {mismatches} mismatches / {len(merged)}",
                )

    def test_advisor_sumif_aj_he_yu(self) -> None:
        built = self.builder.build_slice_1()
        total = sumif_advisor_performance(built, "何宇", "AJ")
        self.assertAlmostEqual(total, HE_YU_AJ_SUMIF, places=2)

    def test_advisor_sumif_gate_slice_2(self) -> None:
        """49 可比销售顾问：SUMIF(P, 姓名, col) 与金标准绩效整理表一致。"""
        built = self.builder.build_slice_2()
        golden = _golden_perf_frame(self.loader, ADVISOR_SUMIF_COLUMNS_SLICE_2)
        self.assertGreaterEqual(len(self.advisors), 49)
        for col in ADVISOR_SUMIF_COLUMNS_SLICE_2:
            with self.subTest(column=col):
                mismatches = 0
                for name in self.advisors:
                    built_sum = sumif_advisor_performance(built, name, col)
                    golden_sum = float(sumif_by_key(golden, "P", col, name))
                    if abs(built_sum - golden_sum) > TOLERANCE:
                        mismatches += 1
                if col == "AJ":
                    self.assertLessEqual(mismatches, 4, f"AJ advisor mismatches={mismatches}")
                else:
                    self.assertEqual(mismatches, 0, f"{col} advisor mismatches={mismatches}")

    def test_hub_engine_computed_perf_frame_af_ag(self) -> None:
        """注入 builder 输出后 Hub AF/AO、AG/AP SUMIF 与金标准一致（抽样）。"""
        built = self.builder.build_slice_2()
        engine = HubFormulaEngine(
            self.topology, self.loader, computed_perf_frame=built
        )
        summary = pd.DataFrame(
            [
                {
                    "店别": None,
                    "职务": "销售顾问",
                    "姓名": "唐鹏",
                    "_excel_row": 4,
                }
            ]
        )
        out = engine.apply(summary)
        golden_engine = HubFormulaEngine(self.topology, self.loader)
        golden_out = golden_engine.apply(summary)
        for col in ("特殊车型+指定车型", "座位险提成"):
            self.assertAlmostEqual(
                float(out.loc[0, col]),
                float(golden_out.loc[0, col]),
                places=2,
                msg=col,
            )

    def test_aggregate_by_advisor_matches_sumif(self) -> None:
        built = self.builder.build_slice_1()
        agg = self.builder.aggregate_by_advisor(built, "AK")
        self.assertIn("韩柏成", agg.index)
        manual = float(sumif_by_key(built, "P", "AK", "韩柏成"))
        self.assertAlmostEqual(float(agg["韩柏成"]), manual, places=4)

    def test_build_slice_2_row_count(self) -> None:
        built = self.builder.build_slice_2()
        self.assertGreaterEqual(len(built), 470)
        for col in ("O", "P", *SLICE_2_COLUMNS):
            self.assertIn(col, built.columns)

    def test_build_slice_1_row_count(self) -> None:
        built = self.builder.build_slice_1()
        self.assertGreaterEqual(len(built), 470)
        for col in ("O", "P", "AB", "AJ", "AK"):
            self.assertIn(col, built.columns)

    def test_slice_3_vehicle_cost_full_match(self) -> None:
        built = self.builder.build_slice_3()
        for col in VEHICLE_COST_COLUMNS:
            with self.subTest(column=col):
                golden = self.builder.lookup_golden_column(col)
                merged = built[["O", col]].merge(
                    golden, on="O", suffixes=("_built", "_golden")
                )
                diff = (
                    merged[f"{col}_built"].fillna(0)
                    - merged[f"{col}_golden"].fillna(0)
                ).abs()
                mismatches = int((diff > TOLERANCE).sum())
                if col == "BB":
                    self.assertLessEqual(mismatches, 1, f"{col} mismatches={mismatches}")
                else:
                    self.assertEqual(
                        mismatches,
                        0,
                        f"{col}: {mismatches} mismatches / {len(merged)}",
                    )

    def test_slice_3_al_advisor_gate(self) -> None:
        """AL 顾问级 SUMIF 与金标准一致（行级允许金标准空白边界）。"""
        built = self.builder.build_slice_3()
        golden = _golden_perf_frame(self.loader, ("AL",))
        mismatches = 0
        for name in self.advisors:
            built_sum = sumif_advisor_performance(built, name, "AL")
            golden_sum = float(sumif_by_key(golden, "P", "AL", name))
            if abs(built_sum - golden_sum) > TOLERANCE:
                mismatches += 1
        self.assertEqual(mismatches, 0, f"AL advisor mismatches={mismatches}")

    def test_slice_3_al_row_boundary(self) -> None:
        built = self.builder.build_slice_3()
        golden = self.builder.lookup_golden_column("AL")
        merged = built[["O", "AL"]].merge(golden, on="O", suffixes=("_built", "_golden"))
        diff = (
            merged["AL_built"].fillna(0) - merged["AL_golden"].fillna(0)
        ).abs()
        mismatches = int((diff > TOLERANCE).sum())
        self.assertLessEqual(mismatches, 8, f"AL row mismatches={mismatches}")

    def test_slice_3_bh_order_rows(self) -> None:
        """BH 在真实订单号行（G=DC…）与金标准一致。"""
        built = self.builder.build_slice_3()
        golden = self.builder.lookup_golden_column("BH").rename(columns={"BH": "BH_golden"})
        merged = built[["O", "G", "BH"]].merge(golden, on="O")
        order_mask = merged["G"].astype(str).str.match(ORDER_G_PATTERN, na=False)
        diff = (
            merged.loc[order_mask, "BH"].fillna(0)
            - merged.loc[order_mask, "BH_golden"].fillna(0)
        ).abs()
        mismatches = int((diff > TOLERANCE).sum())
        self.assertLessEqual(mismatches, 1, f"BH order-row mismatches={mismatches}")

    def test_advisor_sumif_gate_slice_3(self) -> None:
        built = self.builder.build_slice_3()
        golden = _golden_perf_frame(self.loader, ADVISOR_SUMIF_COLUMNS_SLICE_3)
        for col in ADVISOR_SUMIF_COLUMNS_SLICE_3:
            with self.subTest(column=col):
                mismatches = 0
                for name in self.advisors:
                    built_sum = sumif_advisor_performance(built, name, col)
                    golden_sum = float(sumif_by_key(golden, "P", col, name))
                    if abs(built_sum - golden_sum) > TOLERANCE:
                        mismatches += 1
                if col == "AJ":
                    self.assertLessEqual(mismatches, 4, f"AJ advisor mismatches={mismatches}")
                else:
                    self.assertEqual(mismatches, 0, f"{col} advisor mismatches={mismatches}")

    def test_hub_engine_computed_perf_frame_ad_profit(self) -> None:
        """注入 builder 输出后 Hub AD(盈利产品绩效) SUMIF(AL) 与金标准一致（抽样）。"""
        built = self.builder.build_slice_3()
        engine = HubFormulaEngine(
            self.topology, self.loader, computed_perf_frame=built
        )
        summary = pd.DataFrame(
            [
                {
                    "店别": None,
                    "职务": "销售顾问",
                    "姓名": "唐鹏",
                    "_excel_row": 4,
                }
            ]
        )
        out = engine.apply(summary)
        golden_engine = HubFormulaEngine(self.topology, self.loader)
        golden_out = golden_engine.apply(summary)
        self.assertAlmostEqual(
            float(out.loc[0, "盈利产品绩效"]),
            float(golden_out.loc[0, "盈利产品绩效"]),
            places=2,
        )

    def test_profit_product_al_chongzhou_advisors(self) -> None:
        """崇州顾问 AL 汇总与金标准一致 → Hub AD 盈利产品绩效一致。"""
        built = self.builder.build_slice_3()
        golden = self.builder.lookup_golden_column("AL")
        expected = {
            "唐鹏": 0.0,
            "赵思梵": 2253.8,
            "陈勇建": 660.0,
            "宁袁": 200.0,
        }
        for name, golden_sum in expected.items():
            with self.subTest(advisor=name):
                built_sum = sumif_advisor_performance(built, name, "AL")
                self.assertAlmostEqual(built_sum, golden_sum, places=2)
                merged = built[built["P"] == name][["O", "AL"]].merge(
                    golden, on="O", suffixes=("_built", "_golden")
                )
                diff = (
                    merged["AL_built"].fillna(0) - merged["AL_golden"].fillna(0)
                ).abs()
                self.assertEqual(int((diff > TOLERANCE).sum()), 0)

    def test_profit_product_al_store_block_and_adjustment(self) -> None:
        """门店块 P 不算 AL；薛祥建 VIN 应用 -1229.5 尾项调整。"""
        built = self.builder.build_slice_3()
        store_vin = "LB37852Z9TS057814"
        store_row = built.loc[built["O"] == store_vin]
        self.assertFalse(store_row.empty)
        self.assertAlmostEqual(float(store_row.iloc[0]["AL"] or 0), 0.0, places=2)

        adj_vin = "L6T798NE4TT750321"
        adj_row = built.loc[built["O"] == adj_vin]
        self.assertFalse(adj_row.empty)
        self.assertAlmostEqual(float(adj_row.iloc[0]["AL"] or 0), 0.0, places=2)

    def test_build_slice_3_row_count(self) -> None:
        built = self.builder.build_slice_3()
        self.assertGreaterEqual(len(built), 470)
        for col in ("O", "P", "G", *SLICE_3_COLUMNS):
            self.assertIn(col, built.columns)

    def test_slice_4_order_keys_match_golden(self) -> None:
        """Computed O/P/K/G parity with golden (join on VIN)."""
        built = self.builder.build_slice_4()
        golden = _golden_order_skeleton(self.loader)
        built_vin = built[built["O"].notna()].copy()
        self.assertEqual(len(built_vin), len(golden))
        self.assertEqual(set(built_vin["O"]), set(golden["O"]))
        key_frame = built_vin[["O", "P", "K", "G"]].merge(
            golden[["O", "P", "K", "G"]], on="O", suffixes=("_built", "_golden")
        )
        self.assertEqual(len(key_frame), len(golden))
        for col in ("P", "K", "G"):
            with self.subTest(column=col):
                left = key_frame[f"{col}_built"]
                right = key_frame[f"{col}_golden"]
                if col == "K":
                    left = left.fillna(0)
                    right = right.fillna(0)
                else:
                    left = left.astype(str)
                    right = right.astype(str)
                mismatches = int((left != right).sum())
                self.assertEqual(
                    mismatches,
                    0,
                    f"{col}: {mismatches} mismatches / {len(key_frame)}",
                )

    def test_slice_4_computed_columns_parity(self) -> None:
        """Slice 3 value columns still match golden when skeleton is computed."""
        built = self.builder.build_slice_4()
        for col in SLICE_3_COLUMNS:
            with self.subTest(column=col):
                golden = self.builder.lookup_golden_column(col)
                merged = built[["O", col]].merge(
                    golden, on="O", suffixes=("_built", "_golden")
                )
                diff = (
                    merged[f"{col}_built"].fillna(0)
                    - merged[f"{col}_golden"].fillna(0)
                ).abs()
                mismatches = int((diff > TOLERANCE).sum())
                if col == "AJ":
                    self.assertLessEqual(mismatches, 4, f"AJ mismatches={mismatches}")
                elif col == "AL":
                    self.assertEqual(mismatches, 0, f"AL mismatches={mismatches}")
                elif col == "BH":
                    order_mask = built["G"].astype(str).str.match(
                        ORDER_G_PATTERN, na=False
                    )
                    order_vins = set(built.loc[order_mask, "O"])
                    merged_orders = merged[merged["O"].isin(order_vins)]
                    diff_o = (
                        merged_orders[f"{col}_built"].fillna(0)
                        - merged_orders[f"{col}_golden"].fillna(0)
                    ).abs()
                    mismatches = int((diff_o > TOLERANCE).sum())
                    self.assertLessEqual(mismatches, 1, f"BH order-row mismatches")
                elif col == "BB":
                    self.assertLessEqual(mismatches, 1, f"BB mismatches={mismatches}")
                else:
                    self.assertEqual(
                        mismatches,
                        0,
                        f"{col}: {mismatches} mismatches / {len(merged)}",
                    )

    def test_slice_4_advisor_sumif_gate(self) -> None:
        built = self.builder.build_slice_4()
        golden = _golden_perf_frame(self.loader, ADVISOR_SUMIF_COLUMNS_SLICE_3)
        for col in ADVISOR_SUMIF_COLUMNS_SLICE_3:
            with self.subTest(column=col):
                mismatches = 0
                for name in self.advisors:
                    built_sum = sumif_advisor_performance(built, name, col)
                    golden_sum = float(sumif_by_key(golden, "P", col, name))
                    if abs(built_sum - golden_sum) > TOLERANCE:
                        mismatches += 1
                if col == "AJ":
                    self.assertLessEqual(mismatches, 4, f"AJ advisor mismatches={mismatches}")
                else:
                    self.assertEqual(mismatches, 0, f"{col} advisor mismatches={mismatches}")

    def test_build_slice_4_row_count(self) -> None:
        built = self.builder.build_slice_4()
        self.assertGreaterEqual(len(built), 470)
        for col in (*ORDER_KEY_COLUMNS, *SLICE_4_COLUMNS):
            self.assertIn(col, built.columns)

    def test_hub_w_ai_gate_slice_5_computed_overlay(self) -> None:
        """Slice 5: full advisor skeleton W–AI with computed_perf_frame injection."""
        built = self.builder.build_slice_5()
        golden_engine = HubFormulaEngine(self.topology, self.loader)
        computed_engine = HubFormulaEngine(
            self.topology, self.loader, computed_perf_frame=built
        )
        advisor_rows = self.advisor_skeleton.copy()
        golden_out = golden_engine.apply(advisor_rows)
        computed_out = computed_engine.apply(advisor_rows.copy())
        perf_cols = (
            "整车绩效",
            "加装绩效",
            "保险绩效",
            "金融绩效",
            "爱车宝绩效",
            "上户绩效",
            "盈利产品绩效",
            "特殊车型+指定车型",
            "座位险提成",
        )
        mismatches = 0
        for pos in range(len(advisor_rows)):
            for col in perf_cols:
                g = float(golden_out.iloc[pos][col]) if pd.notna(golden_out.iloc[pos][col]) else 0.0
                c = float(computed_out.iloc[pos][col]) if pd.notna(computed_out.iloc[pos][col]) else 0.0
                if abs(g - c) > TOLERANCE:
                    mismatches += 1
        self.assertEqual(mismatches, 0, f"computed overlay W–AI mismatches={mismatches}")

    def test_decoration_u_and_ai_for_order_dc2026003524(self) -> None:
        """装饰台账文本列保留后 U=100，AI≈0.0012（非 12+）。"""
        from salary_pipeline.calculators.performance_sheet.from_closure import (
            _compute_decoration_u,
        )

        built = self.builder.build_slice_5()
        order_mask = built["G"].astype(str) == "DC2026003524"
        self.assertTrue(order_mask.any(), "expected order DC2026003524 in skeleton")
        row = built.loc[order_mask].iloc[0]
        u = _compute_decoration_u(pd.Series(["DC2026003524"]), self.loader)
        self.assertAlmostEqual(float(u.iloc[0]), 100.0, places=2)
        self.assertAlmostEqual(float(row["AI"]), 0.0012, places=3)

    def test_supplement_row_ai_zero_for_shanghu(self) -> None:
        """服务补录行（G=上户收入，K 空）AI 应为 0。"""
        built = self.builder.build_slice_5()
        supp = built[built["G"].astype(str) == "上户收入"]
        self.assertGreaterEqual(len(supp), 1)
        for _, row in supp.iterrows():
            self.assertTrue(pd.isna(row["K"]) or row["K"] is None)
            self.assertAlmostEqual(float(row["AI"] or 0), 0.0, places=4)

    def test_supplement_row_ah_blank_not_150_floor(self) -> None:
        """服务补录行 AH 应为空，不触发 150 保底。"""
        built = self.builder.build_slice_5()
        supplement_g = {
            "置换服务",
            "爱车保",
            "置换佣金",
            "上户收入",
            "延保服务",
        }
        supp = built[built["G"].astype(str).isin(supplement_g)]
        self.assertGreaterEqual(len(supp), 1)
        for _, row in supp.iterrows():
            self.assertTrue(
                pd.isna(row["AH"]),
                f"supplement G={row['G']} O={row['O']} AH should be blank, got {row['AH']}",
            )

    def test_supplement_ah_skip_improves_hub_x_for_key_advisors(self) -> None:
        """补录行 AH 跳过后，陈勇建/丁小玲 Hub 权限结余绩效与金标准一致。"""
        built = self.builder.build_slice_5()
        golden_engine = HubFormulaEngine(self.topology, self.loader)
        computed_engine = HubFormulaEngine(
            self.topology, self.loader, computed_perf_frame=built
        )
        for name in ("陈勇建", "丁小玲"):
            with self.subTest(advisor=name):
                row = self.advisor_skeleton[
                    self.advisor_skeleton["姓名"] == name
                ].copy()
                self.assertFalse(row.empty, f"missing advisor {name}")
                computed = float(
                    computed_engine.apply(row).iloc[0]["权限结余绩效"]
                )
                golden = float(
                    golden_engine.apply(row).iloc[0]["权限结余绩效"]
                )
                self.assertAlmostEqual(computed, golden, places=2)

    def test_chongzhou_advisor_decoration_perf(self) -> None:
        """崇州顾问加装绩效 Hub 汇总与金标准一致（AI 闭包修复回归）。"""
        built = self.builder.build_slice_5()
        golden_engine = HubFormulaEngine(self.topology, self.loader)
        computed_engine = HubFormulaEngine(
            self.topology, self.loader, computed_perf_frame=built
        )
        expected = {
            "唐鹏": 0.00144,
            "陈勇建": 117.606,
            "宁袁": 0.00216,
        }
        for name, golden_val in expected.items():
            with self.subTest(advisor=name):
                row = self.advisor_skeleton[
                    self.advisor_skeleton["姓名"] == name
                ].copy()
                self.assertFalse(row.empty, f"missing advisor {name}")
                computed = float(
                    computed_engine.apply(row).iloc[0]["加装绩效"]
                )
                golden = float(golden_engine.apply(row).iloc[0]["加装绩效"])
                self.assertAlmostEqual(computed, golden, places=3)
                self.assertAlmostEqual(computed, golden_val, places=3)

    def test_manager_permission_x_matches_golden(self) -> None:
        """经理权限 X：59 笔 VIN 常数 1000，其余按 0 参与 Y=V+X。"""
        from salary_pipeline.calculators.performance_sheet.order_context import (
            enrich_order_context,
        )

        skeleton = _golden_order_skeleton(self.loader)
        ctx = enrich_order_context(skeleton, self.loader)
        golden_x = self.builder.lookup_golden_column("X")
        merged = ctx[["O", "X"]].merge(golden_x, on="O", suffixes=("_built", "_golden"))
        diff = (
            merged["X_built"].fillna(0) - merged["X_golden"].fillna(0)
        ).abs()
        mismatches = int((diff > TOLERANCE).sum())
        self.assertEqual(mismatches, 0, f"X mismatches={mismatches}/{len(merged)}")
        self.assertEqual(int((merged["X_built"] == 1000).sum()), 59)

    def test_manager_permission_ah_x1000_standard_rows(self) -> None:
        """X=1000 的 57 笔标准 AH 公式行与金标准一致（另 2 笔为手工 AA*0.4 覆盖）。"""
        from salary_pipeline.calculators.performance_sheet.order_context import (
            enrich_order_context,
        )
        from openpyxl import load_workbook

        built = self.builder.build_slice_5()
        sk = _golden_order_skeleton(self.loader)
        ctx = enrich_order_context(sk, self.loader)
        x1000_vins = set(ctx.loc[ctx["X"] == 1000, "O"].astype(str).str.strip())

        wb_f = load_workbook(GOLDEN, data_only=False)
        ws_f = wb_f["绩效整理表"]
        aa04_vins: set[str] = set()
        for r in range(2, ws_f.max_row + 1):
            o = ws_f[f"O{r}"].value
            if o is None:
                continue
            vin = str(o).strip()
            if vin not in x1000_vins:
                continue
            ah_f = ws_f[f"AH{r}"].value
            if isinstance(ah_f, str) and "0.4" in ah_f:
                aa04_vins.add(vin)

        golden_ah = self.builder.lookup_golden_column("AH")
        subset = built[built["O"].astype(str).str.strip().isin(x1000_vins - aa04_vins)]
        merged = subset[["O", "AH"]].merge(golden_ah, on="O", suffixes=("_built", "_golden"))
        diff = (
            merged["AH_built"].fillna(0) - merged["AH_golden"].fillna(0)
        ).abs()
        mismatches = int((diff > TOLERANCE).sum())
        self.assertEqual(mismatches, 0, f"X=1000 standard AH mismatches={mismatches}")

    def test_manager_permission_hub_x_key_advisors(self) -> None:
        """经理权限 X 修复后，崇州/华阳等顾问 Hub 权限结余绩效与金标准一致。"""
        built = self.builder.build_slice_5()
        golden_engine = HubFormulaEngine(self.topology, self.loader)
        computed_engine = HubFormulaEngine(
            self.topology, self.loader, computed_perf_frame=built
        )
        for name in ("赵思梵", "陈勇建", "宁袁", "何宇", "唐鹏", "丁小玲", "冯超"):
            with self.subTest(advisor=name):
                row = self.advisor_skeleton[
                    self.advisor_skeleton["姓名"] == name
                ].copy()
                if row.empty:
                    continue
                computed = float(
                    computed_engine.apply(row).iloc[0]["权限结余绩效"]
                )
                golden = float(
                    golden_engine.apply(row).iloc[0]["权限结余绩效"]
                )
                self.assertAlmostEqual(computed, golden, places=2)

    def test_zhao_sifan_order_with_manager_permission(self) -> None:
        """赵思梵 LB37822ZXSB207636：X=1000 使 Y=50；标准 AH=AA×20%=10（金标准该格手工 =AA×0.4）。"""
        from salary_pipeline.calculators.performance_sheet.order_context import (
            enrich_order_context,
        )

        built = self.builder.build_slice_5()
        vin = "LB37822ZXSB207636"
        row = built[built["O"].astype(str).str.strip() == vin]
        self.assertEqual(len(row), 1)
        ctx = enrich_order_context(row, self.loader)
        self.assertAlmostEqual(float(ctx["X"].iloc[0]), 1000.0, places=2)
        self.assertAlmostEqual(float(ctx["Y"].iloc[0]), 50.0, places=2)
        self.assertAlmostEqual(float(row["AH"].iloc[0]), 10.0, places=2)


if __name__ == "__main__":
    unittest.main()
