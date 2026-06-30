"""Parity formula-anomaly Excel annotation tests."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

from salary_pipeline.calculators.sales_advisor.parity_annotations import (
    HubCellAnnotation,
    annotations_for_workbook,
    collect_hub_cell_annotations,
    detect_topology_formula_anomalies,
    enrich_cell_mismatches,
    load_annotation_registry,
    lookup_mismatch_root_cause,
)
from salary_pipeline.pipelines.commission_summary import CommissionSummaryBuilder
from salary_pipeline.utils.excel_format import (
    FORMULA_ANOMALY_FILL_RGB,
    add_commission_summary_annotations,
)
from salary_pipeline.validation.parity import CellMismatch


class ParityAnnotationTests(unittest.TestCase):
    def test_load_registry_includes_zhaosifan(self) -> None:
        registry = load_annotation_registry()
        keys = {ann.key() for ann in registry}
        self.assertIn(("赵思梵", "权限结余绩效"), keys)
        zsf = next(a for a in registry if a.name == "赵思梵")
        self.assertIn("LB37822ZXSB207636", zsf.reason)
        self.assertIn("AA×0.4", zsf.reason)

    def test_comment_attached_to_correct_cell(self) -> None:
        ann = HubCellAnnotation(
            name="赵思梵",
            column="权限结余绩效",
            reason="绩效整理表 VIN LB37822ZXSB207636：金标准 AH=AA×0.4，系统 AA×20%",
            golden_value=1484.20,
            computed_value=1474.20,
        )
        df = pd.DataFrame(
            [
                {
                    "店别": "西物",
                    "职务": "销售顾问",
                    "姓名": "赵思梵",
                    "权限结余绩效": 1474.20,
                },
                {
                    "店别": "西物",
                    "职务": "销售顾问",
                    "姓名": "刘波",
                    "权限结余绩效": 100.0,
                },
            ]
        )
        builder = CommissionSummaryBuilder(template_columns=list(df.columns))
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "提成汇总.xlsx"
            builder.export_excel(df, path)

            applied = add_commission_summary_annotations(path, "提成汇总", [ann])
            self.assertEqual(applied, 1)

            wb = load_workbook(path)
            ws = wb["提成汇总"]
            col = list(df.columns).index("权限结余绩效") + 1
            cell = ws.cell(row=3, column=col)
            self.assertEqual(cell.fill.start_color.rgb, FORMULA_ANOMALY_FILL_RGB)
            self.assertIsNotNone(cell.comment)
            assert cell.comment is not None
            self.assertIn("LB37822ZXSB207636", cell.comment.text)
            self.assertIn("1484.2", cell.comment.text)
            self.assertIn("1474.2", cell.comment.text)

            other = ws.cell(row=4, column=col)
            self.assertNotEqual(
                getattr(other.fill.start_color, "rgb", None),
                FORMULA_ANOMALY_FILL_RGB,
            )

    def test_manual_formula_deferred_excluded_from_annotations(self) -> None:
        deferred = {"韩柏成": frozenset({"保险绩效"})}
        with patch(
            "salary_pipeline.calculators.sales_advisor.parity_annotations.load_annotation_registry",
            return_value=[
                HubCellAnnotation(
                    name="韩柏成",
                    column="保险绩效",
                    reason="SUMIFS+600",
                )
            ],
        ):
            with patch(
                "salary_pipeline.calculators.sales_advisor.parity_annotations.detect_topology_formula_anomalies",
                return_value=[],
            ):
                anns = annotations_for_workbook(deferred_cells=deferred)
        keys = {a.key() for a in anns}
        self.assertNotIn(("韩柏成", "保险绩效"), keys)

    def test_deferred_cells_excluded_from_annotations(self) -> None:
        with patch(
            "salary_pipeline.calculators.sales_advisor.parity_annotations.load_annotation_registry",
            return_value=[
                HubCellAnnotation(
                    name="唐操",
                    column="整车绩效",
                    reason="channel mismatch",
                )
            ],
        ):
            with patch(
                "salary_pipeline.calculators.sales_advisor.parity_annotations.detect_topology_formula_anomalies",
                return_value=[],
            ):
                anns = annotations_for_workbook()
        keys = {a.key() for a in anns}
        self.assertNotIn(("唐操", "整车绩效"), keys)

    def test_collect_merges_registry_over_topology(self) -> None:
        registry_ann = HubCellAnnotation(
            name="测试",
            column="权限结余绩效",
            reason="登记原因",
            source="registry",
        )
        topo_ann = HubCellAnnotation(
            name="测试",
            column="权限结余绩效",
            reason="topology 原因",
            source="topology",
        )
        merged = collect_hub_cell_annotations(
            include_topology=False,
        )
        self.assertTrue(any(a.name == "赵思梵" for a in merged))

        manual = collect_hub_cell_annotations(include_topology=False)
        with patch(
            "salary_pipeline.calculators.sales_advisor.parity_annotations.load_annotation_registry",
            return_value=[registry_ann],
        ):
            with patch(
                "salary_pipeline.calculators.sales_advisor.parity_annotations.detect_topology_formula_anomalies",
                return_value=[topo_ann],
            ):
                result = collect_hub_cell_annotations()
        by_key = {a.key(): a for a in result}
        self.assertEqual(by_key[("测试", "权限结余绩效")].reason, "登记原因")
        self.assertEqual(by_key[("测试", "权限结余绩效")].source, "registry")

    def test_topology_ref_skipped_when_hub_row_name_mismatch(self) -> None:
        """Stale hub_excel_row must not attribute #REF! on another person's row."""
        with patch(
            "salary_pipeline.calculators.sales_advisor.parity_annotations._topology_cells",
            return_value={
                "提成汇总!X32": {"formula": "=SUMIFS(绩效整理表!AH:AH,绩效整理表!P:P,#REF!)"},
            },
        ):
            with patch(
                "salary_pipeline.calculators.sales_advisor.parity_annotations.golden_hub_names_by_row",
                return_value={32: "牟春柳"},
            ):
                anns = detect_topology_formula_anomalies(
                    {
                        "roles": [
                            {"name": "刘波", "hub_linked": True, "hub_excel_row": 32},
                        ]
                    },
                    golden_workbook=Path("/tmp/golden.xlsx"),
                )
        self.assertEqual(anns, [])

    def test_topology_ref_kept_when_hub_row_name_matches(self) -> None:
        with patch(
            "salary_pipeline.calculators.sales_advisor.parity_annotations._topology_cells",
            return_value={
                "提成汇总!X32": {"formula": "=SUMIFS(绩效整理表!AH:AH,绩效整理表!P:P,#REF!)"},
            },
        ):
            with patch(
                "salary_pipeline.calculators.sales_advisor.parity_annotations.golden_hub_names_by_row",
                return_value={32: "刘波"},
            ):
                anns = detect_topology_formula_anomalies(
                    {
                        "roles": [
                            {"name": "刘波", "hub_linked": True, "hub_excel_row": 32},
                        ]
                    },
                    golden_workbook=Path("/tmp/golden.xlsx"),
                )
        self.assertEqual(len(anns), 1)
        self.assertEqual(anns[0].name, "刘波")
        self.assertEqual(anns[0].column, "权限结余绩效")

    def test_skip_topology_annotation_when_parity_values_match(self) -> None:
        topo = HubCellAnnotation(
            name="刘波",
            column="权限结余绩效",
            reason="#REF! on wrong row",
            source="topology",
        )
        with patch(
            "salary_pipeline.calculators.sales_advisor.parity_annotations.load_annotation_registry",
            return_value=[],
        ):
            with patch(
                "salary_pipeline.calculators.sales_advisor.parity_annotations.detect_topology_formula_anomalies",
                return_value=[topo],
            ):
                anns = annotations_for_workbook(
                    parity_values={("刘波", "权限结余绩效"): (-1013.25, -1013.25)},
                    deferred_cells={},
                )
        self.assertEqual(anns, [])

    def test_keep_registry_annotation_when_parity_values_match(self) -> None:
        reg_ann = HubCellAnnotation(
            name="赵思梵",
            column="权限结余绩效",
            reason="VIN override",
            source="registry",
        )
        with patch(
            "salary_pipeline.calculators.sales_advisor.parity_annotations.load_annotation_registry",
            return_value=[reg_ann],
        ):
            with patch(
                "salary_pipeline.calculators.sales_advisor.parity_annotations.detect_topology_formula_anomalies",
                return_value=[],
            ):
                anns = annotations_for_workbook(
                    parity_values={("赵思梵", "权限结余绩效"): (100.0, 100.0)},
                    deferred_cells={},
                )
        self.assertEqual(len(anns), 1)
        self.assertEqual(anns[0].source, "registry")

    def test_lookup_liubo_jiazhuange_ref_root_cause(self) -> None:
        mismatch = CellMismatch(
            join_values=(("店别", "西物"), ("职务", "销售顾问"), ("姓名", "刘波")),
            column="加装额",
            golden_value=1300.03,
            computed_value=1700.04,
        )
        with patch(
            "salary_pipeline.calculators.sales_advisor.parity_annotations._topology_cells",
            return_value={
                "提成汇总!J32": {
                    "formula": "=SUMIF(绩效整理表!P:P,#REF!,绩效整理表!S:S)",
                },
            },
        ):
            cause = lookup_mismatch_root_cause(
                mismatch,
                registry={
                    "roles": [
                        {"name": "刘波", "hub_linked": True, "hub_excel_row": 32},
                    ]
                },
            )
        self.assertIn("#REF!", cause)
        self.assertIn("装饰底价(S)", cause)

    def test_enrich_cell_mismatches_attaches_root_cause(self) -> None:
        mismatch = CellMismatch(
            join_values=(("店别", "西物"), ("职务", "销售顾问"), ("姓名", "刘波")),
            column="加装额",
            golden_value=1300.03,
            computed_value=1700.04,
        )
        with patch(
            "salary_pipeline.calculators.sales_advisor.parity_annotations.lookup_mismatch_root_cause",
            return_value="绩效整理表 S 列语义与金标准 SUMIF 源列不一致",
        ):
            enriched = enrich_cell_mismatches([mismatch])
        self.assertEqual(len(enriched), 1)
        self.assertEqual(
            enriched[0].root_cause,
            "绩效整理表 S 列语义与金标准 SUMIF 源列不一致",
        )


if __name__ == "__main__":
    unittest.main()
