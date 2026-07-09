from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from salary_pipeline.pipelines.non_frontline_columns import apply_non_frontline_columns
from salary_pipeline.pipelines.commission_summary import (
    SUMMARY_TEMPLATE_COLUMNS,
    CommissionSummaryBuilder,
)
from salary_pipeline.utils.excel_format import (
    INTEGER_DISPLAY_HEADERS,
    TWO_DECIMAL_FORMAT,
    format_writer_sheet,
)


class ExcelFormatTests(unittest.TestCase):
    def test_format_writer_sheet_applies_two_decimal_display(self) -> None:
        frame = pd.DataFrame(
            {
                "序号": [1],
                "姓名": ["张三"],
                "整车毛利": [1234.5],
                "销量完成率": [0.876],
            }
        )
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "sample.xlsx"
            sheet_name = "提成汇总"
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                frame.to_excel(writer, sheet_name=sheet_name, index=False)
                format_writer_sheet(
                    writer, sheet_name, frame.columns, header_row=1
                )

            wb = load_workbook(path)
            ws = wb[sheet_name]

            self.assertEqual(ws["A2"].number_format, "General")
            self.assertEqual(ws["C2"].number_format, TWO_DECIMAL_FORMAT)
            self.assertEqual(ws["D2"].number_format, TWO_DECIMAL_FORMAT)

    def test_commission_summary_export_formats_numeric_columns(self) -> None:
        summary = pd.DataFrame(
            {
                "序号": [1],
                "店别": ["西物"],
                "职务": ["销售顾问"],
                "姓名": ["李四"],
                "人数": [1],
                "考核量": [10],
                "实际销量": [8],
                "整车毛利": [5000.123],
                "提成合计": [800.1],
            }
        )
        builder = CommissionSummaryBuilder(
            template_columns=list(summary.columns),
        )
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "提成汇总.xlsx"
            builder.export_excel(summary, path)

            wb = load_workbook(path)
            ws = wb["提成汇总"]
            # header row 3, first data row 4
            self.assertEqual(ws["A4"].number_format, "General")
            self.assertEqual(ws["E4"].number_format, "General")
            self.assertEqual(ws["H4"].number_format, TWO_DECIMAL_FORMAT)
            self.assertEqual(ws["I4"].number_format, TWO_DECIMAL_FORMAT)

    def test_commission_summary_export_includes_template_semantic_columns(self) -> None:
        summary = pd.DataFrame(
            {
                "序号": [1],
                "店别": ["财务部"],
                "职务": ["会计"],
                "姓名": ["罗涵"],
                "人数": [1],
                "综合毛利": [742.0],
                "主营单台毛利": [2.0],
                "整车绩效": [2700.0],
                "加装绩效": [1484.0],
            }
        )
        summary = apply_non_frontline_columns(summary)
        builder = CommissionSummaryBuilder()
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "提成汇总.xlsx"
            builder.export_excel(summary, path)
            exported = pd.read_excel(path, header=2, nrows=0)
        self.assertEqual(list(exported.columns), SUMMARY_TEMPLATE_COLUMNS)
        for col in ("台次", "提成系数", "岗位绩效", "业绩绩效1"):
            self.assertIn(col, exported.columns)

    def test_integer_headers_constant(self) -> None:
        self.assertIn("台数", INTEGER_DISPLAY_HEADERS)
        self.assertIn("考核量", INTEGER_DISPLAY_HEADERS)


if __name__ == "__main__":
    unittest.main()
