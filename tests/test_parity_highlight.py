"""Parity mismatch Excel highlighting tests."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from salary_pipeline.pipelines.commission_summary import (
    EXPORT_DATA_START_ROW,
    EXPORT_HEADER_ROW,
    CommissionSummaryBuilder,
)
from salary_pipeline.data_ingestion.data_loader import (
    normalize_header,
    read_computed_summary_excel,
)
from salary_pipeline.calculators.sales_advisor.registry import wa_parity_deferred_cells
from salary_pipeline.calculators.sales_advisor.topology_specs import (
    collect_topology_manual_formula_cells,
    collect_topology_static_fill_cells,
    is_manual_formula_adjustment,
    is_pure_direct_fill_formula,
)
from salary_pipeline.utils.excel_format import (
    FORMULA_ANOMALY_FILL_RGB,
    GOLDEN_STATIC_FILL_RGB,
    MANUAL_DEFERRED_FILL_RGB,
    MANUAL_DEFERRED_FILL_COMMENT,
    PARITY_MISMATCH_FILL_COMMENT,
    PARITY_MISMATCH_FILL_RGB,
    STATIC_FILL_COMMENT,
    add_commission_summary_color_legend,
    format_mismatch_comment_text,
    highlight_commission_summary_deferred_cells,
    highlight_commission_summary_mismatches,
)
from salary_pipeline.validation.parity import (
    CellMismatch,
    CommissionSummaryParity,
    resolve_hub_compare_columns,
)


class ManualFormulaDetectionTests(unittest.TestCase):
    def test_is_pure_direct_fill_formula(self) -> None:
        self.assertTrue(is_pure_direct_fill_formula("=-140"))
        self.assertTrue(is_pure_direct_fill_formula("=100"))
        self.assertTrue(is_pure_direct_fill_formula("=189*20"))
        self.assertTrue(is_pure_direct_fill_formula("=100*0.8"))
        self.assertTrue(is_pure_direct_fill_formula("=500+300"))
        self.assertTrue(is_pure_direct_fill_formula("=462+101+77+102"))
        self.assertFalse(is_pure_direct_fill_formula("=SUMIFS(AJ:AJ,D:D,D3)-100"))
        self.assertFalse(is_pure_direct_fill_formula("=W80*0.8"))
        self.assertFalse(is_pure_direct_fill_formula("100"))

    def test_is_manual_formula_adjustment(self) -> None:
        self.assertFalse(is_manual_formula_adjustment("=-140"))
        self.assertFalse(is_manual_formula_adjustment("=100"))
        self.assertTrue(is_manual_formula_adjustment("=SUMIFS(AJ:AJ,D:D,D3)-100"))
        self.assertTrue(is_manual_formula_adjustment("=AH80-100"))
        self.assertTrue(is_manual_formula_adjustment("=SUMIFS(AJ:AJ,D:D,D3)+600"))
        self.assertFalse(is_manual_formula_adjustment("=189*20"))
        self.assertFalse(is_manual_formula_adjustment("=100*0.8"))
        self.assertFalse(is_manual_formula_adjustment("=500+300"))
        self.assertFalse(is_manual_formula_adjustment("=462+101+77+102"))
        self.assertFalse(is_manual_formula_adjustment("=SUMIFS(AJ:AJ,D:D,D3)"))
        self.assertFalse(is_manual_formula_adjustment("100"))
        self.assertFalse(is_manual_formula_adjustment("=W80*0.8"))

    def test_collect_manual_formula_cells_from_golden_workbook(self) -> None:
        columns = [
            "店别",
            "职务",
            "姓名",
            "权限结余绩效",
            "保险绩效",
            "整车绩效",
        ]
        df = pd.DataFrame(
            [
                {
                    "店别": "西物",
                    "职务": "销售顾问",
                    "姓名": "沈燕1",
                    "权限结余绩效": -140,
                    "保险绩效": 200,
                    "整车绩效": 3500,
                },
                {
                    "店别": "西物",
                    "职务": "销售顾问",
                    "姓名": "韩柏成",
                    "权限结余绩效": 100,
                    "保险绩效": 800,
                    "整车绩效": 1500,
                },
            ]
        )
        builder = CommissionSummaryBuilder(template_columns=columns)
        with tempfile.TemporaryDirectory() as tmp:
            golden_path = Path(tmp) / "golden.xlsx"
            topo_path = Path(tmp) / "topo.json"
            builder.export_excel(df, golden_path)

            wb = load_workbook(golden_path)
            ws = wb["提成汇总"]
            x_col = columns.index("权限结余绩效") + 1
            z_col = columns.index("保险绩效") + 1
            w_col = columns.index("整车绩效") + 1
            ws.cell(row=EXPORT_DATA_START_ROW, column=x_col, value="=-140")
            ws.cell(row=EXPORT_DATA_START_ROW, column=z_col, value="=SUMIFS(AJ:AJ,D:D,D4)+600")
            ws.cell(row=EXPORT_DATA_START_ROW + 1, column=x_col, value="=189*20")
            ws.cell(
                row=EXPORT_DATA_START_ROW + 1,
                column=w_col,
                value="=SUMIFS(绩效整理表!AG:AG,绩效整理表!P:P,D5)",
            )
            wb.save(golden_path)
            wb.close()

            topo_path.write_text(
                '{"cells": {'
                '"提成汇总!W4": {"formula": "=SUMIFS(绩效整理表!AG:AG,绩效整理表!P:P,D4)"}'
                "}}",
                encoding="utf-8",
            )
            manual_cells = collect_topology_manual_formula_cells(
                topology_path=topo_path,
                golden_workbook_path=golden_path,
                header_row=EXPORT_HEADER_ROW,
                data_start_row=EXPORT_DATA_START_ROW,
            )
            self.assertIn(("沈燕1", "销售顾问"), manual_cells)
            self.assertNotIn(
                "权限结余绩效", manual_cells[("沈燕1", "销售顾问")]
            )
            self.assertIn("保险绩效", manual_cells[("沈燕1", "销售顾问")])
            self.assertNotIn(("韩柏成", "销售顾问"), manual_cells)
            self.assertNotIn("整车绩效", manual_cells.get(("韩柏成", "销售顾问"), frozenset()))

    def test_highlight_manual_formula_cells_blue_with_comment(self) -> None:
        columns = ["店别", "职务", "姓名", "权限结余绩效"]
        df = pd.DataFrame(
            [
                {
                    "店别": "西物",
                    "职务": "销售顾问",
                    "姓名": "沈燕1",
                    "权限结余绩效": -140,
                },
            ]
        )
        builder = CommissionSummaryBuilder(template_columns=columns)
        deferred = {"沈燕1": frozenset({"权限结余绩效"})}
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "提成汇总.xlsx"
            builder.export_excel(df, path)

            highlighted = highlight_commission_summary_deferred_cells(
                path,
                "提成汇总",
                deferred,
            )
            self.assertEqual(highlighted, 1)

            wb = load_workbook(path)
            ws = wb["提成汇总"]
            col = columns.index("权限结余绩效") + 1
            cell = ws.cell(row=EXPORT_DATA_START_ROW, column=col)
            self.assertEqual(cell.fill.start_color.rgb, MANUAL_DEFERRED_FILL_RGB)
            self.assertEqual(cell.comment.text, MANUAL_DEFERRED_FILL_COMMENT)


class ParityHighlightTests(unittest.TestCase):
    def test_collect_and_highlight_mismatch_cells(self) -> None:
        join_keys = ["店别", "职务", "姓名"]
        golden = pd.DataFrame(
            [
                {
                    "店别": "西物",
                    "职务": "销售顾问",
                    "姓名": "张三",
                    "整车毛利": 100.0,
                    "整车绩效": 50.0,
                },
                {
                    "店别": "西物",
                    "职务": "销售顾问",
                    "姓名": "李四",
                    "整车毛利": 200.0,
                    "整车绩效": 80.0,
                },
            ]
        )
        computed = golden.copy()
        computed.loc[0, "整车毛利"] = 101.0
        computed.loc[1, "整车绩效"] = 90.0

        checker = CommissionSummaryParity(
            join_keys=join_keys,
            columns=["整车毛利", "整车绩效"],
        )
        mismatches = checker.collect_cell_mismatches(computed, golden)
        self.assertEqual(len(mismatches), 2)
        mismatch_cols = {m.column for m in mismatches}
        self.assertEqual(mismatch_cols, {"整车毛利", "整车绩效"})

        builder = CommissionSummaryBuilder(template_columns=list(computed.columns))
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "提成汇总.xlsx"
            builder.export_excel(computed, path)

            highlighted = highlight_commission_summary_mismatches(
                path,
                "提成汇总",
                mismatches,
                join_keys,
                ["整车毛利", "整车绩效"],
            )
            self.assertEqual(highlighted, 2)

            wb = load_workbook(path)
            ws = wb["提成汇总"]
            self.assertEqual(
                ws[f"D{EXPORT_DATA_START_ROW}"].fill.start_color.rgb,
                PARITY_MISMATCH_FILL_RGB,
            )
            self.assertIsNotNone(ws[f"D{EXPORT_DATA_START_ROW}"].comment)
            assert ws[f"D{EXPORT_DATA_START_ROW}"].comment is not None
            self.assertIn(PARITY_MISMATCH_FILL_COMMENT, ws[f"D{EXPORT_DATA_START_ROW}"].comment.text)
            self.assertIn("金标准: 100", ws[f"D{EXPORT_DATA_START_ROW}"].comment.text)
            self.assertIn("系统: 101", ws[f"D{EXPORT_DATA_START_ROW}"].comment.text)
            self.assertEqual(
                ws[f"E{EXPORT_DATA_START_ROW + 1}"].fill.start_color.rgb,
                PARITY_MISMATCH_FILL_RGB,
            )
            self.assertIsNotNone(ws[f"E{EXPORT_DATA_START_ROW + 1}"].comment)
            self.assertNotEqual(
                ws[f"D{EXPORT_DATA_START_ROW + 1}"].fill.start_color.rgb,
                PARITY_MISMATCH_FILL_RGB,
            )

    def test_wa_parity_deferred_cells_not_collected(self) -> None:
        join_keys = ["店别", "职务", "姓名"]
        golden = pd.DataFrame(
            [
                {
                    "店别": "西物",
                    "职务": "销售顾问",
                    "姓名": "唐操",
                    "整车绩效": 100.0,
                },
            ]
        )
        computed = golden.copy()
        computed.loc[0, "整车绩效"] = 999.0

        checker = CommissionSummaryParity(
            join_keys=join_keys,
            columns=["整车绩效"],
        )
        mismatches = checker.collect_cell_mismatches(computed, golden)
        self.assertEqual(mismatches, [])

    def test_highlight_deferred_cells_by_name_and_role(self) -> None:
        """手工录入格按 姓名+职务 高亮，不依赖 店别。"""
        deferred = wa_parity_deferred_cells()
        self.assertIn("唐操", deferred)
        self.assertIn("整车绩效", deferred["唐操"])

        df = pd.DataFrame(
            [
                {
                    "店别": float("nan"),
                    "职务": "销售顾问",
                    "姓名": "唐操",
                    "整车绩效": 100.0,
                    "保险绩效": 50.0,
                },
                {
                    "店别": "西物",
                    "职务": "销售顾问",
                    "姓名": "刘波",
                    "整车绩效": 200.0,
                    "保险绩效": 80.0,
                },
            ]
        )
        builder = CommissionSummaryBuilder(template_columns=list(df.columns))
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "提成汇总.xlsx"
            builder.export_excel(df, path)

            highlighted = highlight_commission_summary_deferred_cells(
                path,
                "提成汇总",
                deferred,
            )
            self.assertGreaterEqual(highlighted, 1)

            wb = load_workbook(path)
            ws = wb["提成汇总"]
            # 唐操 整车绩效 — deferred (nan 店别 still matches)
            perf_col = list(df.columns).index("整车绩效") + 1
            tang_cell = ws.cell(row=EXPORT_DATA_START_ROW, column=perf_col)
            self.assertEqual(
                tang_cell.fill.start_color.rgb,
                MANUAL_DEFERRED_FILL_RGB,
            )
            # 刘波 — not deferred
            liu_cell = ws.cell(row=EXPORT_DATA_START_ROW + 1, column=perf_col)
            self.assertNotEqual(
                liu_cell.fill.start_color.rgb,
                MANUAL_DEFERRED_FILL_RGB,
            )

    def test_deferred_fill_wins_over_mismatch(self) -> None:
        """Deferred applied after mismatch; manual color takes precedence."""
        join_keys = ["店别", "职务", "姓名"]
        golden = pd.DataFrame(
            [
                {
                    "店别": "西物",
                    "职务": "销售顾问",
                    "姓名": "唐操",
                    "整车绩效": 100.0,
                },
            ]
        )
        computed = golden.copy()
        computed.loc[0, "整车绩效"] = 999.0

        builder = CommissionSummaryBuilder(template_columns=list(computed.columns))
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "提成汇总.xlsx"
            builder.export_excel(computed, path)

            checker = CommissionSummaryParity(
                join_keys=join_keys,
                columns=["整车绩效"],
            )
            mismatches = checker.collect_cell_mismatches(computed, golden)
            self.assertEqual(mismatches, [])

            # Simulate a mismatch highlight on the same cell, then deferred overwrites.
            highlight_commission_summary_mismatches(
                path,
                "提成汇总",
                [
                    CellMismatch(
                        join_values=(("店别", "西物"), ("职务", "销售顾问"), ("姓名", "唐操")),
                        column="整车绩效",
                    )
                ],
                join_keys,
                ["整车绩效"],
            )
            highlight_commission_summary_deferred_cells(
                path,
                "提成汇总",
                wa_parity_deferred_cells(),
            )

            wb = load_workbook(path)
            ws = wb["提成汇总"]
            perf_col = list(computed.columns).index("整车绩效") + 1
            cell = ws.cell(row=EXPORT_DATA_START_ROW, column=perf_col)
            self.assertEqual(cell.fill.start_color.rgb, MANUAL_DEFERRED_FILL_RGB)
            self.assertNotEqual(cell.fill.start_color.rgb, PARITY_MISMATCH_FILL_RGB)

    def test_collect_static_fill_cells_from_golden_workbook(self) -> None:
        """Topology omits formula-less cells; golden workbook is the source of truth."""
        columns = [
            "序号",
            "店别",
            "职务",
            "姓名",
            "人数",
            "考核量",
            "实际销量",
            "销量完成率",
            "整车绩效",
            "加装绩效",
        ]
        df = pd.DataFrame(
            [
                {
                    "序号": 1,
                    "店别": "机场DCC",
                    "职务": "销售助理",
                    "姓名": "王熙鸿",
                    "人数": 1,
                    "考核量": 11,
                    "实际销量": 1,
                    "销量完成率": 1,
                    "整车绩效": 200,
                    "加装绩效": 50,
                },
                {
                    "序号": 2,
                    "店别": "西物-翼真",
                    "职务": "销售顾问",
                    "姓名": "韩柏成",
                    "人数": 1,
                    "考核量": 10,
                    "实际销量": 8,
                    "销量完成率": 0.8,
                    "整车绩效": 1500,
                    "加装绩效": 1000,
                },
            ]
        )
        builder = CommissionSummaryBuilder(template_columns=columns)
        with tempfile.TemporaryDirectory() as tmp:
            golden_path = Path(tmp) / "golden.xlsx"
            topo_path = Path(tmp) / "topo.json"
            builder.export_excel(df, golden_path)

            wb = load_workbook(golden_path)
            ws = wb["提成汇总"]
            ws[f"H{EXPORT_DATA_START_ROW}"] = 1
            y_col = columns.index("加装绩效") + 1
            ws.cell(row=EXPORT_DATA_START_ROW + 1, column=y_col, value=1000)
            wb.save(golden_path)
            wb.close()

            topo_path.write_text(
                '{"cells": {'
                '"提成汇总!W4": {"formula": "=SUMIFS(绩效整理表!AG:AG,绩效整理表!P:P,D4)"},'
                '"提成汇总!W5": {"formula": "=SUMIFS(绩效整理表!AG:AG,绩效整理表!P:P,D5)"}'
                "}}",
                encoding="utf-8",
            )
            static_cells = collect_topology_static_fill_cells(
                topology_path=topo_path,
                golden_workbook_path=golden_path,
                header_row=EXPORT_HEADER_ROW,
                data_start_row=EXPORT_DATA_START_ROW,
            )
            self.assertIn(("王熙鸿", "销售助理"), static_cells)
            self.assertIn("销量完成率", static_cells[("王熙鸿", "销售助理")])
            self.assertIn(("韩柏成", "销售顾问"), static_cells)
            self.assertIn("加装绩效", static_cells[("韩柏成", "销售顾问")])
            self.assertNotIn("整车绩效", static_cells[("韩柏成", "销售顾问")])

    def test_collect_static_includes_pure_arith_and_non_frontline_columns(self) -> None:
        columns = [
            "店别",
            "职务",
            "姓名",
            "综合毛利",
            "主营单台毛利",
            "整车绩效",
            "权限结余绩效",
        ]
        df = pd.DataFrame(
            [
                {
                    "店别": "财务部",
                    "职务": "会计",
                    "姓名": "罗敏",
                    "综合毛利": 742,
                    "主营单台毛利": 2.0,
                    "整车绩效": 5000,
                    "权限结余绩效": 100,
                },
            ]
        )
        builder = CommissionSummaryBuilder(template_columns=columns)
        with tempfile.TemporaryDirectory() as tmp:
            golden_path = Path(tmp) / "golden.xlsx"
            topo_path = Path(tmp) / "topo.json"
            builder.export_excel(df, golden_path)

            wb = load_workbook(golden_path)
            ws = wb["提成汇总"]
            t_col = columns.index("综合毛利") + 1
            ws.cell(row=EXPORT_DATA_START_ROW, column=t_col, value="=462+101+77+102")
            wb.save(golden_path)
            wb.close()

            topo_path.write_text('{"cells": {}}', encoding="utf-8")
            static_cells = collect_topology_static_fill_cells(
                topology_path=topo_path,
                golden_workbook_path=golden_path,
                header_row=EXPORT_HEADER_ROW,
                data_start_row=EXPORT_DATA_START_ROW,
            )
            key = ("罗敏", "会计")
            self.assertIn(key, static_cells)
            self.assertIn("综合毛利", static_cells[key])
            self.assertIn("主营单台毛利", static_cells[key])

            manual_cells = collect_topology_manual_formula_cells(
                topology_path=topo_path,
                golden_workbook_path=golden_path,
                header_row=EXPORT_HEADER_ROW,
                data_start_row=EXPORT_DATA_START_ROW,
            )
            self.assertNotIn("综合毛利", manual_cells.get(key, frozenset()))
            self.assertNotIn("主营单台毛利", manual_cells.get(key, frozenset()))

    def test_pure_arith_static_gray_not_blue_on_highlight(self) -> None:
        columns = [
            "店别",
            "职务",
            "姓名",
            "综合毛利",
            "台次",
            "主营单台毛利",
            "提成系数",
        ]
        df = pd.DataFrame(
            [
                {
                    "店别": "财务部",
                    "职务": "会计",
                    "姓名": "罗敏",
                    "综合毛利": pd.NA,
                    "台次": 742.0,
                    "主营单台毛利": pd.NA,
                    "提成系数": 2.0,
                },
            ]
        )
        builder = CommissionSummaryBuilder(template_columns=columns)
        static_cells = {
            ("罗敏", "会计"): frozenset({"综合毛利", "主营单台毛利"}),
        }
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "提成汇总.xlsx"
            builder.export_excel(df, path)

            highlighted = highlight_commission_summary_deferred_cells(
                path,
                "提成汇总",
                {},
                static_cells=static_cells,
            )
            self.assertEqual(highlighted, 2)

            wb = load_workbook(path)
            ws = wb["提成汇总"]
            taici_col = columns.index("台次") + 1
            coef_col = columns.index("提成系数") + 1
            self.assertEqual(
                ws.cell(row=EXPORT_DATA_START_ROW, column=taici_col).fill.start_color.rgb,
                GOLDEN_STATIC_FILL_RGB,
            )
            self.assertEqual(
                ws.cell(row=EXPORT_DATA_START_ROW, column=coef_col).fill.start_color.rgb,
                GOLDEN_STATIC_FILL_RGB,
            )
            self.assertEqual(
                ws.cell(row=EXPORT_DATA_START_ROW, column=taici_col).comment.text,
                STATIC_FILL_COMMENT,
            )

    def test_sumifs_formula_not_marked_static(self) -> None:
        columns = ["店别", "职务", "姓名", "整车绩效"]
        df = pd.DataFrame(
            [
                {
                    "店别": "西物",
                    "职务": "销售顾问",
                    "姓名": "张三",
                    "整车绩效": 100.0,
                },
            ]
        )
        builder = CommissionSummaryBuilder(template_columns=columns)
        with tempfile.TemporaryDirectory() as tmp:
            golden_path = Path(tmp) / "golden.xlsx"
            topo_path = Path(tmp) / "topo.json"
            builder.export_excel(df, golden_path)

            wb = load_workbook(golden_path)
            ws = wb["提成汇总"]
            w_col = columns.index("整车绩效") + 1
            ws.cell(
                row=EXPORT_DATA_START_ROW,
                column=w_col,
                value="=SUMIFS(绩效整理表!AG:AG,绩效整理表!P:P,D4)",
            )
            wb.save(golden_path)
            wb.close()

            topo_path.write_text(
                '{"cells": {'
                '"提成汇总!W4": {"formula": "=SUMIFS(绩效整理表!AG:AG,绩效整理表!P:P,D4)"}'
                "}}",
                encoding="utf-8",
            )
            static_cells = collect_topology_static_fill_cells(
                topology_path=topo_path,
                golden_workbook_path=golden_path,
                header_row=EXPORT_HEADER_ROW,
                data_start_row=EXPORT_DATA_START_ROW,
            )
            self.assertNotIn("整车绩效", static_cells.get(("张三", "销售顾问"), frozenset()))

    def test_highlight_static_cells_gray_fill_and_comment(self) -> None:
        columns = [
            "店别",
            "职务",
            "姓名",
            "销量完成率",
            "整车绩效",
        ]
        df = pd.DataFrame(
            [
                {
                    "店别": "机场DCC",
                    "职务": "销售助理",
                    "姓名": "王熙鸿",
                    "销量完成率": 1,
                    "整车绩效": 200,
                },
            ]
        )
        builder = CommissionSummaryBuilder(template_columns=columns)
        static_cells = {("王熙鸿", "销售助理"): frozenset({"销量完成率"})}
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "提成汇总.xlsx"
            builder.export_excel(df, path)

            highlighted = highlight_commission_summary_deferred_cells(
                path,
                "提成汇总",
                {},
                static_cells=static_cells,
            )
            self.assertEqual(highlighted, 1)

            wb = load_workbook(path)
            ws = wb["提成汇总"]
            h_col = columns.index("销量完成率") + 1
            cell = ws.cell(row=EXPORT_DATA_START_ROW, column=h_col)
            self.assertEqual(cell.fill.start_color.rgb, GOLDEN_STATIC_FILL_RGB)
            self.assertIsNotNone(cell.comment)
            self.assertEqual(cell.comment.text, STATIC_FILL_COMMENT)

    def test_mismatch_comment_includes_root_cause(self) -> None:
        text = format_mismatch_comment_text(
            golden_value=1300.03,
            computed_value=1700.04,
            root_cause="绩效整理表 S 列语义与金标准 SUMIF 源列不一致",
        )
        self.assertIn("金标准: 1300.03", text)
        self.assertIn("系统: 1700.04", text)
        self.assertIn("原因:", text)
        self.assertIn("S 列语义", text)

    def test_highlight_mismatch_with_root_cause_comment(self) -> None:
        join_keys = ["店别", "职务", "姓名"]
        mismatch = CellMismatch(
            join_values=(("店别", "西物"), ("职务", "销售顾问"), ("姓名", "张三")),
            column="整车毛利",
            golden_value=100.0,
            computed_value=101.0,
            root_cause="F–P 验收层：测试根因",
        )
        builder = CommissionSummaryBuilder(
            template_columns=["店别", "职务", "姓名", "整车毛利"]
        )
        df = pd.DataFrame(
            [{"店别": "西物", "职务": "销售顾问", "姓名": "张三", "整车毛利": 101.0}]
        )
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "提成汇总.xlsx"
            builder.export_excel(df, path)
            highlight_commission_summary_mismatches(
                path,
                "提成汇总",
                [mismatch],
                join_keys,
                ["整车毛利"],
            )
            wb = load_workbook(path)
            ws = wb["提成汇总"]
            col = list(df.columns).index("整车毛利") + 1
            cell = ws.cell(row=EXPORT_DATA_START_ROW, column=col)
            assert cell.comment is not None
            self.assertIn("原因: F–P 验收层：测试根因", cell.comment.text)

    def test_deferred_comment_includes_yaml_reason(self) -> None:
        columns = ["店别", "职务", "姓名", "整车绩效"]
        df = pd.DataFrame(
            [
                {
                    "店别": "西物",
                    "职务": "销售顾问",
                    "姓名": "唐操",
                    "整车绩效": 100.0,
                },
            ]
        )
        builder = CommissionSummaryBuilder(template_columns=columns)
        deferred = {"唐操": frozenset({"整车绩效"})}
        reasons = {
            "唐操": {
                "整车绩效": "渠道 I 个案（L6T78XCZ5TY782006），待核对后修 order_context",
            }
        }
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "提成汇总.xlsx"
            builder.export_excel(df, path)

            highlighted = highlight_commission_summary_deferred_cells(
                path,
                "提成汇总",
                deferred,
                deferred_reasons=reasons,
            )
            self.assertEqual(highlighted, 1)

            wb = load_workbook(path)
            ws = wb["提成汇总"]
            perf_col = columns.index("整车绩效") + 1
            cell = ws.cell(row=EXPORT_DATA_START_ROW, column=perf_col)
            assert cell.comment is not None
            self.assertIn(MANUAL_DEFERRED_FILL_COMMENT, cell.comment.text)
            self.assertIn("渠道 I 个案", cell.comment.text)

    def test_highlight_deferred_cells_blue_fill_and_comment(self) -> None:
        columns = ["店别", "职务", "姓名", "整车绩效"]
        df = pd.DataFrame(
            [
                {
                    "店别": "西物",
                    "职务": "销售顾问",
                    "姓名": "唐操",
                    "整车绩效": 100.0,
                },
            ]
        )
        builder = CommissionSummaryBuilder(template_columns=columns)
        deferred = {"唐操": frozenset({"整车绩效"})}
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "提成汇总.xlsx"
            builder.export_excel(df, path)

            highlighted = highlight_commission_summary_deferred_cells(
                path,
                "提成汇总",
                deferred,
            )
            self.assertEqual(highlighted, 1)

            wb = load_workbook(path)
            ws = wb["提成汇总"]
            perf_col = columns.index("整车绩效") + 1
            cell = ws.cell(row=EXPORT_DATA_START_ROW, column=perf_col)
            self.assertEqual(cell.fill.start_color.rgb, MANUAL_DEFERRED_FILL_RGB)
            self.assertIsNotNone(cell.comment)
            self.assertEqual(cell.comment.text, MANUAL_DEFERRED_FILL_COMMENT)

    def test_static_fill_wins_over_mismatch(self) -> None:
        join_keys = ["店别", "职务", "姓名"]
        columns = ["店别", "职务", "姓名", "加装绩效"]
        golden = pd.DataFrame(
            [
                {
                    "店别": "西物-翼真",
                    "职务": "销售顾问",
                    "姓名": "韩柏成",
                    "加装绩效": 1000.0,
                },
            ]
        )
        computed = golden.copy()
        computed.loc[0, "加装绩效"] = 0.0

        builder = CommissionSummaryBuilder(template_columns=columns)
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "提成汇总.xlsx"
            builder.export_excel(computed, path)

            highlight_commission_summary_mismatches(
                path,
                "提成汇总",
                [
                    CellMismatch(
                        join_values=(
                            ("店别", "西物-翼真"),
                            ("职务", "销售顾问"),
                            ("姓名", "韩柏成"),
                        ),
                        column="加装绩效",
                    )
                ],
                join_keys,
                ["加装绩效"],
            )
            highlight_commission_summary_deferred_cells(
                path,
                "提成汇总",
                {},
                static_cells={("韩柏成", "销售顾问"): frozenset({"加装绩效"})},
            )

            wb = load_workbook(path)
            ws = wb["提成汇总"]
            y_col = columns.index("加装绩效") + 1
            cell = ws.cell(row=EXPORT_DATA_START_ROW, column=y_col)
            self.assertEqual(cell.fill.start_color.rgb, GOLDEN_STATIC_FILL_RGB)

    def test_deferred_wins_over_static_for_same_cell(self) -> None:
        """wa_parity_deferred takes blue fill when cell is also a golden static fill."""
        columns = ["店别", "职务", "姓名", "加装绩效"]
        df = pd.DataFrame(
            [
                {
                    "店别": "西物-翼真",
                    "职务": "销售顾问",
                    "姓名": "韩柏成",
                    "加装绩效": 1000.0,
                },
            ]
        )
        builder = CommissionSummaryBuilder(template_columns=columns)
        deferred = {"韩柏成": frozenset({"加装绩效"})}
        static_cells = {("韩柏成", "销售顾问"): frozenset({"加装绩效"})}
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "提成汇总.xlsx"
            builder.export_excel(df, path)

            highlighted = highlight_commission_summary_deferred_cells(
                path,
                "提成汇总",
                deferred,
                static_cells=static_cells,
            )
            self.assertEqual(highlighted, 1)

            wb = load_workbook(path)
            ws = wb["提成汇总"]
            y_col = columns.index("加装绩效") + 1
            cell = ws.cell(row=EXPORT_DATA_START_ROW, column=y_col)
            self.assertEqual(cell.fill.start_color.rgb, MANUAL_DEFERRED_FILL_RGB)

    def test_cell_mismatch_join_dict(self) -> None:
        mismatch = CellMismatch(
            join_values=(("店别", "西物"), ("职务", "销售顾问"), ("姓名", "张三")),
            column="整车毛利",
        )
        self.assertEqual(
            mismatch.join_dict(),
            {"店别": "西物", "职务": "销售顾问", "姓名": "张三"},
        )

    def test_erwang_blank_ah_skip_for_hub_permission_column(self) -> None:
        """Golden D 二网 → AH blank; computed AH explains Hub 权限结余绩效 diff."""
        from salary_pipeline.validation.golden_perf_skips import (
            hub_parity_skip_erwang_blank_ah,
        )

        join_keys = ["店别", "职务", "姓名"]
        golden = pd.DataFrame(
            [
                {
                    "店别": "武侯DCC",
                    "职务": "销售顾问",
                    "姓名": "蒲喜",
                    "权限结余绩效": -747.3054,
                },
            ]
        )
        computed = golden.copy()
        computed.loc[0, "权限结余绩效"] = -964.0254

        adjustments = {"蒲喜": -216.72}
        checker = CommissionSummaryParity(
            join_keys=join_keys,
            columns=["权限结余绩效"],
            golden_workbook=Path("/tmp/unused-golden.xlsx"),
            computed_perf_path=Path("/tmp/unused-perf.xlsx"),
        )
        checker._erwang_ah_adjustments = adjustments

        mismatches = checker.collect_cell_mismatches(computed, golden)
        self.assertEqual(mismatches, [])
        self.assertTrue(
            hub_parity_skip_erwang_blank_ah(
                "蒲喜",
                "权限结余绩效",
                -747.3054,
                -964.0254,
                adjustments,
            )
        )

    def test_erwang_skip_does_not_apply_without_adjustment(self) -> None:
        join_keys = ["店别", "职务", "姓名"]
        golden = pd.DataFrame(
            [
                {
                    "店别": "武侯DCC",
                    "职务": "销售顾问",
                    "姓名": "刘波",
                    "权限结余绩效": 100.0,
                },
            ]
        )
        computed = golden.copy()
        computed.loc[0, "权限结余绩效"] = 200.0

        checker = CommissionSummaryParity(
            join_keys=join_keys,
            columns=["权限结余绩效"],
        )
        checker._erwang_ah_adjustments = {"蒲喜": -216.72}

        mismatches = checker.collect_cell_mismatches(computed, golden)
        self.assertEqual(len(mismatches), 1)
        self.assertEqual(mismatches[0].column, "权限结余绩效")


class NonFrontlineHighlightTests(unittest.TestCase):
    def test_static_highlight_follows_semantic_column_after_migration(self) -> None:
        columns = [
            "店别",
            "职务",
            "姓名",
            "整车绩效",
            "岗位绩效",
        ]
        df = pd.DataFrame(
            [
                {
                    "店别": "事业部",
                    "职务": "事业部总经理",
                    "姓名": "刘伟生",
                    "整车绩效": pd.NA,
                    "岗位绩效": 6500.0,
                },
            ]
        )
        builder = CommissionSummaryBuilder(template_columns=columns)
        static_cells = {("刘伟生", "事业部总经理"): frozenset({"整车绩效"})}
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "提成汇总.xlsx"
            builder.export_excel(df, path)

            highlighted = highlight_commission_summary_deferred_cells(
                path,
                "提成汇总",
                {},
                static_cells=static_cells,
            )
            self.assertEqual(highlighted, 1)

            wb = load_workbook(path)
            ws = wb["提成汇总"]
            physical_col = columns.index("整车绩效") + 1
            semantic_col = columns.index("岗位绩效") + 1
            self.assertNotEqual(
                ws.cell(row=EXPORT_DATA_START_ROW, column=physical_col).fill.start_color.rgb,
                GOLDEN_STATIC_FILL_RGB,
            )
            self.assertEqual(
                ws.cell(row=EXPORT_DATA_START_ROW, column=semantic_col).fill.start_color.rgb,
                GOLDEN_STATIC_FILL_RGB,
            )

    def test_parity_mismatch_targets_semantic_column(self) -> None:
        join_keys = ["店别", "职务", "姓名"]
        golden = pd.DataFrame(
            [
                {
                    "店别": "事业部",
                    "职务": "事业部总经理",
                    "姓名": "刘伟生",
                    "整车绩效": 6500.0,
                },
            ]
        )
        computed = golden.copy()
        computed["岗位绩效"] = 6400.0
        computed["整车绩效"] = pd.NA

        checker = CommissionSummaryParity(
            join_keys=join_keys,
            columns=["整车绩效"],
        )
        mismatches = checker.collect_cell_mismatches(computed, golden)
        self.assertEqual(len(mismatches), 1)
        self.assertEqual(mismatches[0].column, "岗位绩效")

        builder = CommissionSummaryBuilder(
            template_columns=["店别", "职务", "姓名", "整车绩效", "岗位绩效"]
        )
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "提成汇总.xlsx"
            builder.export_excel(computed, path)
            highlight_commission_summary_mismatches(
                path,
                "提成汇总",
                mismatches,
                join_keys,
                ["整车绩效"],
            )
            wb = load_workbook(path)
            ws = wb["提成汇总"]
            self.assertEqual(
                ws.cell(row=EXPORT_DATA_START_ROW, column=5).fill.start_color.rgb,
                PARITY_MISMATCH_FILL_RGB,
            )
            self.assertNotEqual(
                ws.cell(row=EXPORT_DATA_START_ROW, column=4).fill.start_color.rgb,
                PARITY_MISMATCH_FILL_RGB,
            )


class GatedParityHighlightTests(unittest.TestCase):
    def test_performance_column_scoped_to_family_hub_columns(self) -> None:
        """DCC rows only compare family hub_columns, not all performance_columns."""
        join_keys = ["店别", "职务", "姓名"]
        golden = pd.DataFrame(
            [
                {
                    "店别": "武侯DCC",
                    "职务": "DCC邀约专员",
                    "姓名": "陈文霜",
                    "整车绩效": 4780.0,
                    "加装绩效": 0.0,
                    "保险绩效": 0.0,
                },
            ]
        )
        computed = golden.copy()
        computed.loc[0, "加装绩效"] = pd.NA
        computed.loc[0, "保险绩效"] = pd.NA

        checker = CommissionSummaryParity(
            join_keys=join_keys,
            columns=["整车绩效", "加装绩效", "保险绩效"],
            performance_columns=frozenset(
                {"整车绩效", "加装绩效", "保险绩效"}
            ),
            treat_empty_as_zero=True,
        )
        mismatches = checker.collect_cell_mismatches(computed, golden)
        self.assertEqual(mismatches, [])

    def test_out_of_scope_performance_column_not_collected(self) -> None:
        join_keys = ["店别", "职务", "姓名"]
        golden = pd.DataFrame(
            [
                {
                    "店别": "新媒体销售部",
                    "职务": "主播",
                    "姓名": "王芝婕",
                    "整车绩效": 8672.25,
                    "加装绩效": 0.0,
                },
            ]
        )
        computed = golden.copy()
        computed.loc[0, "加装绩效"] = 999.0

        checker = CommissionSummaryParity(
            join_keys=join_keys,
            columns=["整车绩效", "加装绩效"],
            performance_columns=frozenset({"整车绩效", "加装绩效"}),
        )
        mismatches = checker.collect_cell_mismatches(computed, golden)
        self.assertEqual(mismatches, [])

    def test_empty_equals_zero_for_performance_columns(self) -> None:
        join_keys = ["店别", "职务", "姓名"]
        golden = pd.DataFrame(
            [
                {
                    "店别": "武侯DCC",
                    "职务": "DCC邀约专员",
                    "姓名": "陈文霜",
                    "整车绩效": 4780.0,
                    "整车完成考核": 0.0,
                },
            ]
        )
        computed = golden.copy()
        computed.loc[0, "整车完成考核"] = pd.NA

        checker = CommissionSummaryParity(
            join_keys=join_keys,
            columns=["整车绩效", "整车完成考核"],
            performance_columns=frozenset(
                {"整车绩效", "整车完成考核"}
            ),
            treat_empty_as_zero=True,
        )
        mismatches = checker.collect_cell_mismatches(computed, golden)
        self.assertEqual(mismatches, [])

        checker_strict = CommissionSummaryParity(
            join_keys=join_keys,
            columns=["整车绩效", "整车完成考核"],
            performance_columns=frozenset(
                {"整车绩效", "整车完成考核"}
            ),
            treat_empty_as_zero=False,
        )
        mismatches_strict = checker_strict.collect_cell_mismatches(
            computed, golden
        )
        self.assertEqual(len(mismatches_strict), 1)
        self.assertEqual(mismatches_strict[0].column, "整车完成考核")

    def test_hub_linked_false_advisor_skipped(self) -> None:
        join_keys = ["店别", "职务", "姓名"]
        golden = pd.DataFrame(
            [
                {
                    "店别": "西物",
                    "职务": "销售助理",
                    "姓名": "雷卓远",
                    "整车绩效": 1400.0,
                    "加装绩效": 500.0,
                },
            ]
        )
        computed = golden.copy()
        computed.loc[0, "整车绩效"] = pd.NA
        computed.loc[0, "加装绩效"] = pd.NA

        checker = CommissionSummaryParity(
            join_keys=join_keys,
            columns=["整车绩效", "加装绩效"],
            performance_columns=frozenset({"整车绩效", "加装绩效"}),
        )
        mismatches = checker.collect_cell_mismatches(computed, golden)
        self.assertEqual(mismatches, [])

    def test_metrics_columns_not_gated_by_family(self) -> None:
        """F–P metrics columns compare on all rows regardless of family."""
        join_keys = ["店别", "职务", "姓名"]
        golden = pd.DataFrame(
            [
                {
                    "店别": "新媒体销售部",
                    "职务": "主播",
                    "姓名": "王芝婕",
                    "整车毛利": 100.0,
                },
            ]
        )
        computed = golden.copy()
        computed.loc[0, "整车毛利"] = 200.0

        checker = CommissionSummaryParity(
            join_keys=join_keys,
            columns=["整车毛利", "整车绩效"],
            performance_columns=frozenset({"整车绩效"}),
        )
        mismatches = checker.collect_cell_mismatches(computed, golden)
        self.assertEqual(len(mismatches), 1)
        self.assertEqual(mismatches[0].column, "整车毛利")

    def test_unmapped_role_compares_performance_columns(self) -> None:
        """职务=渠道 等未落入 parity_gate 岗位族时仍比对绩效列。"""
        join_keys = ["店别", "职务", "姓名"]
        golden = pd.DataFrame(
            [
                {
                    "店别": "崇州直营店",
                    "职务": "渠道",
                    "姓名": "余才万3",
                    "整车绩效": 120.0,
                },
            ]
        )
        computed = golden.copy()
        computed.loc[0, "整车绩效"] = pd.NA

        checker = CommissionSummaryParity(
            join_keys=join_keys,
            columns=["整车绩效"],
            performance_columns=frozenset({"整车绩效"}),
            treat_empty_as_zero=True,
        )
        mismatches = checker.collect_cell_mismatches(computed, golden)
        self.assertEqual(len(mismatches), 1)
        self.assertEqual(mismatches[0].column, "整车绩效")
        self.assertEqual(mismatches[0].join_dict()["姓名"], "余才万3")

    def test_resolve_hub_compare_columns_includes_adjustment(self) -> None:
        cols = resolve_hub_compare_columns(
            {
                "columns": ["整车毛利"],
                "performance_columns": ["整车绩效"],
                "adjustment_columns": ["综合项", "04月活动"],
            }
        )
        self.assertEqual(
            cols,
            ["整车毛利", "整车绩效", "综合项", "04月活动"],
        )

    def test_adjustment_columns_compare_without_family_gate(self) -> None:
        join_keys = ["店别", "职务", "姓名"]
        golden = pd.DataFrame(
            [
                {
                    "店别": "西物",
                    "职务": "销售顾问",
                    "姓名": "唐鹏",
                    "综合项": -400.0,
                    "04月活动": 0.0,
                },
            ]
        )
        computed = golden.copy()
        computed.loc[0, "综合项"] = pd.NA

        checker = CommissionSummaryParity(
            join_keys=join_keys,
            columns=["综合项", "04月活动", "整车绩效"],
            performance_columns=frozenset({"整车绩效"}),
            treat_empty_as_zero=True,
        )
        mismatches = checker.collect_cell_mismatches(computed, golden)
        cols = {m.column for m in mismatches}
        self.assertIn("综合项", cols)
        self.assertNotIn("04月活动", cols)

    def test_empty_equals_zero_default_for_metrics_and_adjustment(self) -> None:
        """F–P and adjustment columns use empty≈0 by default (no perf_columns gate)."""
        join_keys = ["店别", "职务", "姓名"]
        golden = pd.DataFrame(
            [
                {
                    "店别": "西物",
                    "职务": "销售顾问",
                    "姓名": "张三",
                    "考核量": 0.0,
                    "04月活动": 0.0,
                },
            ]
        )
        computed = golden.copy()
        computed.loc[0, "考核量"] = pd.NA
        computed.loc[0, "04月活动"] = pd.NA

        checker = CommissionSummaryParity(
            join_keys=join_keys,
            columns=["考核量", "04月活动"],
        )
        mismatches = checker.collect_cell_mismatches(computed, golden)
        self.assertEqual(mismatches, [])

    def test_nonempty_vs_empty_still_mismatches(self) -> None:
        join_keys = ["店别", "职务", "姓名"]
        golden = pd.DataFrame(
            [
                {
                    "店别": "西物",
                    "职务": "销售顾问",
                    "姓名": "李四",
                    "综合项": 120.0,
                },
            ]
        )
        computed = golden.copy()
        computed.loc[0, "综合项"] = pd.NA

        checker = CommissionSummaryParity(
            join_keys=join_keys,
            columns=["综合项"],
        )
        mismatches = checker.collect_cell_mismatches(computed, golden)
        self.assertEqual(len(mismatches), 1)
        self.assertEqual(mismatches[0].column, "综合项")
        self.assertEqual(mismatches[0].golden_value, 120.0)
        self.assertIsNone(mismatches[0].computed_value)


class ColorLegendTests(unittest.TestCase):
    def test_add_commission_summary_color_legend_inserts_row(self) -> None:
        columns = ["店别", "职务", "姓名", "整车绩效"]
        df = pd.DataFrame(
            [{"店别": "西物", "职务": "销售顾问", "姓名": "张三", "整车绩效": 100.0}]
        )
        builder = CommissionSummaryBuilder(template_columns=columns)
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "提成汇总.xlsx"
            builder.export_excel(df, path)

            add_commission_summary_color_legend(path, "提成汇总", insert_at_row=2)

            wb = load_workbook(path)
            ws = wb["提成汇总"]
            self.assertEqual(ws.cell(row=2, column=1).fill.start_color.rgb, GOLDEN_STATIC_FILL_RGB)
            self.assertIn("需要手工填入", ws.cell(row=2, column=2).value)
            self.assertEqual(
                ws.cell(row=2, column=3).fill.start_color.rgb, MANUAL_DEFERRED_FILL_RGB
            )
            self.assertIn("公式含手工", ws.cell(row=2, column=4).value)
            self.assertEqual(
                ws.cell(row=2, column=5).fill.start_color.rgb, PARITY_MISMATCH_FILL_RGB
            )
            self.assertIn("数值不一致", ws.cell(row=2, column=6).value)
            self.assertEqual(
                ws.cell(row=2, column=7).fill.start_color.rgb, FORMULA_ANOMALY_FILL_RGB
            )
            self.assertIn("公式形态异常", ws.cell(row=2, column=8).value)
            self.assertEqual(
                normalize_header(ws.cell(row=EXPORT_HEADER_ROW + 1, column=1).value),
                "店别",
            )


class ReconcileAfterLegendTests(unittest.TestCase):
    def test_computed_highlight_rows_respects_existing_legend(self) -> None:
        from salary_pipeline.pipelines.commission_summary import computed_highlight_rows

        header, data = computed_highlight_rows(
            legend_inserted=False, legend_present=True
        )
        self.assertEqual(header, EXPORT_HEADER_ROW + 1)
        self.assertEqual(data, EXPORT_DATA_START_ROW + 1)

    def test_compare_files_after_legend_insert(self) -> None:
        columns = ["店别", "职务", "姓名", "整车绩效"]
        df = pd.DataFrame(
            [{"店别": "西物", "职务": "销售顾问", "姓名": "张三", "整车绩效": 100.0}]
        )
        builder = CommissionSummaryBuilder(template_columns=columns)
        with tempfile.TemporaryDirectory() as tmp:
            computed = Path(tmp) / "computed.xlsx"
            golden = Path(tmp) / "golden.xlsx"
            builder.export_excel(df, computed)
            builder.export_excel(df, golden)
            add_commission_summary_color_legend(computed, "提成汇总", insert_at_row=2)

            computed_df = read_computed_summary_excel(computed)
            self.assertIn("店别", computed_df.columns)
            self.assertEqual(len(computed_df), 1)

            checker = CommissionSummaryParity(columns=["整车绩效"])
            report = checker.compare_files(
                computed,
                golden,
                "提成汇总",
                header_row=EXPORT_HEADER_ROW,
                data_start_row=EXPORT_DATA_START_ROW,
            )
            self.assertTrue(report.overall_passed)


if __name__ == "__main__":
    unittest.main()
