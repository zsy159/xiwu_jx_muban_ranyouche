#!/usr/bin/env python3
"""Generate 提成汇总-列级计算逻辑.xlsx from system config/docs (no golden values)."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
import yaml
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT = Path(__file__).resolve().parents[1]
OUT_PATH = PROJECT / "docs/iterations/hub/提成汇总-列级计算逻辑.xlsx"

# 2026-05 导出文件 row4 表头 → Excel 列字母（含非一线语义列插入后的物理位置）
EXPORT_HEADER_TO_LETTER: dict[str, str] = {}


def _load_export_letters() -> dict[str, str]:
    hub_path = PROJECT / "output/2026-05/提成汇总.xlsx"
    if not hub_path.exists():
        return {}
    wb = load_workbook(hub_path, read_only=True, data_only=True)
    ws = wb["提成汇总"]
    mapping: dict[str, str] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=4, column=col).value
        if v:
            mapping[str(v).strip()] = get_column_letter(col)
    wb.close()
    return mapping


def _load_yaml(path: Path) -> dict:
    with path.open(encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def _style_sheet(ws, header_row: int = 1) -> None:
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[header_row]:
        if cell.value:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def build_rules_rows(export_letters: dict[str, str]) -> list[dict]:
    """Sheet1: 列级规则总表."""
    rows: list[dict] = []

    def add(
        classic_letter: str,
        header: str,
        logic: str,
        source: str,
        roles: str,
        overlay: str,
        multiplier: str,
        implemented: str,
        note: str = "",
    ) -> None:
        export_col = export_letters.get(header, "")
        rows.append(
            {
                "经典Hub列": classic_letter,
                "导出Excel列": export_col,
                "表头名称": header,
                "计算逻辑（中文）": logic,
                "数据来源表/列": source,
                "岗位族/适用范围": roles,
                "overlay_path": overlay,
                "乘数(H/BA/无)": multiplier,
                "是否已实现": implemented,
                "备注": note,
            }
        )

    # --- F–P 指标列 ---
    add(
        "F", "考核量",
        "按姓名 SUMIF 求和：销售任务及完成率!Y",
        "销售任务及完成率 | 键C(姓名)→值Y(考核量)",
        "全员（按姓名匹配，无任务=0）",
        "hub_metrics_rule_engine",
        "无",
        "✅ 已实现",
        "HubMetricsRuleEngine；与职务无关",
    )
    add(
        "G", "实际销量",
        "按姓名 SUMIF 求和：销售任务及完成率!Z",
        "销售任务及完成率 | 键C(姓名)→值Z(实际销量)",
        "全员",
        "hub_metrics_rule_engine",
        "无",
        "✅ 已实现",
    )
    add(
        "H", "销量完成率",
        "IF(考核量≠0, MIN(实际销量/考核量, 封顶), 0)；封顶按店别分组：默认120%，直营店管理/新媒体销售部/西物-翼真/销售管理部=110%",
        "派生自 F/G；封顶分组见 hub_metrics_rules.yaml cap_overrides",
        "全员",
        "hub_metrics_rule_engine",
        "无（本列是乘数来源）",
        "✅ 已实现",
        "销售顾问 W/Y/Z 的 H 乘数；门店块 W 用 BA 而非 H",
    )
    add(
        "I", "集客达成率",
        "按姓名 INDEX/MATCH 取首个匹配：销售任务及完成率!F",
        "销售任务及完成率 | 键C(姓名)→值F(集客达成率)",
        "全员",
        "hub_metrics_rule_engine",
        "无",
        "✅ 已实现",
    )
    add(
        "J", "加装额",
        "按姓名 SUMIF：绩效整理表!S",
        "绩效整理表(computed_perf_frame) | 键P(顾问)→值S(精品最低价)",
        "全员",
        "hub_metrics_rule_engine",
        "无",
        "✅ 已实现",
    )
    add(
        "K", "加装销量完成率",
        "加装额 / (实际销量 × 1500)；分母为0时取0",
        "派生自 J、G",
        "全员",
        "hub_metrics_rule_engine",
        "无",
        "✅ 已实现",
    )
    add(
        "L", "保险渗透率",
        "SUMIFS(绩效整理表!K, AB>0, P=姓名) / SUMIF(绩效整理表!K, P=姓名)",
        "绩效整理表 | K(台数)、AB(保险返利收入)、P(顾问)",
        "全员",
        "hub_metrics_rule_engine",
        "无",
        "✅ 已实现",
    )
    add(
        "M", "整车毛利",
        "按姓名 SUMIF：绩效整理表!BG",
        "绩效整理表 | 键P→值BG(整车毛利)；底层：终端明细/整车成本/系统销售毛利等",
        "全员",
        "hub_metrics_rule_engine",
        "无",
        "✅ 已实现",
    )
    add(
        "N", "加装毛利",
        "按姓名 SUMIF：绩效整理表!BI",
        "绩效整理表 | 键P→值BI(加装毛利)",
        "全员",
        "hub_metrics_rule_engine",
        "无",
        "✅ 已实现",
    )
    add(
        "O", "保险毛利",
        "按姓名 SUMIF：绩效整理表!AB",
        "绩效整理表 | 键P→值AB(保险返利收入)；底层：保险明细",
        "全员",
        "hub_metrics_rule_engine",
        "无",
        "✅ 已实现",
    )
    add(
        "P", "按揭毛利",
        "按姓名 SUMIF：绩效整理表!AC",
        "绩效整理表 | 键P→值AC(按揭收入)；底层：按揭明细",
        "全员",
        "hub_metrics_rule_engine",
        "无",
        "✅ 已实现",
    )

    # Q–V 毛利派生
    add("", "爱车宝毛利", "按姓名 SUMIF：绩效整理表!AD", "绩效整理表!AD", "全员", "hub_metrics_rule_engine", "无", "✅ 已实现")
    add("", "上户毛利", "按姓名 SUMIF：绩效整理表!AE", "绩效整理表!AE", "全员", "hub_metrics_rule_engine", "无", "✅ 已实现")
    add("", "整车+加装（毛利）", "整车毛利 + 加装毛利 (M+N)", "派生", "全员", "派生", "无", "✅ 已实现")
    add("", "综合毛利", "SUM(M:R) 各毛利列之和", "派生", "全员", "派生", "无", "✅ 已实现")
    add("", "主营单台毛利", "整车+加装毛利 / 实际销量", "派生 S/G", "全员", "派生", "无", "✅ 已实现")
    add("", "综合单台毛利", "综合毛利 / 实际销量", "派生 T/G", "全员", "派生", "无", "✅ 已实现")

    # 非一线语义列 W–AD（导出物理列，经典布局无对应单字母）
    for hdr, logic in [
        ("售后总产值", "非一线管理/support 语义列；从物理毛利列映射，非独立公式"),
        ("配件外销", "非一线语义列"),
        ("售后产值", "非一线语义列"),
        ("出库", "非一线语义列"),
        ("入库", "非一线语义列"),
        ("台次", "非一线语义列"),
        ("提成系数", "非一线语义列"),
        ("提成系数2", "非一线语义列"),
    ]:
        add("", hdr, logic, "非一线配置 non_frontline_roles.yaml", "非一线管理/support", "hub_topology_or_static", "无", "⚠️ 部分实现", "apply_non_frontline_columns 做物理→语义映射")

    for hdr in ("岗位绩效", "业绩绩效", "新能源专项", "业绩绩效1", "业绩绩效2"):
        add(
            "", hdr,
            "非一线管理行：金标准验证为人工填入固定值或引用人工子表，系统留空+灰色标注",
            "无系统公式",
            "非一线管理（销售管理部/事业部/总经办）",
            "manual_semantic",
            "无",
            "❌ 手工填入",
            "delegate: manual_semantic；禁止金标准 bootstrap",
        )

    # W–AI 绩效块（经典 Hub 字母）
    perf_specs = [
        ("W", "整车绩效", "SUM(绩效整理表!AG, P=姓名) × 完成率", "绩效整理表!AG(单台绩效)", "见乘数列", "详见 Sheet3"),
        ("X", "权限结余绩效", "SUM(绩效整理表!AH, P=姓名)", "绩效整理表!AH(闭包列)", "无", ""),
        ("Y", "加装绩效", "SUM(绩效整理表!AI, P=姓名) × H", "绩效整理表!AI(闭包列)", "H", ""),
        ("Z", "保险绩效", "SUM(绩效整理表!AJ, P=姓名) × H（+常数个案）", "绩效整理表!AJ；底层保险明细!BS", "H", ""),
        ("AA", "金融绩效", "SUM(绩效整理表!AK, P=姓名)", "绩效整理表!AK；底层按揭明细!BO", "无", ""),
        ("AB", "爱车宝绩效", "SUM(绩效整理表!AM, P=姓名)", "绩效整理表!AM；底层爱车保!BA", "无", ""),
        ("AC", "上户绩效", "SUM(绩效整理表!AN)+SUM(绩效整理表!AS)", "上户提成!H + 置换服务!BB", "无", ""),
        ("AD", "盈利产品绩效", "SUM(绩效整理表!AL, P=姓名)", "按揭原表!AF + 按揭明细!BR", "无", ""),
        ("AE", "延保提成", "SUM(绩效整理表!AT, P=姓名)", "延保提成!BE（±200规则）", "无", ""),
        ("AF", "特殊车型+指定车型", "SUM(绩效整理表!AQ, P=姓名)", "重功超期+活动/提成标准/系统销售毛利 闭包", "无", ""),
        ("AG", "座位险提成", "SUM(绩效整理表!AO, P=姓名)", "保险明细!BU", "无", ""),
        ("AH", "二手车提成", "SUM(绩效整理表!AR, P=姓名)", "二手置换!AE + 大客户!R", "无", ""),
        ("AI", "玻碎险提成", "SUM(绩效整理表!AP, P=姓名)", "保险明细!BV", "无", ""),
    ]

    advisor_roles = "销售顾问/销售主管/销售助理（统一规则）"
    for letter, header, logic, source, mult, note_extra in perf_specs:
        if header == "整车绩效":
            roles = (
                "销售顾问族：hub_rule_engine_perf；"
                "新媒体：delegate_module(新媒体!Y→AB)；"
                "邀约专员：delegate_module(邀约专员提成!C→AF)；"
                "客户专员：delegate_module(客户部提成按人直引)；"
                "内勤/网络顾问/新媒体翼真：hub_rule_engine_perf"
            )
            overlay = "hub_rule_engine_perf / delegate_module"
            impl = "✅ 已实现"
            note = note_extra
        elif header == "保险绩效":
            roles = advisor_roles + "；招聘族写 Z 列"
            overlay = "hub_rule_engine_perf / delegate_module"
            impl = "✅ 已实现"
            note = "韩柏成(insurance_add模板)：+600 常数"
        else:
            roles = advisor_roles + "；内勤/网络顾问/新媒体翼真部分列"
            overlay = "hub_rule_engine_perf"
            impl = "✅ 已实现"
            note = note_extra if note_extra != "见乘数列" else ""
        add(letter, header, logic, source, roles, overlay, mult, impl, note)

    add("", "提成合计", "SUM(绩效列 W:AI) + 调整列", "派生", "全员", "派生", "无", "⚠️ 部分实现", "调整列多数未接入")

    add(
        "AK", "整车完成考核",
        "按岗位族不同：①邀约崇州杨婷→邀约专员提成!AD；②直营店经理5人→直营店经理提成(财务)!R；③内勤→同店块实际销量×20；④翼真新媒体蒋利→翼真考核子表",
        "邀约专员提成 / 直营店经理提成(财务) / 派生 / 翼真考核",
        "邀约专员/直营店经理/内勤/新媒体翼真",
        "delegate_module / 派生",
        "无",
        "✅ 已实现",
        "经典列 AK；导出列 AX",
    )
    add("", "加装完成考核", "派生 SUM(子行) 或手填常量", "—", "部分岗位", "—", "无", "❌ 未实现", "仅批注登记")

    add(
        "AM", "综合项",
        "设计：综合表!L SUMIF(B, 姓名)",
        "综合表!L",
        "全员",
        "—",
        "无",
        "❌ 未接入生产",
        "sheet_registry 有登记；HubMetrics/HubRule/overlay 均不读取；对账 adjustment_columns 跟踪",
    )
    add(
        "AN", "04月活动",
        "设计：'重功超期+活动'!X SUMIF(Q, 姓名)",
        "重功超期+活动!X",
        "全员",
        "—",
        "无",
        "❌ 未接入生产",
        "同上；parity adjustment_columns 登记",
    )
    add("", "超期", "绩效整理表!AU SUMIF(P, 姓名)", "绩效整理表!AU", "全员", "—", "无", "⚠️ 待确认", "整理表列已实现，Hub 写入待核实")
    add("", "（已发放奖励）", "—", "—", "—", "—", "无", "❌ 手工填入", "需手工填入")
    add("", "交车支出", "—", "—", "—", "—", "无", "❌ 手工填入", "需手工填入")
    add("", "保客考核", "设计：保客考核明细!J SUMIF(E, 姓名)", "保客考核明细!J", "全员", "—", "无", "❌ 未接入生产", "仅批注/对账登记")
    add("", "考核小计", "SUM(考核相关列)", "派生", "全员", "派生", "无", "❌ 未实现")
    add("", "单台提成", "—", "—", "—", "—", "无", "❌ 手工填入")
    add("", "提成毛利占比", "—", "—", "—", "—", "无", "❌ 手工填入")
    add("", "预算单台", "—", "—", "—", "—", "无", "❌ 手工填入")
    add("", "计提单台", "—", "—", "—", "—", "无", "❌ 手工填入")
    add("", "计提金额", "—", "—", "—", "—", "无", "❌ 手工填入")

    return rows


def build_family_rows(hub_perf: dict) -> list[dict]:
    """Sheet2: 岗位族与列映射."""
    rows: list[dict] = []
    for family_id, cfg in hub_perf.get("role_families", {}).items():
        match = cfg.get("match", {})
        match_str = "; ".join(
            f"{k}={v}" if not isinstance(v, list) else f"{k}=[{', '.join(str(x) for x in v)}]"
            for k, v in match.items()
        ) if match else "(delegate)"
        hub_cols = cfg.get("hub_columns") or cfg.get("semantic_columns") or []
        rows.append(
            {
                "岗位族": family_id,
                "匹配条件": match_str,
                "overlay_path": cfg.get("overlay_path", ""),
                "算法/模块": cfg.get("algorithm") or cfg.get("module", ""),
                "规则配置": cfg.get("rule_config", ""),
                "Hub 列": ", ".join(hub_cols) if hub_cols else "(见 delegate)",
                "parity_gate": cfg.get("parity_gate", ""),
                "子表/数据源": _family_source_desc(cfg),
            }
        )
    return rows


def _family_source_desc(cfg: dict) -> str:
    src = cfg.get("source", {})
    if not src:
        return ""
    if "sheet" in src:
        parts = [f"sheet={src['sheet']}"]
        if "name_col" in src:
            parts.append(f"键{src['name_col']}→值{src.get('value_col', '?')}")
        if "team_block_cols" in src:
            parts.append(f"团队块 {src['team_block_cols']}")
        return "; ".join(parts)
    if "workbook" in src:
        return f"workbook={src['workbook']}, sheet={src.get('sheet', '绩效整理表')}"
    return str(src)


def build_vehicle_perf_detail() -> list[dict]:
    """Sheet3: 整车绩效详解."""
    return [
        {"层级": "Hub 提成汇总", "步骤": "1", "项目": "物理列", "说明": "经典 Hub 列 W（2026-05 导出为 AE 列，因 W–AD 插入非一线语义列）"},
        {"层级": "Hub 提成汇总", "步骤": "2", "项目": "汇总公式", "说明": "整车绩效 = SUM(绩效整理表!AG, 键P=顾问姓名) × 完成率乘数"},
        {"层级": "Hub 提成汇总", "步骤": "3a", "项目": "乘数·个人块 personal_h", "说明": "店别不在 store_ba_shops 清单 → × H（销量完成率，F–P 已算，按店别封顶110%/120%）"},
        {"层级": "Hub 提成汇总", "步骤": "3b", "项目": "乘数·门店块 store_ba", "说明": "店别 ∈ {崇州/大邑/邛崃/彭州/华阳直营店} → × BA（销售任务及完成率!AG 合并完成率，按姓名匹配，系统不封顶）"},
        {"层级": "Hub 提成汇总", "步骤": "3c", "项目": "乘数·个案 insurance_add", "说明": "韩柏成（西物-翼真）→ 同 personal_h 乘 H，保险绩效 Z 另 +600"},
        {"层级": "Hub 提成汇总", "步骤": "4", "项目": "岗位族差异", "说明": "销售顾问/主管/助理/内勤/网络顾问/新媒体翼真→HubRuleEngine；新媒体销售部→新媒体!Y→AB SUMIF；DCC邀约→邀约专员提成!C→AF；客户部→按人直引"},
        {"层级": "绩效整理表", "步骤": "5", "项目": "AG 列含义", "说明": "单台绩效（每 VIN 一行）"},
        {"层级": "绩效整理表", "步骤": "6", "项目": "AG 标准公式", "说明": "AG = 提成标准!F(单台基数) × K(台数)；键：店别A × 渠道I × 车种H"},
        {"层级": "绩效整理表", "步骤": "7", "项目": "AG 特例", "说明": "A=武侯自有店 且 H=星越L → AG = 200 × K（不走 lookup）"},
        {"层级": "绩效整理表", "步骤": "8", "项目": "键列 P(顾问)", "说明": "多数来自 系统销售毛利!BJ；部分 VIN 有 p_overrides_by_vin 别名覆盖"},
        {"层级": "绩效整理表", "步骤": "9", "项目": "键列 O(VIN)", "说明": "来自 系统销售毛利!BD"},
        {"层级": "底层表", "步骤": "10", "项目": "提成标准", "说明": "上传账套 sheet「提成标准」≈提成依据·销售提成标准；C(部门)×D(渠道)×E(车种)→F(2026年标准)"},
        {"层级": "底层表", "步骤": "11", "项目": "比对表", "说明": "系统销售毛利!BL → 比对表 → 部门 A（AG lookup 用）"},
        {"层级": "底层表", "步骤": "12", "项目": "系统销售毛利", "说明": "提供 VIN(BD)、顾问(BJ)、车种(D/H)、渠道(E/I)、台数(K)、部门(BL) 等"},
        {"层级": "配置", "步骤": "13", "项目": "代码入口", "说明": "from_closure._compute_ag → PerformanceSheetModule → SalesAdvisorPerformanceModule(HubRuleEngine)"},
        {"层级": "配置", "步骤": "14", "项目": "规则文件", "说明": "hub_column_rules.yaml(销售顾问 family)、hub_performance.yaml、performance_sheet_columns.yaml closure_columns.AG"},
        {"层级": "发薪", "步骤": "15", "项目": "XW提成-发", "说明": "SUMIF(提成汇总!D, 姓名, 提成汇总!W/AE) → 发薪 H 列整车绩效"},
    ]


def main() -> None:
    export_letters = _load_export_letters()
    hub_perf = _load_yaml(PROJECT / "salary_pipeline/config/hub_performance.yaml")

    sheet1 = pd.DataFrame(build_rules_rows(export_letters))
    sheet2 = pd.DataFrame(build_family_rows(hub_perf))
    sheet3 = pd.DataFrame(build_vehicle_perf_detail())

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
        sheet1.to_excel(writer, sheet_name="列级规则总表", index=False)
        sheet2.to_excel(writer, sheet_name="岗位族与列映射", index=False)
        sheet3.to_excel(writer, sheet_name="整车绩效详解", index=False)

    wb = load_workbook(OUT_PATH)
    for name in wb.sheetnames:
        _style_sheet(wb[name])
    wb.save(OUT_PATH)

    print(f"Wrote {OUT_PATH}")
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"  {name}: {ws.max_row - 1} rows")


if __name__ == "__main__":
    main()
