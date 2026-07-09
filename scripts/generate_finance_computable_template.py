#!/usr/bin/env python3
"""Generate 销售账套-财务可计算 workbook (VIN + person Hub inputs + params).

Independent of ingestion / month.yaml. Does not modify 销售账套-财务合并-*.xlsx.

Output default: docs/templates/销售账套-财务可计算-2026-05.xlsx
Config: salary_pipeline/config/finance_computable_template.yaml
"""

from __future__ import annotations

import argparse
import logging
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

PROJECT = Path(__file__).resolve().parents[1]
if str(PROJECT) not in sys.path:
    sys.path.insert(0, str(PROJECT))

from scripts.generate_finance_standard_workbook import (  # noqa: E402
    CONFIG_DIR,
    GROUP_FILL,
    GROUP_FONT,
    HEADER_FILL,
    README_FONT,
    _auto_width,
    _find_golden_workbook,
    _load_yaml,
    _style_header_row,
    _write_exception_sheet,
)
from scripts.generate_finance_consolidated_workbook import (  # noqa: E402
    _commission_basis_workbook,
    _write_role_sheet,
)
from scripts.generate_frontline_finance_template import (  # noqa: E402
    FRONTLINE_EXTRA_GROUPS,
    _closure_perf_cols,
    _select_vin_columns,
    _write_frontline_summary_sheet,
)

DEFAULT_OUT = PROJECT / "docs/templates/销售账套-财务可计算-2026-05.xlsx"
TEMPLATE_CFG = CONFIG_DIR / "finance_computable_template.yaml"
COMMENT_FILL = PatternFill("solid", fgColor="FFF2CC")
EMPHASIS_FILL = PatternFill("solid", fgColor="E2EFDA")

logger = logging.getLogger(__name__)


def _person_comment(col: dict[str, Any]) -> str:
    parts: list[str] = []
    if col.get("hub_column"):
        letter = col.get("hub_letter", "")
        parts.append(f"→ Hub {letter} {col['hub_column']}".strip())
    if col.get("perf_col"):
        parts.append(f"绩效整理表 {col['perf_col']}")
    if col.get("hub_use"):
        parts.append(str(col["hub_use"]))
    if col.get("multiplier"):
        parts.append(f"乘数={col['multiplier']}")
    if col.get("note"):
        parts.append(str(col["note"]))
    return "；".join(parts) if parts else ""


def _write_person_sheet(wb: Workbook, cfg: dict[str, Any]) -> int:
    ps = cfg.get("person_sheet", {})
    columns: list[dict[str, Any]] = list(ps.get("columns", []))
    groups = {g["id"]: g["label"] for g in cfg.get("person_groups", [])}
    sheet_name = ps.get("name", "销售人员")
    ws = wb.create_sheet(sheet_name)

    for col_idx, col in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=groups.get(col.get("group", ""), ""))
    _style_header_row(ws, 1, fill=GROUP_FILL)
    for cell in ws[1]:
        cell.font = GROUP_FONT

    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col["label"])
        cell.fill = EMPHASIS_FILL if col.get("group") == "hub_inputs" else HEADER_FILL
        cell.font = Font(bold=True)

    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=_person_comment(col))
        cell.fill = COMMENT_FILL
        cell.font = Font(italic=True, size=9)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    data_start = int(ps.get("data_start_row", 4))
    for col_idx, col in enumerate(columns, start=1):
        if col.get("default") is not None:
            ws.cell(row=data_start, column=col_idx, value=col["default"])

    ws.freeze_panes = f"A{data_start}"
    ws.row_dimensions[3].height = 48
    _auto_width(ws)
    return len(columns)


def _write_vin_sheet(
    wb: Workbook,
    columns: list[dict[str, Any]],
    groups: dict[str, str],
    *,
    emphasized: set[str],
) -> int:
    count = _write_frontline_summary_sheet(
        wb,
        columns,
        groups,
        uploads_dir=None,
        populate=False,
        sample_rows=0,
    )
    if "汇总数据表" in wb.sheetnames:
        ws = wb["汇总数据表"]
        ws.title = "销售明细"
        for col_idx, col in enumerate(columns, start=1):
            if col.get("id") in emphasized:
                ws.cell(row=2, column=col_idx).fill = EMPHASIS_FILL
    return count


def _build_readme(cfg: dict[str, Any], *, vin_cols: int, person_cols: int) -> list[str]:
    vin_name = cfg.get("vin_sheet", {}).get("name", "销售明细")
    person_name = cfg.get("person_sheet", {}).get("name", "销售人员")
    basis = [s["name"] for s in cfg.get("commission_basis_sheets", {}).get("sheets", [])]
    roles = [s["name"] for s in cfg.get("role_placeholder_sheets", {}).get("sheets", [])]
    deferred = cfg.get("deferred", [])
    hub_cols = cfg.get("trial_hub_columns", [])

    return [
        cfg.get("readme_sheet", {}).get("title", "销售账套财务可计算模板"),
        "",
        "【定位】财务手填 VIN 明细 + 人级合计 → 独立试算脚本产出销售顾问 Hub W–AI 预览。",
        "  不接入生产 ingestion / month.yaml；勿替换 uploads 分表。",
        "  保留 销售账套-财务合并-*.xlsx 不动。",
        "",
        "一、工作簿结构",
        f"  1. {vin_name} — VIN 原子列（{vin_cols} 列），一行一车；本阶段不自动汇总",
        f"  2. {person_name} — 人级合计（{person_cols} 列），财务手填；试算直接读此表",
        f"  3. 提成依据参数 — {', '.join(basis)}",
        f"  4. 岗位直引占位 — {', '.join(roles)}（仅表头）",
        "  5. 例外登记表",
        "",
        "二、本阶段可产出 Hub 列（销售顾问）",
        *[f"  · {c}" for c in hub_cols],
        "",
        "三、延期 / 缺口",
        *[f"  · {d}" for d in deferred],
        "",
        "四、禁止事项",
        "  · 不得抄入金标准/历史提成结果数值",
        "  · 算不出：留空 + 例外登记",
        "",
        "五、生成与试算",
        "  python scripts/generate_finance_computable_template.py --month 2026-05",
        "  python scripts/trial_finance_computable_hub.py \\",
        "    --input docs/templates/销售账套-财务可计算-2026-05.xlsx",
        "",
        "六、填表说明见 docs/templates/销售账套-财务可计算-填表说明.md",
    ]


def _write_readme(wb: Workbook, lines: list[str]) -> None:
    ws = wb.create_sheet("README", 0)
    for row_idx, line in enumerate(lines, start=1):
        cell = ws.cell(row=row_idx, column=1, value=line)
        if row_idx == 1:
            cell.font = README_FONT
    ws.column_dimensions["A"].width = 110


def generate_workbook(
    *,
    month: str,
    out_path: Path,
    uploads_dir: Path | None = None,
) -> dict[str, Any]:
    cfg = _load_yaml(TEMPLATE_CFG)
    col_cfg = _load_yaml(CONFIG_DIR / "canonical_column_registry.yaml")
    hub_rules = _load_yaml(CONFIG_DIR / "hub_column_rules.yaml")
    perf_cfg = _load_yaml(CONFIG_DIR / "performance_sheet_columns.yaml")

    if uploads_dir is None:
        uploads_dir = PROJECT / "data" / "raw" / month / "uploads"

    golden_path = _find_golden_workbook(month)
    basis_path = _commission_basis_workbook(month)

    closure_perf = _closure_perf_cols(hub_rules, perf_cfg)
    vin_columns, groups = _select_vin_columns(col_cfg, closure_perf)
    for g in FRONTLINE_EXTRA_GROUPS:
        groups.setdefault(g["id"], g["label"])

    emphasized = set(cfg.get("vin_sheet", {}).get("emphasized_ids", []))

    wb = Workbook()
    wb.remove(wb.active)

    vin_count = _write_vin_sheet(wb, vin_columns, groups, emphasized=emphasized)
    person_count = _write_person_sheet(wb, cfg)

    for spec in cfg.get("commission_basis_sheets", {}).get("sheets", []):
        _write_role_sheet(
            wb,
            spec,
            uploads_dir=uploads_dir,
            golden_path=golden_path,
            basis_path=basis_path,
            populate=False,
            sample_rows=0,
        )

    for spec in cfg.get("role_placeholder_sheets", {}).get("sheets", []):
        _write_role_sheet(
            wb,
            spec,
            uploads_dir=uploads_dir,
            golden_path=golden_path,
            basis_path=basis_path,
            populate=False,
            sample_rows=0,
        )

    _write_exception_sheet(wb, cfg)

    readme_lines = _build_readme(cfg, vin_cols=vin_count, person_cols=person_count)
    _write_readme(wb, readme_lines)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    return {
        "out_path": str(out_path),
        "month": month,
        "sheet_names": list(wb.sheetnames),
        "vin_column_count": vin_count,
        "person_column_count": person_count,
        "vin_labels": [c["label"] for c in vin_columns],
        "person_labels": [c["label"] for c in cfg["person_sheet"]["columns"]],
        "closure_perf_columns": sorted(closure_perf),
        "trial_hub_columns": list(cfg.get("trial_hub_columns", [])),
        "deferred": list(cfg.get("deferred", [])),
        "golden_reference": str(golden_path) if golden_path else None,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate 财务可计算填写模板")
    parser.add_argument("--month", default="2026-05")
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    parser.add_argument("--uploads-dir", type=Path, default=None)
    parser.add_argument("-v", "--verbose", action="store_true")
    args = parser.parse_args()

    logging.basicConfig(level=logging.DEBUG if args.verbose else logging.INFO)

    if args.out == DEFAULT_OUT and args.month != "2026-05":
        out = PROJECT / f"docs/templates/销售账套-财务可计算-{args.month}.xlsx"
    else:
        out = args.out

    result = generate_workbook(
        month=args.month,
        out_path=out,
        uploads_dir=args.uploads_dir,
    )

    print(f"Wrote: {result['out_path']}")
    print(f"Sheets ({len(result['sheet_names'])}): {', '.join(result['sheet_names'])}")
    print(f"销售明细 columns: {result['vin_column_count']}")
    print(f"销售人员 columns: {result['person_column_count']}")
    print(f"Trial Hub columns: {', '.join(result['trial_hub_columns'])}")
    print("Deferred:")
    for item in result["deferred"]:
        print(f"  - {item}")


if __name__ == "__main__":
    main()
