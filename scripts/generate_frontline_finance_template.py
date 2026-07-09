#!/usr/bin/env python3
"""Generate 一线销售精简财务填写模板 (VIN 汇总 + 最小参数表).

Scope: 销售顾问 Hub W–AI 闭包（绩效整理表 AG–AT + AU 超期）+ 传递闭包底层列 + 键列。
参数表仅含：提成标准、比对表、销售任务及完成率、综合表、重功超期+活动。
另含：人员信息、例外登记表、README、列映射说明。

Output default: docs/templates/销售账套-一线销售-财务填写-2026-05.xlsx
"""

from __future__ import annotations

import argparse
import logging
import sys
from pathlib import Path
from typing import Any

import pandas as pd
import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

PROJECT = Path(__file__).resolve().parents[1]
if str(PROJECT) not in sys.path:
    sys.path.insert(0, str(PROJECT))

from scripts.generate_finance_standard_workbook import (  # noqa: E402
    HEADER_FILL,
    HEADER_FONT,
    GROUP_FILL,
    GROUP_FONT,
    README_FONT,
    CONFIG_DIR,
    _auto_width,
    _build_readme_lines,
    _find_golden_workbook,
    _load_yaml,
    _populate_summary_from_uploads,
    _read_header_template_rows,
    _read_upload_headers,
    _read_upload_sample_rows,
    _resolve_sheet_source,
    _style_header_row,
    _write_exception_sheet,
    _write_header_block,
    _write_param_spec_headers,
)
from scripts.generate_perf_sheet_column_map_xlsx import (  # noqa: E402
    COLUMN_META,
    HUB_LINK,
    build_rows,
    _load_golden_headers,
)

DEFAULT_OUT = PROJECT / "docs/templates/销售账套-一线销售-财务填写-2026-05.xlsx"
MAPPING_XLSX = PROJECT / "docs/iterations/hub/绩效整理表-列取数映射.xlsx"

# Hub 销售顾问 W–AI 对应绩效列 + AU 超期
TARGET_PERF_COLS: frozenset[str] = frozenset(
    {"AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN", "AO", "AP", "AQ", "AR", "AS", "AT", "AU"}
)

# 闭包计算必需的键列 / 中间列（订单骨架 + AU 上游）
SEED_KEY_COLS: frozenset[str] = frozenset(
    {"O", "P", "K", "G", "D", "E", "H", "I", "R", "M", "L", "S", "J", "N", "Q", "F"}
)

MINIMAL_PARAM_SHEETS: tuple[str, ...] = (
    "提成标准",
    "比对表",
    "销售任务及完成率",
    "综合表",
    "重功超期+活动",
)

# 绩效列依赖边（derived / closure 上游）
PERF_DEPENDENCY_EDGES: dict[str, list[str]] = {
    "D": ["A", "I"],
    "E": ["M", "O"],
    "T": ["O"],
    "U": ["G"],
    "V": ["O"],
    "W": ["L", "S", "Y"],
    "Y": ["V", "X"],
    "Z": ["O"],
    "AA": ["Y", "Z"],
    "AT": ["AF"],
    "AU": ["E", "O", "H", "D"],
    "BF": ["AW", "AX", "AY", "AZ", "BA", "BB", "BC", "BE"],
    "BG": ["L", "BF"],
    "BH": ["G"],
    "BI": ["S", "BH"],
    "BJ": ["BG", "BI"],
    "BL": ["BJ", "AB", "AC", "AD", "AE", "AF"],
    "AG": ["H", "I", "A", "C", "K", "D", "E", "BL", "R", "J"],
    "AH": ["AA", "D", "AG", "AI", "AJ", "AK", "AB", "AC", "AD", "AE", "AF"],
    "AI": ["K", "L", "S", "U", "G"],
    "AQ": ["O", "H", "I", "A", "K", "E"],
    "AR": ["O"],
}

# canonical_column_registry 未覆盖、但闭包需要的 VIN 级底层列
FRONTLINE_EXTRA_VIN_COLUMNS: list[dict[str, Any]] = [
    {
        "id": "registration_fee",
        "label": "上户手续费",
        "group": "registration",
        "source_sheet": "上户提成",
        "source_col": "C",
        "join": "vin",
        "source_key_col": "B",
        "perf_col": "AE",
    },
    {
        "id": "registration_perf",
        "label": "上户提成金额",
        "group": "registration",
        "source_sheet": "上户提成",
        "source_col": "H",
        "join": "vin",
        "source_key_col": "B",
        "perf_col": "AN",
    },
    {
        "id": "aicar_income",
        "label": "爱车保收入",
        "group": "aicar",
        "source_sheet": "爱车保",
        "source_col": "K",
        "join": "vin",
        "source_key_col": "F",
        "perf_col": "AD",
    },
    {
        "id": "aicar_commission",
        "label": "爱车宝提成金额",
        "group": "aicar",
        "source_sheet": "爱车保",
        "source_col": "BA",
        "join": "vin",
        "source_key_col": "F",
        "perf_col": "AM",
    },
    {
        "id": "trade_in_service",
        "label": "置换服务提成",
        "group": "trade_in",
        "source_sheet": "置换服务",
        "source_col": "BB",
        "join": "vin",
        "source_key_col": "G",
        "perf_col": "AS",
    },
    {
        "id": "used_car_trade",
        "label": "二手置换合计",
        "group": "trade_in",
        "source_sheet": "二手置换 ",
        "source_col": "AE",
        "join": "vin",
        "source_key_col": "T",
        "perf_col": "AR",
    },
    {
        "id": "key_account_commission",
        "label": "大客户销售提成",
        "group": "trade_in",
        "source_sheet": "大客户",
        "source_col": "R",
        "join": "vin",
        "source_key_col": "O",
        "perf_col": "AR",
    },
    {
        "id": "warranty_income",
        "label": "延保收入",
        "group": "warranty",
        "source_sheet": "延保提成",
        "source_col": "BE",
        "join": "vin",
        "source_key_col": "F",
        "perf_col": "AF",
    },
    {
        "id": "oversell_amount",
        "label": "合计超卖金额",
        "group": "overage",
        "source_sheet": "系统超额",
        "source_col": "P",
        "join": "vin",
        "source_key_col": "W",
        "perf_col": "V",
    },
    {
        "id": "mortgage_over_permission",
        "label": "整车超权限",
        "group": "mortgage",
        "source_sheet": "按揭原表",
        "source_col": "AN",
        "join": "vin",
        "source_key_col": "AC",
        "perf_col": "Z",
    },
]

FRONTLINE_EXTRA_GROUPS: list[dict[str, str]] = [
    {"id": "registration", "label": "上户"},
    {"id": "aicar", "label": "爱车保"},
    {"id": "trade_in", "label": "置换/大客户"},
    {"id": "warranty", "label": "延保"},
    {"id": "overage", "label": "系统超额"},
]

COMMENT_FILL = PatternFill("solid", fgColor="FFF2CC")

logger = logging.getLogger(__name__)


def _expand_perf_closure(
    targets: frozenset[str],
    seeds: frozenset[str],
) -> set[str]:
    seen = set(targets) | set(seeds)
    stack = list(seen)
    while stack:
        letter = stack.pop()
        for dep in PERF_DEPENDENCY_EDGES.get(letter, []):
            if dep not in seen:
                seen.add(dep)
                stack.append(dep)
    return seen


def _hub_target_perf_cols(hub_rules: dict[str, Any]) -> set[str]:
    cols: set[str] = set()
    family = hub_rules.get("role_families", {}).get("销售顾问", {})
    for spec in family.get("columns", []):
        for letter in spec.get("perf_columns", []):
            cols.add(str(letter).strip().upper())
    cols.add("AU")
    return cols


def _closure_perf_cols(
    hub_rules: dict[str, Any],
    perf_cfg: dict[str, Any],
) -> set[str]:
    hub_targets = _hub_target_perf_cols(hub_rules)
    expanded = _expand_perf_closure(frozenset(hub_targets), SEED_KEY_COLS)
    # 对齐 yaml 登记的闭包列
    for letter in (perf_cfg.get("closure_columns") or {}):
        if letter in hub_targets or letter in expanded:
            expanded.add(letter)
    for letter in (perf_cfg.get("detail_sumif_columns") or {}):
        if letter in hub_targets:
            expanded.add(letter)
    return expanded


def _select_vin_columns(
    col_cfg: dict[str, Any],
    closure_perf: set[str],
) -> list[dict[str, Any]]:
    groups = {g["id"]: g["label"] for g in col_cfg.get("groups", [])}
    for g in FRONTLINE_EXTRA_GROUPS:
        groups[g["id"]] = g["label"]

    selected: list[dict[str, Any]] = []
    seen_ids: set[str] = set()

    def add(col: dict[str, Any]) -> None:
        cid = col.get("id", "")
        if cid in seen_ids:
            return
        perf = col.get("perf_col")
        if perf and str(perf) in closure_perf:
            selected.append(col)
            seen_ids.add(cid)
            return
        # 键列 / 无 perf_col 但闭包必需
        if cid in {"vin", "order_no", "advisor", "units"}:
            selected.append(col)
            seen_ids.add(cid)
            return
        if col.get("id") == "vehicle_settle_date":
            selected.append(col)
            seen_ids.add(cid)
            return
        if col.get("id") == "order_type":
            selected.append(col)
            seen_ids.add(cid)

    for col in col_cfg.get("columns", []):
        add(col)
    for col in FRONTLINE_EXTRA_VIN_COLUMNS:
        add(col)

    # 稳定排序：按 group 注册顺序，组内保持 canonical 注册顺序
    group_order = [g["id"] for g in col_cfg.get("groups", [])] + [g["id"] for g in FRONTLINE_EXTRA_GROUPS]
    group_rank = {gid: idx for idx, gid in enumerate(group_order)}
    canonical_rank = {c["id"]: idx for idx, c in enumerate(col_cfg.get("columns", []))}
    extra_rank = {c["id"]: 1000 + idx for idx, c in enumerate(FRONTLINE_EXTRA_VIN_COLUMNS)}

    def sort_key(c: dict[str, Any]) -> tuple[int, int]:
        gid = c.get("group", "")
        cid = c.get("id", "")
        rank_in_group = canonical_rank.get(cid, extra_rank.get(cid, 9999))
        return (group_rank.get(gid, 99), rank_in_group)

    return sorted(selected, key=sort_key), groups


def _comment_for_column(col: dict[str, Any]) -> str:
    sheet = col.get("source_sheet") or "—"
    scol = col.get("source_col") or "—"
    perf = col.get("perf_col") or ""
    perf_part = f" → 绩效{perf}" if perf else ""
    note = col.get("note") or ""
    base = f"{sheet}.{scol}{perf_part}"
    return f"{base}；{note}" if note else base


def _write_summary_comment_row(ws, columns: list[dict[str, Any]]) -> None:
    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=_comment_for_column(col))
        cell.fill = COMMENT_FILL
        cell.font = Font(italic=True, size=9)
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _write_frontline_summary_sheet(
    wb: Workbook,
    columns: list[dict[str, Any]],
    groups: dict[str, str],
    *,
    uploads_dir: Path | None,
    populate: bool,
    sample_rows: int,
) -> int:
    sheet_name = "汇总数据表"
    ws = wb.create_sheet(sheet_name)
    data_start = 4  # row1 group, row2 header, row3 comment, row4+ data

    for col_idx, col in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=groups.get(col.get("group", ""), ""))
    _style_header_row(ws, 1, fill=GROUP_FILL)
    for cell in ws[1]:
        cell.font = GROUP_FONT

    for col_idx, col in enumerate(columns, start=1):
        ws.cell(row=2, column=col_idx, value=col["label"])
    _style_header_row(ws, 2)

    _write_summary_comment_row(ws, columns)

    if populate and uploads_dir is not None:
        temp_wb = Workbook()
        temp_ws = temp_wb.active
        _populate_summary_from_uploads(temp_ws, columns, uploads_dir, sample_rows)
        dest_row = data_start
        for src_row in temp_ws.iter_rows(min_row=3):
            if not any(c.value is not None for c in src_row):
                continue
            for col_idx, cell in enumerate(src_row, start=1):
                if cell.value is not None:
                    ws.cell(row=dest_row, column=col_idx, value=cell.value)
            dest_row += 1

    ws.freeze_panes = f"A{data_start}"
    _auto_width(ws)
    return len(columns)


def _build_frontline_readme(
    *,
    vin_col_count: int,
    closure_perf: set[str],
    param_sheets: list[str],
) -> list[str]:
    perf_sorted = ", ".join(sorted(closure_perf))
    lines = [
        "一线销售精简财务填写模板 — 填账规则",
        "",
        "一、适用范围",
        "  · 一线销售顾问闭包：系统读取本账套后自动生成「绩效整理表」与提成汇总 W–AI",
        "  · 财务仅填写：汇总数据表（VIN 粒度底层）+ 5 张最小参数表 + 人员信息",
        "  · 不含完整 69 列绩效整理表，也不含 34-sheet 标准账套",
        "",
        "二、工作簿结构",
        f"  1. 汇总数据表 — {vin_col_count} 列（闭包底层 + 键列），一行一车",
        "  2. 人员信息 — 店别 / 职务 / 姓名（提成汇总行键）",
        f"  3. 最小参数表 — {len(param_sheets)} 张",
        "  4. 例外登记表 — 系统值与确认值差异登记（不覆盖系统计算）",
        "  5. 列映射说明 — AG–AU 闭包取数链参考",
        "",
        "三、列类型",
        "  · 系统列（汇总数据表）：按 VIN/订单号填写的明细，参与绩效闭包计算",
        "  · 参数列：提成标准、比对表、任务完成率、综合表调整、重功超期活动",
        "  · 例外列：人工确认差异，不写回系统产出",
        "",
        "四、汇总数据表",
        "  · 第 1 行分组、第 2 行列名、第 3 行注释（原表.原列 → 绩效列）",
        "  · VIN码 必填；台数整车订单填 1",
        "  · 闭包绩效列覆盖：" + perf_sorted,
        "",
        "五、最小参数表",
        *[f"  · {name}" for name in param_sheets],
        "",
        "六、禁止事项",
        "  · 不得将金标准/历史提成数值抄入本模板",
        "  · 算不出或规则未覆盖：留空 + 例外登记表说明",
        "",
        "七、重新生成",
        "  python scripts/generate_frontline_finance_template.py --month 2026-05",
        "  python scripts/generate_frontline_finance_template.py --month 2026-05 --populate-from-uploads",
    ]
    return lines


def _write_readme_sheet(wb: Workbook, lines: list[str]) -> None:
    ws = wb.create_sheet("README", 0)
    for row_idx, line in enumerate(lines, start=1):
        cell = ws.cell(row=row_idx, column=1, value=line)
        if row_idx == 1:
            cell.font = README_FONT
    ws.column_dimensions["A"].width = 100


def _param_specs_for_frontline(param_cfg: dict[str, Any]) -> list[dict[str, Any]]:
    specs: list[dict[str, Any]] = []
    for spec in param_cfg.get("tier_a_params", {}).get("sheets", []):
        if spec["name"] in MINIMAL_PARAM_SHEETS:
            specs.append(spec)
    return specs


def _write_param_sheet(
    wb: Workbook,
    spec: dict[str, Any],
    *,
    uploads_dir: Path | None,
    golden_path: Path | None,
    populate: bool,
    sample_rows: int,
) -> None:
    from scripts.generate_finance_standard_workbook import _write_param_sheet as _write_std

    _write_std(
        wb,
        spec,
        uploads_dir=uploads_dir,
        golden_path=golden_path,
        populate=populate,
        sample_rows=sample_rows,
    )


def _write_personnel_sheet(
    wb: Workbook,
    param_cfg: dict[str, Any],
    *,
    uploads_dir: Path | None,
    golden_path: Path | None,
    populate: bool,
    sample_rows: int,
) -> None:
    skeleton_specs = param_cfg.get("skeleton", {}).get("sheets", [])
    if not skeleton_specs:
        return
    _write_param_sheet(
        wb,
        skeleton_specs[0],
        uploads_dir=uploads_dir,
        golden_path=golden_path,
        populate=populate,
        sample_rows=sample_rows,
    )


def _write_mapping_reference_sheet(
    wb: Workbook,
    hub_perf: set[str],
    perf_cfg: dict[str, Any],
) -> None:
    headers = _load_golden_headers()
    rows = build_rows(headers, perf_cfg)
    # AG–AU Hub 目标列（一线销售闭包输出链），不含中间派生列
    subset = [r for r in rows if r["绩效列字母"] in hub_perf]
    subset.sort(key=lambda r: r["绩效列字母"])

    ws = wb.create_sheet("列映射说明")
    if not subset:
        ws.cell(row=1, column=1, value="（无映射数据）")
        return

    df = pd.DataFrame(subset)
    preferred = [
        "绩效列字母",
        "绩效列中文名",
        "来源类型",
        "来源表",
        "来源列",
        "匹配键",
        "计算公式简述",
        "是否闭包列",
        "Hub下游",
        "实现状态",
    ]
    cols = [c for c in preferred if c in df.columns]
    df = df[cols]

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    _style_header_row(ws, 1)
    ws.freeze_panes = "A2"
    _auto_width(ws)


def generate_workbook(
    *,
    month: str,
    out_path: Path,
    populate: bool = False,
    uploads_dir: Path | None = None,
    sample_rows: int = 5,
) -> dict[str, Any]:
    col_cfg = _load_yaml(CONFIG_DIR / "canonical_column_registry.yaml")
    param_cfg = _load_yaml(CONFIG_DIR / "canonical_param_sheets.yaml")
    hub_rules = _load_yaml(CONFIG_DIR / "hub_column_rules.yaml")
    perf_cfg = _load_yaml(CONFIG_DIR / "performance_sheet_columns.yaml")

    if uploads_dir is None:
        uploads_dir = PROJECT / "data" / "raw" / month / "uploads"

    golden_path = _find_golden_workbook(month)
    closure_perf = _closure_perf_cols(hub_rules, perf_cfg)
    hub_perf_set = _hub_target_perf_cols(hub_rules)
    vin_columns, groups = _select_vin_columns(col_cfg, closure_perf)
    param_specs = _param_specs_for_frontline(param_cfg)

    wb = Workbook()
    wb.remove(wb.active)

    readme_lines = _build_frontline_readme(
        vin_col_count=len(vin_columns),
        closure_perf=closure_perf,
        param_sheets=[s["name"] for s in param_specs],
    )
    _write_readme_sheet(wb, readme_lines)

    summary_col_count = _write_frontline_summary_sheet(
        wb,
        vin_columns,
        groups,
        uploads_dir=uploads_dir if populate else None,
        populate=populate,
        sample_rows=sample_rows,
    )

    _write_personnel_sheet(
        wb,
        param_cfg,
        uploads_dir=uploads_dir,
        golden_path=golden_path,
        populate=populate,
        sample_rows=sample_rows,
    )

    for spec in param_specs:
        _write_param_sheet(
            wb,
            spec,
            uploads_dir=uploads_dir,
            golden_path=golden_path,
            populate=populate,
            sample_rows=sample_rows,
        )

    _write_exception_sheet(wb, param_cfg)
    _write_mapping_reference_sheet(wb, hub_perf_set, perf_cfg)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    return {
        "out_path": str(out_path),
        "month": month,
        "sheet_count": len(wb.sheetnames),
        "sheet_names": list(wb.sheetnames),
        "summary_column_count": summary_col_count,
        "closure_perf_columns": sorted(closure_perf),
        "hub_perf_columns": sorted(hub_perf_set),
        "param_sheets": [s["name"] for s in param_specs],
        "populated_from_uploads": populate,
        "golden_reference": str(golden_path) if golden_path else None,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate 一线销售精简财务填写模板")
    parser.add_argument("--month", default="2026-05")
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    parser.add_argument("--populate-from-uploads", action="store_true")
    parser.add_argument("--uploads-dir", type=Path, default=None)
    parser.add_argument("--sample-rows", type=int, default=5)
    parser.add_argument("-v", "--verbose", action="store_true")
    args = parser.parse_args()

    logging.basicConfig(level=logging.DEBUG if args.verbose else logging.INFO)

    if args.out == DEFAULT_OUT and args.month != "2026-05":
        out = PROJECT / f"docs/templates/销售账套-一线销售-财务填写-{args.month}.xlsx"
    else:
        out = args.out

    result = generate_workbook(
        month=args.month,
        out_path=out,
        populate=args.populate_from_uploads,
        uploads_dir=args.uploads_dir,
        sample_rows=args.sample_rows,
    )

    print(f"Wrote: {result['out_path']}")
    print(f"Total sheets: {result['sheet_count']}")
    print(f"汇总数据表 columns: {result['summary_column_count']}")
    print(f"Closure perf columns ({len(result['closure_perf_columns'])}): {', '.join(result['closure_perf_columns'])}")
    print(f"Hub target perf ({len(result['hub_perf_columns'])}): {', '.join(result['hub_perf_columns'])}")
    print("Param sheets:")
    for name in result["param_sheets"]:
        print(f"  - {name}")
    if result.get("golden_reference"):
        print(f"Golden header reference: {result['golden_reference']}")


if __name__ == "__main__":
    main()
