#!/usr/bin/env python3
"""Generate 绩效整理表-列取数映射.xlsx from yaml + performance_sheet calculators."""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

PROJECT = Path(__file__).resolve().parents[1]
if str(PROJECT) not in sys.path:
    sys.path.insert(0, str(PROJECT))

import pandas as pd
import yaml
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

CONFIG_PATH = PROJECT / "salary_pipeline/config/performance_sheet_columns.yaml"
GOLDEN_PATH = (
    PROJECT / "data/raw/2026-05/燃油车-2026年05月西物超市销售提成(终)(1).xlsx"
)
OUT_PATH = PROJECT / "docs/iterations/hub/绩效整理表-列取数映射.xlsx"

# Fallback when golden workbook is unavailable
FALLBACK_HEADERS: tuple[tuple[str, str], ...] = (
    ("A", "指标汇总部门"),
    ("B", "排放标准"),
    ("C", "指标汇总车型"),
    ("D", "渠道"),
    ("E", "库存天数"),
    ("F", "购进公司"),
    ("G", "订单号"),
    ("H", "车种"),
    ("I", "销售渠道"),
    ("J", "车型"),
    ("K", "台数"),
    ("L", "订单合计(含税)"),
    ("M", "结算日期"),
    ("N", "车主名称"),
    ("O", "VIN码"),
    ("P", "销售顾问"),
    ("Q", "审核人"),
    ("R", "部门"),
    ("S", "精品最低价金额"),
    ("T", "车架号后8位"),
    ("U", "不提成精品"),
    ("V", "整车节约"),
    ("W", "整车最低售价"),
    ("X", "经理权限"),
    ("Y", "整车实际节约"),
    ("Z", "盈利性产品按揭权限扣除"),
    ("AA", "实际整车节约权限"),
    ("AB", "保险返利收入"),
    ("AC", "按揭收入"),
    ("AD", "爱车宝收入"),
    ("AE", "上户收入"),
    ("AF", "延保收入"),
    ("AG", "单台绩效"),
    ("AH", "整车超额"),
    ("AI", "加装绩效"),
    ("AJ", "保险提成"),
    ("AK", "按揭提成"),
    ("AL", "盈利产品"),
    ("AM", "爱车宝"),
    ("AN", "上户绩效"),
    ("AO", "座位险绩效"),
    ("AP", "玻璃险绩效"),
    ("AQ", "特殊车型追加绩效（追加+质损车）"),
    ("AR", "二手车置换"),
    ("AS", "置换服务"),
    ("AT", "延保提成"),
    ("AU", "超期追加"),
    ("AV", "合计"),
    ("AW", "出厂指导价/扣除降价补差"),
    ("AX", "提车现返"),
    ("AY", "合同履约"),
    ("AZ", "项目金额附加费"),
    ("BA", "广告返利"),
    ("BB", "提车奖励当月返"),
    ("BC", "当月返利"),
    ("BD", "促销奖"),
    ("BE", "代交车佣金"),
    ("BF", "整车采购净价"),
    ("BG", "裸车毛利"),
    ("BH", "装饰成本"),
    ("BI", "装饰毛利"),
    ("BJ", "主营业务毛利"),
    ("BK", "上户代收代付部分"),
    ("BL", "综合毛利"),
    ("BM", "不含税毛利"),
    ("BN", "不含安心包提成"),
    ("BO", "安心保"),
    ("BP", "爱车保"),
    ("BQ", "上户量"),
)

# Per-column metadata (letter → fields). Slice/status merged from yaml where present.
COLUMN_META: dict[str, dict[str, Any]] = {
    "A": {
        "source_type": "lookup",
        "source_sheets": "比对表",
        "source_cols": "A(系统部门)→B(指标汇总部门)",
        "match_key": "R=系统销售毛利!BL(部门)",
        "formula": "比对表按部门映射指标汇总部门",
        "is_closure": "否",
        "slice": "8",
        "status": "implemented",
        "module": "order_context.py",
    },
    "B": {
        "source_type": "constant",
        "source_sheets": "—",
        "source_cols": "—",
        "match_key": "—",
        "formula": "常量「整车订单」",
        "is_closure": "否",
        "slice": "8",
        "status": "implemented",
        "module": "from_order_context_export.py",
    },
    "C": {
        "source_type": "lookup",
        "source_sheets": "比对表",
        "source_cols": "D(系统车种)→E(指标汇总车型)",
        "match_key": "H=系统销售毛利!D(车种)",
        "formula": "比对表按车种映射指标汇总车型",
        "is_closure": "否",
        "slice": "8",
        "status": "implemented",
        "module": "order_context.py",
    },
    "D": {
        "source_type": "derived",
        "source_sheets": "系统销售毛利; 比对表",
        "source_cols": "A(指标汇总部门); I=系统销售毛利!E(销售渠道)",
        "match_key": "O(VIN)",
        "formula": "由A+I派生渠道标签（直营店店面/二网/分公司后缀等）",
        "is_closure": "是",
        "slice": "8",
        "status": "implemented",
        "module": "order_context.py",
    },
    "E": {
        "source_type": "derived",
        "source_sheets": "系统销售毛利; 整车成本",
        "source_cols": "BA(结算日期); B(入库/结算日期)",
        "match_key": "O(VIN)",
        "formula": "INT(系统销售毛利!BA) − INDEX(整车成本!B, MATCH(O, K))",
        "is_closure": "否",
        "slice": "7",
        "status": "implemented",
        "module": "from_overdue_stock.py",
    },
    "F": {
        "source_type": "lookup",
        "source_sheets": "工厂购进",
        "source_cols": "CA(付款单位)",
        "match_key": "O(VIN)=C(VIN码)",
        "formula": "INDEX/MATCH 按VIN取购进公司",
        "is_closure": "否",
        "slice": "8",
        "status": "implemented",
        "module": "from_order_context_export.py",
    },
    "G": {
        "source_type": "direct",
        "source_sheets": "系统销售毛利",
        "source_cols": "B(订单号)",
        "match_key": "—",
        "formula": "当月含整车订单筛选后取订单号",
        "is_closure": "否",
        "slice": "4",
        "status": "implemented",
        "module": "order_skeleton.py",
    },
    "H": {
        "source_type": "direct",
        "source_sheets": "系统销售毛利",
        "source_cols": "D(车种)",
        "match_key": "O(VIN)=BD",
        "formula": "按VIN关联取车种",
        "is_closure": "是",
        "slice": "8",
        "status": "implemented",
        "module": "order_context.py",
    },
    "I": {
        "source_type": "direct",
        "source_sheets": "系统销售毛利",
        "source_cols": "E(销售渠道)",
        "match_key": "O(VIN)=BD",
        "formula": "按VIN关联取销售渠道",
        "is_closure": "是",
        "slice": "8",
        "status": "implemented",
        "module": "order_context.py",
    },
    "J": {
        "source_type": "direct",
        "source_sheets": "系统销售毛利",
        "source_cols": "F(车型)",
        "match_key": "O(VIN)=BD",
        "formula": "按VIN关联取车型",
        "is_closure": "否",
        "slice": "8",
        "status": "implemented",
        "module": "order_context.py",
    },
    "K": {
        "source_type": "constant",
        "source_sheets": "—",
        "source_cols": "—",
        "match_key": "—",
        "formula": "默认1；supplement_rows补录行可为空",
        "is_closure": "否",
        "slice": "4",
        "status": "implemented",
        "module": "order_skeleton.py",
    },
    "L": {
        "source_type": "direct",
        "source_sheets": "系统销售毛利",
        "source_cols": "AO(订单合计含税)",
        "match_key": "O(VIN)=BD",
        "formula": "按VIN取订单合计",
        "is_closure": "是",
        "slice": "8",
        "status": "implemented",
        "module": "order_context.py",
    },
    "M": {
        "source_type": "direct",
        "source_sheets": "系统销售毛利",
        "source_cols": "BA(结算日期)",
        "match_key": "O(VIN)=BD",
        "formula": "按VIN取结算日期",
        "is_closure": "否",
        "slice": "8",
        "status": "implemented",
        "module": "order_context.py",
    },
    "N": {
        "source_type": "direct",
        "source_sheets": "系统销售毛利",
        "source_cols": "BC(车主名称)",
        "match_key": "O(VIN)=BD",
        "formula": "按VIN取车主名称",
        "is_closure": "否",
        "slice": "8",
        "status": "implemented",
        "module": "order_context.py",
    },
    "O": {
        "source_type": "direct",
        "source_sheets": "系统销售毛利",
        "source_cols": "BD(VIN码)",
        "match_key": "—",
        "formula": "当月含整车订单骨架主键；多数SUMIF criteria列",
        "is_closure": "否",
        "slice": "4",
        "status": "implemented",
        "module": "order_skeleton.py",
    },
    "P": {
        "source_type": "lookup",
        "source_sheets": "系统销售毛利",
        "source_cols": "BJ(销售顾问)",
        "match_key": "O(VIN)=BD",
        "formula": "INDEX(BJ,MATCH(O,BD))+p_overrides_by_vin别名覆盖",
        "is_closure": "否",
        "slice": "4",
        "status": "implemented",
        "module": "order_skeleton.py",
    },
    "Q": {
        "source_type": "direct",
        "source_sheets": "系统销售毛利",
        "source_cols": "BK(审核人)",
        "match_key": "O(VIN)=BD",
        "formula": "按VIN取审核人",
        "is_closure": "否",
        "slice": "8",
        "status": "implemented",
        "module": "order_context.py",
    },
    "R": {
        "source_type": "direct",
        "source_sheets": "系统销售毛利",
        "source_cols": "BL(部门)",
        "match_key": "O(VIN)=BD",
        "formula": "按VIN取部门→比对表键",
        "is_closure": "是",
        "slice": "8",
        "status": "implemented",
        "module": "order_context.py",
    },
    "S": {
        "source_type": "direct",
        "source_sheets": "系统销售毛利",
        "source_cols": "BQ(精品最低价金额)",
        "match_key": "O(VIN)=BD",
        "formula": "按VIN取精品最低价",
        "is_closure": "是",
        "slice": "8",
        "status": "implemented",
        "module": "order_context.py",
    },
    "T": {
        "source_type": "derived",
        "source_sheets": "—",
        "source_cols": "O(VIN码)",
        "match_key": "O",
        "formula": "VIN后8位",
        "is_closure": "否",
        "slice": "8",
        "status": "implemented",
        "module": "from_order_context_export.py",
    },
    "U": {
        "source_type": "sumif",
        "source_sheets": "装饰台账",
        "source_cols": "AR(最低价含税金额); H(精品名称); N(订单号); M",
        "match_key": "G(订单号)",
        "formula": "按精品名称规则多段SUMIFS汇总不提成精品",
        "is_closure": "是",
        "slice": "8",
        "status": "implemented",
        "module": "from_closure.py",
    },
    "V": {
        "source_type": "lookup",
        "source_sheets": "系统超额",
        "source_cols": "P(合计超卖金额)",
        "match_key": "O(VIN)=W(VIN码)",
        "formula": "INDEX(P,MATCH(O,W,0))",
        "is_closure": "是",
        "slice": "8",
        "status": "implemented",
        "module": "order_context.py",
    },
    "W": {
        "source_type": "derived",
        "source_sheets": "系统销售毛利",
        "source_cols": "L(AO订单合计); S(BQ精品最低价); Y",
        "match_key": "O",
        "formula": "(L−S)−Y 整车最低售价",
        "is_closure": "是",
        "slice": "8",
        "status": "implemented",
        "module": "from_order_context_export.py",
    },
    "X": {
        "source_type": "constant",
        "source_sheets": "performance_sheet_columns.yaml",
        "source_cols": "manager_permission_by_vin",
        "match_key": "O(VIN)",
        "formula": "按VIN配置常数(样本月59笔=1000)",
        "is_closure": "是",
        "slice": "8",
        "status": "implemented",
        "module": "order_context.py",
    },
    "Y": {
        "source_type": "derived",
        "source_sheets": "系统超额; yaml",
        "source_cols": "V; X",
        "match_key": "O",
        "formula": "V+X 整车实际节约",
        "is_closure": "是",
        "slice": "8",
        "status": "implemented",
        "module": "order_context.py",
    },
    "Z": {
        "source_type": "sumif",
        "source_sheets": "按揭原表",
        "source_cols": "AN(整车超权限)",
        "match_key": "O(VIN)=AC(车架号)",
        "formula": "SUMIF(按揭原表!AC,O,AN)",
        "is_closure": "是",
        "slice": "8",
        "status": "implemented",
        "module": "order_context.py",
    },
    "AA": {
        "source_type": "derived",
        "source_sheets": "—",
        "source_cols": "Y; Z",
        "match_key": "O",
        "formula": "若Z≠0则0，否则Y−Z",
        "is_closure": "是",
        "slice": "8",
        "status": "implemented",
        "module": "order_context.py",
    },
    "AB": {
        "source_type": "sumif",
        "source_sheets": "保险明细",
        "source_cols": "BP(返利金额)",
        "match_key": "O(VIN)=D(车架号)",
        "formula": "SUMIF(保险明细!D,O,BP)",
        "is_closure": "否",
        "slice": "1",
        "status": "implemented",
        "module": "from_insurance.py",
    },
    "AC": {
        "source_type": "lookup",
        "source_sheets": "按揭明细",
        "source_cols": "Z(服务费合计)",
        "match_key": "O(VIN)=G(VIN码)",
        "formula": "INDEX/MATCH按VIN取按揭收入",
        "is_closure": "否",
        "slice": "8",
        "status": "implemented",
        "module": "from_income.py",
    },
    "AD": {
        "source_type": "sumif",
        "source_sheets": "爱车保",
        "source_cols": "K(产品名称/收入)",
        "match_key": "O(VIN)=F(VIN)",
        "formula": "SUMIF(爱车保!F,O,K)",
        "is_closure": "否",
        "slice": "8",
        "status": "implemented",
        "module": "from_income.py",
    },
    "AE": {
        "source_type": "sumif",
        "source_sheets": "上户提成",
        "source_cols": "C(手续费/收入)",
        "match_key": "O(VIN)=B(VIN码)",
        "formula": "SUMIF(上户提成!B,O,C)",
        "is_closure": "否",
        "slice": "8",
        "status": "implemented",
        "module": "from_income.py",
    },
    "AF": {
        "source_type": "sumif",
        "source_sheets": "延保提成",
        "source_cols": "BE(延保收入)",
        "match_key": "O(VIN)=F(VIN)",
        "formula": "SUMIF(延保提成!F,O,BE)",
        "is_closure": "否",
        "slice": "8",
        "status": "implemented",
        "module": "from_income.py",
    },
    "AG": {
        "source_type": "closure",
        "source_sheets": "系统销售毛利; 比对表; 提成标准",
        "source_cols": "D/E/BL; A/B/D/E; F(2026年标准)",
        "match_key": "O; H/I/A/K",
        "formula": "提成标准!F×K；特例武侯自有店+星越L→200×K",
        "is_closure": "是",
        "slice": "5",
        "status": "implemented",
        "module": "from_closure.py",
    },
    "AH": {
        "source_type": "closure",
        "source_sheets": "系统销售毛利; 比对表; 装饰台账; 保险明细; 按揭明细; 提成标准",
        "source_cols": "闭包链AG/AI/AJ/AK; D; AA",
        "match_key": "O; D渠道",
        "formula": "整车超额：AA×比例分支+perf_sum保底150等",
        "is_closure": "是",
        "slice": "5",
        "status": "implemented",
        "module": "from_closure.py",
    },
    "AI": {
        "source_type": "closure",
        "source_sheets": "系统销售毛利; 装饰台账",
        "source_cols": "AO/BQ; AR(按G汇总)",
        "match_key": "O; G; K",
        "formula": "K=0→L×12%；else (S−U)×12%",
        "is_closure": "是",
        "slice": "5",
        "status": "implemented",
        "module": "from_closure.py",
    },
    "AJ": {
        "source_type": "sumif",
        "source_sheets": "保险明细",
        "source_cols": "BS(提成金额)",
        "match_key": "O(VIN)=D(车架号)",
        "formula": "SUMIF(保险明细!D,O,BS)",
        "is_closure": "否",
        "slice": "1",
        "status": "implemented",
        "module": "from_insurance.py",
    },
    "AK": {
        "source_type": "sumif",
        "source_sheets": "按揭明细",
        "source_cols": "BO(基础绩效)",
        "match_key": "O(VIN)=G(VIN码)",
        "formula": "SUMIF(按揭明细!G,O,BO)",
        "is_closure": "否",
        "slice": "1",
        "status": "implemented",
        "module": "from_mortgage.py",
    },
    "AL": {
        "source_type": "sumif",
        "source_sheets": "按揭原表; 按揭明细",
        "source_cols": "AF(盈利性奖励); BR(吉致额外追加)",
        "match_key": "O(VIN)=AC/G",
        "formula": "SUMIF(按揭原表!AC,O,AF)+SUMIF(按揭明细!G,O,BR)；门店P跳过+按VIN调整",
        "is_closure": "否",
        "slice": "3",
        "status": "implemented",
        "module": "from_mortgage.py",
    },
    "AM": {
        "source_type": "sumif",
        "source_sheets": "爱车保",
        "source_cols": "BA(爱车宝提成金额)",
        "match_key": "O(VIN)=F(VIN)",
        "formula": "SUMIF(爱车保!F,O,BA)",
        "is_closure": "是",
        "slice": "5",
        "status": "implemented",
        "module": "from_closure.py",
    },
    "AN": {
        "source_type": "sumif",
        "source_sheets": "上户提成",
        "source_cols": "H(提成金额)",
        "match_key": "O(VIN)=B(VIN码)",
        "formula": "SUMIF(上户提成!B,O,H)",
        "is_closure": "是",
        "slice": "5",
        "status": "implemented",
        "module": "from_closure.py",
    },
    "AO": {
        "source_type": "sumif",
        "source_sheets": "保险明细",
        "source_cols": "BU(座位险提成)",
        "match_key": "O(VIN)=D(车架号)",
        "formula": "SUMIF(保险明细!D,O,BU)",
        "is_closure": "否",
        "slice": "2",
        "status": "implemented",
        "module": "from_insurance.py",
    },
    "AP": {
        "source_type": "sumif",
        "source_sheets": "保险明细",
        "source_cols": "BV(玻璃险提成)",
        "match_key": "O(VIN)=D(车架号)",
        "formula": "SUMIF(保险明细!D,O,BV)",
        "is_closure": "否",
        "slice": "2",
        "status": "implemented",
        "module": "from_insurance.py",
    },
    "AQ": {
        "source_type": "closure",
        "source_sheets": "重功超期+活动; 提成标准; 系统销售毛利; 比对表",
        "source_cols": "N(销售追加奖励); H(标准); D/E/BL",
        "match_key": "O(VIN)=E; H/I/A/K",
        "formula": "SUMIF(重功超期+活动!E,O,N)+提成标准!H×K",
        "is_closure": "是",
        "slice": "5",
        "status": "implemented",
        "module": "from_closure.py",
    },
    "AR": {
        "source_type": "sumif",
        "source_sheets": "二手置换 ; 大客户",
        "source_cols": "AE(合计); R(销售提成)",
        "match_key": "O(VIN)=T/O(车架号)",
        "formula": "SUMIF(二手置换!T,O,AE)+SUMIF(大客户!O,O,R)",
        "is_closure": "是",
        "slice": "5",
        "status": "implemented",
        "module": "from_closure.py",
    },
    "AS": {
        "source_type": "sumif",
        "source_sheets": "置换服务",
        "source_cols": "BB(提成金额)",
        "match_key": "O(VIN)=G(VIN)",
        "formula": "SUMIF(置换服务!G,O,BB)",
        "is_closure": "是",
        "slice": "5",
        "status": "implemented",
        "module": "from_closure.py",
    },
    "AT": {
        "source_type": "derived",
        "source_sheets": "延保提成",
        "source_cols": "BE(经AF)",
        "match_key": "O",
        "formula": "AF<0→-200；AF>0→200；否则0",
        "is_closure": "否",
        "slice": "6",
        "status": "implemented",
        "module": "from_warranty.py",
    },
    "AU": {
        "source_type": "derived",
        "source_sheets": "topology; 提成依据",
        "source_cols": "超期政策表",
        "match_key": "O(VIN); E(库存天数)",
        "formula": "按topology/VIN超期政策×库存天数E",
        "is_closure": "否",
        "slice": "6/7",
        "status": "implemented",
        "module": "from_overdue_stock.py",
    },
    "AV": {
        "source_type": "derived",
        "source_sheets": "—",
        "source_cols": "AG:AU",
        "match_key": "—",
        "formula": "SUM(AG:AU) 绩效合计",
        "is_closure": "否",
        "slice": "8",
        "status": "implemented",
        "module": "from_derived.py",
    },
    "AW": {
        "source_type": "lookup",
        "source_sheets": "整车成本",
        "source_cols": "R(出厂指导价)",
        "match_key": "O(VIN)=K(全车架号)",
        "formula": "INDEX(R,MATCH(O,K,0))",
        "is_closure": "否",
        "slice": "3",
        "status": "implemented",
        "module": "from_vehicle_cost.py",
    },
    "AX": {
        "source_type": "lookup",
        "source_sheets": "整车成本",
        "source_cols": "S(提车现返)",
        "match_key": "O(VIN)=K",
        "formula": "INDEX(S,MATCH(O,K,0))",
        "is_closure": "否",
        "slice": "3",
        "status": "implemented",
        "module": "from_vehicle_cost.py",
    },
    "AY": {
        "source_type": "lookup",
        "source_sheets": "整车成本",
        "source_cols": "T(合同履约奖)",
        "match_key": "O(VIN)=K",
        "formula": "INDEX(T,MATCH(O,K,0))",
        "is_closure": "否",
        "slice": "3",
        "status": "implemented",
        "module": "from_vehicle_cost.py",
    },
    "AZ": {
        "source_type": "lookup",
        "source_sheets": "整车成本",
        "source_cols": "U(项目金额附加费)",
        "match_key": "O(VIN)=K",
        "formula": "INDEX(U,MATCH(O,K,0))",
        "is_closure": "否",
        "slice": "3",
        "status": "implemented",
        "module": "from_vehicle_cost.py",
    },
    "BA": {
        "source_type": "lookup",
        "source_sheets": "整车成本",
        "source_cols": "V(广告返利)",
        "match_key": "O(VIN)=K",
        "formula": "INDEX(V,MATCH(O,K,0))",
        "is_closure": "否",
        "slice": "3",
        "status": "implemented",
        "module": "from_vehicle_cost.py",
    },
    "BB": {
        "source_type": "lookup",
        "source_sheets": "整车成本",
        "source_cols": "W(提车奖励当月返)",
        "match_key": "O(VIN)=K",
        "formula": "INDEX(W,MATCH(O,K,0))",
        "is_closure": "否",
        "slice": "3",
        "status": "implemented",
        "module": "from_vehicle_cost.py",
    },
    "BC": {
        "source_type": "sumif",
        "source_sheets": "终端明细表",
        "source_cols": "P(金额)",
        "match_key": "O(VIN)=C(车辆VIN码); K≠0",
        "formula": "−IF(K≠0,SUMIFS(终端明细表!P,C,O),0)",
        "is_closure": "否",
        "slice": "6",
        "status": "implemented",
        "module": "from_terminal.py",
    },
    "BD": {
        "source_type": "—",
        "source_sheets": "—",
        "source_cols": "—",
        "match_key": "—",
        "formula": "金标准无公式，系统未实现",
        "is_closure": "否",
        "slice": "—",
        "status": "未实现",
        "module": "—",
        "notes": "促销奖：导出留空，需手工填入",
    },
    "BE": {
        "source_type": "lookup",
        "source_sheets": "系统二手车降价",
        "source_cols": "BE(佣金去税)",
        "match_key": "O(VIN)=CL(VIN码)",
        "formula": "INDEX/MATCH按VIN取代交车佣金",
        "is_closure": "否",
        "slice": "8",
        "status": "implemented",
        "module": "from_derived.py",
    },
    "BF": {
        "source_type": "derived",
        "source_sheets": "—",
        "source_cols": "AW:BE",
        "match_key": "—",
        "formula": "SUM(AW:BE) 整车采购净价",
        "is_closure": "否",
        "slice": "8",
        "status": "implemented",
        "module": "from_derived.py",
    },
    "BG": {
        "source_type": "derived",
        "source_sheets": "系统销售毛利",
        "source_cols": "L; BF",
        "match_key": "O",
        "formula": "L−BF 裸车毛利",
        "is_closure": "否",
        "slice": "8",
        "status": "implemented",
        "module": "from_derived.py",
    },
    "BH": {
        "source_type": "sumif",
        "source_sheets": "装饰台账",
        "source_cols": "AK(含税成本价)",
        "match_key": "G(订单号)=N(订单号)",
        "formula": "SUMIF(装饰台账!N,G,AK)",
        "is_closure": "否",
        "slice": "3",
        "status": "implemented",
        "module": "from_decoration.py",
    },
    "BI": {
        "source_type": "derived",
        "source_sheets": "系统销售毛利; 装饰台账",
        "source_cols": "S(BQ); BH",
        "match_key": "O/G",
        "formula": "S−BH 装饰毛利",
        "is_closure": "否",
        "slice": "8",
        "status": "implemented",
        "module": "from_derived.py",
    },
    "BJ": {
        "source_type": "derived",
        "source_sheets": "—",
        "source_cols": "BG; BI",
        "match_key": "—",
        "formula": "BG+BI 主营业务毛利",
        "is_closure": "否",
        "slice": "8",
        "status": "implemented",
        "module": "from_derived.py",
    },
    "BK": {
        "source_type": "—",
        "source_sheets": "—",
        "source_cols": "—",
        "match_key": "—",
        "formula": "金标准无系统取数规则",
        "is_closure": "否",
        "slice": "—",
        "status": "未实现",
        "module": "—",
        "notes": "上户代收代付部分：导出留空",
    },
    "BL": {
        "source_type": "derived",
        "source_sheets": "—",
        "source_cols": "BJ; AB:AF",
        "match_key": "—",
        "formula": "BJ+SUM(AB:AF) 综合毛利",
        "is_closure": "否",
        "slice": "8",
        "status": "implemented",
        "module": "from_derived.py",
    },
    "BM": {
        "source_type": "derived",
        "source_sheets": "—",
        "source_cols": "BJ; AB:AF",
        "match_key": "—",
        "formula": "BJ/1.13+收入列/1.06 不含税毛利",
        "is_closure": "否",
        "slice": "8",
        "status": "implemented",
        "module": "from_derived.py",
    },
    "BN": {
        "source_type": "derived",
        "source_sheets": "—",
        "source_cols": "AG:AT",
        "match_key": "—",
        "formula": "SUM(AG:AT) 不含安心包提成",
        "is_closure": "否",
        "slice": "8",
        "status": "implemented",
        "module": "from_derived.py",
    },
    "BO": {
        "source_type": "—",
        "source_sheets": "—",
        "source_cols": "—",
        "match_key": "—",
        "formula": "金标准无系统取数规则",
        "is_closure": "否",
        "slice": "—",
        "status": "未实现",
        "module": "—",
        "notes": "安心保：导出留空",
    },
    "BP": {
        "source_type": "—",
        "source_sheets": "—",
        "source_cols": "—",
        "match_key": "—",
        "formula": "金标准无系统取数规则(与AM爱车宝不同列)",
        "is_closure": "否",
        "slice": "—",
        "status": "未实现",
        "module": "—",
        "notes": "爱车保(列名重复语义)：导出留空",
    },
    "BQ": {
        "source_type": "—",
        "source_sheets": "—",
        "source_cols": "—",
        "match_key": "—",
        "formula": "金标准无系统取数规则",
        "is_closure": "否",
        "slice": "—",
        "status": "未实现",
        "module": "—",
        "notes": "上户量：导出留空",
    },
}

HUB_LINK: dict[str, str] = {
    "AG": "Hub W 整车绩效",
    "AH": "Hub Y 整车超额",
    "AI": "Hub AA 加装绩效",
    "AJ": "Hub Z 保险绩效",
    "AK": "Hub AB 按揭提成",
    "AL": "Hub AD 盈利产品",
    "AM": "Hub AB 爱车宝",
    "AN": "Hub AC 上户绩效(之一)",
    "AS": "Hub AC 上户绩效(之二)",
    "AO": "Hub AG 座位险",
    "AP": "Hub AI 玻璃险",
    "AQ": "Hub AF 特殊车型",
    "AR": "Hub AH 二手车",
    "AT": "Hub AE 延保提成",
    "AB": "Hub O 保险毛利",
    "AC": "Hub P 按揭毛利",
    "BG": "Hub M 整车毛利",
    "BI": "Hub N 加装毛利",
    "AU": "Hub AO 超期(经SUMIF P)",
    "S": "Hub J 加装额",
    "K": "Hub L 保险渗透率",
}


def _load_yaml(path: Path) -> dict[str, Any]:
    with path.open(encoding="utf-8") as handle:
        return yaml.safe_load(handle) or {}


def _load_golden_headers() -> tuple[tuple[str, str], ...]:
    if not GOLDEN_PATH.exists():
        return FALLBACK_HEADERS
    from salary_pipeline.data_ingestion.performance_sheet_golden_scan import (
        load_golden_column_headers,
    )

    spec = load_golden_column_headers(GOLDEN_PATH, header_row=2)
    return spec or FALLBACK_HEADERS


def _merge_yaml_slices(cfg: dict[str, Any]) -> None:
    """Overlay slice/status from performance_sheet_columns.yaml where registered."""
    for letter, spec in (cfg.get("detail_sumif_columns") or {}).items():
        if letter in COLUMN_META:
            COLUMN_META[letter]["slice"] = str(spec.get("slice", COLUMN_META[letter]["slice"]))
            COLUMN_META[letter]["status"] = spec.get("status", COLUMN_META[letter]["status"])
    for letter, spec in (cfg.get("closure_columns") or {}).items():
        if letter in COLUMN_META:
            COLUMN_META[letter]["slice"] = str(spec.get("slice", COLUMN_META[letter]["slice"]))
            COLUMN_META[letter]["status"] = spec.get("status", COLUMN_META[letter]["status"])
            if spec.get("formula"):
                COLUMN_META[letter]["yaml_formula"] = spec["formula"]
            sources = spec.get("sources") or spec.get("source_sheet")
            if sources:
                COLUMN_META[letter]["yaml_sources"] = sources


def build_rows(headers: tuple[tuple[str, str], ...], cfg: dict[str, Any]) -> list[dict[str, Any]]:
    _merge_yaml_slices(cfg)
    billing = (cfg.get("order_skeleton") or {}).get("billing_month", "2026-05")
    rows: list[dict[str, Any]] = []
    for letter, name_cn in headers:
        meta = COLUMN_META.get(letter, {})
        rows.append(
            {
                "账期样板": billing,
                "绩效列字母": letter,
                "绩效列中文名": name_cn,
                "来源类型": meta.get("source_type", "—"),
                "来源表": meta.get("source_sheets", "—"),
                "来源列": meta.get("source_cols", "—"),
                "匹配键": meta.get("match_key", "—"),
                "计算公式简述": meta.get("formula", "—"),
                "是否闭包列": meta.get("is_closure", "—"),
                "Slice/Phase": meta.get("slice", "—"),
                "实现状态": meta.get("status", "—"),
                "代码模块": meta.get("module", "—"),
                "Hub下游": HUB_LINK.get(letter, ""),
                "备注": meta.get("notes", ""),
            }
        )
    return rows


def _direct_underlying_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Sheet3: deduplicated 表.列 for non-derived sources."""
    seen: set[str] = set()
    out: list[dict[str, Any]] = []
    direct_types = {"direct", "sumif", "lookup", "constant"}
    for row in rows:
        if row["来源类型"] not in direct_types:
            continue
        sheets = str(row["来源表"]).split(";")
        cols = str(row["来源列"]).split(";")
        for sheet in sheets:
            sheet = sheet.strip()
            if not sheet or sheet in {"—", "performance_sheet_columns.yaml", "topology; 提成依据"}:
                continue
            for col_part in cols:
                col_part = col_part.strip()
                if not col_part or col_part == "—":
                    continue
                # strip parenthetical descriptions for dedupe key
                base_col = col_part.split("(")[0].strip()
                if "→" in base_col:
                    parts = [p.strip() for p in base_col.split("→")]
                    for p in parts:
                        key = f"{sheet}.{p}"
                        if key not in seen:
                            seen.add(key)
                            out.append(
                                {
                                    "来源表": sheet,
                                    "来源列": p,
                                    "整理表引用列": row["绩效列字母"],
                                    "整理表列名": row["绩效列中文名"],
                                    "匹配键": row["匹配键"],
                                }
                            )
                    continue
                key = f"{sheet}.{base_col}"
                if key not in seen:
                    seen.add(key)
                    out.append(
                        {
                            "来源表": sheet,
                            "来源列": col_part,
                            "整理表引用列": row["绩效列字母"],
                            "整理表列名": row["绩效列中文名"],
                            "匹配键": row["匹配键"],
                        }
                    )
    out.sort(key=lambda r: (r["来源表"], r["来源列"]))
    return out


def _style_sheet(ws, header_row: int = 1) -> None:
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[header_row]:
        if cell.value:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 48)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def generate(out_path: Path | None = None) -> Path:
    out = out_path or OUT_PATH
    out.parent.mkdir(parents=True, exist_ok=True)
    cfg = _load_yaml(CONFIG_PATH)
    headers = _load_golden_headers()
    rows = build_rows(headers, cfg)

    df_all = pd.DataFrame(rows)
    df_by_source = df_all.sort_values(["来源表", "绩效列字母"])
    df_direct = pd.DataFrame(_direct_underlying_rows(rows))

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df_all.to_excel(writer, sheet_name="完整列映射", index=False)
        df_by_source.to_excel(writer, sheet_name="按来源表分组", index=False)
        df_direct.to_excel(writer, sheet_name="直接底层列清单", index=False)

    wb = load_workbook(out)
    for name in wb.sheetnames:
        _style_sheet(wb[name])
    wb.save(out)
    return out


def main() -> None:
    path = generate()
    cfg = _load_yaml(CONFIG_PATH)
    headers = _load_golden_headers()
    rows = build_rows(headers, cfg)
    unclear = [r for r in rows if r["实现状态"] == "未实现" or r["来源类型"] == "—"]
    print(f"Wrote {path}")
    print(f"Rows (完整列映射): {len(rows)}")
    print(f"Direct underlying entries: {len(_direct_underlying_rows(rows))}")
    if unclear:
        print("Unclear / unimplemented columns:")
        for r in unclear:
            print(f"  {r['绩效列字母']} {r['绩效列中文名']}: {r['备注'] or r['计算公式简述']}")


if __name__ == "__main__":
    main()
