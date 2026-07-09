#!/usr/bin/env python3
"""Generate 财务合并填写模板 (销售 + 提成依据 + 岗位直引).

One workbook for finance trial fill — NOT wired into ingestion (uploads/*.xlsx unchanged).

Output default: docs/templates/销售账套-财务合并-2026-05.xlsx
Config: salary_pipeline/config/finance_consolidated_template.yaml
"""

from __future__ import annotations

import argparse
import logging
import sys
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook

PROJECT = Path(__file__).resolve().parents[1]
if str(PROJECT) not in sys.path:
    sys.path.insert(0, str(PROJECT))

from scripts.generate_finance_standard_workbook import (  # noqa: E402
    CONFIG_DIR,
    _auto_width,
    _build_readme_lines,
    _find_golden_workbook,
    _load_yaml,
    _populate_summary_from_uploads,
    _resolve_sheet_source,
    _style_header_row,
    _write_exception_sheet,
    _write_header_block,
    _write_param_sheet,
    _write_param_spec_headers,
    _write_readme_sheet,
    HEADER_FILL,
    GROUP_FILL,
    GROUP_FONT,
    README_FONT,
)

DEFAULT_OUT = PROJECT / "docs/templates/销售账套-财务合并-2026-05.xlsx"
TEMPLATE_CFG = CONFIG_DIR / "finance_consolidated_template.yaml"

logger = logging.getLogger(__name__)


def _commission_basis_workbook(month: str) -> Path | None:
    """提成依据.xlsx under data/raw/{month}/."""
    for candidate in (
        PROJECT / "data" / "raw" / month / "提成依据.xlsx",
        PROJECT / "data" / "raw" / month / ".staging" / "uploads" / "提成依据.xlsx",
        PROJECT / "data" / "raw" / month / "uploads" / "提成依据.xlsx",
    ):
        if candidate.exists():
            return candidate
    return None


def _build_consolidated_readme(cfg: dict[str, Any], col_count: int) -> list[str]:
    sales = cfg.get("sales_sheet", {})
    basis = cfg.get("commission_basis_sheets", {})
    roles = cfg.get("role_direct_fetch_sheets", {})
    exc = cfg.get("exception_sheet", {})

    basis_names = [s["name"] for s in basis.get("sheets", [])]
    role_names = [s["name"] for s in roles.get("sheets", [])]

    return [
        cfg.get("readme_sheet", {}).get("title", "销售账套财务合并模板 — 填账规则"),
        "",
        "【过渡说明】本模板仅供财务试填与列位熟悉；生产流水线仍读取",
        "  data/raw/<月>/uploads/*.xlsx，勿将本文件直接替换合并账套。",
        "",
        "一、工作簿结构",
        f"  1. {sales.get('name', '销售')} — VIN 粒度主数据（{col_count} 列）",
        "  2. 人员信息 — 提成汇总行键（店别/职务/姓名）",
        f"  3. {basis.get('label', '提成依据参数')} — {len(basis_names)} 张",
        *[f"     · {n}" for n in basis_names],
        f"  4. {roles.get('label', '岗位直引')} — {len(role_names)} 张（有对应人员再填）",
        *[f"     · {n}" for n in role_names],
        f"  5. {exc.get('name', '例外登记表')} — 人工确认差异",
        "",
        "二、与 uploads 原名对照",
        "  · 销售 ← 汇总数据表概念（原 系统销售毛利 + 保险/按揭/成本 等列合并）",
        "  · 提成依据 ← 提成依据.xlsx·销售提成标准",
        "  · 经理 ← 直营店经理提成 (财务)",
        "  · 客服专员 ← 客户部提成",
        "  · 按揭专员 ← 按揭绩效",
        "",
        "三、禁止事项",
        "  · 不得抄入金标准/历史提成数值",
        "  · 算不出：留空 + 例外登记表",
        "",
        "四、重新生成",
        "  python scripts/generate_finance_consolidated_workbook.py --month 2026-05",
    ]


def _write_sales_sheet(
    wb: Workbook,
    col_cfg: dict[str, Any],
    *,
    sheet_name: str,
    uploads_dir: Path | None,
    populate: bool,
    sample_rows: int,
) -> int:
    """VIN sheet — same columns as 汇总数据表, renamed to 销售."""
    from openpyxl.utils import get_column_letter

    columns: list[dict[str, Any]] = col_cfg.get("columns", [])
    groups: dict[str, str] = {g["id"]: g["label"] for g in col_cfg.get("groups", [])}

    ws = wb.create_sheet(sheet_name)
    labels = [c["label"] for c in columns]

    for col_idx, col in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=groups.get(col.get("group", ""), ""))
    for cell in ws[1]:
        cell.fill = GROUP_FILL
        cell.font = GROUP_FONT

    for col_idx, label in enumerate(labels, start=1):
        ws.cell(row=2, column=col_idx, value=label)
    _style_header_row(ws, 2)

    for col_idx, col in enumerate(columns, start=1):
        note_parts = []
        if col.get("source_sheet"):
            note_parts.append(f"原{col['source_sheet']}.{col.get('source_col', '')}")
        if col.get("perf_col"):
            note_parts.append(f"→绩效{col['perf_col']}")
        if note_parts:
            ws.cell(row=3, column=col_idx, value=" ".join(note_parts))

    if populate and uploads_dir is not None:
        _populate_summary_from_uploads(ws, columns, uploads_dir, sample_rows)

    ws.freeze_panes = "A4"
    _auto_width(ws)
    return len(columns)


def _write_role_sheet(
    wb: Workbook,
    spec: dict[str, Any],
    *,
    uploads_dir: Path | None,
    golden_path: Path | None,
    basis_path: Path | None,
    populate: bool,
    sample_rows: int,
) -> None:
    """Role / param sheet; resolve headers from uploads, golden, or 提成依据.xlsx."""
    source_sheet = spec.get("source_sheet", spec["name"])
    from_basis = spec.get("source_workbook") == "提成依据"

    if from_basis and basis_path is not None:
        from scripts.generate_finance_standard_workbook import _resolve_inner_sheet_name

        source_path = basis_path
        inner_sheet = _resolve_inner_sheet_name(basis_path, source_sheet)
    else:
        source_path, inner_sheet = _resolve_sheet_source(
            source_sheet,
            uploads_dir=uploads_dir,
            golden_path=golden_path,
            alias=spec.get("sheet_alias"),
        )

    ws = wb.create_sheet(spec["name"])

    from scripts.generate_finance_standard_workbook import (
        _read_header_template_rows,
        _read_upload_headers,
        _read_upload_sample_rows,
    )
    from openpyxl.utils.dataframe import dataframe_to_rows

    header_row = int(spec.get("header_row", 1))
    template_rows = int(spec.get("header_template_rows", 0))
    data_start_row = 2

    if source_path is not None:
        if template_rows > 0:
            block = _read_header_template_rows(
                source_path,
                sheet_name=inner_sheet,
                num_rows=template_rows,
            )
            if block:
                written = _write_header_block(ws, block, style_last=True)
                data_start_row = written + 1
            else:
                data_start_row = _write_param_spec_headers(ws, spec) + 1
        else:
            headers = _read_upload_headers(
                source_path,
                sheet_name=inner_sheet,
                header_row=header_row,
            )
            if headers and any(headers):
                for col_idx, header in enumerate(headers, start=1):
                    if header:
                        ws.cell(row=1, column=col_idx, value=header)
                _style_header_row(ws, 1)
                data_start_row = 2
            else:
                data_start_row = _write_param_spec_headers(ws, spec) + 1

        if populate:
            sample = _read_upload_sample_rows(
                source_path,
                sheet_name=inner_sheet,
                header_row=header_row if template_rows == 0 else template_rows,
                max_rows=sample_rows,
            )
            for r_idx, row in enumerate(
                dataframe_to_rows(sample, index=False, header=False),
                start=data_start_row,
            ):
                for c_idx, value in enumerate(row, start=1):
                    if value is not None and str(value) not in {"", "nan"}:
                        ws.cell(row=r_idx, column=c_idx, value=value)
    else:
        data_start_row = _write_param_spec_headers(ws, spec) + 1

    freeze_row = max(data_start_row, 2)
    ws.freeze_panes = f"A{freeze_row}"
    _auto_width(ws)


def generate_workbook(
    *,
    month: str,
    out_path: Path,
    populate: bool = False,
    uploads_dir: Path | None = None,
    sample_rows: int = 5,
) -> dict[str, Any]:
    cfg = _load_yaml(TEMPLATE_CFG)
    col_cfg = _load_yaml(CONFIG_DIR / "canonical_column_registry.yaml")

    if uploads_dir is None:
        uploads_dir = PROJECT / "data" / "raw" / month / "uploads"

    golden_path = _find_golden_workbook(month)
    basis_path = _commission_basis_workbook(month)

    wb = Workbook()
    wb.remove(wb.active)

    sales_name = cfg.get("sales_sheet", {}).get("name", "销售")
    col_count = 0

    readme_lines = _build_consolidated_readme(cfg, 0)
    _write_readme_sheet(wb, readme_lines)

    # skeleton
    sk = cfg.get("skeleton", {}).get("sheets", [{}])[0]
    _write_param_sheet(
        wb,
        {"name": sk.get("name", "人员信息"), "header_row": 1, "key_labels": sk.get("key_labels", [])},
        uploads_dir=uploads_dir,
        golden_path=golden_path,
        populate=populate,
        sample_rows=sample_rows,
    )

    col_count = _write_sales_sheet(
        wb,
        col_cfg,
        sheet_name=sales_name,
        uploads_dir=uploads_dir if populate else None,
        populate=populate,
        sample_rows=sample_rows,
    )

    basis_names: list[str] = []
    for spec in cfg.get("commission_basis_sheets", {}).get("sheets", []):
        _write_role_sheet(
            wb,
            spec,
            uploads_dir=uploads_dir,
            golden_path=golden_path,
            basis_path=basis_path,
            populate=populate,
            sample_rows=sample_rows,
        )
        basis_names.append(spec["name"])

    role_names: list[str] = []
    for spec in cfg.get("role_direct_fetch_sheets", {}).get("sheets", []):
        _write_role_sheet(
            wb,
            spec,
            uploads_dir=uploads_dir,
            golden_path=golden_path,
            basis_path=basis_path,
            populate=populate,
            sample_rows=sample_rows,
        )
        role_names.append(spec["name"])

    _write_exception_sheet(wb, cfg)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    return {
        "out_path": str(out_path),
        "month": month,
        "sheet_count": len(wb.sheetnames),
        "sheet_names": list(wb.sheetnames),
        "sales_column_count": col_count,
        "basis_sheets": basis_names,
        "role_sheets": role_names,
        "populated_from_uploads": populate,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate 销售账套-财务合并 template workbook")
    parser.add_argument("--month", default="2026-05")
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    parser.add_argument("--populate-from-uploads", action="store_true")
    parser.add_argument("--uploads-dir", type=Path, default=None)
    parser.add_argument("--sample-rows", type=int, default=5)
    parser.add_argument("-v", "--verbose", action="store_true")
    args = parser.parse_args()

    logging.basicConfig(level=logging.DEBUG if args.verbose else logging.INFO)

    if args.out == DEFAULT_OUT and args.month != "2026-05":
        out = PROJECT / f"docs/templates/销售账套-财务合并-{args.month}.xlsx"
    else:
        out = args.out

    result = generate_workbook(
        month=args.month,
        out_path=out,
        populate=args.populate_from_uploads,
        uploads_dir=args.uploads_dir,
        sample_rows=args.sample_rows,
    )

    print(f"Wrote: {result['out_path']}")
    print(f"Sheets ({result['sheet_count']}): {', '.join(result['sheet_names'])}")
    print(f"销售 columns: {result['sales_column_count']}")


if __name__ == "__main__":
    main()
