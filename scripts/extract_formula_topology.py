#!/usr/bin/env python3
"""Extract Excel formula cells and dependency topology into JSON."""

from __future__ import annotations

import argparse
import io
import json
import re
import sys
from collections import defaultdict, deque
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula

try:
    import msoffcrypto
except ImportError:  # pragma: no cover
    msoffcrypto = None


EXTERNAL_REF_PATTERN = re.compile(
    r"\[[0-9]+\][^!,\)]+!",
    re.IGNORECASE,
)
EXTERNAL_RANGE_PATTERN = re.compile(
    r"(\[[0-9]+\][^!,\)]+)!"
    r"(\$?[A-Z]{1,3}(?::\$?[A-Z]{1,3}|\$?\d{1,7}(?::\$?[A-Z]{1,3}\$?\d{1,7})?))",
    re.IGNORECASE,
)
QUOTED_SHEET_RANGE_PATTERN = re.compile(
    r"'([^']+)'!"
    r"(\$?[A-Z]{1,3}(?::\$?[A-Z]{1,3}|\$?\d{1,7}(?::\$?[A-Z]{1,3}\$?\d{1,7})?))",
    re.IGNORECASE,
)
LOCAL_CELL_PATTERN = re.compile(
    r"(?<![A-Z0-9_$])(\$?[A-Z]{1,3}\$?\d{1,7})"
    r"(?::(\$?[A-Z]{1,3}\$?\d{1,7}))?",
    re.IGNORECASE,
)
LOCAL_COLUMN_PATTERN = re.compile(
    r"(?<![A-Z0-9_$])(\$?[A-Z]{1,3}:\$?[A-Z]{1,3})(?!\d)",
    re.IGNORECASE,
)
STRING_LITERAL_PATTERN = re.compile(r'"(?:[^"]|"")*"')


def normalize_formula_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, (ArrayFormula, DataTableFormula)):
        text = value.text
    elif isinstance(value, str):
        text = value
    else:
        text = str(value)

    text = text.strip()
    if not text:
        return None
    if not text.startswith("="):
        text = f"={text}"
    return text


def strip_string_literals(formula: str) -> str:
    return STRING_LITERAL_PATTERN.sub('""', formula)


def cell_key(sheet: str, coordinate: str) -> str:
    return f"{sheet}!{coordinate.upper()}"


def normalize_address(address: str) -> str:
    return address.replace("$", "").upper()


def build_unquoted_sheet_pattern(sheet_names: set[str]) -> re.Pattern[str] | None:
    """Match sheet-qualified refs like 销售任务及完成率!C:C (no quotes)."""
    valid_names = sorted(
        (name for name in sheet_names if name and "!" not in name),
        key=len,
        reverse=True,
    )
    if not valid_names:
        return None
    joined = "|".join(re.escape(name) for name in valid_names)
    return re.compile(
        rf"(?P<sheet>{joined})!"
        r"(?P<range>"
        r"\$?[A-Z]{1,3}(?::\$?[A-Z]{1,3}|\$?\d{1,7}(?::\$?[A-Z]{1,3}\$?\d{1,7})?)"
        r")",
        re.IGNORECASE,
    )


def add_range_dep(
    range_deps: list[str],
    sheet: str,
    range_ref: str,
) -> None:
    range_key = f"{sheet}!{normalize_address(range_ref)}"
    if range_key not in range_deps:
        range_deps.append(range_key)


def mask_sheet_qualified_references(formula: str, sheet_names: set[str]) -> str:
    """Remove external / quoted / unquoted sheet refs before local cell parsing."""
    masked = EXTERNAL_RANGE_PATTERN.sub(lambda _: " " * 20, formula)
    masked = QUOTED_SHEET_RANGE_PATTERN.sub(lambda _: " " * 20, masked)
    unquoted_sheet_pattern = build_unquoted_sheet_pattern(sheet_names)
    if unquoted_sheet_pattern:
        masked = unquoted_sheet_pattern.sub(lambda _: " " * 20, masked)
    return masked


def parse_formula_references(
    formula: str,
    current_sheet: str,
    sheet_names: set[str],
) -> tuple[list[str], list[str], list[str]]:
    """Return (cell_deps, range_deps, external_refs)."""
    cleaned = strip_string_literals(formula)
    cell_deps: list[str] = []
    range_deps: list[str] = []
    external_refs: list[str] = []

    for match in EXTERNAL_REF_PATTERN.finditer(cleaned):
        token = match.group(0).rstrip("!")
        if token not in external_refs:
            external_refs.append(token)

    for match in EXTERNAL_RANGE_PATTERN.finditer(cleaned):
        external_sheet, range_ref = match.groups()
        external_token = f"{external_sheet}!{normalize_address(range_ref)}"
        if external_token not in external_refs:
            external_refs.append(external_token)

    for match in QUOTED_SHEET_RANGE_PATTERN.finditer(cleaned):
        sheet, range_ref = match.groups()
        if ":" in range_ref and not re.search(r"\d", range_ref):
            add_range_dep(range_deps, sheet, range_ref)
        elif ":" in range_ref:
            add_range_dep(range_deps, sheet, range_ref)
        else:
            key = cell_key(sheet, normalize_address(range_ref))
            if key not in cell_deps:
                cell_deps.append(key)

    unquoted_sheet_pattern = build_unquoted_sheet_pattern(sheet_names)
    if unquoted_sheet_pattern:
        for match in unquoted_sheet_pattern.finditer(cleaned):
            sheet = match.group("sheet")
            range_ref = match.group("range")
            if ":" in range_ref and not re.search(r"\d", range_ref):
                add_range_dep(range_deps, sheet, range_ref)
            elif ":" in range_ref:
                add_range_dep(range_deps, sheet, range_ref)
            else:
                key = cell_key(sheet, normalize_address(range_ref))
                if key not in cell_deps:
                    cell_deps.append(key)

    local_source = mask_sheet_qualified_references(cleaned, sheet_names)

    for match in LOCAL_CELL_PATTERN.finditer(local_source):
        start, end = match.groups()
        if "#REF" in match.group(0).upper():
            continue
        start_addr = normalize_address(start)
        if end:
            add_range_dep(range_deps, current_sheet, f"{start_addr}:{normalize_address(end)}")
        else:
            key = cell_key(current_sheet, start_addr)
            if key not in cell_deps:
                cell_deps.append(key)

    for match in LOCAL_COLUMN_PATTERN.finditer(local_source):
        column_range = match.group(1)
        add_range_dep(range_deps, current_sheet, column_range)

    return cell_deps, range_deps, external_refs


def is_formula_cell(cell: Cell) -> bool:
    if cell.data_type == "f":
        return True
    value = cell.value
    if isinstance(value, (ArrayFormula, DataTableFormula)):
        return True
    return isinstance(value, str) and value.startswith("=")


def load_workbook_from_path(path: Path, password: str | None) -> tuple[Any, list[str]]:
    warnings: list[str] = []

    try:
        return openpyxl.load_workbook(path, data_only=False, keep_links=True), warnings
    except Exception as first_error:
        if msoffcrypto is None:
            raise first_error

        with path.open("rb") as handle:
            office_file = msoffcrypto.OfficeFile(handle)
            if not office_file.is_encrypted():
                raise first_error

            if not password:
                raise ValueError(
                    f"{path.name} is encrypted; provide --password to decrypt."
                ) from first_error

            decrypted = io.BytesIO()
            office_file.load_key(password=password)
            office_file.decrypt(decrypted)
            decrypted.seek(0)
            warnings.append(f"Decrypted {path.name} with supplied password.")
            return openpyxl.load_workbook(decrypted, data_only=False, keep_links=True), warnings


def topological_sort(cells: dict[str, dict[str, Any]]) -> tuple[list[str], list[str]]:
    graph: dict[str, set[str]] = defaultdict(set)
    indegree: dict[str, int] = {key: 0 for key in cells}

    for key, info in cells.items():
        for dep in info["depends_on"]:
            if dep in cells and dep != key:
                graph[dep].add(key)
                indegree[key] += 1

    queue = deque(sorted(node for node, degree in indegree.items() if degree == 0))
    order: list[str] = []

    while queue:
        node = queue.popleft()
        order.append(node)
        for neighbor in sorted(graph[node]):
            indegree[neighbor] -= 1
            if indegree[neighbor] == 0:
                queue.append(neighbor)

    warnings: list[str] = []
    if len(order) != len(cells):
        remaining = sorted(set(cells) - set(order))
        warnings.append(
            "Cycle or unresolved dependency among formula cells: "
            + ", ".join(remaining[:20])
            + (" ..." if len(remaining) > 20 else "")
        )
    return order, warnings


def extract_workbook_topology(path: Path, password: str | None) -> dict[str, Any]:
    workbook, load_warnings = load_workbook_from_path(path, password)
    sheet_names = set(workbook.sheetnames)
    cells: dict[str, dict[str, Any]] = {}
    warnings = list(load_warnings)
    external_dependencies: list[str] = []

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        for row in worksheet.iter_rows():
            for cell in row:
                if not isinstance(cell, Cell) or not is_formula_cell(cell):
                    continue

                formula = normalize_formula_text(cell.value)
                if not formula:
                    warnings.append(
                        f"Skipped empty formula at {cell_key(sheet_name, cell.coordinate)}"
                    )
                    continue

                coordinate = cell.coordinate.upper()
                key = cell_key(sheet_name, coordinate)
                depends_on, depends_on_ranges, external_refs = parse_formula_references(
                    formula, sheet_name, sheet_names
                )

                for external in external_refs:
                    if external not in external_dependencies:
                        external_dependencies.append(external)

                if "#REF!" in formula.upper():
                    warnings.append(f"Broken reference (#REF!) in {key}: {formula}")

                formula_type = "normal"
                if isinstance(cell.value, ArrayFormula):
                    formula_type = "array"
                elif isinstance(cell.value, DataTableFormula):
                    formula_type = "data_table"

                cells[key] = {
                    "sheet": sheet_name,
                    "coordinate": coordinate,
                    "formula": formula,
                    "formula_type": formula_type,
                    "depends_on": depends_on,
                    "depends_on_ranges": depends_on_ranges,
                    "is_computed": True,
                }

    execution_order, sort_warnings = topological_sort(cells)
    warnings.extend(sort_warnings)

    return {
        "meta": {
            "source_file": path.name,
            "source_path": str(path.resolve()),
            "extracted_at": datetime.now(timezone.utc).isoformat(),
            "sheet_names": workbook.sheetnames,
            "formula_count": len(cells),
        },
        "cells": cells,
        "execution_order": execution_order,
        "external_dependencies": external_dependencies,
        "warnings": warnings,
    }


def iter_workbook_paths(directory: Path) -> list[Path]:
    return sorted(
        path
        for path in directory.iterdir()
        if path.is_file() and path.suffix.lower() == ".xlsx" and not path.name.startswith("~$")
    )


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Extract formula dependency topology from Excel workbooks."
    )
    parser.add_argument(
        "--dir",
        default=None,
        help="含 .xlsx 的目录（默认 data/raw/<month> 或 data/raw）",
    )
    parser.add_argument(
        "--month",
        default=None,
        help="月份子目录，如 2026-05（与 --dir 二选一，默认扫描 data/raw 下全部月份）",
    )
    parser.add_argument(
        "--output-dir",
        default=None,
        help="拓扑 JSON 输出目录（默认 data/topology/<month> 或与输入同月）",
    )
    parser.add_argument(
        "--password",
        "-p",
        default=None,
        help="Password for encrypted workbooks.",
    )
    args = parser.parse_args()

    project_root = Path(__file__).resolve().parents[1]
    raw_root = project_root / "data" / "raw"
    topo_root = project_root / "data" / "topology"

    if args.dir:
        input_dir = Path(args.dir).resolve()
        output_dir = (
            Path(args.output_dir).resolve()
            if args.output_dir
            else topo_root / input_dir.name
        )
        input_dirs = [input_dir]
    elif args.month:
        input_dir = raw_root / args.month
        output_dir = (
            Path(args.output_dir).resolve()
            if args.output_dir
            else topo_root / args.month
        )
        input_dirs = [input_dir]
    else:
        input_dirs = sorted(p for p in raw_root.iterdir() if p.is_dir())
        if not input_dirs:
            input_dirs = [raw_root]
        output_dir = topo_root

    failures = 0
    for input_dir in input_dirs:
        out = output_dir / input_dir.name if len(input_dirs) > 1 else output_dir
        out.mkdir(parents=True, exist_ok=True)
        workbook_paths = iter_workbook_paths(input_dir)
        if not workbook_paths:
            print(f"No .xlsx in {input_dir}", file=sys.stderr)
            continue
        for path in workbook_paths:
            out_path = out / f"{path.stem}.topology.json"
            try:
                topology = extract_workbook_topology(path, args.password)
                out_path.write_text(
                    json.dumps(topology, ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )
                print(
                    f"[ok] {path.name} -> {out_path.relative_to(project_root)} "
                    f"({topology['meta']['formula_count']} formulas)"
                )
                for warning in topology["warnings"]:
                    print(f"  warning: {warning}")
            except Exception as exc:
                failures += 1
                print(f"[fail] {path.name}: {exc}", file=sys.stderr)

    if failures:
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
