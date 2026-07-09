#!/usr/bin/env python3
"""Build 绩效整理表-精简 template: merged field row + source sub-headers (如图)."""

from __future__ import annotations

import re
import sys
from pathlib import Path

PROJECT = Path(__file__).resolve().parents[1]
if str(PROJECT) not in sys.path:
    sys.path.insert(0, str(PROJECT))

from copy import copy

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter

TARGET = PROJECT / "docs/templates/绩效整理表-精简/绩效整理表-系统生成-精简.xlsx"
UPLOADS = PROJECT / "data/raw/2026-05/uploads"
SALES_WORKBOOK = PROJECT / "data/raw/2026-05/燃油车-2026年05月西物超市销售提成(终)(1).xlsx"

TITLE_ROW = 1
LEGEND_ROW = 2
FIELD_TITLE_ROW = 3
SUBHEADER_ROW = 4
DATA_START_ROW = 5

SLIM_SOURCE: dict[str, str] = {
    "指标汇总部门": "比对表!A→B",
    "渠道": "派生: 指标汇总部门+销售渠道→渠道规则",
    "库存天数": "派生: 系统销售毛利!BA−整车成本!B",
    "订单号": "系统销售毛利!B",
    "车种": "系统销售毛利!D",
    "销售渠道": "系统销售毛利!E",
    "台数": "派生: 默认1; supplement_rows可空",
    "订单合计(含税)": "系统销售毛利!AO",
    "VIN码": "系统销售毛利!BD",
    "销售顾问": "系统销售毛利!BJ",
    "部门": "系统销售毛利!BL",
    "精品最低价金额": "系统销售毛利!BQ",
    "不提成精品": "装饰台账!AR(按订单号汇总)",
    "整车节约": "系统超额!P",
    "经理权限": "配置: manager_permission_by_vin",
    "整车实际节约": "派生: 整车节约+经理权限",
    "盈利性产品按揭权限扣除": "按揭原表!AN",
    "实际整车节约权限": "派生: 盈利性产品按揭权限扣除≠0→0 else 整车实际节约−盈利性产品按揭权限扣除",
    "保险返利收入": "保险明细!BP",
    "按揭收入": "按揭明细!Z",
    "爱车宝收入": "爱车保!K",
    "上户收入": "上户提成!C",
    "延保收入": "延保提成!BE",
    "单台绩效": "派生: 提成标准!F×台数",
    "整车超额": "派生: 实际整车节约权限×比例分支",
    "加装绩效": "派生: 台数=0→订单合计×12%; else (精品最低价金额−不提成精品)×12%",
    "保险提成": "保险明细!BS",
    "按揭提成": "按揭明细!BO",
    "盈利产品": "按揭原表!AF+按揭明细!BR",
    "爱车宝": "爱车保!BA",
    "上户绩效": "上户提成!H",
    "座位险绩效": "保险明细!BU",
    "玻璃险绩效": "保险明细!BV",
    "特殊车型追加绩效（追加+质损车）": "重功超期+活动!N+提成标准!H×台数",
    "二手车置换": "二手置换!(AC+AD)+大客户!R",
    "置换服务": "置换服务!BB",
    "延保提成": "派生: 延保收入分段(-200/0/200)",
    "超期追加": "派生: 超期政策×库存天数",
    "出厂指导价/扣除降价补差": "整车成本!R",
    "提车现返": "整车成本!S",
    "合同履约": "整车成本!T",
    "项目金额附加费": "整车成本!U",
    "广告返利": "整车成本!V",
    "提车奖励当月返": "整车成本!W",
    "当月返利": "终端明细表!P",
    "促销奖": "需手工填入",
    "代交车佣金": "系统二手车降价!BE",
    "裸车毛利": "派生: 订单合计(含税)−SUM(出厂指导价:代交车佣金)",
    "装饰成本": "装饰台账!AK",
    "装饰毛利": "派生: 精品最低价金额−装饰成本",
    "主营业务毛利": "派生: 裸车毛利+装饰毛利",
    "上户代收代付部分": "需手工填入",
    "安心保": "需手工填入",
    "爱车保": "需手工填入",
    "上户量": "需手工填入",
}

# 全字段子列（与 salary_pipeline 计算逻辑对齐；优先于自动推断）
SUBHEADER_OVERRIDES: dict[str, list[str]] = {
    "指标汇总部门": [
        "系统销售毛利-部门（BL列）",
        "比对表-系统部门（A列）→指标汇总部门（B列）",
        "通过Vin码匹配",
    ],
    "渠道": [
        "表内派生：指标汇总部门+销售渠道→渠道规则",
    ],
    "库存天数": [
        "系统销售毛利-结算日期（BA列）",
        "整车成本-结算日期（B列）",
        "通过Vin码匹配",
        "表内派生：INT(订单结算日−成本结算日)",
    ],
    "订单号": [
        "系统销售毛利-订单号（B列）",
        "通过Vin码匹配",
    ],
    "车种": [
        "系统销售毛利-车种（D列）",
        "通过Vin码匹配",
    ],
    "销售渠道": [
        "系统销售毛利-销售渠道（E列）",
        "通过Vin码匹配",
    ],
    "台数": [
        "表内派生：默认1; supplement_rows可空",
    ],
    "订单合计(含税)": [
        "系统销售毛利-订单合计(含税)（AO列）",
        "通过Vin码匹配",
    ],
    "VIN码": [
        "系统销售毛利-VIN码（BD列）",
        "通过Vin码匹配",
    ],
    "销售顾问": [
        "系统销售毛利-销售顾问（BJ列）",
        "通过Vin码匹配",
    ],
    "部门": [
        "系统销售毛利-部门（BL列）",
        "通过Vin码匹配",
    ],
    "精品最低价金额": [
        "系统销售毛利-精品最低价金额（BQ列）",
        "通过Vin码匹配",
    ],
    "不提成精品": [
        "装饰台账-最低价含税金额（AR列）",
        "按订单号汇总",
    ],
    "整车节约": [
        "系统超额-合计超卖金额（P列）",
        "通过Vin码匹配",
    ],
    "经理权限": [
        "配置：manager_permission_by_vin",
    ],
    "整车实际节约": [
        "整车节约",
        "经理权限",
        "整车节约+经理权限",
    ],
    "盈利性产品按揭权限扣除": [
        "按揭原表-整车超权限（AN列）",
        "通过Vin码匹配",
    ],
    "实际整车节约权限": [
        "表内派生：盈利性产品按揭权限扣除≠0→0 else 整车实际节约−盈利性产品按揭权限扣除",
    ],
    "保险返利收入": [
        "保险明细-应结算服务费-合计（AN列）",
        "保险明细-返利金额（BP列）",
        "通过Vin码匹配",
        "BP=SUMIF(VIN,AN)",
    ],
    "按揭收入": [
        "按揭明细-服务费合计（Z列）",
        "通过Vin码匹配",
    ],
    "爱车宝收入": [
        "爱车保-产品名称（K列）",
        "通过Vin码匹配",
    ],
    "上户收入": [
        "上户提成-手续费（C列）",
        "通过Vin码匹配",
    ],
    "延保收入": [
        "延保提成-延保收入（BE列）",
        "通过Vin码匹配",
    ],
    "单台绩效": [
        "提成标准-单台绩效（F列）",
        "台数",
        "提成标准F×台数",
    ],
    "整车超额": [
        "表内派生：实际整车节约权限×比例分支",
    ],
    "加装绩效": [
        "表内派生：台数=0→订单合计×12%",
        "表内派生：else (精品最低价金额−不提成精品)×12%",
    ],
    "保险提成": [
        "保险明细-购买险种-商业险保费（AE列）",
        "保险明细-直营店保险额外提成-保险额外提成（BT列）",
        "通过Vin码匹配",
        "AE*0.03+BT",
    ],
    "按揭提成": [
        "按揭明细-基础绩效（BO列）",
        "通过Vin码匹配",
    ],
    "盈利产品": [
        "按揭原表-盈利性奖励（AF列）",
        "按揭明细-吉致额外追加奖励（BR列）",
        "通过Vin码匹配",
        "AF+BR",
    ],
    "爱车宝": [
        "爱车保-VIN（F列）",
        "爱车保-产品名称（K列）",
        "爱车保-实际金额（Z列）",
        "通过Vin码匹配",
        "分段规则(熊猫/畅行保C/置换无忧/悦行保A/金额档)",
    ],
    "上户绩效": [
        "上户提成-手续费（C列）",
        "上户提成-福袋（D列）",
        "上户提成-VIN码（B列）",
        "上户提成-车型（M列）",
        "通过Vin码匹配",
        "IF(G=0,0,IF(G>=330,G*0.14+53.8,(G-330)*0.14))",
    ],
    "座位险绩效": [
        "保险明细-15/单-座位险提成（BU列）",
        "通过Vin码匹配",
        "IF(BF>0,15,0)",
    ],
    "玻璃险绩效": [
        "保险明细-玻璃险-BG（BG列）",
        "通过Vin码匹配",
        "BG*0.08",
    ],
    "特殊车型追加绩效（追加+质损车）": [
        "重功超期+活动-奖励明细（N列）",
        "提成标准-政策追加2月（H列）",
        "通过Vin码匹配",
        "N+提成标准H×台数",
    ],
    "二手车置换": [
        "二手置换-公司留存补贴（Y列）",
        "二手置换-结余提成（AC列）",
        "Y*0.14",
        "二手置换-毛利（O列）",
        "二手置换-置换提成（AD列）",
        "IF(O>=1000,O*0.14,IF(O>=500,100,0))",
        "大客户-销售提成（R列）",
        "通过Vin码匹配",
        "AC+AD+大客户R",
    ],
    "置换服务": [
        "置换服务-服务费金额（AH列）",
        "通过Vin码匹配",
        "AH*0.14",
    ],
    "延保提成": [
        "表内派生：延保收入分段(-200/0/200)",
    ],
    "超期追加": [
        "表内派生：超期政策×库存天数",
    ],
    "出厂指导价/扣除降价补差": [
        "整车成本-出厂指导价（R列）",
        "通过Vin码匹配",
    ],
    "提车现返": [
        "整车成本-提车现返（S列）",
        "通过Vin码匹配",
    ],
    "合同履约": [
        "整车成本-合同履约奖（T列）",
        "通过Vin码匹配",
    ],
    "项目金额附加费": [
        "整车成本-项目金额附加费（U列）",
        "通过Vin码匹配",
    ],
    "广告返利": [
        "整车成本-广告返利（V列）",
        "通过Vin码匹配",
    ],
    "提车奖励当月返": [
        "整车成本-提车奖励当月返（W列）",
        "通过Vin码匹配",
    ],
    "当月返利": [
        "终端明细表-金额（P列）",
        "通过Vin码匹配",
        "表内派生：台数≠0时取负SUMIFS",
    ],
    "促销奖": [
        "需手工填入",
    ],
    "代交车佣金": [
        "系统二手车降价-佣金（去税）（BE列）",
        "通过Vin码匹配",
    ],
    "裸车毛利": [
        "表内派生：订单合计(含税)−SUM(出厂指导价:代交车佣金)",
    ],
    "装饰成本": [
        "装饰台账-含税成本价（AK列）",
        "按订单号汇总",
    ],
    "装饰毛利": [
        "精品最低价金额",
        "装饰成本",
        "精品最低价金额−装饰成本",
    ],
    "主营业务毛利": [
        "裸车毛利",
        "装饰毛利",
        "裸车毛利+装饰毛利",
    ],
    "上户代收代付部分": [
        "需手工填入",
    ],
    "安心保": [
        "需手工填入",
    ],
    "爱车保": [
        "需手工填入",
    ],
    "上户量": [
        "需手工填入",
    ],
}

# Hub 输入 — 从 uploads 镜像源表版式（表头 + 公式 + 版式；数值格留空供填数）
HUB_SOURCE_SHEETS: list[tuple[str, str]] = [
    ("销售任务及完成率", "销售任务及完成率.xlsx"),
    ("新媒体", "新媒体.xlsx"),
    ("邀约专员提成", "邀约专员提成.xlsx"),
    ("客户部提成", "客户部提成.xlsx"),
    ("直营店经理提成 (财务)", "直营店经理提成 (财务).xlsx"),
    ("招聘", "招聘.xlsx"),
    ("综合表", "综合表.xlsx"),
    ("重功超期+活动", "重功超期+活动.xlsx"),
    ("保客考核明细", "保客考核明细.xlsx"),
    ("直营店交车", "直营店交车.xlsx"),
    ("翼真考核", "翼真考核.xlsx"),
    ("二手置换 ", "二手置换.xlsx"),  # 尾空格与源账套 sheet 名一致
    ("按揭绩效", "按揭绩效.xlsx"),
]

# 发薪边表 — 从主销售账套按 sheet 名镜像（基本工资 / 银河代发）
PAYOUT_SOURCE_SHEETS: tuple[str, ...] = (
    "西物基本",
    "直营店基本",
    "超市基本",
    "银河B直营店提成",
    "银河A提成- 渠道+直营店",
)

SHEET_HEADER_ROW: dict[str, int] = {
    "保险明细": 3,
    "重功超期+活动": 3,
    "系统销售毛利": 2,
    "装饰台账": 2,
    "上户提成": 3,
    "按揭原表": 2,
}
# 个别列的表头不在默认行（如按揭原表 AN 在第3行）
SHEET_COL_HEADER_ROW: dict[tuple[str, str], int] = {
    ("按揭原表", "AN"): 3,
}
SHEET_GROUP_ROW: dict[str, int] = {"保险明细": 2}
# 这些表第2行已是字段名，不再用第1行合并分组
SHEET_SKIP_GROUP: frozenset[str] = frozenset({
    "系统销售毛利",
    "整车成本",
    "装饰台账",
    "终端明细表",
    "比对表",
    "系统超额",
    "系统二手车降价",
    "上户提成",
    "爱车保",
    "延保提成",
    "按揭明细",
    "按揭原表",
    "二手置换",
    "大客户",
    "置换服务",
})
# 按 VIN/车架号匹配的源表（整理表 O 列 ↔ 下列键列）
VIN_MATCH_SHEETS: frozenset[str] = frozenset({
    "系统销售毛利",
    "系统超额",
    "按揭明细",
    "按揭原表",
    "上户提成",
    "延保提成",
    "爱车保",
    "保险明细",
})
SHEET_FILE: dict[str, str] = {
    "保险明细": "保险明细.xlsx",
    "按揭明细": "按揭明细.xlsx",
    "按揭原表": "按揭原表 09.49.39.xlsx",
    "系统销售毛利": "系统销售毛利.xlsx",
    "整车成本": "整车成本.xlsx",
    "装饰台账": "装饰台账.xlsx",
    "终端明细表": "终端明细表.xlsx",
    "比对表": "比对表.xlsx",
    "上户提成": "上户提成.xlsx",
    "爱车保": "爱车保.xlsx",
    "延保提成": "延保提成.xlsx",
    "二手置换": "二手置换.xlsx",
    "大客户": "大客户.xlsx",
    "置换服务": "置换服务.xlsx",
    "重功超期+活动": "重功超期+活动.xlsx",
    "提成标准": "提成标准.xlsx",
    "系统超额": "系统超额.xlsx",
    "系统二手车降价": "系统二手车降价.xlsx",
    "工厂购进": "工厂购进.xlsx",
}

_ARITH_OPS = re.compile(r"[\+\-\*/]")


class SheetMeta:
    def __init__(self) -> None:
        self._books: dict[str, object] = {}
        self._group: dict[tuple[str, str], str] = {}
        self._header: dict[tuple[str, str], str] = {}
        self._formula: dict[tuple[str, str], str | None] = {}

    def _book(self, sheet: str):
        if sheet not in self._books:
            fname = SHEET_FILE.get(sheet)
            if not fname or not (UPLOADS / fname).exists():
                return None
            self._books[sheet] = load_workbook(UPLOADS / fname, read_only=True, data_only=False)
        return self._books[sheet]

    def group(self, sheet: str, col: str) -> str:
        key = (sheet, col.upper())
        if key in self._group:
            return self._group[key]
        wb = self._book(sheet)
        if wb is None:
            return ""
        ws = wb.active
        gr = SHEET_GROUP_ROW.get(sheet, 1)
        hr = SHEET_HEADER_ROW.get(sheet, 1)
        idx = column_index_from_string(col)
        # forward-fill group labels on group row
        current = ""
        label = ""
        for c in range(1, min(ws.max_column, idx) + 1):
            v = ws.cell(gr, c).value
            if v and str(v).strip():
                current = str(v).strip()
            if c == idx:
                label = current
        self._group[key] = label
        return label

    def header(self, sheet: str, col: str) -> str:
        key = (sheet, col.upper())
        if key in self._header:
            return self._header[key]
        wb = self._book(sheet)
        if wb is None:
            return col
        ws = wb.active
        hr = SHEET_COL_HEADER_ROW.get(key, SHEET_HEADER_ROW.get(sheet, 1))
        text = ws.cell(hr, column_index_from_string(col)).value
        label = str(text).strip() if text else col
        self._header[key] = label
        return label

    def formula(self, sheet: str, col: str) -> str | None:
        key = (sheet, col.upper())
        if key in self._formula:
            return self._formula[key]
        wb = self._book(sheet)
        if wb is None:
            self._formula[key] = None
            return None
        ws = wb.active
        hr = SHEET_HEADER_ROW.get(sheet, 1)
        idx = column_index_from_string(col)
        found = None
        for r in range(hr + 1, min(hr + 10, ws.max_row + 1)):
            val = ws.cell(r, idx).value
            if isinstance(val, str) and val.startswith("="):
                found = val
                break
        self._formula[key] = found
        return found

    def format_col(self, sheet: str, col: str) -> str:
        field = self.header(sheet, col)
        if sheet in SHEET_SKIP_GROUP:
            return f"{sheet}-{field}（{col}列）"
        group = self.group(sheet, col)
        if group and group != field:
            return f"{sheet}-{group}-{field}（{col}列）"
        return f"{sheet}-{field}（{col}列）"

    def close(self) -> None:
        for wb in self._books.values():
            wb.close()


def _parse_sheet_col(text: str) -> tuple[str, str] | None:
    cleaned = text.replace(" ", "")
    m = re.search(r"([^!\s(]+)!([A-Z]{1,3})", cleaned)
    if not m or m.group(1).startswith("["):
        return None
    return m.group(1), m.group(2).upper()


def _is_sumif_only(formula: str) -> bool:
    up = formula.upper().replace(" ", "")
    return up.startswith("=SUMIF(") and _ARITH_OPS.search(up[7:]) is None


def _sumif_value_col(formula: str) -> str | None:
    m = re.search(
        r"SUMIF\s*\([^,]+,[^,]+,\s*([A-Z]{1,3})\s*(?::\1)?\s*\)",
        formula,
        re.I,
    )
    return m.group(1).upper() if m else None


def _formula_leaf_cols(sheet: str, col: str, meta: SheetMeta, seen: set[tuple[str, str]]) -> list[str]:
    key = (sheet, col.upper())
    if key in seen:
        return []
    seen.add(key)
    formula = meta.formula(sheet, col)
    if not formula:
        return [meta.format_col(sheet, col)]
    if _is_sumif_only(formula):
        tgt = _sumif_value_col(formula)
        if tgt:
            return _formula_leaf_cols(sheet, tgt, meta, seen)
        return [meta.format_col(sheet, col)]
    # arithmetic / IF — operands are leaves
    leaves: list[str] = []
    for ref in re.findall(r"\b([A-Z]{1,3})\d+", formula.upper()):
        if ref in {"IF", "AND", "OR", "NOT", "INDEX", "MATCH"}:
            continue
        leaf_key = (sheet, ref)
        if leaf_key in seen:
            continue
        sub = _formula_leaf_cols(sheet, ref, meta, seen)
        for s in sub:
            if s not in leaves:
                leaves.append(s)
    return leaves or [meta.format_col(sheet, col)]


def _compact_formula(formula: str) -> str:
    """AE4*0.03+BT4 → AE*0.03+BT"""
    expr = formula.lstrip("=").upper()
    expr = re.sub(r"([A-Z]{1,3})\d+", r"\1", expr)
    return expr


def build_subheaders(field: str, immediate: str, meta: SheetMeta) -> list[str]:
    if field in SUBHEADER_OVERRIDES:
        return list(SUBHEADER_OVERRIDES[field])

    if immediate.startswith("需手工填入"):
        return ["需手工填入"]
    if immediate.startswith("配置:"):
        return [immediate.replace("配置:", "配置：").strip()]
    if immediate.startswith("派生:"):
        return [f"表内派生：{immediate.removeprefix('派生:').strip()}"]

    if "→" in immediate and "!" not in immediate.split("→")[-1]:
        return [immediate]

    # compound external A+B
    if "+" in immediate and "!" in immediate:
        parts = [p.strip() for p in immediate.split("+")]
        out: list[str] = []
        for p in parts:
            ref = _parse_sheet_col(p)
            if ref:
                out.extend(_formula_leaf_cols(ref[0], ref[1], meta, set()))
        return _dedupe(out) or [immediate]

    ref = _parse_sheet_col(immediate)
    if not ref:
        return [immediate]

    sheet, col = ref
    formula = meta.formula(sheet, col)

    # 源表列含四则公式：叶子 + Vin匹配 + 公式
    if formula and not _is_sumif_only(formula) and _ARITH_OPS.search(formula):
        leaves = _formula_leaf_cols(sheet, col, meta, set())
        out = list(leaves)
        out.append("通过Vin码匹配")
        out.append(_compact_formula(formula))
        return out

    # 源表列仅 SUMIF 汇总：只列叶子字段
    if formula and _is_sumif_only(formula):
        leaves = _formula_leaf_cols(sheet, col, meta, set())
        return leaves or [meta.format_col(sheet, col)]

    # 直引值列（按 VIN/车架号匹配的源表）
    if sheet in VIN_MATCH_SHEETS:
        return _with_vin_match(sheet, [meta.format_col(sheet, col)])
    return [meta.format_col(sheet, col)]


def _with_vin_match(sheet: str, subs: list[str]) -> list[str]:
    if sheet in VIN_MATCH_SHEETS and not any("Vin码" in s or "订单号" in s for s in subs):
        subs = list(subs)
        subs.append("通过Vin码匹配")
    return subs


def _dedupe(items: list[str]) -> list[str]:
    out: list[str] = []
    for x in items:
        if x not in out:
            out.append(x)
    return out


def _lineage_styles() -> tuple[
    PatternFill,
    Font,
    tuple[PatternFill, PatternFill],
    tuple[PatternFill, PatternFill],
]:
    title_fill = PatternFill("solid", fgColor="4472C4")
    title_font = Font(color="FFFFFF", bold=True, size=12)
    field_fills = (
        PatternFill("solid", fgColor="D9E1F2"),
        PatternFill("solid", fgColor="E2EFDA"),
    )
    sub_fills = (
        PatternFill("solid", fgColor="EDF2FA"),
        PatternFill("solid", fgColor="F1FAF1"),
    )
    return title_fill, title_font, field_fills, sub_fills


def _write_lineage_sheet(
    ws,
    *,
    title: str,
    legend: str,
    fields: dict[str, list[str]],
    title_fill: PatternFill,
    title_font: Font,
    field_fills: tuple[PatternFill, PatternFill],
    sub_fills: tuple[PatternFill, PatternFill],
) -> int:
    """Write rows 1–4 lineage layout; return total physical column count."""
    col = 1
    field_names = list(fields.keys())
    for idx, field in enumerate(field_names):
        subs = fields[field]
        width = max(len(subs), 1)
        start, end = col, col + width - 1
        field_fill = field_fills[idx % 2]
        sub_fill = sub_fills[idx % 2]

        ws.merge_cells(
            start_row=FIELD_TITLE_ROW,
            start_column=start,
            end_row=FIELD_TITLE_ROW,
            end_column=end,
        )
        for c in range(start, end + 1):
            top = ws.cell(FIELD_TITLE_ROW, c)
            if c == start:
                top.value = field
            top.fill = field_fill
            top.font = Font(bold=True, size=11)
            top.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for i in range(width):
            c = start + i
            text = subs[i] if i < len(subs) else ""
            sub = ws.cell(SUBHEADER_ROW, c, text)
            sub.fill = sub_fill
            sub.font = Font(size=10)
            sub.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(c)].width = max(16, min(len(text) + 2, 36))

        col = end + 1

    total = col - 1
    ws.merge_cells(start_row=TITLE_ROW, start_column=1, end_row=TITLE_ROW, end_column=total)
    t = ws.cell(TITLE_ROW, 1, title)
    t.fill = title_fill
    t.font = title_font
    t.alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=LEGEND_ROW, start_column=1, end_row=LEGEND_ROW, end_column=total)
    ws.cell(LEGEND_ROW, 1, legend).alignment = Alignment(wrap_text=True)

    ws.freeze_panes = ws.cell(DATA_START_ROW, 1)
    ws.row_dimensions[FIELD_TITLE_ROW].height = 28
    ws.row_dimensions[SUBHEADER_ROW].height = 48
    return total


def _copy_worksheet(src_ws, dst_ws) -> tuple[int, int]:
    """Mirror source sheet layout into target (values, styles, merges)."""
    max_row = src_ws.max_row or 1
    max_col = src_ws.max_column or 1
    for row in src_ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            dst_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                dst_cell.font = copy(cell.font)
                dst_cell.border = copy(cell.border)
                dst_cell.fill = copy(cell.fill)
                dst_cell.number_format = copy(cell.number_format)
                dst_cell.protection = copy(cell.protection)
                dst_cell.alignment = copy(cell.alignment)
    for merged in list(src_ws.merged_cells.ranges):
        dst_ws.merge_cells(str(merged))
    for col, dim in src_ws.column_dimensions.items():
        if dim.width is not None:
            dst_ws.column_dimensions[col].width = dim.width
    for row, dim in src_ws.row_dimensions.items():
        if dim.height is not None:
            dst_ws.row_dimensions[row].height = dim.height
    if src_ws.freeze_panes:
        dst_ws.freeze_panes = src_ws.freeze_panes
    return max_row, max_col


def _clear_literal_data(ws, *, max_row: int, max_col: int) -> int:
    """Clear numeric business values; keep formulas and text labels."""
    cleared = 0
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            value = cell.value
            if value is None:
                continue
            if isinstance(value, str):
                continue
            cell.value = None
            cleared += 1
    return cleared


def _append_source_sheet(
    target_wb: Workbook,
    sheet_name: str,
    *,
    uploads_dir: Path | None = None,
    upload_file: str | None = None,
    workbook_path: Path | None = None,
    source_sheet: str | None = None,
) -> tuple[int, int, int]:
    """Copy source sheet into workbook; return (rows, cols, cleared_cells)."""
    if workbook_path is not None:
        src_path = workbook_path
        src_sheet_name = source_sheet or sheet_name
    else:
        if uploads_dir is None or upload_file is None:
            raise ValueError("uploads_dir and upload_file required when workbook_path omitted")
        src_path = uploads_dir / upload_file
        src_sheet_name = None

    if not src_path.exists():
        raise FileNotFoundError(src_path)

    src_wb = load_workbook(src_path, data_only=False)
    try:
        if src_sheet_name is not None:
            if src_sheet_name not in src_wb.sheetnames:
                raise KeyError(f"{src_sheet_name!r} not in {src_path.name}")
            src_ws = src_wb[src_sheet_name]
        else:
            src_ws = src_wb.active
        dst_ws = target_wb.create_sheet(sheet_name)
        rows, cols = _copy_worksheet(src_ws, dst_ws)
        cleared = _clear_literal_data(dst_ws, max_row=rows, max_col=cols)
        return rows, cols, cleared
    finally:
        src_wb.close()


def rebuild(path: Path) -> None:
    meta = SheetMeta()
    fields = list(SLIM_SOURCE.keys())
    groups = {f: build_subheaders(f, SLIM_SOURCE[f], meta) for f in fields}
    meta.close()

    title_fill, title_font, field_fills, sub_fills = _lineage_styles()
    perf_legend = (
        "第3行=整理表字段（合并）；第4行=源表叶子字段/匹配方式/公式；第5行起填数据。"
        f" 同工作簿另含 {len(HUB_SOURCE_SHEETS) + len(PAYOUT_SOURCE_SHEETS)} 张源表"
        "（Hub/岗位 + 发薪边表；版式来自 uploads/主账套，公式保留、数值留空）。"
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "绩效整理表"
    perf_cols = _write_lineage_sheet(
        ws,
        title="2026-05 销售绩效溯源工作簿（精简整理表 + Hub输入）",
        legend=perf_legend,
        fields=groups,
        title_fill=title_fill,
        title_font=title_font,
        field_fills=field_fills,
        sub_fills=sub_fills,
    )

    source_sheet_count = 0
    for sheet_name, upload_file in HUB_SOURCE_SHEETS:
        try:
            rows, cols, cleared = _append_source_sheet(
                wb,
                sheet_name,
                uploads_dir=UPLOADS,
                upload_file=upload_file,
            )
        except FileNotFoundError:
            print(f"  ! skip {sheet_name!r}: missing {upload_file}")
            continue
        source_sheet_count += 1
        print(
            f"  + sheet {sheet_name!r}: source={upload_file} "
            f"size={rows}x{cols} cleared_literals={cleared}"
        )

    if SALES_WORKBOOK.exists():
        for payout_sheet in PAYOUT_SOURCE_SHEETS:
            try:
                rows, cols, cleared = _append_source_sheet(
                    wb,
                    payout_sheet,
                    workbook_path=SALES_WORKBOOK,
                    source_sheet=payout_sheet,
                )
            except KeyError as exc:
                print(f"  ! skip {payout_sheet!r}: {exc}")
                continue
            source_sheet_count += 1
            print(
                f"  + sheet {payout_sheet!r}: source={SALES_WORKBOOK.name} "
                f"size={rows}x{cols} cleared_literals={cleared}"
            )
    else:
        print(f"  ! skip payout sheets: missing {SALES_WORKBOOK}")

    wb.save(path)
    print(
        f"wrote {path} perf_fields={len(fields)} perf_cols={perf_cols} "
        f"source_sheets={source_sheet_count} total_sheets={len(wb.sheetnames)}"
    )
    _print_audit(groups)


def _print_audit(groups: dict[str, list[str]]) -> None:
    """Print column-by-column audit for review."""
    col = 1
    print("\n=== 绩效整理表精简版 逐列审阅 ===")
    for field, subs in groups.items():
        width = max(len(subs), 1)
        end = col + width - 1
        span = f"{col}" if width == 1 else f"{col}-{end}"
        print(f"\n[{span}] {field} ({width}子列)")
        for i, s in enumerate(subs, 1):
            print(f"  {i}. {s}")
        col = end + 1
    print(f"\n合计: {len(groups)} 字段, {col - 1} 物理列")


if __name__ == "__main__":
    rebuild(TARGET)
