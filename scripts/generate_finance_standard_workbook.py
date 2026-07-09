#!/usr/bin/env python3
"""Generate finance standard workbook template (销售账套-标准-YYYY-MM-模板.xlsx).

Workbook sections (see canonical_param_sheets.yaml):
  - README (填账规则)
  - 汇总数据表 (VIN grain, canonical_column_registry.yaml)
  - 人员信息 (skeleton row keys)
  - tier_a_params / payout_side_tables / role_family_optional sheets
  - 例外登记表 (governance template)

Header structure may be copied from monthly uploads or golden reference workbook
(structure only — never bootstrap metric values from golden parity workbook).
"""

from __future__ import annotations

import argparse
import logging
from pathlib import Path
from typing import Any

import pandas as pd
import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

PROJECT = Path(__file__).resolve().parents[1]
CONFIG_DIR = PROJECT / "salary_pipeline" / "config"
DEFAULT_OUT = PROJECT / "docs/templates/销售账套-标准-2026-05-模板.xlsx"

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
GROUP_FILL = PatternFill("solid", fgColor="D9E1F2")
HEADER_FONT = Font(bold=True, color="FFFFFF")
GROUP_FONT = Font(bold=True)
README_FONT = Font(bold=True, size=12)

SHEET_CATEGORIES: tuple[str, ...] = (
    "skeleton",
    "tier_a_params",
    "payout_side_tables",
    "role_family_optional",
)


def _load_yaml(path: Path) -> dict[str, Any]:
    with path.open(encoding="utf-8") as handle:
        return yaml.safe_load(handle) or {}


def _style_header_row(ws, row: int, *, fill: PatternFill = HEADER_FILL) -> None:
    for cell in ws[row]:
        if cell.value:
            cell.fill = fill
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_width(ws, min_width: int = 10, max_width: int = 48) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[letter].width = min(max(max_len + 2, min_width), max_width)


def _find_golden_workbook(month: str) -> Path | None:
    """Pick the multi-sheet reference workbook under data/raw/{month}/."""
    month_dir = PROJECT / "data" / "raw" / month
    if not month_dir.is_dir():
        return None
    candidates: list[tuple[int, Path]] = []
    for path in month_dir.glob("*.xlsx"):
        if "合并" in path.name or "标准" in path.name:
            continue
        try:
            xl = pd.ExcelFile(path, engine="openpyxl")
            candidates.append((len(xl.sheet_names), path))
        except Exception:
            continue
    if not candidates:
        return None
    candidates.sort(key=lambda item: item[0], reverse=True)
    return candidates[0][1]


def _workbook_has_sheet(workbook_path: Path, sheet_name: str) -> bool:
    try:
        xl = pd.ExcelFile(workbook_path, engine="openpyxl")
    except Exception:
        return False
    normalized = sheet_name.strip()
    return any(s.strip() == normalized for s in xl.sheet_names)


def _resolve_inner_sheet_name(workbook_path: Path, sheet_name: str) -> str | None:
    try:
        xl = pd.ExcelFile(workbook_path, engine="openpyxl")
    except Exception:
        return None
    normalized = sheet_name.strip()
    for candidate in xl.sheet_names:
        if candidate.strip() == normalized:
            return candidate
    return xl.sheet_names[0] if len(xl.sheet_names) == 1 else None


def _upload_search_dirs(uploads_dir: Path | None) -> list[Path]:
    if uploads_dir is None:
        return []
    dirs = [uploads_dir]
    staging = uploads_dir.parent / ".staging" / "uploads"
    if staging.is_dir() and staging not in dirs:
        dirs.append(staging)
    return dirs


def _resolve_sheet_source(
    sheet_name: str,
    *,
    uploads_dir: Path | None,
    golden_path: Path | None,
    alias: str | None = None,
) -> tuple[Path | None, str | None]:
    """Return (workbook_path, inner_sheet_name) for header/template copy."""
    for search_dir in _upload_search_dirs(uploads_dir):
        for candidate in (sheet_name, alias):
            if not candidate:
                continue
            upload_path = search_dir / f"{candidate.strip()}.xlsx"
            if upload_path.exists():
                inner = _resolve_inner_sheet_name(upload_path, sheet_name)
                return upload_path, inner

    if golden_path is not None and _workbook_has_sheet(golden_path, sheet_name):
        inner = _resolve_inner_sheet_name(golden_path, sheet_name)
        return golden_path, inner
    return None, None


def _read_upload_headers(
    source_path: Path,
    *,
    sheet_name: str | None,
    header_row: int = 1,
) -> list[str]:
    """Read header row text only (no golden bootstrap of values)."""
    try:
        kwargs: dict[str, Any] = {"header": None, "engine": "openpyxl"}
        if sheet_name:
            kwargs["sheet_name"] = sheet_name
        raw = pd.read_excel(source_path, nrows=header_row + 2, **kwargs)
        idx = max(header_row - 1, 0)
        if raw.shape[0] <= idx:
            return []
        return [str(v).strip() if pd.notna(v) else "" for v in raw.iloc[idx].tolist()]
    except Exception as exc:
        logger.warning("Cannot read headers from %s: %s", source_path, exc)
        return []


def _read_header_template_rows(
    source_path: Path,
    *,
    sheet_name: str | None,
    num_rows: int,
) -> list[list[Any]]:
    try:
        kwargs: dict[str, Any] = {"header": None, "engine": "openpyxl", "nrows": num_rows}
        if sheet_name:
            kwargs["sheet_name"] = sheet_name
        raw = pd.read_excel(source_path, **kwargs)
        return [list(row) for _, row in raw.iterrows()]
    except Exception as exc:
        logger.warning("Cannot read header template from %s: %s", source_path, exc)
        return []


def _read_upload_sample_rows(
    source_path: Path,
    *,
    sheet_name: str | None,
    header_row: int = 1,
    max_rows: int = 5,
) -> pd.DataFrame:
    """Read a few data rows from upload for structure demo (not golden workbook)."""
    try:
        kwargs: dict[str, Any] = {
            "header": header_row - 1,
            "engine": "openpyxl",
        }
        if sheet_name:
            kwargs["sheet_name"] = sheet_name
        df = pd.read_excel(source_path, **kwargs)
        return df.head(max_rows)
    except Exception as exc:
        logger.warning("Cannot read sample rows from %s: %s", source_path, exc)
        return pd.DataFrame()


def _sheet_specs(param_cfg: dict[str, Any]) -> list[tuple[str, dict[str, Any]]]:
    specs: list[tuple[str, dict[str, Any]]] = []
    for category in SHEET_CATEGORIES:
        block = param_cfg.get(category, {})
        for spec in block.get("sheets", []):
            specs.append((category, spec))
    return specs


def _build_readme_lines(param_cfg: dict[str, Any], col_cfg: dict[str, Any]) -> list[str]:
    col_count = len(col_cfg.get("columns", []))
    skeleton_names = [s["name"] for s in param_cfg.get("skeleton", {}).get("sheets", [])]
    tier_a_names = [s["name"] for s in param_cfg.get("tier_a_params", {}).get("sheets", [])]
    payout_names = [s["name"] for s in param_cfg.get("payout_side_tables", {}).get("sheets", [])]
    role_names = [s["name"] for s in param_cfg.get("role_family_optional", {}).get("sheets", [])]
    exc = param_cfg.get("exception_sheet", {})
    payout_note = param_cfg.get("payout_side_tables", {}).get("note", "")

    lines = [
        "销售账套标准模板 — 填账规则",
        "",
        "一、工作簿结构",
        f"  1. 汇总数据表 — VIN 粒度主数据（{col_count} 列），一行一车/一单",
        f"  2. 人员信息 — 提成汇总行键（店别/职务/姓名）",
        f"  3. Tier A 参数表 — Hub/闭包必需 {len(tier_a_names)} 张",
        f"  4. 发薪边表 — HR/财务维护 {len(payout_names)} 张（不参与提成汇总生成）",
        f"  5. 岗位族可选表 — {len(role_names)} 张（有对应人员时再填）",
        f"  6. {exc.get('name', '例外登记表')} — 人工确认差异登记",
        "",
        "二、列类型说明",
        "  · 系统列（汇总数据表）：财务按 VIN 填写的明细，系统直接读取参与绩效/毛利计算",
        "  · 参数列（Tier A）：按姓名/部门/车型等维度的标准、任务、闭包输入",
        "  · 发薪边表：基本工资、社保公积金、银河代发放、扣款等，供发薪表 SUMIF",
        "  · 例外列（例外登记表）：系统值与确认值不一致时登记，不覆盖系统计算",
        "",
        "三、汇总数据表填账要点",
        "  · 键列 VIN码 必填且唯一（同月内）",
        "  · 台数：整车订单填 1；纯服务补录可留空",
        "  · 装饰成本按订单号关联，须与订单号列一致",
        "  · 分组行（第 1 行）仅作阅读辅助，勿删改结构",
        "",
        "四、人员信息（骨架）",
        *[f"  · {name}" for name in skeleton_names],
        "  · 合并账套无「提成汇总」sheet 时，系统从此表读取行键",
        "",
        "五、Tier A 必需参数表",
        *[f"  · {name}" for name in tier_a_names],
        "",
        "六、发薪边表（HR/财务）",
    ]
    if payout_note:
        lines.append(f"  · {payout_note}")
    lines.extend(f"  · {name}" for name in payout_names)
    lines.extend(
        [
            "",
            "七、岗位族可选参数表",
            *[f"  · {name}" for name in role_names],
            "",
            "八、禁止事项",
            "  · 不得将金标准/历史提成结果当作「正确答案」抄入本模板",
            "  · 系统算不出或规则未覆盖的格：留空 + 例外登记表说明",
            "",
            "九、重新生成模板",
            "  python scripts/generate_finance_standard_workbook.py --month 2026-05",
            "  python scripts/generate_finance_standard_workbook.py --month 2026-05 --populate-from-uploads",
        ]
    )
    return lines


def _write_readme_sheet(wb: Workbook, lines: list[str]) -> None:
    ws = wb.create_sheet("README", 0)
    for row_idx, line in enumerate(lines, start=1):
        cell = ws.cell(row=row_idx, column=1, value=line)
        if row_idx == 1:
            cell.font = README_FONT
    ws.column_dimensions["A"].width = 100


def _write_summary_sheet(
    wb: Workbook,
    col_cfg: dict[str, Any],
    *,
    uploads_dir: Path | None,
    populate: bool,
    sample_rows: int,
) -> int:
    sheet_name = col_cfg.get("sheet", "汇总数据表")
    columns: list[dict[str, Any]] = col_cfg.get("columns", [])
    groups: dict[str, str] = {g["id"]: g["label"] for g in col_cfg.get("groups", [])}

    ws = wb.create_sheet(sheet_name)
    labels = [c["label"] for c in columns]

    for col_idx, col in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=groups.get(col.get("group", ""), ""))
    _style_header_row(ws, 1, fill=GROUP_FILL)
    for cell in ws[1]:
        cell.font = GROUP_FONT

    for col_idx, label in enumerate(labels, start=1):
        ws.cell(row=2, column=col_idx, value=label)
    _style_header_row(ws, 2)

    if populate and uploads_dir is not None:
        _populate_summary_from_uploads(ws, columns, uploads_dir, sample_rows)

    ws.freeze_panes = "A3"
    _auto_width(ws)
    return len(columns)


def _col_letter_index(letter: str) -> int:
    from openpyxl.utils import column_index_from_string

    return column_index_from_string(letter.upper()) - 1


def _populate_summary_from_uploads(
    ws,
    columns: list[dict[str, Any]],
    uploads_dir: Path,
    sample_rows: int,
) -> None:
    """Join upload sheets on VIN for demo rows — structure only."""
    vin_col_def = next((c for c in columns if c.get("id") == "vin"), None)
    if vin_col_def is None:
        return

    sg_path = uploads_dir / "系统销售毛利.xlsx"
    if not sg_path.exists():
        return

    try:
        sg_raw = pd.read_excel(sg_path, header=None, engine="openpyxl")
        sg_data = sg_raw.iloc[2:].copy()
        vin_idx = _col_letter_index(vin_col_def["source_col"])
        sg_data = sg_data[sg_data.iloc[:, vin_idx].notna()].head(sample_rows)
        if sg_data.empty:
            return
    except Exception as exc:
        logger.warning("Summary populate skipped: %s", exc)
        return

    source_cache: dict[str, pd.DataFrame] = {"系统销售毛利": sg_raw}

    def load_source(sheet: str) -> pd.DataFrame:
        if sheet in source_cache:
            return source_cache[sheet]
        path = uploads_dir / f"{sheet}.xlsx"
        if not path.exists():
            return pd.DataFrame()
        raw = pd.read_excel(path, header=None, engine="openpyxl")
        source_cache[sheet] = raw
        return raw

    for row_offset, (_, sg_row) in enumerate(sg_data.iterrows()):
        excel_row = 3 + row_offset
        vin = str(sg_row.iloc[vin_idx]).strip()
        order_no = ""
        order_def = next((c for c in columns if c.get("id") == "order_no"), None)
        if order_def:
            order_no = str(sg_row.iloc[_col_letter_index(order_def["source_col"])]).strip()

        for col_idx, col_def in enumerate(columns, start=1):
            cid = col_def.get("id")
            if cid == "units":
                ws.cell(row=excel_row, column=col_idx, value=1)
                continue

            source_sheet = col_def.get("source_sheet")
            source_col = col_def.get("source_col")
            if not source_sheet or not source_col:
                continue

            raw = load_source(source_sheet)
            if raw.empty:
                continue

            join = col_def.get("join", "vin")
            key_col = col_def.get("source_key_col") or (
                source_col if join == "vin" and source_sheet != "系统销售毛利" else None
            )

            value = None
            if source_sheet == "系统销售毛利":
                value = sg_row.iloc[_col_letter_index(source_col)]
            elif join == "order_no" and order_no:
                key_idx = _col_letter_index(key_col or source_col)
                val_idx = _col_letter_index(source_col)
                data = raw.iloc[1:] if source_sheet == "装饰台账" else raw.iloc[0:]
                matches = data[data.iloc[:, key_idx].astype(str).str.strip() == order_no]
                if not matches.empty:
                    value = matches.iloc[0, val_idx]
            elif key_col and vin:
                key_idx = _col_letter_index(key_col)
                val_idx = _col_letter_index(source_col)
                skip = 2 if source_sheet in {"系统销售毛利", "系统二手车降价"} else 1
                data = raw.iloc[skip:]
                matches = data[data.iloc[:, key_idx].astype(str).str.strip() == vin]
                if not matches.empty:
                    value = matches.iloc[0, val_idx]

            if value is not None and str(value).strip() not in {"", "nan"}:
                ws.cell(row=excel_row, column=col_idx, value=value)


def _write_param_spec_headers(ws, spec: dict[str, Any]) -> int:
    cols: list[str] = []
    cols.extend(spec.get("key_labels") or spec.get("key_cols") or [])
    cols.extend(spec.get("value_labels") or spec.get("value_cols") or [])
    if not cols:
        cols = ["（请按原表结构填写）"]
    for col_idx, label in enumerate(cols, start=1):
        ws.cell(row=1, column=col_idx, value=label)
    _style_header_row(ws, 1)
    return 1


def _write_header_block(ws, rows: list[list[Any]], *, style_last: bool = True) -> int:
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            if pd.notna(value):
                ws.cell(row=row_idx, column=col_idx, value=value)
    if style_last and rows:
        _style_header_row(ws, len(rows))
    return len(rows)


def _write_param_sheet(
    wb: Workbook,
    spec: dict[str, Any],
    *,
    uploads_dir: Path | None,
    golden_path: Path | None,
    populate: bool,
    sample_rows: int,
) -> None:
    name = spec["name"]
    header_row = int(spec.get("header_row", 1))
    template_rows = int(spec.get("header_template_rows", 0))
    ws = wb.create_sheet(name)

    source_path, inner_sheet = _resolve_sheet_source(
        name,
        uploads_dir=uploads_dir,
        golden_path=golden_path,
        alias=spec.get("sheet_alias"),
    )

    data_start_row = 2
    if source_path is not None:
        if template_rows > 0:
            block = _read_header_template_rows(
                source_path,
                sheet_name=inner_sheet,
                num_rows=template_rows,
            )
            if block:
                written = _write_header_block(ws, block, style_last=True)
                data_start_row = written + 1
            else:
                data_start_row = _write_param_spec_headers(ws, spec) + 1
        else:
            headers = _read_upload_headers(
                source_path,
                sheet_name=inner_sheet,
                header_row=header_row,
            )
            if headers and any(headers):
                for col_idx, header in enumerate(headers, start=1):
                    if header:
                        ws.cell(row=1, column=col_idx, value=header)
                _style_header_row(ws, 1)
                data_start_row = 2
            else:
                data_start_row = _write_param_spec_headers(ws, spec) + 1

        if populate:
            sample = _read_upload_sample_rows(
                source_path,
                sheet_name=inner_sheet,
                header_row=header_row if template_rows == 0 else template_rows,
                max_rows=sample_rows,
            )
            for r_idx, row in enumerate(dataframe_to_rows(sample, index=False, header=False), start=data_start_row):
                for c_idx, value in enumerate(row, start=1):
                    if pd.notna(value):
                        ws.cell(row=r_idx, column=c_idx, value=value)
    else:
        if template_rows > 0:
            data_start_row = _write_param_spec_headers(ws, spec) + 1
        else:
            data_start_row = _write_param_spec_headers(ws, spec) + 1

    freeze_row = max(data_start_row, 2)
    ws.freeze_panes = f"A{freeze_row}"
    _auto_width(ws)


def _write_exception_sheet(wb: Workbook, param_cfg: dict[str, Any]) -> None:
    exc = param_cfg.get("exception_sheet", {})
    name = exc.get("name", "例外登记表")
    columns = exc.get("columns", [])
    ws = wb.create_sheet(name)
    for col_idx, col in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=col.get("label", col.get("id", "")))
    _style_header_row(ws, 1)
    ws.freeze_panes = "A2"
    _auto_width(ws)


def generate_workbook(
    *,
    month: str,
    out_path: Path,
    populate: bool = False,
    uploads_dir: Path | None = None,
    sample_rows: int = 5,
) -> dict[str, Any]:
    col_cfg = _load_yaml(CONFIG_DIR / "canonical_column_registry.yaml")
    param_cfg = _load_yaml(CONFIG_DIR / "canonical_param_sheets.yaml")

    if uploads_dir is None:
        uploads_dir = PROJECT / "data" / "raw" / month / "uploads"

    golden_path = _find_golden_workbook(month)

    wb = Workbook()
    wb.remove(wb.active)

    readme_lines = _build_readme_lines(param_cfg, col_cfg)
    _write_readme_sheet(wb, readme_lines)

    summary_col_count = _write_summary_sheet(
        wb,
        col_cfg,
        uploads_dir=uploads_dir if populate else None,
        populate=populate,
        sample_rows=sample_rows,
    )

    category_sheet_names: dict[str, list[str]] = {cat: [] for cat in SHEET_CATEGORIES}
    for category, spec in _sheet_specs(param_cfg):
        _write_param_sheet(
            wb,
            spec,
            uploads_dir=uploads_dir,
            golden_path=golden_path,
            populate=populate,
            sample_rows=sample_rows,
        )
        category_sheet_names[category].append(spec["name"])

    _write_exception_sheet(wb, param_cfg)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    all_param_sheets = [name for names in category_sheet_names.values() for name in names]
    return {
        "out_path": str(out_path),
        "month": month,
        "sheet_count": len(wb.sheetnames),
        "sheet_names": list(wb.sheetnames),
        "summary_column_count": summary_col_count,
        "skeleton_sheets": category_sheet_names["skeleton"],
        "tier_a_param_sheets": category_sheet_names["tier_a_params"],
        "payout_side_tables": category_sheet_names["payout_side_tables"],
        "role_family_optional_sheets": category_sheet_names["role_family_optional"],
        "all_param_sheets": all_param_sheets,
        "populated_from_uploads": populate,
        "golden_reference": str(golden_path) if golden_path else None,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate 销售账套-标准 finance template workbook")
    parser.add_argument("--month", default="2026-05", help="Example month label (default: 2026-05)")
    parser.add_argument(
        "--out",
        type=Path,
        default=DEFAULT_OUT,
        help="Output .xlsx path",
    )
    parser.add_argument(
        "--populate-from-uploads",
        action="store_true",
        help="Copy header structure + a few sample rows from data/raw/{month}/uploads",
    )
    parser.add_argument(
        "--uploads-dir",
        type=Path,
        default=None,
        help="Override uploads directory (default: data/raw/{month}/uploads)",
    )
    parser.add_argument(
        "--sample-rows",
        type=int,
        default=5,
        help="Max demo rows per sheet when --populate-from-uploads (default: 5)",
    )
    parser.add_argument("-v", "--verbose", action="store_true")
    args = parser.parse_args()

    logging.basicConfig(level=logging.DEBUG if args.verbose else logging.INFO)

    if args.out == DEFAULT_OUT and args.month != "2026-05":
        out = PROJECT / f"docs/templates/销售账套-标准-{args.month}-模板.xlsx"
    else:
        out = args.out

    result = generate_workbook(
        month=args.month,
        out_path=out,
        populate=args.populate_from_uploads,
        uploads_dir=args.uploads_dir,
        sample_rows=args.sample_rows,
    )

    print(f"Wrote: {result['out_path']}")
    print(f"Total sheets: {result['sheet_count']}")
    print(f"汇总数据表 columns: {result['summary_column_count']}")
    if result.get("golden_reference"):
        print(f"Golden header reference: {result['golden_reference']}")
    print("Skeleton:")
    for name in result["skeleton_sheets"]:
        print(f"  - {name}")
    print("Tier A param sheets:")
    for name in result["tier_a_param_sheets"]:
        print(f"  - {name}")
    print("Payout side tables:")
    for name in result["payout_side_tables"]:
        print(f"  - {name}")
    print("Role-family optional:")
    for name in result["role_family_optional_sheets"]:
        print(f"  - {name}")


if __name__ == "__main__":
    main()
